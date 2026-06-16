from __future__ import annotations

import json
import logging
import os
import base64
import hashlib
import secrets
import uuid
import unicodedata
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union, cast

import bcrypt
import httpx
from fastapi import (
    APIRouter,
    BackgroundTasks,
    FastAPI,
    File,
    Form,
    HTTPException,
    Request,
    Response,
    UploadFile,
)
import re
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from bson import ObjectId
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, ConfigDict, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from openpyxl import Workbook, load_workbook

from backend.services.audit import AuditService
from backend.services.cash import CashService
from backend.services.pin_policy import PinPolicyService

logger = logging.getLogger("erp")
logging.basicConfig(level=logging.INFO)

from fastapi.middleware.cors import CORSMiddleware

app = FastAPI()

# CORS para desarrollo local
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)
api_router = APIRouter(prefix="/api")

# Basic API root for health/checks
@api_router.get("/")
async def api_root():
    return JSONResponse({"message": "MUNDO DE ACCESORIOS ERP API", "version": os.environ.get("APP_VERSION", "dev")})


# Drafts backup endpoints - simple storage in `drafts_backup` collection
@api_router.get("/drafts/backup")
async def get_drafts_backup():
    doc = await db.drafts_backup.find_one({"_id": "backup"})
    entries = doc.get("entries") if doc else []
    return JSONResponse({"entries": entries})


@api_router.post("/drafts/backup")
async def post_drafts_backup(request: Request):
    payload = await request.json()
    entries = payload.get("entries") if isinstance(payload, dict) else None
    if entries is None:
        return JSONResponse({"detail": "No entries provided"}, status_code=400)
    await db.drafts_backup.update_one({"_id": "backup"}, {"$set": {"entries": entries}}, upsert=True)
    return JSONResponse({"status": "ok"})


@api_router.delete("/drafts/backup")
async def delete_drafts_backup():
    await db.drafts_backup.delete_one({"_id": "backup"})
    return JSONResponse({"status": "deleted"})

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
MONGO_DB = os.environ.get("MONGO_DB", os.environ.get("DB_NAME", "mc-larens2_erp"))
client = AsyncIOMotorClient(MONGO_URL)
db = client[MONGO_DB]

STRIPE_API_KEY = os.environ.get("STRIPE_API_KEY", "")
ROOT_DIR = Path(__file__).resolve().parent
CORE_SEED_FILE = ROOT_DIR / "data" / "seeds" / "core_seed.json"
BACKUP_SCHEMA_VERSION = 2

HYPERVISOR_FULL_ROLES = {"gerencia", "programador"}
HYPERVISOR_READONLY_ROLES = {"recursos_humanos"}


def get_hypervisor_access_mode(role: Optional[str]) -> Optional[str]:
    role_value = str(role or "").strip().lower()
    if role_value in HYPERVISOR_FULL_ROLES:
        return "full"
    if role_value in HYPERVISOR_READONLY_ROLES:
        return "readonly"
    return None


def normalize_excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, ObjectId):
        return str(value)
    if isinstance(value, str):
        # Mitigate spreadsheet formula injection vectors.
        if value and value[0] in ("=", "+", "-", "@", "\t"):
            return "'" + value
    return value


BACKUP_SCOPE_COLLECTIONS: Dict[str, List[str]] = {
    "users": ["users"],
    "customers_with_vehicles": ["customers", "vehicles"],
    "inventory": ["inventory"],
    "permissions": ["user_permissions", "role_permissions"],
    "movements": ["inventory_movements", "product_transfers", "dispatch_orders"],
    "registros": ["notifications", "audit_logs", "hypervisor_events", "price_history"],
}


def parse_scope_values(scopes: Optional[str]) -> List[str]:
    if not scopes:
        return []
    values = [s.strip().lower() for s in scopes.split(",") if s.strip()]
    dedup: List[str] = []
    for value in values:
        if value not in dedup:
            dedup.append(value)
    return dedup


def parse_filter_values(raw_values: Optional[str]) -> List[str]:
    if not raw_values:
        return []
    values = [v.strip() for v in raw_values.split(",") if v and v.strip()]
    dedup: List[str] = []
    for value in values:
        if value not in dedup:
            dedup.append(value)
    return dedup


def parse_string_bool(value: Optional[str], default: bool = False) -> bool:
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def parse_bool_like(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "on", "si", "sí"}:
        return True
    if text in {"0", "false", "no", "off"}:
        return False
    return default


def read_backup_manifest_metadata(workbook: Any) -> Dict[str, Any]:
    """Extract manifest metadata from secure backups; returns defaults for legacy files."""
    manifest = workbook["__MANIFEST"] if "__MANIFEST" in workbook.sheetnames else None
    if manifest is None:
        return {
            "format": "LEGACY",
            "schema_version": 1,
            "selected_scopes": ["all"],
            "manifest_checksum": None,
        }

    metadata: Dict[str, Any] = {
        "format": "LEGACY",
        "schema_version": 1,
        "selected_scopes": ["all"],
        "manifest_checksum": None,
    }
    for row in manifest.iter_rows(min_row=2, values_only=True):
        key = row[0] if len(row) > 0 else None
        value = row[1] if len(row) > 1 else None
        if not key or str(key).strip() == "---":
            break
        key_text = str(key).strip()
        if key_text == "format":
            metadata["format"] = value
        elif key_text == "schema_version":
            try:
                metadata["schema_version"] = int(value)
            except Exception:
                metadata["schema_version"] = 1
        elif key_text == "selected_scopes":
            try:
                parsed = json.loads(str(value or "[]"))
                metadata["selected_scopes"] = parsed if isinstance(parsed, list) else ["all"]
            except Exception:
                metadata["selected_scopes"] = ["all"]
        elif key_text == "manifest_checksum":
            metadata["manifest_checksum"] = value

    return metadata


def migrate_backup_row_for_schema(
    schema_version: int,
    collection_name: str,
    doc: Dict[str, Any],
) -> Dict[str, Any]:
    """Migrate row payload from older backup schemas into current document shape."""
    migrated = dict(doc)

    # Schema v1 -> v2 compatibility normalizations.
    if schema_version <= 1:
        if collection_name == "users":
            if "is_active" not in migrated:
                migrated["is_active"] = True
            migrated["is_active"] = parse_bool_like(migrated.get("is_active"), default=True)
            if "is_pin_user" in migrated:
                migrated["is_pin_user"] = parse_bool_like(migrated.get("is_pin_user"), default=False)

        if collection_name == "customers":
            if "is_active" not in migrated:
                migrated["is_active"] = True
            migrated["is_active"] = parse_bool_like(migrated.get("is_active"), default=True)

        if collection_name == "vehicles":
            year_value = migrated.get("year")
            if isinstance(year_value, str) and year_value.isdigit():
                migrated["year"] = int(year_value)

        if collection_name == "inventory":
            min_stock = migrated.get("min_stock")
            if min_stock in (None, ""):
                migrated["min_stock"] = 0
            elif isinstance(min_stock, str):
                try:
                    migrated["min_stock"] = int(float(min_stock))
                except Exception:
                    migrated["min_stock"] = 0

    return migrated


def get_backup_targets(
    selected_scopes: List[str],
    available_collections: List[str],
) -> Dict[str, Any]:
    available_set = set(available_collections)
    selected = [s for s in selected_scopes if s in BACKUP_SCOPE_COLLECTIONS or s == "brands_models"]
    if not selected:
        return {
            "mode": "all",
            "collections": sorted(available_set),
            "include_brands_models": True,
            "selected_scopes": ["all"],
        }

    collection_set: set[str] = set()
    include_brands_models = False
    for scope in selected:
        if scope == "brands_models":
            include_brands_models = True
            continue
        for collection_name in BACKUP_SCOPE_COLLECTIONS.get(scope, []):
            if collection_name in available_set:
                collection_set.add(collection_name)

    return {
        "mode": "scoped",
        "collections": sorted(collection_set),
        "include_brands_models": include_brands_models,
        "selected_scopes": selected,
    }


def compute_sheet_checksum(rows: List[Dict[str, Any]]) -> str:
    payload = json.dumps(rows, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def maybe_parse_json_like(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text:
        return value
    if (text.startswith("{") and text.endswith("}")) or (text.startswith("[") and text.endswith("]")):
        try:
            return json.loads(text)
        except Exception:
            return value
    return value


async def build_brands_models_rows() -> List[Dict[str, Any]]:
    rows = await db.vehicles.find({}, {"_id": 0, "brand": 1, "model": 1, "vehicle_type": 1}).to_list(50000)
    seen: set[tuple[str, str, str]] = set()
    output: List[Dict[str, Any]] = []
    for row in rows:
        brand = str(row.get("brand") or "").strip()
        model = str(row.get("model") or "").strip()
        vehicle_type = str(row.get("vehicle_type") or "").strip()
        if not brand or not model:
            continue
        key = (brand.lower(), model.lower(), vehicle_type.lower())
        if key in seen:
            continue
        seen.add(key)
        output.append({"brand": brand, "model": model, "vehicle_type": vehicle_type})
    output.sort(key=lambda x: (str(x.get("brand") or "").lower(), str(x.get("model") or "").lower()))
    return output


async def record_hypervisor_event(
    request: Request,
    actor: Optional[User],
    action: str,
    entity_type: str,
    entity_id: Optional[str] = None,
    details: Optional[Dict[str, Any]] = None,
    status_code: Optional[int] = None,
) -> None:
    try:
        event_doc: Dict[str, Any] = {
            "event_id": f"hv_{uuid.uuid4().hex[:12]}",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "actor_user_id": actor.user_id if actor else None,
            "actor_name": actor.name if actor else None,
            "actor_role": actor.role if actor else None,
            "action": action,
            "entity_type": entity_type,
            "entity_id": entity_id,
            "path": request.url.path,
            "method": request.method,
            "status_code": status_code,
            "ip": request.client.host if request.client else None,
            "details": details or {},
        }
        await db.hypervisor_events.insert_one(event_doc)
    except Exception:
        logger.exception("Failed to record HyperVisor event")


async def require_hypervisor_access(request: Request, write: bool = False) -> User:
    user = await require_auth(request)
    mode = get_hypervisor_access_mode(user.role)
    if not mode:
        raise HTTPException(status_code=403, detail="Forbidden")
    if write and mode != "full":
        raise HTTPException(status_code=403, detail="Forbidden")
    return user

PRODUCT_CATEGORIES = {
    "audio": {"name": "Audio", "subcategories": ["Parlantes", "Subwoofers"]},
    "security": {"name": "Seguridad", "subcategories": ["Alarmas", "Cámaras"]},
    "tint": {"name": "Polarizados", "subcategories": ["Películas", "Accesorios"]},
    "accessories": {"name": "Accesorios", "subcategories": ["Interior", "Exterior"]},
}

audit_service = AuditService(db, logger)
pin_policy_service = PinPolicyService(db, logger)
cash_service = CashService(db, logger)

# Central roles catalog (frontend can fetch via /api/roles)
ROLES = {
    "gerencia": {"label": "Gerencia", "color": "purple"},
    "recursos_humanos": {"label": "Recursos Humanos", "color": "violet"},
    "supervisor": {"label": "Supervisor", "color": "blue"},
    "cajero": {"label": "Cajero", "color": "emerald"},
    "ventas": {"label": "Ventas", "color": "green"},
    "jefe_vendedores": {"label": "Jefe de Vendedores", "color": "teal"},
    "jefe_tienda": {"label": "Jefe de Tienda", "color": "cyan"},
    "electrico": {"label": "Eléctrico", "color": "indigo"},
    "polarizador": {"label": "Polarizador", "color": "pink"},
    "transporte": {"label": "Transporte", "color": "orange"},
    "bodegas": {"label": "Bodegas", "color": "yellow"},
    "instalaciones": {"label": "Instalaciones", "color": "red"},
    "programador": {"label": "Programador", "color": "slate"},
}

BRANCH_SERVICE_POLICY_DEFAULTS: Dict[str, Dict[str, Any]] = {
    "branch_main": {
        "installations_enabled": True,
        "tint_enabled": True,
        "local_delivery_enabled": True,
        "departmental_delivery_enabled": True,
    },
    "branch_north": {
        "installations_enabled": False,
        "tint_enabled": False,
        "local_delivery_enabled": True,
        "departmental_delivery_enabled": True,
    },
    "branch_south": {
        "installations_enabled": False,
        "tint_enabled": False,
        "local_delivery_enabled": True,
        "departmental_delivery_enabled": True,
    },
}

ROLE_EQUIVALENCE: Dict[str, str] = {
    "recursos_humanos": "gerencia",
    "jefe_vendedores": "supervisor",
    "jefe_tienda": "supervisor",
}

ROLE_KEY_RE = re.compile(r"^[a-z][a-z0-9_]{2,39}$")


def resolve_effective_role(role: Optional[str]) -> str:
    if not role:
        return ""
    return ROLE_EQUIVALENCE.get(role, role)


def allow_rrhh_pin_management() -> bool:
    return os.environ.get("ALLOW_RRHH_PIN_MANAGEMENT", "false").strip().lower() in {
        "1",
        "true",
        "yes",
        "on",
    }


def can_manage_other_users_pin(actor: User) -> bool:
    if actor.role == "gerencia":
        return True
    if actor.role == "recursos_humanos" and allow_rrhh_pin_management():
        return True
    return False


def can_manage_login_pin(actor: User) -> bool:
    return actor.role in {"gerencia", "recursos_humanos", "programador"}


async def get_roles_catalog() -> Dict[str, Dict[str, str]]:
    roles = dict(ROLES)
    custom_roles = await db.custom_roles.find({}, {"_id": 0, "role": 1, "label": 1, "color": 1}).to_list(1000)
    for role_doc in custom_roles:
        role_key = (role_doc or {}).get("role")
        if not role_key or not isinstance(role_key, str):
            continue
        roles[role_key] = {
            "label": (role_doc or {}).get("label") or role_key.replace("_", " ").title(),
            "color": (role_doc or {}).get("color") or "gray",
        }
    return roles

PERMISSION_ACTIONS = ["create", "view", "edit", "delete"]

PERMISSIONS_CATALOG: Dict[str, Dict[str, Any]] = {
    "general": {
        "label": "General",
        "functions": {
            "dashboard": "Dashboard",
            "notifications": "Notificaciones",
            "tutorials": "Tutoriales",
        },
    },
    "ventas": {
        "label": "Ventas",
        "functions": {
            "sales": "Ventas",
            "quotations": "Cotizaciones",
            "credits": "Créditos",
            "returns": "Devoluciones",
            "approvals": "Aprobaciones",
        },
    },
    "clientes": {
        "label": "Clientes",
        "functions": {
            "customers": "Clientes",
            "vehicles": "Vehículos",
            "followups": "Seguimientos",
        },
    },
    "inventario": {
        "label": "Inventario",
        "functions": {
            "catalog": "Catálogo",
            "samples": "Muestras",
            "inventory": "Inventario",
            "dispatch": "Despacho",
            "warehouses": "Bodegas",
            "promotions": "Promociones",
        },
    },
    "operaciones": {
        "label": "Operaciones",
        "functions": {
            "work_orders": "Órdenes de Trabajo",
            "quality_control": "Control de Calidad",
            "kds": "KDS",
            "deliveries": "Entregas",
            "calendar": "Calendario",
            "tint_orders": "Polarizados",
            "warranties": "Garantías",
        },
    },
    "administracion": {
        "label": "Administración",
        "functions": {
            "reports": "Reportes",
            "branches": "Sucursales",
            "human_resources": "Recursos Humanos",
            "users": "Usuarios",
            "settings": "Configuración",
            "system_settings": "Sistema",
        },
    },
}

FUNCTION_ALLOWED_ROLES: Dict[str, List[str]] = {
    "dashboard": ["all"],
    "notifications": ["gerencia", "supervisor", "ventas", "cajero", "jefe_vendedores", "jefe_tienda"],
    "tutorials": ["all"],
    "sales": ["gerencia", "supervisor", "ventas", "cajero", "jefe_vendedores", "jefe_tienda"],
    "quotations": ["gerencia", "supervisor", "ventas", "cajero", "jefe_vendedores", "jefe_tienda"],
    "credits": ["gerencia", "supervisor", "ventas", "cajero", "jefe_vendedores", "jefe_tienda"],
    "returns": ["gerencia", "supervisor", "ventas", "cajero", "jefe_vendedores", "jefe_tienda"],
    "approvals": ["gerencia", "supervisor"],
    "customers": ["gerencia", "supervisor", "ventas", "cajero", "jefe_vendedores", "jefe_tienda"],
    "vehicles": ["gerencia", "supervisor", "ventas", "cajero", "instalaciones", "jefe_vendedores", "jefe_tienda"],
    "followups": ["gerencia", "supervisor", "ventas", "cajero", "jefe_vendedores", "jefe_tienda"],
    "catalog": ["gerencia", "supervisor", "ventas", "jefe_vendedores", "jefe_tienda"],
    "samples": ["gerencia", "supervisor", "ventas", "jefe_vendedores", "jefe_tienda"],
    "inventory": ["gerencia", "supervisor", "bodegas", "jefe_tienda", "ventas", "cajero", "jefe_vendedores"],
    "dispatch": ["gerencia", "supervisor", "bodegas", "jefe_tienda"],
    "warehouses": ["gerencia", "supervisor", "jefe_tienda", "ventas", "cajero", "jefe_vendedores"],
    "promotions": ["gerencia", "supervisor", "jefe_vendedores", "jefe_tienda"],
    "work_orders": ["gerencia", "supervisor", "instalaciones"],
    "quality_control": ["gerencia", "supervisor"],
    "kds": ["all"],
    "deliveries": ["gerencia", "supervisor", "transporte"],
    "calendar": ["gerencia", "supervisor", "instalaciones"],
    "tint_orders": ["gerencia", "supervisor", "instalaciones"],
    "warranties": ["gerencia", "supervisor", "instalaciones"],
    "reports": ["gerencia", "supervisor", "jefe_vendedores", "jefe_tienda"],
    "branches": ["gerencia"],
    "human_resources": ["gerencia", "recursos_humanos", "supervisor"],
    "users": ["gerencia"],
    "settings": ["gerencia"],
    "system_settings": ["gerencia"],
}

ROLE_WRITE_ALLOWED_FUNCTIONS: Dict[str, set[str]] = {
    "ventas": {"sales", "quotations", "credits", "returns", "customers", "vehicles", "followups"},
    "cajero": {"sales", "quotations", "credits", "returns", "customers", "vehicles", "followups"},
    "jefe_vendedores": {"sales", "quotations", "credits", "returns", "customers", "vehicles", "followups", "catalog", "samples"},
    "jefe_tienda": {"sales", "quotations", "credits", "returns", "customers", "vehicles", "followups", "inventory", "dispatch", "catalog", "samples", "promotions"},
    "bodegas": {"inventory", "dispatch"},
}

ROLE_PERMISSION_FLOORS: Dict[str, Dict[str, Dict[str, bool]]] = {
    "programador": {
        "users": {"view": True, "create": True, "edit": True, "delete": True},
    },
    "bodegas": {
        "inventory": {"view": True, "create": True, "edit": True},
        "dispatch": {"view": True, "create": True, "edit": True},
    }
}


def apply_role_permission_floor(matrix: Dict[str, Any], role: str) -> Dict[str, Any]:
    normalized = normalize_permission_matrix(matrix)
    floor_cfg = ROLE_PERMISSION_FLOORS.get(resolve_effective_role(role), {})
    if not floor_cfg:
        return normalized

    for function_key, required_actions in floor_cfg.items():
        for module_key in PERMISSIONS_CATALOG.keys():
            module_data = normalized.get(module_key, {}) or {}
            function_data = module_data.get(function_key)
            if not isinstance(function_data, dict):
                continue
            for action, required in required_actions.items():
                if required:
                    function_data[action] = True

    return normalized

PERMISSION_METHOD_TO_ACTION: Dict[str, str] = {
    "GET": "view",
    "POST": "create",
    "PUT": "edit",
    "PATCH": "edit",
    "DELETE": "delete",
}

PERMISSION_ROUTE_MAP: List[tuple[str, str]] = [
    ("/api/notifications", "notifications"),
    ("/api/followups", "followups"),
    ("/api/sales", "sales"),
    ("/api/quotations", "quotations"),
    ("/api/credit", "credits"),
    ("/api/returns", "returns"),
    ("/api/approvals", "approvals"),
    ("/api/customers", "customers"),
    ("/api/vehicles", "vehicles"),
    ("/api/catalog", "catalog"),
    ("/api/samples", "samples"),
    ("/api/inventory", "inventory"),
    ("/api/dispatch", "dispatch"),
    ("/api/warehouses", "warehouses"),
    ("/api/promotions", "promotions"),
    ("/api/work-orders", "work_orders"),
    ("/api/quality-control", "quality_control"),
    ("/api/kds", "kds"),
    ("/api/deliveries", "deliveries"),
    ("/api/calendar", "calendar"),
    ("/api/tint", "tint_orders"),
    ("/api/warranties", "warranties"),
    ("/api/reports", "reports"),
    ("/api/branches", "branches"),
    ("/api/hr", "human_resources"),
    ("/api/permissions", "users"),
    ("/api/users", "users"),
    ("/api/settings", "settings"),
    ("/api/system-settings", "system_settings"),
]

PERMISSION_ENFORCEMENT_EXACT_PATHS = {
    "/api/",
    "/api",
    "/api/health",
    "/api/roles",
    "/api/auth/pin/login",
    "/api/auth/pin/users",
    "/api/auth/logout",
    "/api/auth/me",
    "/api/permissions/me",
}

PERMISSION_ENFORCEMENT_EXEMPT_PREFIXES = [
    "/api/auth/",
    "/api/test/",
]

SESSION_LOCK_EXEMPT_PATHS = {
    "/api/auth/me",
    "/api/auth/logout",
    "/api/auth/session/lock",
    "/api/auth/session/unlock",
    "/api/auth/pin/login",
    "/api/auth/pin/users",
    "/api/health",
    "/api",
    "/api/",
}


def build_empty_permission_matrix() -> Dict[str, Any]:
    matrix: Dict[str, Any] = {}
    for module_key, module_cfg in PERMISSIONS_CATALOG.items():
        matrix[module_key] = {}
        for function_key in module_cfg.get("functions", {}).keys():
            matrix[module_key][function_key] = {
                "create": False,
                "view": False,
                "edit": False,
                "delete": False,
            }
    return matrix


def normalize_permission_matrix(raw_permissions: Any) -> Dict[str, Any]:
    normalized = build_empty_permission_matrix()
    if not isinstance(raw_permissions, dict):
        return normalized

    for module_key, module_cfg in PERMISSIONS_CATALOG.items():
        module_raw = raw_permissions.get(module_key, {})
        if not isinstance(module_raw, dict):
            continue
        for function_key in module_cfg.get("functions", {}).keys():
            function_raw = module_raw.get(function_key, {})
            if not isinstance(function_raw, dict):
                continue
            for action in PERMISSION_ACTIONS:
                normalized[module_key][function_key][action] = bool(
                    function_raw.get(action, False)
                )

    return normalized


def build_default_role_permissions(role: str) -> Dict[str, Any]:
    matrix = build_empty_permission_matrix()
    effective_role = resolve_effective_role(role)

    for module_key, module_cfg in PERMISSIONS_CATALOG.items():
        for function_key in module_cfg.get("functions", {}).keys():
            allowed_roles = FUNCTION_ALLOWED_ROLES.get(function_key, [])
            effective_allowed_roles = [resolve_effective_role(item) for item in allowed_roles]
            can_view = "all" in allowed_roles or effective_role in effective_allowed_roles

            if effective_role == "gerencia":
                matrix[module_key][function_key] = {
                    "create": True,
                    "view": True,
                    "edit": True,
                    "delete": True,
                }
            elif effective_role == "supervisor":
                matrix[module_key][function_key] = {
                    "create": can_view,
                    "view": can_view,
                    "edit": can_view,
                    "delete": False,
                }
            elif effective_role == "programador":
                matrix[module_key][function_key] = {
                    "create": True,
                    "view": True,
                    "edit": True,
                    "delete": False,
                }
            else:
                allow_write = function_key in ROLE_WRITE_ALLOWED_FUNCTIONS.get(effective_role, set())
                matrix[module_key][function_key] = {
                    "create": can_view and allow_write,
                    "view": can_view,
                    "edit": can_view and allow_write,
                    "delete": False,
                }

    return matrix


def merge_permission_matrix(base: Dict[str, Any], extra: Dict[str, Any]) -> Dict[str, Any]:
    merged = normalize_permission_matrix(base)
    if not isinstance(extra, dict):
        return merged

    for module_key, module_cfg in PERMISSIONS_CATALOG.items():
        extra_module = extra.get(module_key)
        if not isinstance(extra_module, dict):
            continue
        for function_key in module_cfg.get("functions", {}).keys():
            extra_function = extra_module.get(function_key)
            if not isinstance(extra_function, dict):
                continue
            for action in PERMISSION_ACTIONS:
                if action in extra_function:
                    merged[module_key][function_key][action] = bool(
                        extra_function.get(action)
                    )

    return merged


async def get_effective_role_permissions(role: str) -> Dict[str, Any]:
    defaults = build_default_role_permissions(role)
    custom_doc = await db.role_permissions.find_one({"role": role}, {"_id": 0})
    if not custom_doc:
        return apply_role_permission_floor(defaults, role)
    merged = merge_permission_matrix(defaults, custom_doc.get("permissions", {}))
    return apply_role_permission_floor(merged, role)


async def get_effective_permissions_for_user(user: User) -> Dict[str, Any]:
    role_effective = await get_effective_role_permissions(user.role)
    user_custom_doc = await db.user_permissions.find_one(
        {"user_id": user.user_id}, {"_id": 0}
    )
    if not user_custom_doc:
        return apply_role_permission_floor(role_effective, user.role)
    merged = merge_permission_matrix(role_effective, user_custom_doc.get("permissions", {}))
    return apply_role_permission_floor(merged, user.role)


def get_permission_value(
    matrix: Dict[str, Any], function_key: str, action: str
) -> bool:
    for module_cfg in PERMISSIONS_CATALOG.values():
        for candidate in module_cfg.get("functions", {}).keys():
            if candidate != function_key:
                continue
            for module_key in PERMISSIONS_CATALOG.keys():
                if candidate in (matrix.get(module_key, {}) or {}):
                    return bool(
                        (matrix.get(module_key, {}) or {})
                        .get(candidate, {})
                        .get(action, False)
                    )
    return False


def match_permission_function_for_path(path: str) -> Optional[str]:
    for prefix, function_key in PERMISSION_ROUTE_MAP:
        if path.startswith(prefix):
            return function_key
    return None

# Roles endpoint
@api_router.get("/roles")
async def get_roles():
    return JSONResponse(await get_roles_catalog())


@api_router.post("/roles")
async def create_role(payload: Dict[str, Any], request: Request):
    actor = await require_roles(request, ["gerencia"])

    data = payload or {}
    role_raw = data.get("role")
    label_raw = data.get("label")
    color_raw = data.get("color")

    if not isinstance(role_raw, str) or not role_raw.strip():
        raise HTTPException(status_code=400, detail="El campo 'role' es requerido")

    role_key = role_raw.strip().lower().replace("-", " ")
    role_key = re.sub(r"\s+", "_", role_key)

    if not ROLE_KEY_RE.match(role_key):
        raise HTTPException(status_code=400, detail="Rol inválido. Usa solo letras minúsculas, números y guion bajo")

    if role_key in ROLES:
        raise HTTPException(status_code=400, detail="El rol ya existe en el sistema")

    existing = await db.custom_roles.find_one({"role": role_key}, {"_id": 1})
    if existing:
        raise HTTPException(status_code=400, detail="El rol ya existe")

    label = label_raw.strip() if isinstance(label_raw, str) and label_raw.strip() else role_key.replace("_", " ").title()
    color = color_raw.strip() if isinstance(color_raw, str) and color_raw.strip() else "gray"

    now_iso = datetime.now(timezone.utc).isoformat()
    await db.custom_roles.insert_one(
        {
            "role": role_key,
            "label": label,
            "color": color,
            "created_at": now_iso,
            "created_by": actor.user_id,
            "updated_at": now_iso,
            "updated_by": actor.user_id,
        }
    )

    return {"message": "Rol creado", "role": role_key, "label": label, "color": color}


@api_router.delete("/roles/{role}")
async def delete_role(role: str, request: Request):
    await require_roles(request, ["gerencia"])

    role_key = (role or "").strip().lower()
    if not role_key:
        raise HTTPException(status_code=400, detail="Rol inválido")

    if role_key in ROLES:
        raise HTTPException(status_code=400, detail="No se puede eliminar un rol base del sistema")

    if role_key in ROLE_EQUIVALENCE or role_key in ROLE_EQUIVALENCE.values():
        raise HTTPException(status_code=400, detail="No se puede eliminar un rol reservado")

    in_use = await db.users.count_documents({"role": role_key}, limit=1)
    if in_use:
        raise HTTPException(status_code=400, detail="No se puede eliminar el rol porque está asignado a usuarios")

    delete_result = await db.custom_roles.delete_one({"role": role_key})
    if delete_result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Rol no encontrado")

    await db.role_permissions.delete_many({"role": role_key})
    return {"message": "Rol eliminado", "role": role_key}


@api_router.get("/permissions/catalog")
async def get_permissions_catalog(request: Request):
    await require_roles(request, ["gerencia", "programador"])
    return {
        "actions": PERMISSION_ACTIONS,
        "modules": PERMISSIONS_CATALOG,
    }


@api_router.get("/permissions/me")
async def get_my_permissions(request: Request):
    user = await require_auth(request)
    effective = await get_effective_permissions_for_user(user)
    return {
        "user_id": user.user_id,
        "role": user.role,
        "effective_permissions": effective,
    }


@api_router.get("/permissions/roles")
async def get_role_permissions(request: Request):
    await require_roles(request, ["gerencia", "programador"])

    result: Dict[str, Any] = {}
    roles_catalog = await get_roles_catalog()
    for role_key in roles_catalog.keys():
        result[role_key] = await get_effective_role_permissions(role_key)

    return result


@api_router.put("/permissions/roles/{role}")
async def update_role_permissions(role: str, payload: Dict[str, Any], request: Request):
    user = await require_roles(request, ["gerencia"])

    roles_catalog = await get_roles_catalog()
    if role not in roles_catalog:
        raise HTTPException(status_code=404, detail="Rol no encontrado")

    matrix = normalize_permission_matrix((payload or {}).get("permissions", {}))
    now_iso = datetime.now(timezone.utc).isoformat()

    await db.role_permissions.update_one(
        {"role": role},
        {
            "$set": {
                "role": role,
                "permissions": matrix,
                "updated_at": now_iso,
                "updated_by": user.user_id,
            }
        },
        upsert=True,
    )

    return {"message": "Permisos de rol actualizados", "role": role}


@api_router.get("/permissions/users/{user_id}")
async def get_user_permissions(user_id: str, request: Request):
    await require_roles(request, ["gerencia"])

    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    role = user_doc.get("role", "ventas")
    role_effective = await get_effective_role_permissions(role)

    user_custom_doc = await db.user_permissions.find_one({"user_id": user_id}, {"_id": 0})
    user_custom = normalize_permission_matrix((user_custom_doc or {}).get("permissions", {}))
    effective = (
        merge_permission_matrix(role_effective, user_custom_doc.get("permissions", {}))
        if user_custom_doc
        else role_effective
    )

    return {
        "user_id": user_id,
        "role": role,
        "role_permissions": role_effective,
        "user_permissions": user_custom,
        "effective_permissions": effective,
    }


@api_router.put("/permissions/users/{user_id}")
async def update_user_permissions(user_id: str, payload: Dict[str, Any], request: Request):
    actor = await require_roles(request, ["gerencia"])

    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0, "user_id": 1})
    if not user_doc:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    matrix = normalize_permission_matrix((payload or {}).get("permissions", {}))
    now_iso = datetime.now(timezone.utc).isoformat()

    await db.user_permissions.update_one(
        {"user_id": user_id},
        {
            "$set": {
                "user_id": user_id,
                "permissions": matrix,
                "updated_at": now_iso,
                "updated_by": actor.user_id,
            }
        },
        upsert=True,
    )

    return {"message": "Permisos de usuario actualizados", "user_id": user_id}


@api_router.delete("/permissions/users/{user_id}")
async def reset_user_permissions(user_id: str, request: Request):
    await require_roles(request, ["gerencia"])
    await db.user_permissions.delete_one({"user_id": user_id})
    return {"message": "Permisos de usuario restablecidos", "user_id": user_id}


# Validate incoming requests that create/update PIN users to enforce simple
# server-side constraints for `name` and `phone` before route handlers run.
PHONE_RE = re.compile(r"^\d{4}-\d{4}$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+$")
VIN_RE = re.compile(r"^[A-HJ-NPR-Z0-9]{17}$")


@app.middleware("http")
async def validate_pin_user_requests(request: Request, call_next):
    path = request.url.path or ""
    method = request.method or ""

    # Validate creation of PIN users: POST /api/users/pin
    try:
        if path == "/api/users/pin" and method.upper() == "POST":
            body_bytes = await request.body()
            if not body_bytes:
                return JSONResponse({"detail": "Cuerpo vacío"}, status_code=400)
            try:
                payload = json.loads(body_bytes.decode("utf-8"))
            except Exception:
                return JSONResponse({"detail": "JSON inválido"}, status_code=400)

            name = payload.get("name")
            if not name or not isinstance(name, str) or not name.strip():
                return JSONResponse({"detail": "El campo 'name' es requerido"}, status_code=400)

            phone = payload.get("phone")
            if phone is not None and phone != "":
                if not isinstance(phone, str) or not PHONE_RE.match(phone):
                    return JSONResponse({"detail": "El campo 'phone' debe tener formato ####-####"}, status_code=400)

            # restore body for downstream handlers
            request._body = body_bytes
    except Exception:
        # Don't block request processing on unexpected middleware errors
        pass

    return await call_next(request)


@app.middleware("http")
async def enforce_session_lock(request: Request, call_next):
    path = request.url.path or ""
    method = (request.method or "GET").upper()

    if not path.startswith("/api"):
        return await call_next(request)

    if method == "OPTIONS":
        return await call_next(request)

    if path in SESSION_LOCK_EXEMPT_PATHS:
        return await call_next(request)

    token: Optional[str] = None
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.lower().startswith("bearer "):
        token = auth_header.split(" ", 1)[1].strip()
    if not token:
        token = request.cookies.get("session_token")

    if not token:
        return await call_next(request)

    session_doc = await db.sessions.find_one({"session_token": token}, {"_id": 0, "locked": 1})
    if not session_doc:
        return await call_next(request)

    if bool(session_doc.get("locked")):
        return JSONResponse(
            {
                "detail": {
                    "message": "Sesión bloqueada en servidor",
                    "code": "SESSION_LOCKED",
                }
            },
            status_code=423,
        )

    return await call_next(request)


@app.middleware("http")
async def enforce_runtime_permissions(request: Request, call_next):
    path = request.url.path or ""
    method = (request.method or "GET").upper()

    if not path.startswith("/api"):
        return await call_next(request)

    if path in PERMISSION_ENFORCEMENT_EXACT_PATHS:
        return await call_next(request)

    if any(path.startswith(prefix) for prefix in PERMISSION_ENFORCEMENT_EXEMPT_PREFIXES):
        return await call_next(request)

    action = PERMISSION_METHOD_TO_ACTION.get(method)
    if not action:
        return await call_next(request)

    function_key = match_permission_function_for_path(path)
    if not function_key:
        return await call_next(request)

    token: Optional[str] = None
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.lower().startswith("bearer "):
        token = auth_header.split(" ", 1)[1].strip()
    if not token:
        token = request.cookies.get("session_token")

    if not token:
        return await call_next(request)

    try:
        user_doc = await _get_user_by_session(token)
    except Exception:
        user_doc = None

    if not user_doc:
        return await call_next(request)

    try:
        user = User(**user_doc)
        effective = await get_effective_permissions_for_user(user)
        if not get_permission_value(effective, function_key, action):
            return JSONResponse(
                {
                    "detail": "Permiso denegado",
                    "permission": {
                        "function": function_key,
                        "action": action,
                    },
                },
                status_code=403,
            )
    except Exception:
        return JSONResponse({"detail": "Error validando permisos"}, status_code=500)

    return await call_next(request)


@app.middleware("http")
async def hypervisor_runtime_audit(request: Request, call_next):
    path = request.url.path or ""
    method = (request.method or "GET").upper()

    response = await call_next(request)

    if not path.startswith("/api"):
        return response

    if method not in {"POST", "PUT", "PATCH", "DELETE"}:
        return response

    if path.startswith("/api/auth/pin/login"):
        return response

    actor: Optional[User] = None
    token: Optional[str] = None
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.lower().startswith("bearer "):
        token = auth_header.split(" ", 1)[1].strip()
    if not token:
        token = request.cookies.get("session_token")

    if token:
        try:
            user_doc = await _get_user_by_session(token)
            if user_doc:
                actor = User(**user_doc)
        except Exception:
            actor = None

    path_parts = [p for p in path.split("/") if p]
    entity_type = path_parts[1] if len(path_parts) > 1 else "system"
    entity_id = path_parts[2] if len(path_parts) > 2 else None

    await record_hypervisor_event(
        request=request,
        actor=actor,
        action=f"request_{method.lower()}",
        entity_type=entity_type,
        entity_id=entity_id,
        details={
            "auto": True,
            "query": dict(request.query_params),
        },
        status_code=response.status_code,
    )

    return response


class FlexibleModel(BaseModel):
    model_config = ConfigDict(extra="allow")


class User(FlexibleModel):
    user_id: str
    name: str
    role: str = "ventas"
    email: Optional[str] = None
    branch_id: Optional[str] = None
    warehouse_id: Optional[str] = None


class PinLoginRequest(FlexibleModel):
    pin: str
    user_id: Optional[str] = None


class SessionUnlockRequest(FlexibleModel):
    pin: str


class Warehouse(FlexibleModel):
    pass


class Product(FlexibleModel):
    pass


class ProductCreate(FlexibleModel):
    pass


class ProductUpdate(FlexibleModel):
    pass


class Customer(FlexibleModel):
    pass


class CustomerCreate(FlexibleModel):
    pass


class Vehicle(FlexibleModel):
    pass


class VehicleCreate(FlexibleModel):
    pass


class VehicleUpdate(FlexibleModel):
    pass


class InventoryUpdate(FlexibleModel):
    pass


class QuotationItem(FlexibleModel):
    product_id: str
    product_name: str
    quantity: int
    unit_price: float
    with_installation: bool = False
    discount: float = 0.0
    subtotal: float = 0.0


class Quotation(FlexibleModel):
    quotation_id: Optional[str] = None
    customer_id: str
    customer_name: str
    branch_id: str
    salesperson_id: str
    salesperson_name: str
    vehicle_id: Optional[str] = None
    warehouse_id: Optional[str] = None
    items: List[Dict[str, Any]] = Field(default_factory=list)
    subtotal: float = 0.0
    tax: float = 0.0
    discount: float = 0.0
    discount_percent: float = 0.0
    apply_iva: bool = True
    iva_rate: float = 12.0
    currency: str = "USD"
    exchange_rate: Optional[float] = None
    discount_codes: List[str] = Field(default_factory=list)
    payment_type: str = "cash"
    payment_method: str = "cash"
    credit_days: Optional[int] = None
    discounts_blocked_by_method: bool = False
    total: float = 0.0
    valid_until: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    notes: Optional[str] = None
    status: str = "pending"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class QuotationCreate(FlexibleModel):
    customer_id: str
    items: List[Dict[str, Any]] = Field(default_factory=list)
    currency: Optional[str] = None
    exchange_rate: Optional[float] = None
    apply_iva: Optional[bool] = None
    iva_rate: Optional[float] = None
    discount: float = 0.0
    discount_codes: Optional[List[str]] = None
    payment_type: Optional[str] = None
    payment_method: Optional[str] = None
    credit_days: Optional[int] = None
    vehicle_id: Optional[str] = None
    warehouse_id: Optional[str] = None
    valid_days: int = 7
    notes: Optional[str] = None


class ThemeSettings(FlexibleModel):
    mode: Optional[str] = None
    skin: Optional[str] = None
    custom: Optional[Dict[str, str]] = None


class AppearanceSettings(FlexibleModel):
    watermark_opacity: Optional[float] = None


class VehicleSettingsNamePayload(FlexibleModel):
    name: str


class VehicleSettingsYearPayload(FlexibleModel):
    year: str


class VehicleSettingsVariationPayload(FlexibleModel):
    value: str


class VehicleSettingsColorPayload(FlexibleModel):
    value: str


class BillingExchangeUpdatePayload(FlexibleModel):
    official_rate: float


class BillingIvaUpdatePayload(FlexibleModel):
    iva_rate: float


class SellerGlobalDiscountPolicyPayload(FlexibleModel):
    enabled: bool = True
    max_percent: float = 5.0
    max_amount_nio: float = 1000.0


class BillingExchangeRulePayload(FlexibleModel):
    name: str
    cadence: str  # daily | weekly | monthly | custom
    rate: float
    start_at: Optional[datetime] = None
    end_at: Optional[datetime] = None
    active: bool = True


class BillingCancelReasonPayload(FlexibleModel):
    reason: str
    active: bool = True
    sort_order: Optional[int] = None


class SaleRequestPayload(FlexibleModel):
    reason: str


class InvoiceEditItemPayload(FlexibleModel):
    product_id: str
    quantity: int
    unit_price: float
    discount: float = 0.0


class InvoiceEditPayload(FlexibleModel):
    items: List[InvoiceEditItemPayload] = Field(default_factory=list)
    notes: Optional[str] = None


class SaleItem(FlexibleModel):
    product_id: str
    product_name: str
    quantity: int
    unit_price: float
    discount: float = 0.0
    subtotal: float = 0.0
    warehouse_id: Optional[str] = None
    installation_type: Optional[str] = None
    with_installation: bool = False
    display_note: Optional[str] = None


class Sale(FlexibleModel):
    sale_id: Optional[str] = None
    invoice_number: str
    quotation_id: Optional[str] = None
    customer_id: str
    customer_name: str
    branch_id: str
    salesperson_id: str
    salesperson_name: str
    items: List[Dict[str, Any]] = Field(default_factory=list)
    subtotal: float = 0.0
    tax: float = 0.0
    discount: float = 0.0
    total: float = 0.0
    payment_type: str = "cash"
    payment_status: str = "pending"
    sale_channel: str = "minorista"
    credit_due_date: Optional[datetime] = None
    delivery_required: bool = False
    delivery_address: Optional[str] = None
    delivery_status: Optional[str] = None
    notes: Optional[str] = None
    has_installation: bool = False
    cash_session_id: Optional[str] = None
    payment_method: Optional[str] = None
    iva_rate: float = 0.12
    iva_amount: float = 0.0
    total_legal: float = 0.0
    discounts_applied_amount: float = 0.0
    discounts_blocked_by_method: bool = False
    retention_rate: float = 0.0
    retention_amount: float = 0.0
    net_to_collect: float = 0.0
    print_format: str = "thermal80"
    retention_receipt_required: bool = False
    pos_bank_withholding_expected: float = 0.0
    commercial_terms_locked: bool = False
    settlement_warnings: List[str] = Field(default_factory=list)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class SaleCreate(FlexibleModel):
    customer_id: str
    quotation_id: Optional[str] = None
    items: List[Dict[str, Any]] = Field(default_factory=list)
    discount: float = 0.0
    payment_type: str = "cash"
    sale_channel: Optional[str] = None
    credit_days: Optional[int] = None
    manager_authorization_code: Optional[str] = None
    delivery_required: bool = False
    delivery_address: Optional[str] = None
    notes: Optional[str] = None
    vehicle_id: Optional[str] = None
    cash_session_id: Optional[str] = None
    payment_method: Optional[str] = None
    idempotency_key: Optional[str] = None


class DraftEntryPayload(FlexibleModel):
    name: Optional[str] = None
    snapshot: Optional[Dict[str, Any]] = None


class DraftStatePayload(FlexibleModel):
    active_draft_id: Optional[str] = None


class SaleSettlementPreviewRequest(FlexibleModel):
    customer_id: Optional[str] = None
    items: List[Dict[str, Any]] = Field(default_factory=list)
    subtotal: Optional[float] = None
    discount_percent: float = 0.0
    discounts_amount: float = 0.0
    promotions_amount: float = 0.0
    payment_method: str = "cash"
    print_format: str = "thermal80"  # letter | thermal80
    apply_iva: Optional[bool] = None
    retention_profile: Optional[str] = None  # general | agricola | exento
    retention_rate_hint: Optional[float] = None  # 0.01 | 0.02


class SaleCommercialTermsUpdate(FlexibleModel):
    payment_method: str
    print_format: str = "thermal80"
    apply_iva: Optional[bool] = None
    retention_profile: Optional[str] = None
    retention_rate_hint: Optional[float] = None
    lock_terms: bool = True


class CashierCollectRequest(FlexibleModel):
    amount: float = 0.0
    payment_method: Optional[str] = None
    reference: Optional[str] = None
    notes: Optional[str] = None
    idempotency_key: Optional[str] = None
    received_amount: Optional[float] = None
    force_remove_discount: bool = False
    pagos: List[MixedPaymentItem] = Field(default_factory=list)
    autorizacion_descuento_pos: Optional[PosDiscountAuthorization] = None


class MixedPaymentItem(FlexibleModel):
    metodo: str
    moneda: str = "NIO"
    monto_origen: float
    tasa_cambio: Optional[float] = None
    monto_cordobas: Optional[float] = None
    referencia_bancaria: Optional[str] = None
    notas_auditoria: Optional[str] = None


class PosDiscountAuthorization(FlexibleModel):
    autorizado_por: str
    justificacion_interna: str
    mostrar_al_cliente: bool = False


class FacturacionPagarRequest(FlexibleModel):
    factura_id: str
    total_original: Optional[float] = None
    descuento_aplicado: Optional[float] = None
    pagos: List[MixedPaymentItem] = Field(default_factory=list)
    autorizacion_descuento_pos: Optional[PosDiscountAuthorization] = None
    idempotency_key: Optional[str] = None
    force_remove_discount: bool = False
    notas: Optional[str] = None


class CashierInvoiceCollectRequest(FlexibleModel):
    sesion_id: str
    amount: float = 0.0
    payment_method: Optional[str] = None
    reference: Optional[str] = None
    notes: Optional[str] = None
    idempotency_key: Optional[str] = None
    received_amount: Optional[float] = None
    force_remove_discount: bool = False
    pagos: List[MixedPaymentItem] = Field(default_factory=list)
    autorizacion_descuento_pos: Optional[PosDiscountAuthorization] = None


class CashierInvoiceCancelRequest(FlexibleModel):
    motivo: str
    justificacion_interna: str
    autorizado_por: Optional[str] = None


def _normalize_method_name(value: Any) -> str:
    method = str(value or "").strip().lower()
    aliases = {
        "efectivo": "cash",
        "cash": "cash",
        "transfer": "transfer",
        "transferencia": "transfer",
        "bank_transfer": "transfer",
        "wire": "transfer",
        "card": "card",
        "tarjeta": "card",
        "credit_card": "card",
        "debit_card": "card",
        "tarjeta_credito": "card",
        "tarjeta_debito": "card",
    }
    return aliases.get(method, method or "cash")


def _currency_code(value: Any) -> str:
    code = str(value or "NIO").strip().upper()
    return code if code in {"NIO", "USD"} else "NIO"


def _round4(value: Union[int, float]) -> float:
    return round(float(value or 0.0), 4)


def _round2(value: Union[int, float]) -> float:
    return round(float(value or 0.0), 2)


def _is_card_method(value: Any) -> bool:
    return _normalize_method_name(value) == "card"


def _is_discount_allowed(method: str) -> bool:
    normalized = _normalize_method_name(method)
    return normalized in {"cash", "transfer"}


async def _get_seller_global_discount_policy() -> Dict[str, Any]:
    default_policy = {
        "type": "seller_global_discount_policy",
        "enabled": True,
        "max_percent": 5.0,
        "max_amount_nio": 1000.0,
        "updated_at": None,
    }

    doc = await db.settings.find_one({"type": "seller_global_discount_policy"}, {"_id": 0})
    if not doc:
        return default_policy

    enabled = bool(doc.get("enabled", True))
    try:
        max_percent = float(doc.get("max_percent", 5.0))
    except (TypeError, ValueError):
        max_percent = 5.0
    try:
        max_amount_nio = float(doc.get("max_amount_nio", 1000.0))
    except (TypeError, ValueError):
        max_amount_nio = 1000.0

    max_percent = max(0.0, min(100.0, max_percent))
    max_amount_nio = max(0.0, max_amount_nio)

    return {
        "type": "seller_global_discount_policy",
        "enabled": enabled,
        "max_percent": max_percent,
        "max_amount_nio": max_amount_nio,
        "updated_at": doc.get("updated_at"),
    }


async def _enforce_seller_global_discount_limits(
    actor: User,
    subtotal: float,
    discount_percent: float,
    currency: str = "USD",
    exchange_rate: Optional[float] = None,
) -> None:
    effective_role = resolve_effective_role(getattr(actor, "role", None))
    if effective_role != "ventas":
        return

    requested_percent = max(float(discount_percent or 0.0), 0.0)
    if requested_percent <= 0:
        return

    policy = await _get_seller_global_discount_policy()
    if not bool(policy.get("enabled", True)):
        return

    max_percent = float(policy.get("max_percent") or 0.0)
    max_amount_nio = float(policy.get("max_amount_nio") or 0.0)

    if requested_percent > max_percent + 1e-9:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "SELLER_GLOBAL_DISCOUNT_LIMIT_EXCEEDED",
                "message": f"El descuento global para vendedores no puede superar {max_percent:.2f}%.",
                "policy": policy,
                "requested_discount_percent": round(requested_percent, 4),
            },
        )

    normalized_currency = _currency_code(currency)
    subtotal_amount = max(float(subtotal or 0.0), 0.0)
    requested_discount_amount = subtotal_amount * (requested_percent / 100.0)

    if normalized_currency == "NIO":
        discount_amount_nio = requested_discount_amount
        rate = float(exchange_rate or 0.0)
        if rate <= 0:
            rate = await _get_usd_to_nio_rate()
        max_in_currency = max_amount_nio
    else:
        rate = float(exchange_rate or 0.0)
        if rate <= 0:
            rate = await _get_usd_to_nio_rate()
        discount_amount_nio = requested_discount_amount * rate
        max_in_currency = (max_amount_nio / rate) if rate > 0 else 0.0

    if discount_amount_nio > (max_amount_nio + 0.009):
        max_label = "C$" if normalized_currency == "NIO" else "USD"
        raise HTTPException(
            status_code=400,
            detail={
                "code": "SELLER_GLOBAL_DISCOUNT_LIMIT_EXCEEDED",
                "message": "El descuento global para vendedores no puede superar 5% ni C$1,000.00 (o su equivalente en USD).",
                "policy": policy,
                "requested_discount_percent": round(requested_percent, 4),
                "requested_discount_amount": round(requested_discount_amount, 2),
                "requested_discount_amount_nio": round(discount_amount_nio, 2),
                "currency": normalized_currency,
                "max_discount_amount_in_currency": round(max_in_currency, 2),
                "max_discount_amount_nio": round(max_amount_nio, 2),
                "max_discount_amount_label": f"{max_label} {round(max_in_currency, 2):.2f}",
            },
        )


def _normalize_print_format(value: Any) -> str:
    v = str(value or "thermal80").strip().lower()
    if v in {"letter", "thermal80"}:
        return v
    return "thermal80"


def _normalize_retention_profile(value: Any) -> str:
    profile = str(value or "general").strip().lower()
    aliases = {
        "general": "general",
        "agricola": "agricola",
        "agrícola": "agricola",
        "exento": "exento",
        "exempt": "exento",
    }
    return aliases.get(profile, "general")


def _extract_retention_profile_from_customer(customer: Optional[Dict[str, Any]]) -> str:
    if not customer:
        return "general"
    if bool(customer.get("is_exempt") or customer.get("retention_exempt")):
        return "exento"
    if bool(customer.get("is_agricola") or customer.get("is_agricultural")):
        return "agricola"
    candidate = customer.get("retention_profile") or customer.get("ir_retention_profile") or customer.get("tax_profile")
    return _normalize_retention_profile(candidate)


def _retention_rate_for_settlement(
    subtotal_after_discount: float,
    profile: str,
    retention_rate_hint: Optional[float] = None,
) -> float:
    if subtotal_after_discount < 1000.0:
        return 0.0

    if retention_rate_hint in {0.01, 0.02}:
        return float(retention_rate_hint)

    normalized_profile = _normalize_retention_profile(profile)
    if normalized_profile == "exento":
        return 0.0
    if normalized_profile == "agricola":
        return 0.01
    return 0.02


def _compute_items_subtotal(items: List[Dict[str, Any]]) -> float:
    subtotal = 0.0
    for item in items:
        qty = float(item.get("quantity") or 0.0)
        unit = float(item.get("unit_price") or 0.0)
        line_subtotal = item.get("subtotal")
        if line_subtotal is not None:
            try:
                subtotal += float(line_subtotal)
                continue
            except Exception:
                pass
        subtotal += qty * unit
    return round(subtotal, 2)


def _build_sale_settlement(
    *,
    subtotal_base: float,
    discount_percent: float,
    discounts_amount: float,
    promotions_amount: float,
    payment_method: str,
    print_format: str,
    apply_iva: Optional[bool],
    iva_rate_percent: float,
    retention_profile: str,
    retention_rate_hint: Optional[float],
) -> Dict[str, Any]:
    warnings: List[str] = []

    method = _normalize_method_name(payment_method)
    fmt = _normalize_print_format(print_format)

    subtotal_base = round(max(float(subtotal_base or 0.0), 0.0), 2)
    discount_percent = max(float(discount_percent or 0.0), 0.0)
    discounts_amount = max(float(discounts_amount or 0.0), 0.0)
    promotions_amount = max(float(promotions_amount or 0.0), 0.0)

    discounts_from_percent = round(subtotal_base * (discount_percent / 100.0), 2)
    discounts_requested_total = round(discounts_from_percent + discounts_amount + promotions_amount, 2)

    if _is_discount_allowed(method):
        discounts_applied = min(discounts_requested_total, subtotal_base)
        discounts_removed = 0.0
        discounts_blocked = False
    else:
        discounts_applied = 0.0
        discounts_removed = min(discounts_requested_total, subtotal_base)
        discounts_blocked = discounts_removed > 0
        warnings.append("Con tarjeta no aplican descuentos ni promociones; se recalculará el total.")

    subtotal_after_discount = round(max(subtotal_base - discounts_applied, 0.0), 2)

    retention_rate = _retention_rate_for_settlement(
        subtotal_after_discount,
        retention_profile,
        retention_rate_hint=retention_rate_hint,
    )
    retention_amount = round(subtotal_after_discount * retention_rate, 2)
    retention_required = retention_amount > 0

    if retention_required and fmt != "letter":
        fmt = "letter"
        warnings.append("Se forzó factura carta porque hay retención IR (1%/2%).")

    normalized_iva_rate = max(float(iva_rate_percent or 0.0), 0.0) / 100.0

    if fmt == "letter":
        iva_rate = normalized_iva_rate
        iva_forced = True
    else:
        iva_forced = False
        iva_rate = normalized_iva_rate if bool(apply_iva) else 0.0

    iva_amount = round(subtotal_after_discount * iva_rate, 2)
    total_legal = round(subtotal_after_discount + iva_amount, 2)
    net_to_collect = round(total_legal - retention_amount, 2)

    card_withholding_expected = round(total_legal * 0.015, 2) if _is_card_method(method) else 0.0

    if discounts_applied > 0:
        discount_view = "applied"
    elif discounts_removed > 0:
        discount_view = "removed_by_payment_method"
    else:
        discount_view = "none"

    return {
        "payment_method": method,
        "print_format": fmt,
        "subtotal_base": subtotal_base,
        "discounts_requested_total": discounts_requested_total,
        "discounts_applied_amount": round(discounts_applied, 2),
        "discounts_removed_amount": round(discounts_removed, 2),
        "discounts_blocked_by_method": discounts_blocked,
        "discount_view": discount_view,
        "subtotal_after_discount": subtotal_after_discount,
        "iva_rate": iva_rate,
        "iva_amount": iva_amount,
        "iva_forced": iva_forced,
        "total_legal": total_legal,
        "retention_profile": _normalize_retention_profile(retention_profile),
        "retention_rate": retention_rate,
        "retention_amount": retention_amount,
        "retention_receipt_required": retention_required,
        "net_to_collect": net_to_collect,
        "pos_bank_withholding_expected": card_withholding_expected,
        "warnings": warnings,
    }


class CheckoutRequest(FlexibleModel):
    sale_id: str
    origin_url: str


class WorkOrder(FlexibleModel):
    work_order_id: Optional[str] = None
    sale_id: Optional[str] = None
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    vehicle_id: Optional[str] = None
    vehicle_info: Optional[str] = None
    branch_id: Optional[str] = None
    items: List[Dict[str, Any]] = Field(default_factory=list)
    status: str = "pending"
    priority: str = "normal"
    estimated_time: Optional[int] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class WorkOrderCreate(FlexibleModel):
    sale_id: Optional[str] = None
    customer_id: Optional[str] = None
    vehicle_id: Optional[str] = None
    items: List[Dict[str, Any]] = Field(default_factory=list)
    priority: str = "normal"
    estimated_time: Optional[int] = None
    notes: Optional[str] = None


class WorkOrderStatusUpdate(FlexibleModel):
    status: Optional[str] = None
    technician_id: Optional[str] = None
    quality_score: Optional[float] = None
    quality_notes: Optional[str] = None
    notes: Optional[str] = None


class QualityControlCreate(FlexibleModel):
    work_order_id: str
    overall_rating: Optional[float] = None
    cleanliness_rating: Optional[float] = None
    functionality_rating: Optional[float] = None
    finish_rating: Optional[float] = None
    safety_rating: Optional[float] = None
    checklist: Optional[List[Dict[str, Any]]] = None
    comments: Optional[str] = None
    photos: Optional[List[str]] = None
    approved: bool = False


class Promotion(FlexibleModel):
    pass


def hash_pin(pin: str) -> str:
    return bcrypt.hashpw(pin.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


LOGIN_PIN_LENGTH = 8
ATTENDANCE_PIN_LENGTH = 4
PIN_LENGTH = ATTENDANCE_PIN_LENGTH


def compute_pin_index(pin: str) -> str:
    return hashlib.sha256(pin.encode("utf-8")).hexdigest()


def verify_pin_hash(pin: str, pin_hash: Optional[str]) -> bool:
    if not pin_hash:
        return False
    try:
        return bcrypt.checkpw(pin.encode("utf-8"), pin_hash.encode("utf-8"))
    except Exception:
        return False


def get_attendance_pin_hash(user_doc: Dict[str, Any]) -> Optional[str]:
    return user_doc.get("attendance_pin_hash") or user_doc.get("pin_hash")


def get_login_pin_hash(user_doc: Dict[str, Any]) -> Optional[str]:
    return user_doc.get("login_pin_hash")


def is_valid_login_pin(pin: Optional[str]) -> bool:
    return isinstance(pin, str) and pin.isdigit() and len(pin) == LOGIN_PIN_LENGTH


def is_valid_attendance_pin(pin: Optional[str]) -> bool:
    return isinstance(pin, str) and pin.isdigit() and len(pin) == ATTENDANCE_PIN_LENGTH


def generate_kiosk_pin() -> str:
    return "".join([str(secrets.randbelow(10)) for _ in range(ATTENDANCE_PIN_LENGTH)])


async def touch_kiosk_pin_sync_marker(actor_user_id: Optional[str] = None) -> str:
    marker = datetime.now(timezone.utc).isoformat()
    await db.hr_runtime_meta.update_one(
        {"key": "pin_directory_sync"},
        {
            "$set": {
                "key": "pin_directory_sync",
                "marker": marker,
                "updated_at": marker,
                "updated_by": actor_user_id or "system",
            }
        },
        upsert=True,
    )
    return marker


async def ensure_runtime_indexes() -> None:
    """Create indexes that keep critical workflows idempotent and fast."""
    try:
        await db.dispatch_orders.create_index("sale_id", unique=True, sparse=True)
    except Exception as exc:
        logger.warning("Could not create dispatch_orders.sale_id unique index: %s", exc)

    try:
        await db.notifications.create_index("dedupe_key", unique=True, sparse=True)
    except Exception as exc:
        logger.warning("Could not create notifications.dedupe_key unique index: %s", exc)

    try:
        await db.sales.create_index(
            [("salesperson_id", 1), ("idempotency_key", 1)],
            unique=True,
            partialFilterExpression={"idempotency_key": {"$type": "string"}},
        )
    except Exception as exc:
        logger.warning(
            "Could not create sales idempotency unique index (salesperson_id,idempotency_key): %s",
            exc,
        )

    try:
        await db.hypervisor_events.create_index("timestamp")
        await db.hypervisor_events.create_index([("actor_user_id", 1), ("timestamp", -1)])
        await db.hypervisor_events.create_index([("entity_type", 1), ("timestamp", -1)])
        await db.hypervisor_events.create_index([("action", 1), ("timestamp", -1)])
    except Exception as exc:
        logger.warning("Could not create HyperVisor indexes: %s", exc)

    try:
        await db.price_history.create_index([("product_id", 1), ("changed_at", -1)])
        await db.price_history.create_index("changed_by")
    except Exception as exc:
        logger.warning("Could not create price history indexes: %s", exc)

    try:
        await db.user_drafts.create_index(
            [("user_id", 1), ("flow", 1), ("draft_id", 1)],
            unique=True,
        )
        await db.user_drafts.create_index([("user_id", 1), ("flow", 1), ("updated_at", -1)])
        await db.user_draft_state.create_index(
            [("user_id", 1), ("flow", 1)],
            unique=True,
        )
    except Exception as exc:
        logger.warning("Could not create user draft indexes: %s", exc)


async def apply_core_seed_data(
    overwrite: bool = False,
    only_when_empty: bool = False,
) -> Dict[str, Any]:
    """Apply core seed data from file.

    - overwrite=False: only insert missing rows (safe for button-driven recovery)
    - overwrite=True: upsert and update existing rows by unique key
    - only_when_empty=True: skip unless users/customers/inventory are empty
    """
    report: Dict[str, Any] = {
        "applied": False,
        "reason": None,
        "file": str(CORE_SEED_FILE),
        "overwrite": overwrite,
        "collections": {},
    }

    if only_when_empty:
        users_count = await db.users.estimated_document_count()
        customers_count = await db.customers.estimated_document_count()
        inventory_count = await db.inventory.estimated_document_count()
        if users_count > 0 or customers_count > 0 or inventory_count > 0:
            report["reason"] = (
                f"skipped_not_empty(users={users_count},customers={customers_count},inventory={inventory_count})"
            )
            return report

    if not CORE_SEED_FILE.exists():
        report["reason"] = "seed_file_not_found"
        return report

    payload = json.loads(CORE_SEED_FILE.read_text(encoding="utf-8"))
    collections = payload.get("collections") if isinstance(payload, dict) else None
    if not isinstance(collections, dict):
        report["reason"] = "invalid_seed_format"
        return report

    unique_keys_by_collection: Dict[str, List[str]] = {
        "users": ["user_id"],
        "customers": ["customer_id"],
        "inventory": ["product_id", "warehouse_id"],
        "products": ["product_id"],
        "vehicles": ["vehicle_id"],
        "warehouses": ["warehouse_id"],
        "branches": ["branch_id"],
    }

    report["applied"] = True
    for collection_name, docs in collections.items():
        if not isinstance(collection_name, str) or not isinstance(docs, list):
            continue

        keys = unique_keys_by_collection.get(collection_name, [])
        inserted = 0
        updated = 0
        skipped = 0

        for raw_doc in docs:
            if not isinstance(raw_doc, dict):
                skipped += 1
                continue

            doc = dict(raw_doc)
            doc.pop("_id", None)

            filt: Dict[str, Any] = {}
            if keys:
                for key in keys:
                    value = doc.get(key)
                    if value is None:
                        filt = {}
                        break
                    filt[key] = value

            if not filt:
                skipped += 1
                continue

            existing = await db[collection_name].find_one(filt, {"_id": 1})
            if existing and not overwrite:
                skipped += 1
                continue

            if existing and overwrite:
                await db[collection_name].update_one(filt, {"$set": doc})
                updated += 1
            else:
                await db[collection_name].insert_one(doc)
                inserted += 1

        report["collections"][collection_name] = {
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "total_seed_rows": len(docs),
        }

    return report


async def load_core_seed_if_empty() -> None:
    """Load core business data from seed file when database is empty."""
    autoload_enabled = os.environ.get("AUTOLOAD_CORE_SEED", "true").strip().lower() in {
        "1",
        "true",
        "yes",
        "on",
    }
    if not autoload_enabled:
        logger.info("Core seed autoload disabled by AUTOLOAD_CORE_SEED")
        return

    try:
        core_report = await apply_core_seed_data(overwrite=False, only_when_empty=True)
        reason = core_report.get("reason")
        if reason and str(reason).startswith("skipped_not_empty"):
            logger.info("Skipping core seed autoload (%s)", reason)
            return
        if reason:
            logger.info("Core seed autoload skipped (%s)", reason)
            return
        logger.info("Core seed autoload completed: %s", core_report.get("collections"))
    except Exception:
        logger.exception("Failed autoloading core seed data from %s", CORE_SEED_FILE)


@app.on_event("startup")
async def seed_default_pin_user() -> None:
    await ensure_runtime_indexes()
    await load_core_seed_if_empty()
    try:
        xinon_email = os.environ.get("DEFAULT_PIN_USER_EMAIL", "xinon@local")
        xinon_attendance_pin = os.environ.get("DEFAULT_PIN_USER_PIN", "0101")
        xinon_login_pin = os.environ.get("DEFAULT_LOGIN_PIN_USER_PIN", "01010101")
        now_iso = datetime.now(timezone.utc).isoformat()

        xinon = await db.users.find_one(
            {"$or": [{"email": xinon_email}, {"name": "Xinon"}]}, {"_id": 0}
        )

        if not xinon:
            xinon_id = f"user_{uuid.uuid4().hex[:12]}"
            await db.users.insert_one(
                {
                    "user_id": xinon_id,
                    "email": xinon_email,
                    "name": "Xinon",
                    "role": "gerencia",
                    "is_active": True,
                    "is_pin_user": True,
                    "attendance_pin_hash": hash_pin(xinon_attendance_pin),
                    "attendance_pin_index": compute_pin_index(xinon_attendance_pin),
                    "attendance_pin_last_set_at": now_iso,
                    "kiosk_pin_plain": xinon_attendance_pin,
                    "login_pin_hash": hash_pin(xinon_login_pin),
                    "login_pin_index": compute_pin_index(xinon_login_pin),
                    "login_pin_last_set_at": now_iso,
                    "pin_hash": hash_pin(xinon_attendance_pin),
                    "pin_index": compute_pin_index(xinon_attendance_pin),
                    "pin_last_set_at": now_iso,
                    "failed_pin_attempts": 0,
                    "pin_lockout_until": None,
                    "created_at": now_iso,
                }
            )
            logger.info("Seeded default PIN user Xinon")
            # re-fetch the user we just inserted so later accessors are not None
            xinon = await db.users.find_one({"$or": [{"email": xinon_email}, {"name": "Xinon"}]}, {"_id": 0})
            # continue to also ensure test users for each role exist
        if not xinon:
            logger.warning("Default PIN user Xinon not found after seeding; skipping update")
            return

        updates: Dict[str, Any] = {
            "name": "Xinon",
            "role": "gerencia",
            "is_active": True,
            "is_pin_user": True,
            "pin_lockout_until": None,
            "failed_pin_attempts": 0,
        }

        attendance_pin_hash = get_attendance_pin_hash(xinon)
        if not attendance_pin_hash or not verify_pin_hash(xinon_attendance_pin, attendance_pin_hash):
            updates["attendance_pin_hash"] = hash_pin(xinon_attendance_pin)
            updates["attendance_pin_index"] = compute_pin_index(xinon_attendance_pin)
            updates["attendance_pin_last_set_at"] = now_iso
            updates["kiosk_pin_plain"] = xinon_attendance_pin
            updates["pin_hash"] = updates["attendance_pin_hash"]
            updates["pin_index"] = updates["attendance_pin_index"]
            updates["pin_last_set_at"] = now_iso

        login_pin_hash = get_login_pin_hash(xinon)
        if not login_pin_hash or not verify_pin_hash(xinon_login_pin, login_pin_hash):
            updates["login_pin_hash"] = hash_pin(xinon_login_pin)
            updates["login_pin_index"] = compute_pin_index(xinon_login_pin)
            updates["login_pin_last_set_at"] = now_iso

        await db.users.update_one({"user_id": xinon.get("user_id")}, {"$set": updates})
        logger.info("Ensured default PIN user Xinon is configured")
    except Exception:
        logger.exception("Failed to seed default PIN user Xinon")

    # Additionally, ensure there is at least one test PIN user per role.
    # Assign deterministic test PINs: 111111, 222222, 333333, ... based on role order.
    # Can be disabled with ENABLE_TEST_PIN_SEEDING=false
    enable_test_pin_seeding = os.environ.get("ENABLE_TEST_PIN_SEEDING", "true").strip().lower() in {
        "1",
        "true",
        "yes",
        "on",
    }
    if not enable_test_pin_seeding:
        logger.info("Test PIN user seeding disabled by ENABLE_TEST_PIN_SEEDING")
        return

    try:
        for idx, (role_key, role_def) in enumerate(ROLES.items(), start=1):
            attendance_pin = f"{idx % 10000:04d}"
            login_pin = f"{idx % 100000000:08d}"
            test_email = f"test_{role_key}@local"
            display_name = f"Test {role_def.get('label', role_key)}"

            existing = await db.users.find_one(
                {"$or": [{"email": test_email}, {"name": display_name}]}, {"_id": 0}
            )
            if existing:
                # Ensure PIN is up-to-date for existing test users
                try:
                    updates: Dict[str, Any] = {}
                    if not verify_pin_hash(attendance_pin, get_attendance_pin_hash(existing)):
                        attendance_pin_hash = hash_pin(attendance_pin)
                        attendance_pin_index = compute_pin_index(attendance_pin)
                        updates["attendance_pin_hash"] = attendance_pin_hash
                        updates["attendance_pin_index"] = attendance_pin_index
                        updates["attendance_pin_last_set_at"] = datetime.now(timezone.utc).isoformat()
                        updates["kiosk_pin_plain"] = attendance_pin
                        updates["pin_hash"] = attendance_pin_hash
                        updates["pin_index"] = attendance_pin_index
                        updates["pin_last_set_at"] = updates["attendance_pin_last_set_at"]
                    if not verify_pin_hash(login_pin, get_login_pin_hash(existing)):
                        updates["login_pin_hash"] = hash_pin(login_pin)
                        updates["login_pin_index"] = compute_pin_index(login_pin)
                        updates["login_pin_last_set_at"] = datetime.now(timezone.utc).isoformat()
                    if updates:
                        await db.users.update_one(
                            {"user_id": existing.get("user_id")},
                            {"$set": updates}
                        )
                        logger.info(f"Updated PIN for existing test user {test_email}")
                except Exception:
                    logger.exception("Failed to update PIN for existing test user")
                continue

            user_id = f"user_{uuid.uuid4().hex[:10]}"
            now_iso = datetime.now(timezone.utc).isoformat()
            await db.users.insert_one(
                {
                    "user_id": user_id,
                    "email": test_email,
                    "name": display_name,
                    "role": role_key,
                    "is_active": True,
                    "is_pin_user": True,
                    "attendance_pin_hash": hash_pin(attendance_pin),
                    "attendance_pin_index": compute_pin_index(attendance_pin),
                    "attendance_pin_last_set_at": now_iso,
                    "kiosk_pin_plain": attendance_pin,
                    "login_pin_hash": hash_pin(login_pin),
                    "login_pin_index": compute_pin_index(login_pin),
                    "login_pin_last_set_at": now_iso,
                    "pin_hash": hash_pin(attendance_pin),
                    "pin_index": compute_pin_index(attendance_pin),
                    "pin_last_set_at": now_iso,
                    "failed_pin_attempts": 0,
                    "pin_lockout_until": None,
                    "created_at": now_iso,
                }
            )
            logger.info(
                f"Seeded test PIN user for role {role_key} ({test_email}) with attendance PIN {attendance_pin} and login PIN {login_pin}"
            )

        # Seed one humanized leader per branch for each requested role.
        branch_docs = await db.branches.find({}, {"_id": 0, "branch_id": 1, "name": 1}).to_list(100)
        if not branch_docs:
            branch_docs = [
                {"branch_id": "branch_main", "name": "Mundo de Accesorios"},
                {"branch_id": "branch_north", "name": "TopCar El Calvario"},
                {"branch_id": "branch_south", "name": "TopCar La Tigre"},
            ]

        leader_templates = {
            "jefe_vendedores": [
                "Carlos Mena",
                "Andrea Ruiz",
                "Mauricio Pineda",
                "Laura Brenes",
            ],
            "jefe_tienda": [
                "Patricia Rivas",
                "Miguel Larios",
                "Daniela Solis",
                "Roberto Palma",
            ],
        }

        for role_key, names in leader_templates.items():
            for idx, branch in enumerate(branch_docs):
                branch_id = str(branch.get("branch_id") or "").strip()
                if not branch_id:
                    continue
                branch_slug = re.sub(r"[^a-z0-9]+", "_", branch_id.lower()).strip("_") or "branch"
                display_name = names[idx % len(names)]
                attendance_pin = f"{(8400 + idx + (100 if role_key == 'jefe_tienda' else 0)) % 10000:04d}"
                login_pin = f"{(84000000 + idx + (1000000 if role_key == 'jefe_tienda' else 0)) % 100000000:08d}"
                email = f"{role_key}_{branch_slug}@local"

                existing = await db.users.find_one(
                    {
                        "$or": [
                            {"email": email},
                            {"role": role_key, "branch_id": branch_id},
                        ]
                    },
                    {"_id": 0},
                )

                leader_doc: Dict[str, Any] = {
                    "name": display_name,
                    "role": role_key,
                    "branch_id": branch_id,
                    "email": email,
                    "is_active": True,
                    "is_pin_user": True,
                    "failed_pin_attempts": 0,
                    "pin_lockout_until": None,
                }

                if existing:
                    updates = dict(leader_doc)
                    if not verify_pin_hash(attendance_pin, get_attendance_pin_hash(existing)):
                        updates["attendance_pin_hash"] = hash_pin(attendance_pin)
                        updates["attendance_pin_index"] = compute_pin_index(attendance_pin)
                        updates["attendance_pin_last_set_at"] = datetime.now(timezone.utc).isoformat()
                        updates["kiosk_pin_plain"] = attendance_pin
                        updates["pin_hash"] = updates["attendance_pin_hash"]
                        updates["pin_index"] = updates["attendance_pin_index"]
                        updates["pin_last_set_at"] = updates["attendance_pin_last_set_at"]
                    if not verify_pin_hash(login_pin, get_login_pin_hash(existing)):
                        updates["login_pin_hash"] = hash_pin(login_pin)
                        updates["login_pin_index"] = compute_pin_index(login_pin)
                        updates["login_pin_last_set_at"] = datetime.now(timezone.utc).isoformat()

                    await db.users.update_one({"user_id": existing.get("user_id")}, {"$set": updates})
                    continue

                now_iso = datetime.now(timezone.utc).isoformat()
                await db.users.insert_one(
                    {
                        "user_id": f"user_{uuid.uuid4().hex[:10]}",
                        **leader_doc,
                        "created_at": now_iso,
                        "attendance_pin_hash": hash_pin(attendance_pin),
                        "attendance_pin_index": compute_pin_index(attendance_pin),
                        "attendance_pin_last_set_at": now_iso,
                        "kiosk_pin_plain": attendance_pin,
                        "login_pin_hash": hash_pin(login_pin),
                        "login_pin_index": compute_pin_index(login_pin),
                        "login_pin_last_set_at": now_iso,
                        "pin_hash": hash_pin(attendance_pin),
                        "pin_index": compute_pin_index(attendance_pin),
                        "pin_last_set_at": now_iso,
                    }
                )
                logger.info(
                    "Seeded branch leader user role=%s branch=%s email=%s",
                    role_key,
                    branch_id,
                    email,
                )
    except Exception:
        logger.exception("Failed to seed test PIN users for roles")


async def _get_user_by_session(session_token: str) -> Optional[Dict[str, Any]]:
    session = await db.sessions.find_one({"session_token": session_token}, {"_id": 0})
    if not session:
        return None
    user_id = session.get("user_id")
    if not user_id:
        return None
    return await db.users.find_one({"user_id": user_id}, {"_id": 0})


def _extract_session_token(request: Request) -> Optional[str]:
    token: Optional[str] = None
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.lower().startswith("bearer "):
        token = auth_header.split(" ", 1)[1].strip()
    if not token:
        token = request.cookies.get("session_token")
    return token


async def _validate_login_pin_for_user(
    user_doc: Dict[str, Any],
    pin: str,
    request: Request,
) -> None:
    now = datetime.now(timezone.utc)
    user_id = user_doc.get("user_id")

    lockout_step_size = 3
    base_lockout_seconds = 30

    def compute_lockout_seconds_for_failed_attempts(failed_attempts: int) -> int:
        if failed_attempts < lockout_step_size:
            return 0
        if failed_attempts % lockout_step_size != 0:
            return 0
        step = failed_attempts // lockout_step_size
        return base_lockout_seconds * (2 ** (step - 1))

    lockout_until = user_doc.get("pin_lockout_until")
    if lockout_until:
        try:
            lockout_dt = datetime.fromisoformat(str(lockout_until).replace("Z", "+00:00"))
            if lockout_dt > now:
                policy_seconds = int(max(0, (lockout_dt - now).total_seconds()))
                detail = {
                    "message": "PIN bloqueado. Intente más tarde",
                    "remaining_attempts": 0,
                    "failed_attempts": int(user_doc.get("failed_pin_attempts", 0)),
                    "lockout_until": lockout_until,
                    "lockout_seconds": policy_seconds,
                }
                raise HTTPException(status_code=403, detail=detail)
        except ValueError:
            pass

    if not verify_pin_hash(pin, get_login_pin_hash(user_doc)):
        failed = int(user_doc.get("failed_pin_attempts", 0)) + 1
        update: Dict[str, Any] = {"failed_pin_attempts": failed}
        lockout_seconds = compute_lockout_seconds_for_failed_attempts(failed)
        if lockout_seconds > 0:
            update["pin_lockout_until"] = (now + timedelta(seconds=lockout_seconds)).isoformat()

        await db.users.update_one({"user_id": user_id}, {"$set": update})
        await audit_service.log_pin_auth_attempt(user_id, request.client.host if request.client else "unknown", False)
        remaining = 0
        if lockout_seconds <= 0:
            remaining = lockout_step_size - (failed % lockout_step_size)
            if remaining <= 0:
                remaining = lockout_step_size
        detail = {
            "message": "PIN incorrecto",
            "remaining_attempts": remaining,
            "failed_attempts": failed,
            "lockout_until": update.get("pin_lockout_until", None),
            "lockout_seconds": lockout_seconds,
        }
        if lockout_seconds > 0:
            raise HTTPException(status_code=403, detail=detail)
        raise HTTPException(status_code=401, detail=detail)

    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"failed_pin_attempts": 0, "pin_lockout_until": None}},
    )
    await audit_service.log_pin_auth_attempt(user_id, request.client.host if request.client else "unknown", True)


async def require_auth(request: Request) -> User:
    token = None
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.lower().startswith("bearer "):
        token = auth_header.split(" ", 1)[1].strip()
    if not token:
        token = request.cookies.get("session_token")
    if not token:
        raise HTTPException(status_code=401, detail="Unauthorized")
    user_doc = await _get_user_by_session(token)
    if not user_doc:
        raise HTTPException(status_code=401, detail="Invalid session")
    return User(**user_doc)


async def require_roles(request: Request, roles: List[str]) -> User:
    user = await require_auth(request)
    user_effective_role = resolve_effective_role(user.role)
    allowed_effective_roles = [resolve_effective_role(role) for role in roles]
    if user_effective_role not in allowed_effective_roles:
        raise HTTPException(status_code=403, detail="Forbidden")
    return user


def _sanitize_user_doc(user_doc: Dict[str, Any]) -> Dict[str, Any]:
    sanitized = dict(user_doc)
    for key in (
        "pin_hash",
        "failed_pin_attempts",
        "pin_lockout_until",
        "pin_last_set_at",
    ):
        sanitized.pop(key, None)
    return sanitized


async def _create_session_response(user_doc: Dict[str, Any]) -> JSONResponse:
    user_id = user_doc["user_id"]

    # Single-session policy: invalidate any previous sessions for this user
    await db.sessions.delete_many({"user_id": user_id})

    session_token = secrets.token_hex(16)
    await db.sessions.insert_one(
        {
            "session_token": session_token,
            "user_id": user_id,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        }
    )
    payload = {"user": _sanitize_user_doc(user_doc), "session_token": session_token}
    response = JSONResponse(payload)
    # Use host-only cookie to support localhost and LAN-IP access.
    response.set_cookie(
        "session_token",
        session_token,
        httponly=True,
        samesite="lax",
        secure=False,  # Cambia a True solo si usas HTTPS
        path="/",
    )
    return response


class TestSessionRequest(FlexibleModel):
    email: Optional[str] = None
    name: Optional[str] = None
    role: Optional[str] = None


@api_router.post("/test/create-session")
async def create_test_session(payload: Optional[TestSessionRequest] = None):
    if os.environ.get("ENABLE_TEST_ENDPOINTS", "false").lower() not in ("1", "true", "yes"):
        raise HTTPException(status_code=404, detail="Not Found")
    data = payload.model_dump() if payload else {}
    email = data.get("email") or "test.admin@mundodeaccesorios.com"
    name = data.get("name") or "Test Admin"
    role = data.get("role") or "gerencia"

    existing = await db.users.find_one({"email": email}, {"_id": 0})
    user_id = existing.get("user_id") if existing else f"user_{uuid.uuid4().hex[:10]}"
    user_doc = {
        "user_id": user_id,
        "email": email,
        "name": name,
        "role": role,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.update_one({"user_id": user_id}, {"$set": user_doc}, upsert=True)

    session_response = await _create_session_response(user_doc)
    return session_response


_DRAFT_FLOW_ALIASES = {
    "sale": "sale",
    "sales": "sale",
    "venta": "sale",
    "ventas": "sale",
    "quotation": "quotation",
    "quotations": "quotation",
    "quote": "quotation",
    "quotes": "quotation",
    "cotizacion": "quotation",
    "cotizaciones": "quotation",
    "cotización": "quotation",
}

_DRAFT_FLOW_LABELS = {
    "sale": "Venta",
    "quotation": "Cotización",
}


def _normalize_draft_flow(flow: str) -> str:
    normalized = _DRAFT_FLOW_ALIASES.get(str(flow or "").strip().lower())
    if not normalized:
        raise HTTPException(status_code=400, detail="Invalid draft flow")
    return normalized


def _default_draft_name(flow: str, draft_id: str) -> str:
    label = _DRAFT_FLOW_LABELS.get(flow, "Borrador")
    suffix = str(draft_id or "").split("_")[-1][:5]
    return f"{label} {suffix}".strip()


def _serialize_user_draft(doc: Dict[str, Any]) -> Dict[str, Any]:
    flow = _normalize_draft_flow(str(doc.get("flow") or "sale"))
    draft_id = str(doc.get("draft_id") or "").strip()
    raw_snapshot = doc.get("snapshot")
    snapshot: Dict[str, Any] = raw_snapshot if isinstance(raw_snapshot, dict) else {}
    updated_at = (
        str(doc.get("updated_at") or "").strip()
        or str(snapshot.get("updatedAt") or "").strip()
        or str(doc.get("created_at") or "").strip()
        or datetime.now(timezone.utc).isoformat()
    )
    return {
        "id": draft_id,
        "name": str(doc.get("name") or _default_draft_name(flow, draft_id)),
        "updatedAt": updated_at,
        "snapshot": snapshot,
    }


@api_router.get("/drafts/{flow}")
async def list_user_drafts(flow: str, request: Request):
    user = await require_auth(request)
    normalized_flow = _normalize_draft_flow(flow)
    docs = await db.user_drafts.find(
        {"user_id": user.user_id, "flow": normalized_flow},
        {"_id": 0},
    ).sort("updated_at", -1).to_list(100)
    drafts = [_serialize_user_draft(doc) for doc in docs if doc.get("draft_id")]
    state_doc = await db.user_draft_state.find_one(
        {"user_id": user.user_id, "flow": normalized_flow},
        {"_id": 0, "active_draft_id": 1},
    )
    active_draft_id = str((state_doc or {}).get("active_draft_id") or "").strip() or None
    if active_draft_id and not any(draft["id"] == active_draft_id for draft in drafts):
        active_draft_id = None
    if not active_draft_id and drafts:
        active_draft_id = drafts[0]["id"]
    return {
        "flow": normalized_flow,
        "active_draft_id": active_draft_id,
        "drafts": drafts,
    }


@api_router.put("/drafts/{flow}/state")
async def update_user_draft_state(flow: str, payload: DraftStatePayload, request: Request):
    user = await require_auth(request)
    normalized_flow = _normalize_draft_flow(flow)
    active_draft_id = str(payload.active_draft_id or "").strip() or None
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.user_draft_state.update_one(
        {"user_id": user.user_id, "flow": normalized_flow},
        {
            "$set": {
                "user_id": user.user_id,
                "flow": normalized_flow,
                "active_draft_id": active_draft_id,
                "updated_at": now_iso,
            },
            "$setOnInsert": {"created_at": now_iso},
        },
        upsert=True,
    )
    return {"active_draft_id": active_draft_id}


@api_router.put("/drafts/{flow}/{draft_id}")
async def upsert_user_draft(flow: str, draft_id: str, payload: DraftEntryPayload, request: Request):
    user = await require_auth(request)
    normalized_flow = _normalize_draft_flow(flow)
    cleaned_draft_id = str(draft_id or "").strip()
    if not cleaned_draft_id:
        raise HTTPException(status_code=400, detail="Draft id is required")

    snapshot = payload.snapshot if isinstance(payload.snapshot, dict) else {}
    now_iso = str(snapshot.get("updatedAt") or "").strip() or datetime.now(timezone.utc).isoformat()

    existing = await db.user_drafts.find_one(
        {"user_id": user.user_id, "flow": normalized_flow, "draft_id": cleaned_draft_id},
        {"_id": 0, "name": 1},
    )
    draft_name = str(
        payload.name or (existing or {}).get("name") or _default_draft_name(normalized_flow, cleaned_draft_id)
    ).strip()

    await db.user_drafts.update_one(
        {"user_id": user.user_id, "flow": normalized_flow, "draft_id": cleaned_draft_id},
        {
            "$set": {
                "user_id": user.user_id,
                "flow": normalized_flow,
                "draft_id": cleaned_draft_id,
                "name": draft_name,
                "snapshot": snapshot,
                "updated_at": now_iso,
            },
            "$setOnInsert": {"created_at": now_iso},
        },
        upsert=True,
    )

    saved = await db.user_drafts.find_one(
        {"user_id": user.user_id, "flow": normalized_flow, "draft_id": cleaned_draft_id},
        {"_id": 0},
    )
    return _serialize_user_draft(saved or {
        "flow": normalized_flow,
        "draft_id": cleaned_draft_id,
        "name": draft_name,
        "snapshot": snapshot,
        "updated_at": now_iso,
    })


@api_router.delete("/drafts/{flow}/{draft_id}")
async def delete_user_draft(flow: str, draft_id: str, request: Request):
    user = await require_auth(request)
    normalized_flow = _normalize_draft_flow(flow)
    cleaned_draft_id = str(draft_id or "").strip()
    if not cleaned_draft_id:
        raise HTTPException(status_code=400, detail="Draft id is required")

    await db.user_drafts.delete_one(
        {"user_id": user.user_id, "flow": normalized_flow, "draft_id": cleaned_draft_id}
    )

    state_doc = await db.user_draft_state.find_one(
        {"user_id": user.user_id, "flow": normalized_flow},
        {"_id": 0, "active_draft_id": 1},
    )
    current_active_id = str((state_doc or {}).get("active_draft_id") or "").strip() or None
    next_active_id = current_active_id
    if current_active_id == cleaned_draft_id:
        next_draft = await db.user_drafts.find_one(
            {"user_id": user.user_id, "flow": normalized_flow},
            {"_id": 0, "draft_id": 1},
            sort=[("updated_at", -1)],
        )
        next_active_id = str((next_draft or {}).get("draft_id") or "").strip() or None
        now_iso = datetime.now(timezone.utc).isoformat()
        await db.user_draft_state.update_one(
            {"user_id": user.user_id, "flow": normalized_flow},
            {
                "$set": {
                    "user_id": user.user_id,
                    "flow": normalized_flow,
                    "active_draft_id": next_active_id,
                    "updated_at": now_iso,
                },
                "$setOnInsert": {"created_at": now_iso},
            },
            upsert=True,
        )

    return {"status": "deleted", "active_draft_id": next_active_id}


@api_router.get("/auth/me")
async def get_auth_me(request: Request):
    user = await require_auth(request)
    token = _extract_session_token(request)
    session_locked = False
    if token:
        session_doc = await db.sessions.find_one(
            {"session_token": token},
            {"_id": 0, "locked": 1},
        )
        session_locked = bool((session_doc or {}).get("locked"))
    payload = user.model_dump()
    payload["session_locked"] = session_locked
    return payload


@api_router.post("/auth/logout")
async def auth_logout(request: Request):
    token = _extract_session_token(request)

    if token:
        await db.sessions.delete_many({"session_token": token})

    response = JSONResponse({"message": "Sesión cerrada"})
    response.delete_cookie("session_token")
    return response


@api_router.post("/auth/session/lock")
async def lock_current_session(request: Request):
    user = await require_auth(request)
    token = _extract_session_token(request)
    if not token:
        raise HTTPException(status_code=401, detail="Unauthorized")

    result = await db.sessions.update_one(
        {"session_token": token, "user_id": user.user_id},
        {
            "$set": {
                "locked": True,
                "locked_at": datetime.now(timezone.utc).isoformat(),
            }
        },
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=401, detail="Invalid session")

    return {"locked": True}


@api_router.post("/auth/session/unlock")
async def unlock_current_session(payload: SessionUnlockRequest, request: Request):
    pin = payload.pin.strip()
    if not is_valid_login_pin(pin):
        raise HTTPException(status_code=400, detail="PIN inválido")

    user = await require_auth(request)
    token = _extract_session_token(request)
    if not token:
        raise HTTPException(status_code=401, detail="Unauthorized")

    user_doc = await db.users.find_one(
        {
            "user_id": user.user_id,
            "is_pin_user": True,
            "is_active": True,
        },
        {"_id": 0},
    )
    if not user_doc:
        raise HTTPException(status_code=404, detail="Usuario no encontrado o inactivo")

    await _validate_login_pin_for_user(user_doc, pin, request)

    result = await db.sessions.update_one(
        {"session_token": token, "user_id": user.user_id},
        {
            "$set": {
                "locked": False,
                "unlocked_at": datetime.now(timezone.utc).isoformat(),
            }
        },
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=401, detail="Invalid session")

    return {"locked": False}


@api_router.get("/auth/pin/users")
async def get_pin_users():
    users = await db.users.find(
        {"is_pin_user": True, "is_active": True},
        {
            "_id": 0,
            "pin_hash": 0,
            "pin_index": 0,
            "pin_last_set_at": 0,
            "attendance_pin_hash": 0,
            "attendance_pin_index": 0,
            "attendance_pin_last_set_at": 0,
            "login_pin_hash": 0,
            "login_pin_last_set_at": 0,
            "kiosk_pin_plain": 0,
            "failed_pin_attempts": 0,
            "pin_lockout_until": 0,
        },
    ).to_list(500)
    return users


@api_router.post("/auth/pin/login")

async def login_with_pin(payload: PinLoginRequest, request: Request):
    pin = payload.pin.strip()
    if not is_valid_login_pin(pin):
        raise HTTPException(status_code=400, detail="PIN inválido")

    user_doc: Optional[Dict[str, Any]] = None
    # Si se provee user_id, se valida contra ese usuario para poder aplicar
    # la politica progresiva de bloqueo por intentos fallidos.
    if payload.user_id:
        user_doc = await db.users.find_one({
            "user_id": payload.user_id,
            "is_pin_user": True,
            "is_active": True
        }, {
            "_id": 0
        })
        if user_doc:
            await _validate_login_pin_for_user(user_doc, pin, request)
    else:
        user_doc = None
        async for u in db.users.find({"is_pin_user": True, "is_active": True}, {"_id": 0}):
            if verify_pin_hash(pin, get_login_pin_hash(u)):
                user_doc = u
                break

    if not user_doc:
        await audit_service.log_pin_auth_attempt(None, request.client.host if request.client else "unknown", False)
        raise HTTPException(status_code=401, detail="PIN incorrecto")

    return await _create_session_response(user_doc)


# Basic users list (sanitized) - required by frontend admin pages
@api_router.get("/users")
async def list_users(request: Request):
    # Gerencia and Programador may list users in the admin UI
    await require_roles(request, ["gerencia", "programador"])
    users = await db.users.find(
        {},
        {
            "_id": 0,
            "pin_hash": 0,
            "pin_index": 0,
            "pin_last_set_at": 0,
            "attendance_pin_hash": 0,
            "attendance_pin_index": 0,
            "attendance_pin_last_set_at": 0,
            "login_pin_hash": 0,
            "login_pin_index": 0,
            "login_pin_last_set_at": 0,
            "kiosk_pin_plain": 0,
            "failed_pin_attempts": 0,
            "pin_lockout_until": 0,
        },
    ).to_list(1000)
    return users


@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, payload: Dict[str, Any], request: Request):
    await require_roles(request, ["gerencia", "programador", "recursos_humanos"])

    data = payload or {}
    name = data.get("name")
    last_name = data.get("last_name")
    email = data.get("email")
    phone = data.get("phone")
    role = data.get("role")
    if not isinstance(role, str) or not role.strip():
        raise HTTPException(status_code=400, detail="El campo 'role' es requerido")

    if not isinstance(name, str) or not name.strip():
        raise HTTPException(status_code=400, detail="El campo 'name' es requerido")

    if not isinstance(last_name, str) or not last_name.strip():
        raise HTTPException(status_code=400, detail="El campo 'last_name' es requerido")

    email_normalized: Optional[str] = None
    if email not in (None, ""):
        if not isinstance(email, str):
            raise HTTPException(status_code=400, detail="El campo 'email' no es válido")
        email_normalized = email.strip().lower()
        if not EMAIL_RE.match(email_normalized):
            raise HTTPException(status_code=400, detail="El campo 'email' no es válido")

        existing_email_user = await db.users.find_one(
            {"email": email_normalized, "user_id": {"$ne": user_id}},
            {"_id": 0, "user_id": 1},
        )
        if existing_email_user:
            raise HTTPException(status_code=400, detail="El correo ya está en uso por otro usuario")

    if not isinstance(phone, str) or not PHONE_RE.match(phone):
        raise HTTPException(status_code=400, detail="El campo 'phone' es requerido y debe tener formato 0000-0000")
    phone_normalized: str = phone

    role_key = role.strip().lower()
    roles_catalog = await get_roles_catalog()
    if role_key not in roles_catalog:
        raise HTTPException(status_code=400, detail="Rol inválido")

    branch_id = data.get("branch_id")
    warehouse_id = data.get("warehouse_id")
    if not isinstance(branch_id, str) or not branch_id.strip() or branch_id == "none":
        raise HTTPException(status_code=400, detail="El campo 'branch_id' es requerido")

    update_data: Dict[str, Any] = {
        "name": name.strip(),
        "last_name": last_name.strip(),
        "email": email_normalized,
        "phone": phone_normalized,
        "role": role_key,
        "branch_id": branch_id.strip(),
        "warehouse_id": warehouse_id if warehouse_id not in ("", "none") else None,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    result = await db.users.update_one({"user_id": user_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    user_doc = await db.users.find_one(
        {"user_id": user_id},
        {
            "_id": 0,
            "pin_hash": 0,
            "pin_index": 0,
            "pin_last_set_at": 0,
            "attendance_pin_hash": 0,
            "attendance_pin_index": 0,
            "attendance_pin_last_set_at": 0,
            "login_pin_hash": 0,
            "login_pin_index": 0,
            "login_pin_last_set_at": 0,
            "kiosk_pin_plain": 0,
            "failed_pin_attempts": 0,
            "pin_lockout_until": 0,
        },
    )
    if not user_doc:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return user_doc


# Create a PIN user (POST /api/users/pin)
@api_router.post("/users/pin")
async def create_pin_user(payload: Dict[str, Any], request: Request):
    await require_roles(request, ["gerencia", "programador", "recursos_humanos"])

    data = payload or {}
    name = data.get("name")
    if not name or not isinstance(name, str) or not name.strip():
        raise HTTPException(status_code=400, detail="El campo 'name' es requerido")

    last_name = data.get("last_name")
    if not last_name or not isinstance(last_name, str) or not last_name.strip():
        raise HTTPException(status_code=400, detail="El campo 'last_name' es requerido")

    phone = data.get("phone")
    if not isinstance(phone, str) or not PHONE_RE.match(phone):
        raise HTTPException(status_code=400, detail="El campo 'phone' es requerido y debe tener formato 0000-0000")

    email = data.get("email")
    role = data.get("role")
    if not isinstance(role, str) or not role.strip():
        raise HTTPException(status_code=400, detail="El campo 'role' es requerido")

    branch_id = data.get("branch_id")
    if not isinstance(branch_id, str) or not branch_id.strip() or branch_id == "none":
        raise HTTPException(status_code=400, detail="El campo 'branch_id' es requerido")

    warehouse_id = data.get("warehouse_id")
    attendance_pin = data.get("pin")
    login_pin = data.get("login_pin")
    if not is_valid_login_pin(login_pin):
        raise HTTPException(status_code=400, detail="El campo 'login_pin' es requerido y debe ser de 8 dígitos")
    login_pin_value = str(login_pin)

    roles_catalog = await get_roles_catalog()
    role_key = role.strip().lower()
    if role_key not in roles_catalog:
        raise HTTPException(status_code=400, detail="Rol inválido")

    email_normalized: Optional[str] = None
    if email not in (None, ""):
        if not isinstance(email, str):
            raise HTTPException(status_code=400, detail="El campo 'email' no es válido")
        email_normalized = email.strip().lower()
        if not EMAIL_RE.match(email_normalized):
            raise HTTPException(status_code=400, detail="El campo 'email' no es válido")
        existing_email_user = await db.users.find_one({"email": email_normalized}, {"_id": 0, "user_id": 1})
        if existing_email_user:
            raise HTTPException(status_code=400, detail="El correo ya está en uso por otro usuario")

    user_id = f"user_{uuid.uuid4().hex[:12]}"
    now_iso = datetime.now(timezone.utc).isoformat()

    doc: Dict[str, Any] = {
        "user_id": user_id,
        "name": name.strip(),
        "last_name": last_name.strip(),
        "phone": phone,
        "email": email_normalized,
        "role": role_key,
        "branch_id": branch_id.strip(),
        "warehouse_id": warehouse_id,
        "is_active": True,
        "is_pin_user": True,
        "created_at": now_iso,
    }

    attendance_pin_value: Optional[str] = str(attendance_pin) if is_valid_attendance_pin(attendance_pin) else None

    try:
        if attendance_pin_value:
            attendance_index = compute_pin_index(attendance_pin_value)
            existing_attendance = await db.users.find_one(
                {
                    "is_pin_user": True,
                    "$or": [
                        {"attendance_pin_index": attendance_index},
                        {"pin_index": attendance_index},
                    ],
                },
                {"_id": 0, "user_id": 1},
            )
            if existing_attendance:
                raise HTTPException(status_code=400, detail="El PIN de marcación ya está en uso")
        else:
            max_attempts = 20
            attendance_index = ""
            for _ in range(max_attempts):
                candidate_pin = generate_kiosk_pin()
                candidate_index = compute_pin_index(candidate_pin)
                existing_attendance = await db.users.find_one(
                    {
                        "is_pin_user": True,
                        "$or": [
                            {"attendance_pin_index": candidate_index},
                            {"pin_index": candidate_index},
                        ],
                    },
                    {"_id": 0, "user_id": 1},
                )
                if not existing_attendance:
                    attendance_pin_value = candidate_pin
                    attendance_index = candidate_index
                    break
            if not attendance_pin_value:
                raise HTTPException(status_code=500, detail="No se pudo generar PIN de marcación único")

        login_index = compute_pin_index(login_pin_value)
        existing_login = await db.users.find_one(
            {"is_pin_user": True, "login_pin_index": login_index},
            {"_id": 0, "user_id": 1},
        )
        if existing_login:
            raise HTTPException(status_code=400, detail="El PIN de inicio de sesión ya está en uso")
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=500, detail="Error verificando PIN existente")

    attendance_hash = hash_pin(attendance_pin_value)
    doc["attendance_pin_hash"] = attendance_hash
    doc["attendance_pin_index"] = attendance_index
    doc["attendance_pin_last_set_at"] = now_iso
    doc["kiosk_pin_plain"] = attendance_pin_value
    doc["pin_hash"] = attendance_hash
    doc["pin_index"] = attendance_index
    doc["pin_last_set_at"] = now_iso

    doc["login_pin_hash"] = hash_pin(login_pin_value)
    doc["login_pin_index"] = login_index
    doc["login_pin_last_set_at"] = now_iso

    await db.users.insert_one(doc)
    # Return sanitized document
    doc.pop("pin_hash", None)
    doc.pop("attendance_pin_hash", None)
    doc.pop("login_pin_hash", None)
    doc.pop("_id", None)
    await touch_kiosk_pin_sync_marker(user_id)
    return doc


@api_router.put("/users/{user_id}/pin")
async def update_user_pin(user_id: str, payload: Dict[str, Any], request: Request):
    """Update attendance PIN (4 digits) for a user."""
    current_user = await require_auth(request)

    # Only gerencia (or RRHH when feature flag is enabled) or the user themselves can change PIN
    if current_user.user_id != user_id and not can_manage_other_users_pin(current_user):
        raise HTTPException(status_code=403, detail="No tienes permiso para cambiar este PIN")

    new_pin = (payload or {}).get("new_pin")
    if not is_valid_attendance_pin(new_pin):
        raise HTTPException(status_code=400, detail="El PIN de marcación debe ser de 4 dígitos")
    new_pin_value = str(new_pin)

    # Ensure the new PIN is not already used by another PIN user
    try:
        existing = await db.users.find_one(
            {
                "is_pin_user": True,
                "user_id": {"$ne": user_id},
                "$or": [
                    {"attendance_pin_index": compute_pin_index(new_pin_value)},
                    {"pin_index": compute_pin_index(new_pin_value)},
                ],
            },
            {"_id": 0, "user_id": 1},
        )
        if existing:
            raise HTTPException(status_code=400, detail="El PIN de marcación ya está en uso")
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=500, detail="Error verificando PIN existente")

    result = await db.users.update_one(
        {"user_id": user_id, "is_pin_user": True},
        {
            "$set": {
                "attendance_pin_hash": hash_pin(new_pin_value),
                "attendance_pin_index": compute_pin_index(new_pin_value),
                "attendance_pin_last_set_at": datetime.now(timezone.utc).isoformat(),
                "kiosk_pin_plain": new_pin_value,
                "pin_hash": hash_pin(new_pin_value),
                "pin_index": compute_pin_index(new_pin_value),
                "pin_last_set_at": datetime.now(timezone.utc).isoformat(),
                "failed_pin_attempts": 0,
            },
            "$unset": {"pin_lockout_until": ""},
        },
    )

    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado o no es usuario PIN")

    await touch_kiosk_pin_sync_marker(current_user.user_id)

    return {"message": "PIN de marcación actualizado exitosamente"}


@api_router.get("/users/pin/kiosk-table")
async def get_kiosk_pins_table(request: Request):
    current_user = await require_auth(request)
    if not can_manage_login_pin(current_user) and not can_manage_other_users_pin(current_user):
        raise HTTPException(status_code=403, detail="No tienes permiso para ver PIN Kiosko")

    users = await db.users.find(
        {"is_pin_user": True, "is_active": True},
        {
            "_id": 0,
            "user_id": 1,
            "name": 1,
            "last_name": 1,
            "role": 1,
            "branch_id": 1,
            "kiosk_pin_plain": 1,
            "attendance_pin_last_set_at": 1,
            "pin_last_set_at": 1,
        },
    ).to_list(3000)
    return [
        {
            "user_id": item.get("user_id"),
            "name": item.get("name"),
            "last_name": item.get("last_name"),
            "role": item.get("role"),
            "branch_id": item.get("branch_id"),
            "kiosk_pin": item.get("kiosk_pin_plain"),
            "pin_last_set_at": item.get("attendance_pin_last_set_at") or item.get("pin_last_set_at"),
        }
        for item in users
    ]


@api_router.post("/users/pin/kiosk/seed")
async def seed_kiosk_pins_for_testing(payload: Dict[str, Any], request: Request):
    current_user = await require_auth(request)
    if not can_manage_login_pin(current_user) and not can_manage_other_users_pin(current_user):
        raise HTTPException(status_code=403, detail="No tienes permiso para generar PIN Kiosko")

    reset_all = bool((payload or {}).get("reset_all", True))
    now_iso = datetime.now(timezone.utc).isoformat()

    users = await db.users.find(
        {"is_pin_user": True},
        {
            "_id": 0,
            "user_id": 1,
            "name": 1,
            "role": 1,
            "branch_id": 1,
            "is_active": 1,
            "kiosk_pin_plain": 1,
            "attendance_pin_index": 1,
            "pin_index": 1,
        },
    ).to_list(5000)

    used_indexes = {
        str(item.get("attendance_pin_index") or item.get("pin_index"))
        for item in users
        if item.get("attendance_pin_index") or item.get("pin_index")
    }

    table_rows: List[Dict[str, Any]] = []
    for item in users:
        if not item.get("is_active"):
            continue
        current_pin = item.get("kiosk_pin_plain")
        should_regenerate = reset_all or not current_pin

        pin_value = str(current_pin) if current_pin else ""
        if should_regenerate:
            attempts = 0
            while True:
                attempts += 1
                if attempts > 20000:
                    raise HTTPException(status_code=500, detail="No se pudo generar un PIN único")
                candidate = generate_kiosk_pin()
                candidate_index = compute_pin_index(candidate)
                if candidate_index not in used_indexes:
                    pin_value = candidate
                    used_indexes.add(candidate_index)
                    break
        pin_index = compute_pin_index(pin_value)

        await db.users.update_one(
            {"user_id": item.get("user_id")},
            {
                "$set": {
                    "attendance_pin_hash": hash_pin(pin_value),
                    "attendance_pin_index": pin_index,
                    "attendance_pin_last_set_at": now_iso,
                    "kiosk_pin_plain": pin_value,
                    "pin_hash": hash_pin(pin_value),
                    "pin_index": pin_index,
                    "pin_last_set_at": now_iso,
                }
            },
        )

        table_rows.append(
            {
                "user_id": item.get("user_id"),
                "name": item.get("name"),
                "last_name": item.get("last_name"),
                "role": item.get("role"),
                "branch_id": item.get("branch_id"),
                "kiosk_pin": pin_value,
                "pin_last_set_at": now_iso,
            }
        )

    await touch_kiosk_pin_sync_marker(current_user.user_id)

    return {
        "message": "PIN Kiosko generados correctamente",
        "count": len(table_rows),
        "rows": table_rows,
    }


@api_router.put("/users/{user_id}/login-pin")
async def update_user_login_pin(user_id: str, payload: Dict[str, Any], request: Request):
    """Update login PIN (8 digits). Only RRHH, Gerencia or Programador roles are allowed."""
    current_user = await require_auth(request)
    if not can_manage_login_pin(current_user):
        raise HTTPException(status_code=403, detail="No tienes permiso para cambiar el PIN de inicio de sesión")

    new_pin = (payload or {}).get("new_pin")
    if not is_valid_login_pin(new_pin):
        raise HTTPException(status_code=400, detail="El PIN de inicio de sesión debe ser de 8 dígitos")
    new_pin_value = str(new_pin)

    try:
        existing = await db.users.find_one(
            {"is_pin_user": True, "user_id": {"$ne": user_id}, "login_pin_index": compute_pin_index(new_pin_value)},
            {"_id": 0, "user_id": 1},
        )
        if existing:
            raise HTTPException(status_code=400, detail="El PIN de inicio de sesión ya está en uso")
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=500, detail="Error verificando PIN existente")

    result = await db.users.update_one(
        {"user_id": user_id, "is_pin_user": True},
        {
            "$set": {
                "login_pin_hash": hash_pin(new_pin_value),
                "login_pin_index": compute_pin_index(new_pin_value),
                "login_pin_last_set_at": datetime.now(timezone.utc).isoformat(),
                "failed_pin_attempts": 0,
            },
            "$unset": {"pin_lockout_until": ""},
        },
    )

    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado o no es usuario PIN")

    return {"message": "PIN de inicio de sesión actualizado exitosamente"}


# ============ TECHNICIANS CRUD (admin) ============


@api_router.get("/technicians")
async def list_technicians(request: Request):
    """List all technicians (admin only)"""
    await require_roles(request, ["gerencia"])
    # Return users flagged as technicians or with role indicating installation
    technicians = await db.users.find(
        {"$or": [{"is_technician": True}, {"role": {"$in": ["instalaciones", "electrico", "instalador", "polarizador"]}}]},
        {
            "_id": 0,
            "pin_hash": 0,
            "pin_index": 0,
            "pin_last_set_at": 0,
            "attendance_pin_hash": 0,
            "attendance_pin_index": 0,
            "attendance_pin_last_set_at": 0,
            "login_pin_hash": 0,
            "login_pin_index": 0,
            "login_pin_last_set_at": 0,
            "failed_pin_attempts": 0,
            "pin_lockout_until": 0,
        },
    ).to_list(1000)
    return technicians


@api_router.post("/technicians")
async def create_technician(payload: Dict[str, Any], request: Request):
    """Create a new technician user (admin only)"""
    await require_roles(request, ["gerencia"])
    data = payload or {}
    name = data.get("name")
    email = data.get("email")
    specialty = data.get("specialty")

    if not name or not isinstance(name, str) or not name.strip():
        raise HTTPException(status_code=400, detail="El campo 'name' es requerido")
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    now_iso = datetime.now(timezone.utc).isoformat()

    doc = {
        "user_id": user_id,
        "name": name.strip(),
        "email": email,
        "role": "instalaciones",
        "specialty": specialty,
        "is_active": True,
        "is_technician": True,
        "created_at": now_iso,
    }

    await db.users.insert_one(doc)
    doc.pop("_id", None)
    # hide sensitive fields
    doc.pop("pin_hash", None)
    return doc


@api_router.put("/technicians/{user_id}")
async def update_technician(user_id: str, updates: Dict[str, Any], request: Request):
    """Update technician fields (admin only)"""
    await require_roles(request, ["gerencia"])
    # Only allow name and specialty updates for now
    allowed = {k: v for k, v in (updates or {}).items() if k in ("name", "specialty", "email")}
    if not allowed:
        raise HTTPException(status_code=400, detail="No valid fields to update")

    result = await db.users.update_one({"user_id": user_id, "is_technician": True}, {"$set": allowed})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Technician not found")

    tech = await db.users.find_one(
        {"user_id": user_id},
        {
            "_id": 0,
            "pin_hash": 0,
            "pin_index": 0,
            "pin_last_set_at": 0,
            "attendance_pin_hash": 0,
            "attendance_pin_index": 0,
            "attendance_pin_last_set_at": 0,
            "login_pin_hash": 0,
            "login_pin_index": 0,
            "login_pin_last_set_at": 0,
            "failed_pin_attempts": 0,
            "pin_lockout_until": 0,
        },
    )
    return tech


@api_router.delete("/technicians/{user_id}")
async def delete_technician(user_id: str, request: Request):
    """Delete a technician user (admin only)"""
    await require_roles(request, ["gerencia"])
    result = await db.users.delete_one({"user_id": user_id, "is_technician": True})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Technician not found")
    return {"message": "Technician deleted", "user_id": user_id}


@api_router.delete("/users/pin/{user_id}")
async def delete_pin_user(user_id: str, request: Request):
    """Delete a PIN user (gerencia only)"""
    current_user = await require_roles(request, ["gerencia", "programador"])

    if current_user.user_id == user_id:
        raise HTTPException(status_code=400, detail="No puedes eliminar tu propio usuario")

    result = await db.users.delete_one({"user_id": user_id, "is_pin_user": True})

    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Usuario PIN no encontrado")

    # Also delete their sessions if present
    try:
        await db.sessions.delete_many({"user_id": user_id})
    except Exception:
        pass
    try:
        await db.user_sessions.delete_many({"user_id": user_id})
    except Exception:
        pass

    await touch_kiosk_pin_sync_marker(current_user.user_id)

    return {"message": "Usuario PIN eliminado"}


@api_router.post("/users/{user_id}/pin/reset")
async def reset_user_pin(user_id: str, payload: Dict[str, Any], request: Request):
    """Admin endpoint to reset a user's login PIN. Returns the new PIN to the caller (admin)."""
    current_user = await require_auth(request)
    if not can_manage_login_pin(current_user):
        raise HTTPException(status_code=403, detail="No tienes permiso para cambiar este PIN")

    new_pin = (payload or {}).get("new_pin")
    if new_pin:
        if not is_valid_login_pin(new_pin):
            raise HTTPException(status_code=400, detail="El PIN debe ser de 8 dígitos")
    else:
        new_pin = "".join([str(secrets.randbelow(10)) for _ in range(LOGIN_PIN_LENGTH)])
    new_pin_value = str(new_pin)

    result = await db.users.update_one(
        {"user_id": user_id},
        {
            "$set": {
                "login_pin_hash": hash_pin(new_pin_value),
                "login_pin_index": compute_pin_index(new_pin_value),
                "is_pin_user": True,
                "login_pin_last_set_at": datetime.now(timezone.utc).isoformat(),
                "failed_pin_attempts": 0,
            },
            "$unset": {"pin_lockout_until": ""},
        },
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    # Audit the reset
    try:
        await db.pin_auth_logs.insert_one(
            {
                "user_id": user_id,
                "admin": current_user.user_id,
                "action": "pin_reset",
                "timestamp": datetime.now(timezone.utc).isoformat(),
            }
        )
    except Exception:
        logger.exception("Failed to write pin reset audit record")

    return {"message": "PIN de inicio reseteado exitosamente", "new_pin": new_pin_value}


_MODEL_STUBS = [
    "Warehouse",
    "Product",
    "ProductCreate",
    "ProductUpdate",
    "Customer",
    "CustomerCreate",
    "Vehicle",
    "VehicleCreate",
    "VehicleUpdate",
    "InventoryUpdate",
    "UserCreate",
    "UserUpdate",
    "PinUserCreate",
    "PinLoginRequest",
    "Quotation",
    "QuotationCreate",
    "QuotationItem",
    "Sale",
    "SaleCreate",
    "SaleItem",
    "SalePayment",
    "WorkOrder",
    "WorkOrderCreate",
    "Branch",
    "BranchCreate",
    "TechnicianCreate",
    "TechnicianUpdate",
    "Technician",
    "CashSessionCreate",
    "CashSessionClose",
    "CashSession",
    "InventoryTransferRequest",
    "InventoryTransferApproval",
    "AuditLogFilter",
    "QualityControlCreate",
    "QualityControlUpdate",
    "PinPolicyUpdate",
]

for _name in _MODEL_STUBS:
    if _name not in globals():
        globals()[_name] = type(_name, (FlexibleModel,), {})
# ============ CUSTOMERS ============


@api_router.get("/warehouses")
async def get_warehouses(request: Request):
    await require_auth(request)
    warehouses = await db.warehouses.find({}, {"_id": 0}).to_list(100)
    return warehouses


async def get_branch_with_policy(branch_id: Optional[str]) -> Dict[str, Any]:
    resolved_branch_id = str(branch_id or "branch_main")
    branch_doc = await db.branches.find_one({"branch_id": resolved_branch_id}, {"_id": 0})
    branch = dict(branch_doc or {})
    branch.setdefault("branch_id", resolved_branch_id)

    base_policy = dict(
        BRANCH_SERVICE_POLICY_DEFAULTS.get(
            resolved_branch_id,
            {
                "installations_enabled": True,
                "tint_enabled": True,
                "local_delivery_enabled": True,
                "departmental_delivery_enabled": True,
            },
        )
    )
    custom_policy = branch.get("service_policy") if isinstance(branch.get("service_policy"), dict) else {}
    base_policy.update(custom_policy or {})
    branch["service_policy"] = base_policy
    return branch


async def ensure_branch_service_enabled(branch_id: Optional[str], service_key: str, error_message: str) -> Dict[str, Any]:
    branch = await get_branch_with_policy(branch_id)
    policy = branch.get("service_policy") or {}
    if not bool(policy.get(service_key, False)):
        raise HTTPException(status_code=400, detail=error_message)
    return branch


async def get_default_warehouse_for_branch(branch_id: Optional[str]) -> Optional[Dict[str, Any]]:
    resolved_branch_id = str(branch_id or "")
    if not resolved_branch_id:
        return None
    return await db.warehouses.find_one(
        {"branch_id": resolved_branch_id, "is_active": True},
        {"_id": 0},
        sort=[("warehouse_id", 1)],
    )


async def ensure_warehouse_belongs_to_branch(warehouse_id: str, branch_id: Optional[str]) -> Dict[str, Any]:
    warehouse = await db.warehouses.find_one({"warehouse_id": warehouse_id}, {"_id": 0})
    if not warehouse:
        raise HTTPException(status_code=404, detail=f"Bodega {warehouse_id} no encontrada")

    resolved_branch_id = str(branch_id or "")
    warehouse_branch_id = str(warehouse.get("branch_id") or "")
    if resolved_branch_id and warehouse_branch_id and warehouse_branch_id != resolved_branch_id:
        raise HTTPException(
            status_code=400,
            detail=f"La bodega {warehouse_id} no pertenece a la sucursal del usuario",
        )
    return warehouse


@api_router.get("/branches")
async def get_branches(request: Request):
    """Return list of branches (sucursales)."""
    await require_auth(request)
    branches = await db.branches.find({}, {"_id": 0}).to_list(100)
    return branches


@api_router.get("/categories")
async def get_categories(request: Request):
    await require_auth(request)
    categories: Dict[str, Dict[str, List[str]]] = {}
    vehicle_type_set: set[str] = set()
    window_options_set: set[str] = set()

    cursor = db.products.find(
        {},
        {"_id": 0, "category": 1, "subcategory": 1, "compatibility": 1, "vehicle_types": 1},
    )
    async for product in cursor:
        category = product.get("category") or "otros"
        subcategory = product.get("subcategory") or "General"
        categories.setdefault(category, {"subcategories": []})
        if subcategory not in categories[category]["subcategories"]:
            categories[category]["subcategories"].append(subcategory)

        compatibility = product.get("compatibility") or {}
        for vehicle_type in compatibility.get("vehicle_types") or []:
            if vehicle_type:
                vehicle_type_set.add(vehicle_type)
        # collect window options if present (e.g., for polarizados services)
        for wo in product.get("window_options") or []:
            if wo:
                window_options_set.add(wo)
        for vehicle_type in product.get("vehicle_types") or []:
            if vehicle_type:
                vehicle_type_set.add(vehicle_type)

    try:
        vehicle_values = await db.vehicles.distinct("vehicle_type")
        for value in vehicle_values or []:
            if value:
                vehicle_type_set.add(value)
    except Exception:
        pass

    for key in categories:
        categories[key]["subcategories"].sort()

    return {
        "categories": categories,
        "vehicle_types": sorted(vehicle_type_set),
        "window_options": sorted(window_options_set),
    }


@api_router.post("/warehouses")
async def create_warehouse(warehouse: Warehouse, request: Request):
    await require_roles(request, ["gerencia"])
    doc = warehouse.model_dump()
    await db.warehouses.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/warehouses/{warehouse_id}")
async def update_warehouse(warehouse_id: str, updates: Dict[str, Any], request: Request):
    """Actualizar campos de una bodega existente."""
    await require_roles(request, ["gerencia"])
    result = await db.warehouses.update_one({"warehouse_id": warehouse_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    return {"message": "Warehouse updated"}


@api_router.delete("/warehouses/{warehouse_id}")
async def delete_warehouse(warehouse_id: str, request: Request):
    """Eliminar una bodega por su warehouse_id."""
    await require_roles(request, ["gerencia"])
    result = await db.warehouses.delete_one({"warehouse_id": warehouse_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    return {"message": "Warehouse deleted"}


# ============ PRODUCTS ============


@api_router.get("/products")
async def get_products(
    request: Request,
    category: Optional[str] = None,
    subcategory: Optional[str] = None,
    product_type: Optional[str] = None,
    search: Optional[str] = None,
):
    await require_auth(request)
    query: dict[str, Any] = {
        "$or": [
            {"is_active": True},
            {"is_active": {"$exists": False}},
        ]
    }
    if category:
        query["category"] = category
    if subcategory:
        query["subcategory"] = subcategory
    if product_type:
        query["product_type"] = product_type
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"sku": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}},
        ]
    products = await db.products.find(query, {"_id": 0}).to_list(1000)

    # Ensure all products have installation_type (migration for legacy products)
    for product in products:
        if "installation_type" not in product or not product.get("installation_type"):
            product["installation_type"] = "optional"  # Default for legacy products

        # Ensure price tiers exist: precio1, precio2, precio3, precio_vip
        # Backwards-compatible: use existing `price` as precio1 if tiers missing
        try:
            base_price = float(product.get("price", 0) or 0)
        except Exception:
            base_price = 0.0
        if "precio1" not in product:
            product["precio1"] = round(base_price, 2)
        if "precio2" not in product:
            product["precio2"] = round(base_price * 1.05, 2)
        if "precio3" not in product:
            product["precio3"] = round(base_price * 1.1, 2)
        if "precio_vip" not in product:
            product["precio_vip"] = round(base_price * 0.9, 2)

        # Keep `price` field pointing to precio1 for UI compatibility (catalog shows precio1)
        product["price"] = product.get("precio1", base_price)

        raw_low_stock_threshold = product.get("low_stock_threshold", 5)
        try:
            normalized_low_stock_threshold = int(float(raw_low_stock_threshold))
        except Exception:
            normalized_low_stock_threshold = 5
        product["low_stock_threshold"] = max(1, normalized_low_stock_threshold)

    return products


@api_router.post("/products")
async def create_product(product_data: ProductCreate, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor", "bodegas", "jefe_tienda"])
    product = Product(**product_data.model_dump())
    doc = product.model_dump()

    raw_initial_stock = doc.pop("initial_stock", 0)
    raw_initial_warehouse_id = doc.pop("initial_warehouse_id", None)
    try:
        initial_stock = max(0, int(float(raw_initial_stock or 0)))
    except Exception:
        initial_stock = 0
    # Ensure created_at exists and is ISO string
    created_at_val = doc.get("created_at")
    if not created_at_val:
        created_at_dt = datetime.now(timezone.utc)
    else:
        created_at_dt = created_at_val
        if isinstance(created_at_dt, str):
            try:
                created_at_dt = datetime.fromisoformat(created_at_dt)
            except Exception:
                created_at_dt = datetime.now(timezone.utc)

    doc["created_at"] = created_at_dt.isoformat()

    # Ensure a stable product_id is present for tests/consumers
    if not doc.get("product_id"):
        doc["product_id"] = f"product_{uuid.uuid4().hex[:12]}"

    if str(doc.get("category", "")).lower() == "polarizados":
        doc["installation_type"] = "required"
        doc["installation_required"] = True

    # Ensure price tiers exist for newly created products
    try:
        base_price_new = float(doc.get("precio1", doc.get("price", 0)) or 0)
    except Exception:
        base_price_new = 0.0
    try:
        precio1 = float(doc.get("precio1", base_price_new) or base_price_new)
    except Exception:
        precio1 = base_price_new
    try:
        precio2 = float(doc.get("precio2", precio1 * 1.05) or (precio1 * 1.05))
    except Exception:
        precio2 = precio1 * 1.05
    try:
        precio3 = float(doc.get("precio3", precio1 * 1.1) or (precio1 * 1.1))
    except Exception:
        precio3 = precio1 * 1.1

    doc["precio1"] = round(precio1, 2)
    doc["precio2"] = round(precio2, 2)
    doc["precio3"] = round(precio3, 2)
    doc["precio_vip"] = round(base_price_new * 0.9, 2)
    doc["price"] = doc["precio1"]

    raw_low_stock_threshold = doc.get("low_stock_threshold", 5)
    try:
        normalized_low_stock_threshold = int(float(raw_low_stock_threshold))
    except Exception:
        normalized_low_stock_threshold = 5
    doc["low_stock_threshold"] = max(1, normalized_low_stock_threshold)

    await db.products.insert_one(doc)

    if initial_stock > 0:
        warehouse_id = str(raw_initial_warehouse_id or user.warehouse_id or "").strip()
        if not warehouse_id:
            fallback = await db.warehouses.find_one({}, {"_id": 0, "warehouse_id": 1})
            warehouse_id = str((fallback or {}).get("warehouse_id") or "")

        if warehouse_id:
            inv_filter = {"product_id": doc["product_id"], "warehouse_id": warehouse_id}
            existing_inventory = await db.inventory.find_one(inv_filter, {"_id": 0, "inventory_id": 1})
            if existing_inventory:
                await db.inventory.update_one(
                    inv_filter,
                    {
                        "$inc": {"quantity": int(initial_stock)},
                        "$set": {"last_updated": datetime.now(timezone.utc).isoformat()},
                    },
                )
                inventory_id = existing_inventory.get("inventory_id")
            else:
                inventory_id = f"inv_{uuid.uuid4().hex[:8]}"
                await db.inventory.insert_one(
                    {
                        "inventory_id": inventory_id,
                        "product_id": doc["product_id"],
                        "warehouse_id": warehouse_id,
                        "quantity": int(initial_stock),
                        "min_stock": int(doc.get("low_stock_threshold") or 5),
                        "last_updated": datetime.now(timezone.utc).isoformat(),
                    }
                )

            await audit_service.log_inventory_movement(
                product_id=doc["product_id"],
                warehouse_id=warehouse_id,
                quantity_change=int(initial_stock),
                reason="initial_stock_product_create",
                actor=user,
                branch_id=user.branch_id,
                reference_id=inventory_id,
                metadata={"sku": doc.get("sku")},
            )

    # Return authoritative stored document when possible
    stored = await db.products.find_one({"product_id": doc["product_id"]}, {"_id": 0})
    if not stored:
        doc.pop("_id", None)
        logger.info(f"create_product returning (fallback): {list(doc.keys())}")
        return doc

    logger.info(f"create_product returning stored doc for {stored.get('product_id')}")
    return stored


@api_router.get("/products/{product_id}")
async def get_product(product_id: str, request: Request):
    await require_auth(request)
    product = await db.products.find_one({"product_id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Ensure installation_type exists (migration for legacy products)
    if "installation_type" not in product or not product.get("installation_type"):
        product["installation_type"] = "optional"

    # Ensure price tiers exist on single product fetch as well
    try:
        base_price = float(product.get("price", 0) or 0)
    except Exception:
        base_price = 0.0
    if "precio1" not in product:
        product["precio1"] = round(base_price, 2)
    if "precio2" not in product:
        product["precio2"] = round(base_price * 1.05, 2)
    if "precio3" not in product:
        product["precio3"] = round(base_price * 1.1, 2)
    if "precio_vip" not in product:
        product["precio_vip"] = round(base_price * 0.9, 2)

    product["price"] = product.get("precio1", base_price)

    raw_low_stock_threshold = product.get("low_stock_threshold", 5)
    try:
        normalized_low_stock_threshold = int(float(raw_low_stock_threshold))
    except Exception:
        normalized_low_stock_threshold = 5
    product["low_stock_threshold"] = max(1, normalized_low_stock_threshold)

    return product


@api_router.put("/products/{product_id}")
async def update_product(product_id: str, updates: Dict[str, Any], request: Request):
    actor = await require_roles(request, ["gerencia", "supervisor", "bodegas"])
    if str(updates.get("category", "")).lower() == "polarizados":
        updates["installation_type"] = "required"
        updates["installation_required"] = True

    if "precio1" in updates and "price" not in updates:
        updates["price"] = updates["precio1"]
    if "price" in updates and "precio1" not in updates:
        updates["precio1"] = updates["price"]

    if "low_stock_threshold" in updates:
        try:
            updates["low_stock_threshold"] = max(1, int(float(updates.get("low_stock_threshold", 5))))
        except Exception:
            updates["low_stock_threshold"] = 5

    before = await db.products.find_one({"product_id": product_id}, {"_id": 0})
    result = await db.products.update_one({"product_id": product_id}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")

    tracked_price_fields = ["price", "precio1", "precio2", "precio3", "precio_vip", "cost"]
    price_changes: List[Dict[str, Any]] = []
    before_doc = before or {}
    for field in tracked_price_fields:
        if field not in updates:
            continue
        old_value = before_doc.get(field)
        new_value = updates.get(field)
        if old_value != new_value:
            price_changes.append(
                {
                    "field": field,
                    "old_value": old_value,
                    "new_value": new_value,
                }
            )

    if price_changes:
        history_doc = {
            "history_id": f"ph_{uuid.uuid4().hex[:12]}",
            "product_id": product_id,
            "changes": price_changes,
            "changed_at": datetime.now(timezone.utc).isoformat(),
            "changed_by": actor.user_id,
            "changed_by_name": actor.name,
            "changed_by_role": actor.role,
        }
        await db.price_history.insert_one(history_doc)

        await record_hypervisor_event(
            request=request,
            actor=actor,
            action="price_update",
            entity_type="products",
            entity_id=product_id,
            details={"changes": price_changes},
            status_code=200,
        )

    return {"message": "Product updated"}


@api_router.get("/hypervisor/access")
async def hypervisor_access(request: Request):
    user = await require_auth(request)
    mode = get_hypervisor_access_mode(user.role)
    return {
        "enabled": bool(mode),
        "mode": mode,
        "user_id": user.user_id,
        "role": user.role,
    }


@api_router.get("/hypervisor/events")
async def get_hypervisor_events(
    request: Request,
    limit: int = 200,
    actor_user_id: Optional[str] = None,
    actor_name: Optional[str] = None,
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    action: Optional[str] = None,
    from_timestamp: Optional[str] = None,
    to_timestamp: Optional[str] = None,
):
    user = await require_hypervisor_access(request, write=False)
    mode = get_hypervisor_access_mode(user.role)

    safe_limit = max(1, min(limit, 1000))
    conditions: List[Dict[str, Any]] = []

    def add_partial_multi_filter(field_name: str, raw_values: Optional[str]) -> None:
        values = parse_filter_values(raw_values)
        if not values:
            return
        if len(values) == 1:
            conditions.append({field_name: {"$regex": re.escape(values[0]), "$options": "i"}})
            return
        conditions.append(
            {
                "$or": [
                    {field_name: {"$regex": re.escape(value), "$options": "i"}}
                    for value in values
                ]
            }
        )

    if actor_user_id:
        conditions.append({"actor_user_id": actor_user_id})

    add_partial_multi_filter("actor_name", actor_name)
    add_partial_multi_filter("entity_type", entity_type)
    add_partial_multi_filter("entity_id", entity_id)
    add_partial_multi_filter("action", action)

    timestamp_filter: Dict[str, str] = {}
    if from_timestamp:
        try:
            dt_from = datetime.fromisoformat(from_timestamp.replace("Z", "+00:00"))
            if dt_from.tzinfo is None:
                dt_from = dt_from.replace(tzinfo=timezone.utc)
            else:
                dt_from = dt_from.astimezone(timezone.utc)
            timestamp_filter["$gte"] = dt_from.isoformat()
        except ValueError:
            raise HTTPException(status_code=400, detail="from_timestamp inválido")
    if to_timestamp:
        try:
            dt_to = datetime.fromisoformat(to_timestamp.replace("Z", "+00:00"))
            if dt_to.tzinfo is None:
                dt_to = dt_to.replace(tzinfo=timezone.utc)
            else:
                dt_to = dt_to.astimezone(timezone.utc)
            timestamp_filter["$lte"] = dt_to.isoformat()
        except ValueError:
            raise HTTPException(status_code=400, detail="to_timestamp inválido")
    if timestamp_filter:
        conditions.append({"timestamp": timestamp_filter})

    query: Dict[str, Any] = {"$and": conditions} if conditions else {}

    events = await db.hypervisor_events.find(query, {"_id": 0}).sort("timestamp", -1).to_list(safe_limit)
    return {
        "mode": mode,
        "count": len(events),
        "events": events,
    }


@api_router.get("/hypervisor/filter-options")
async def get_hypervisor_filter_options(request: Request, limit: int = 500):
    await require_hypervisor_access(request, write=False)
    safe_limit = max(50, min(limit, 2000))

    users_docs = await db.users.find({}, {"_id": 0, "name": 1}).sort("name", 1).to_list(safe_limit)
    user_names: List[str] = []
    for doc in users_docs:
        name = str(doc.get("name") or "").strip()
        if name and name not in user_names:
            user_names.append(name)

    actions = await db.hypervisor_events.distinct("action")
    entity_types = await db.hypervisor_events.distinct("entity_type")

    recent_ids_docs = await (
        db.hypervisor_events.find({"entity_id": {"$nin": [None, "", "-"]}}, {"_id": 0, "entity_id": 1})
        .sort("timestamp", -1)
        .to_list(safe_limit)
    )
    entity_ids: List[str] = []
    for doc in recent_ids_docs:
        entity_id = str(doc.get("entity_id") or "").strip()
        if entity_id and entity_id not in entity_ids:
            entity_ids.append(entity_id)

    return {
        "users": user_names,
        "actions": sorted([str(v) for v in actions if v]),
        "entity_types": sorted([str(v) for v in entity_types if v]),
        "entity_ids": entity_ids,
    }


@api_router.get("/hypervisor/summary")
async def get_hypervisor_summary(request: Request):
    user = await require_hypervisor_access(request, write=False)
    mode = get_hypervisor_access_mode(user.role)

    now = datetime.now(timezone.utc)
    day_ago = (now - timedelta(days=1)).isoformat()
    week_ago = (now - timedelta(days=7)).isoformat()

    total_events = await db.hypervisor_events.count_documents({})
    events_24h = await db.hypervisor_events.count_documents({"timestamp": {"$gte": day_ago}})
    events_7d = await db.hypervisor_events.count_documents({"timestamp": {"$gte": week_ago}})
    price_changes_7d = await db.price_history.count_documents({"changed_at": {"$gte": week_ago}})

    top_actions = await db.hypervisor_events.aggregate(
        [
            {"$match": {"timestamp": {"$gte": week_ago}}},
            {"$group": {"_id": "$action", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 10},
        ]
    ).to_list(20)

    return {
        "mode": mode,
        "total_events": total_events,
        "events_24h": events_24h,
        "events_7d": events_7d,
        "price_changes_7d": price_changes_7d,
        "top_actions": [
            {"action": row.get("_id"), "count": row.get("count", 0)} for row in top_actions
        ],
    }


@api_router.get("/backup/excel")
async def export_full_backup_excel(
    request: Request,
    scopes: Optional[str] = None,
    secure: bool = True,
):
    await require_roles(request, ["gerencia", "programador", "recursos_humanos"])

    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    available_collections = sorted(await db.list_collection_names())
    selected_scopes = parse_scope_values(scopes)
    targets = get_backup_targets(selected_scopes, available_collections)

    collections = cast(List[str], targets.get("collections") or [])
    include_brands_models = bool(targets.get("include_brands_models"))
    used_sheet_names: set[str] = set()
    manifest_rows: List[Dict[str, Any]] = []

    for collection_name in collections:
        sheet_base = collection_name[:31] or "sheet"
        sheet_name = sheet_base
        suffix = 1
        while sheet_name in used_sheet_names:
            suffix += 1
            sheet_name = f"{sheet_base[:28]}_{suffix}"
        used_sheet_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)

        docs = await db[collection_name].find({}, {"_id": 0}).to_list(200000)
        if not docs:
            ws.append(["info"])
            ws.append(["No data"])
            manifest_rows.append(
                {
                    "sheet": sheet_name,
                    "source": collection_name,
                    "rows": 0,
                    "checksum": compute_sheet_checksum([]),
                }
            )
            continue

        all_keys: List[str] = []
        key_set: set[str] = set()
        for doc in docs:
            if not isinstance(doc, dict):
                continue
            for key in doc.keys():
                if key not in key_set:
                    key_set.add(key)
                    all_keys.append(key)

        export_rows: List[Dict[str, Any]] = []
        for doc in docs:
            clean_doc = {k: normalize_excel_value(doc.get(k)) for k in all_keys}
            row_payload = json.dumps(clean_doc, ensure_ascii=False, sort_keys=True)
            row_checksum = hashlib.sha256(row_payload.encode("utf-8")).hexdigest()
            clean_doc["_row_payload"] = row_payload
            clean_doc["_row_checksum"] = row_checksum
            export_rows.append(clean_doc)

        header = list(all_keys) + ["_row_payload", "_row_checksum"]
        ws.append(header)
        for doc in export_rows:
            row = [doc.get(k) for k in header]
            ws.append(row)

        manifest_rows.append(
            {
                "sheet": sheet_name,
                "source": collection_name,
                "rows": len(export_rows),
                "checksum": compute_sheet_checksum(export_rows),
            }
        )

    if include_brands_models:
        sheet_name = "brands_models"
        if sheet_name in used_sheet_names:
            sheet_name = f"brands_models_{uuid.uuid4().hex[:4]}"
        used_sheet_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)
        rows = await build_brands_models_rows()
        if not rows:
            ws.append(["info"])
            ws.append(["No data"])
            manifest_rows.append(
                {
                    "sheet": sheet_name,
                    "source": "derived:brands_models",
                    "rows": 0,
                    "checksum": compute_sheet_checksum([]),
                }
            )
        else:
            header = ["brand", "model", "vehicle_type", "_row_payload", "_row_checksum"]
            ws.append(header)
            secured_rows: List[Dict[str, Any]] = []
            for row in rows:
                secure_row = {
                    "brand": normalize_excel_value(row.get("brand")),
                    "model": normalize_excel_value(row.get("model")),
                    "vehicle_type": normalize_excel_value(row.get("vehicle_type")),
                }
                row_payload = json.dumps(secure_row, ensure_ascii=False, sort_keys=True)
                secure_row["_row_payload"] = row_payload
                secure_row["_row_checksum"] = hashlib.sha256(row_payload.encode("utf-8")).hexdigest()
                secured_rows.append(secure_row)
                ws.append([secure_row.get(k) for k in header])

            manifest_rows.append(
                {
                    "sheet": sheet_name,
                    "source": "derived:brands_models",
                    "rows": len(secured_rows),
                    "checksum": compute_sheet_checksum(secured_rows),
                }
            )

    if secure:
        manifest_sheet_name = "__MANIFEST"
        manifest = wb.create_sheet(title=manifest_sheet_name)
        manifest.append(["key", "value"])
        manifest.append(["format", "ERP_SECURE_BACKUP_V1"])
        manifest.append(["schema_version", BACKUP_SCHEMA_VERSION])
        manifest.append(["generated_at", datetime.now(timezone.utc).isoformat()])
        manifest.append(["selected_scopes", json.dumps(targets.get("selected_scopes") or [], ensure_ascii=False)])
        manifest.append(["sheet_count", len(manifest_rows)])
        manifest_checksum = hashlib.sha256(
            json.dumps(manifest_rows, ensure_ascii=False, sort_keys=True).encode("utf-8")
        ).hexdigest()
        manifest.append(["manifest_checksum", manifest_checksum])
        manifest.append(["---", "---"])
        manifest.append(["sheet", "source"])  # marker row
        for row in manifest_rows:
            manifest.append(
                [
                    json.dumps(row, ensure_ascii=False),
                    row.get("checksum", ""),
                ]
            )
        manifest.sheet_state = "hidden"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    suffix = "secure" if secure else "plain"
    file_name = f"erp_backup_{suffix}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={file_name}"},
    )


@api_router.post("/backup/excel/import")
async def import_backup_excel(
    request: Request,
    file: UploadFile = File(...),
    scopes: Optional[str] = Form(None),
    overwrite: Optional[str] = Form("false"),
):
    actor = await require_hypervisor_access(request, write=True)

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El respaldo debe ser un archivo .xlsx")

    payload = await file.read()
    if not payload:
        raise HTTPException(status_code=400, detail="Archivo vacío")

    try:
        wb = load_workbook(filename=BytesIO(payload), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel")

    manifest_metadata = read_backup_manifest_metadata(wb)
    schema_version = int(manifest_metadata.get("schema_version") or 1)
    if schema_version > BACKUP_SCHEMA_VERSION:
        raise HTTPException(
            status_code=400,
            detail=(
                f"El respaldo usa schema_version={schema_version} y este sistema soporta hasta {BACKUP_SCHEMA_VERSION}."
            ),
        )

    selected_scopes = parse_scope_values(scopes)
    targets = get_backup_targets(selected_scopes, sorted(await db.list_collection_names()))
    allowed_collections = set(cast(List[str], targets.get("collections") or []))
    include_brands_models = bool(targets.get("include_brands_models")) or not selected_scopes
    overwrite_enabled = parse_string_bool(overwrite, default=False)

    unique_keys_by_collection: Dict[str, List[str]] = {
        "users": ["user_id"],
        "customers": ["customer_id"],
        "vehicles": ["vehicle_id"],
        "products": ["product_id"],
        "inventory": ["product_id", "warehouse_id"],
        "branches": ["branch_id"],
        "warehouses": ["warehouse_id"],
        "user_permissions": ["user_id"],
        "role_permissions": ["role"],
        "inventory_movements": ["movement_id"],
        "product_transfers": ["transfer_id"],
        "dispatch_orders": ["dispatch_id"],
        "notifications": ["notification_id"],
        "hypervisor_events": ["event_id"],
        "price_history": ["history_id"],
        "audit_logs": ["audit_id"],
    }

    report: Dict[str, Any] = {
        "overwrite": overwrite_enabled,
        "selected_scopes": targets.get("selected_scopes") or [],
        "detected_schema_version": schema_version,
        "supported_schema_version": BACKUP_SCHEMA_VERSION,
        "manifest_format": manifest_metadata.get("format"),
        "sheets": {},
    }

    for ws in wb.worksheets:
        name = ws.title
        if name == "__MANIFEST":
            continue

        collection_name = name
        if collection_name not in allowed_collections and not (include_brands_models and name.startswith("brands_models")):
            continue

        rows_iter = list(ws.iter_rows(values_only=True))
        if not rows_iter:
            continue
        header_row = rows_iter[0]
        if not header_row:
            continue
        headers = [str(h) if h is not None else "" for h in header_row]
        if len(headers) == 1 and headers[0] == "info":
            continue

        inserted = 0
        updated = 0
        skipped = 0
        invalid_checksum = 0

        for values in rows_iter[1:]:
            doc = {headers[idx]: values[idx] if idx < len(values) else None for idx in range(len(headers))}
            if not any(v is not None and str(v) != "" for v in doc.values()):
                continue

            row_checksum = str(doc.pop("_row_checksum", "") or "")
            row_payload = doc.pop("_row_payload", None)

            parsed_doc: Dict[str, Any]
            if isinstance(row_payload, str) and row_payload.strip():
                try:
                    parsed_doc = cast(Dict[str, Any], json.loads(row_payload))
                except Exception:
                    parsed_doc = {
                        k: maybe_parse_json_like(v) for k, v in doc.items() if k and not k.startswith("__")
                    }
            else:
                parsed_doc = {
                    k: maybe_parse_json_like(v) for k, v in doc.items() if k and not k.startswith("__")
                }

            if row_checksum:
                if isinstance(row_payload, str) and row_payload.strip():
                    payload_row = row_payload
                else:
                    payload_row = json.dumps(parsed_doc, ensure_ascii=False, sort_keys=True)
                expected = hashlib.sha256(payload_row.encode("utf-8")).hexdigest()
                if expected != row_checksum:
                    # Backward compatibility for old backups that did not include _row_payload:
                    # keep import path and report warning instead of blocking all rows.
                    invalid_checksum += 1

            parsed_doc = migrate_backup_row_for_schema(schema_version, collection_name, parsed_doc)

            if name.startswith("brands_models"):
                brand = str(parsed_doc.get("brand") or "").strip()
                model = str(parsed_doc.get("model") or "").strip()
                if not brand or not model:
                    skipped += 1
                    continue
                filt = {"brand": brand, "model": model}
                existing = await db.vehicle_catalog.find_one(filt, {"_id": 1})
                if existing and not overwrite_enabled:
                    skipped += 1
                    continue
                payload_doc = {
                    "brand": brand,
                    "model": model,
                    "vehicle_type": str(parsed_doc.get("vehicle_type") or "").strip(),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }
                if existing:
                    await db.vehicle_catalog.update_one(filt, {"$set": payload_doc})
                    updated += 1
                else:
                    await db.vehicle_catalog.insert_one(payload_doc)
                    inserted += 1
                continue

            keys = unique_keys_by_collection.get(collection_name, [])
            filt: Dict[str, Any] = {}
            for key in keys:
                value = parsed_doc.get(key)
                if value in (None, ""):
                    filt = {}
                    break
                filt[key] = value

            if not filt:
                skipped += 1
                continue

            existing = await db[collection_name].find_one(filt, {"_id": 1})
            if existing and not overwrite_enabled:
                skipped += 1
                continue

            if existing:
                await db[collection_name].update_one(filt, {"$set": parsed_doc})
                updated += 1
            else:
                await db[collection_name].insert_one(parsed_doc)
                inserted += 1

        report["sheets"][name] = {
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "invalid_checksum": invalid_checksum,
        }

    await record_hypervisor_event(
        request=request,
        actor=actor,
        action="backup_import",
        entity_type="backup",
        entity_id=file.filename,
        details={
            "selected_scopes": targets.get("selected_scopes") or [],
            "overwrite": overwrite_enabled,
            "report": report,
        },
        status_code=200,
    )

    return {
        "message": "Respaldo importado",
        **report,
    }


@api_router.get("/products/{product_id}/check-compatibility/{vehicle_id}")
async def check_product_compatibility(
    product_id: str, vehicle_id: str, request: Request
):
    """Check if a product is compatible with a specific vehicle"""
    await require_auth(request)

    product = await db.products.find_one({"product_id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    vehicle = await db.vehicles.find_one({"vehicle_id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")

    compatibility = product.get("compatibility")

    # If no compatibility restrictions, it's compatible with all
    if not compatibility:
        return {
            "compatible": True,
            "message": "Producto universal, compatible con todos los vehículos",
            "product": product["name"],
            "vehicle": f"{vehicle['brand']} {vehicle['model']} {vehicle['year']}",
        }

    is_compatible = True
    reasons = []

    # Check brand compatibility
    if compatibility.get("brands") and len(compatibility["brands"]) > 0:
        if vehicle["brand"] not in compatibility["brands"]:
            is_compatible = False
            reasons.append(
                f"Marca no compatible. Compatible con: {', '.join(compatibility['brands'])}"
            )

    # Check model compatibility
    if compatibility.get("models") and len(compatibility["models"]) > 0:
        if vehicle["model"] not in compatibility["models"]:
            is_compatible = False
            reasons.append(
                f"Modelo no compatible. Compatible con: {', '.join(compatibility['models'])}"
            )

    # Check year compatibility
    year_from = compatibility.get("year_from")
    year_to = compatibility.get("year_to")
    if year_from and vehicle["year"] < year_from:
        is_compatible = False
        reasons.append(f"Año muy antiguo. Compatible desde año {year_from}")
    if year_to and vehicle["year"] > year_to:
        is_compatible = False
        reasons.append(f"Año muy reciente. Compatible hasta año {year_to}")

    # Check vehicle type (for polarizados)
    if compatibility.get("vehicle_types") and len(compatibility["vehicle_types"]) > 0:
        vehicle_type = vehicle.get("vehicle_type", "Sedán")
        if vehicle_type not in compatibility["vehicle_types"]:
            is_compatible = False
            reasons.append(
                f"Tipo de vehículo no compatible. Compatible con: {', '.join(compatibility['vehicle_types'])}"
            )

    return {
        "compatible": is_compatible,
        "message": "Producto compatible" if is_compatible else "Producto NO compatible",
        "reasons": reasons,
        "product": product["name"],
        "vehicle": f"{vehicle['brand']} {vehicle['model']} {vehicle['year']} ({vehicle.get('vehicle_type', 'N/A')})",
    }


@api_router.post("/products/check-compatibility-batch")
async def check_products_compatibility_batch(
    request: Request, product_ids: List[str], vehicle_id: str
):
    """Check compatibility of multiple products with a vehicle (for cart validation)"""
    await require_auth(request)

    vehicle = await db.vehicles.find_one({"vehicle_id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")

    results = []

    for product_id in product_ids:
        product = await db.products.find_one({"product_id": product_id}, {"_id": 0})
        if not product:
            results.append(
                {
                    "product_id": product_id,
                    "compatible": False,
                    "error": "Producto no encontrado",
                }
            )
            continue

        compatibility = product.get("compatibility")

        # If no compatibility restrictions, it's compatible with all
        if not compatibility:
            results.append(
                {
                    "product_id": product_id,
                    "product_name": product["name"],
                    "compatible": True,
                    "message": "Compatible (universal)",
                }
            )
            continue

        is_compatible = True
        reasons = []

        # Check brand
        if compatibility.get("brands") and len(compatibility["brands"]) > 0:
            if vehicle["brand"] not in compatibility["brands"]:
                is_compatible = False
                reasons.append(
                    f"Marca no compatible ({', '.join(compatibility['brands'])})"
                )

        # Check model
        if compatibility.get("models") and len(compatibility["models"]) > 0:
            if vehicle["model"] not in compatibility["models"]:
                is_compatible = False
                reasons.append(
                    f"Modelo no compatible ({', '.join(compatibility['models'])})"
                )

        # Check year
        year_from = compatibility.get("year_from")
        year_to = compatibility.get("year_to")
        if year_from and vehicle["year"] < year_from:
            is_compatible = False
            reasons.append(f"Año muy antiguo (desde {year_from})")
        if year_to and vehicle["year"] > year_to:
            is_compatible = False
            reasons.append(f"Año muy nuevo (hasta {year_to})")

        # Check vehicle type
        if (
            compatibility.get("vehicle_types")
            and len(compatibility["vehicle_types"]) > 0
        ):
            vehicle_type = vehicle.get("vehicle_type", "")
            if vehicle_type not in compatibility["vehicle_types"]:
                is_compatible = False
                reasons.append(
                    f"Tipo de vehículo no compatible ({', '.join(compatibility['vehicle_types'])})"
                )

        results.append(
            {
                "product_id": product_id,
                "product_name": product["name"],
                "compatible": is_compatible,
                "message": "Compatible" if is_compatible else "No compatible",
                "reasons": reasons,
            }
        )

    # Summary
    compatible_count = sum(1 for r in results if r.get("compatible"))
    incompatible_count = len(results) - compatible_count

    return {
        "vehicle": {
            "vehicle_id": vehicle["vehicle_id"],
            "description": f"{vehicle['brand']} {vehicle['model']} {vehicle['year']} - {vehicle['plate']}",
        },
        "results": results,
        "summary": {
            "total": len(results),
            "compatible": compatible_count,
            "incompatible": incompatible_count,
            "all_compatible": incompatible_count == 0,
        },
    }


@api_router.post("/products/{product_id}/images")
async def add_product_image(product_id: str, image_url: str, request: Request):
    """Add an image URL to a product"""
    await require_roles(request, ["gerencia", "supervisor", "bodegas"])

    result = await db.products.update_one(
        {"product_id": product_id}, {"$push": {"images": image_url}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Image added"}


# ============ CUSTOMERS ============


@api_router.get("/customers")
async def get_customers(request: Request, search: Optional[str] = None):
    await require_auth(request)
    query: dict[str, Any] = {"is_active": True}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
        ]
    customers = await db.customers.find(query, {"_id": 0}).to_list(1000)
    return customers


@api_router.post("/customers")
async def create_customer(customer_data: CustomerCreate, request: Request):
    await require_auth(request)
    customer = Customer(**customer_data.model_dump())
    doc = customer.model_dump()
    customer_type = str(doc.get("customer_type") or "").strip().lower()
    tax_id = str(doc.get("tax_id") or "").strip()

    # Company customers must always provide RUC.
    if customer_type in {"empresa", "company", "juridica", "juridico"} and not tax_id:
        raise HTTPException(status_code=422, detail="RUC es requerido para clientes empresa")

    if "tax_id" in doc:
        doc["tax_id"] = tax_id

    # Ensure created_at exists and is ISO string
    created_at_val = doc.get("created_at")
    if not created_at_val:
        created_at_dt = datetime.now(timezone.utc)
    else:
        created_at_dt = created_at_val
        if isinstance(created_at_dt, str):
            try:
                created_at_dt = datetime.fromisoformat(created_at_dt)
            except Exception:
                created_at_dt = datetime.now(timezone.utc)

    doc["created_at"] = created_at_dt.isoformat()
    # Ensure a stable customer_id is present for consumers/tests
    if not doc.get("customer_id"):
        doc["customer_id"] = f"customer_{uuid.uuid4().hex[:12]}"

    # Ensure new customers are active by default so they appear in listings
    if doc.get("is_active") is None:
        doc["is_active"] = True

    if not isinstance(doc.get("salesperson_history"), list):
        doc["salesperson_history"] = []
    if not isinstance(doc.get("branch_visit_history"), list):
        doc["branch_visit_history"] = []
    if not isinstance(doc.get("customer_segments"), list):
        doc["customer_segments"] = ["minorista"]

    # insert and then return the authoritative stored doc (without _id)
    await db.customers.insert_one(doc)
    stored = await db.customers.find_one({"customer_id": doc["customer_id"]}, {"_id": 0})
    if not stored:
        # fallback to the local doc if find fails for any reason
        doc.pop("_id", None)
        logger.info(f"create_customer returning (fallback): {list(doc.keys())}")
        return doc

    logger.info(f"create_customer returning stored doc for {stored.get('customer_id')}")
    return stored


@api_router.get("/customers/{customer_id}")
async def get_customer(customer_id: str, request: Request):
    await require_auth(request)
    customer = await db.customers.find_one({"customer_id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return customer


@api_router.put("/customers/{customer_id}")
async def update_customer(customer_id: str, updates: Dict[str, Any], request: Request):
    """Update customer fields. Only 'gerencia' and 'supervisor' may update directly.
    Other roles must create an approval instead (enforced at API level)."""
    # enforce role-based direct modifications
    await require_roles(request, ["gerencia", "supervisor"])
    result = await db.customers.update_one({"customer_id": customer_id}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    # audit
    try:
        user = await require_auth(request)
        await audit_service.log_audit_event(
            action="edit_customer",
            actor_id=user.user_id,
            actor_name=user.name,
            actor_role=user.role,
            entity="customer",
            entity_id=customer_id,
            metadata={"changes": updates},
        )
    except Exception:
        logger.exception("Failed to write audit log for update_customer")
    return {"message": "Customer updated"}


# ============ VEHICLES ============


@api_router.get("/vehicles")
async def get_vehicles(request: Request, customer_id: Optional[str] = None):
    await require_auth(request)
    query: dict[str, Any] = {}
    if customer_id:
        query["customer_id"] = customer_id
    vehicles = await db.vehicles.find(query, {"_id": 0}).to_list(1000)
    return vehicles


@api_router.put("/vehicles/{vehicle_id}")
async def update_vehicle(
    vehicle_id: str, vehicle_data: VehicleUpdate, request: Request
):
    # only supervisors/gerencia can update vehicles directly
    await require_roles(request, ["gerencia", "supervisor"])
    existing = await db.vehicles.find_one({"vehicle_id": vehicle_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Vehicle not found")

    update_doc = {k: v for k, v in vehicle_data.model_dump().items() if v is not None}
    if not update_doc:
        return existing

    await db.vehicles.update_one({"vehicle_id": vehicle_id}, {"$set": update_doc})
    vehicle = await db.vehicles.find_one({"vehicle_id": vehicle_id}, {"_id": 0})
    # audit
    try:
        user = await require_auth(request)
        await audit_service.log_audit_event(
            action="edit_vehicle",
            actor_id=user.user_id,
            actor_name=user.name,
            actor_role=user.role,
            entity="vehicle",
            entity_id=vehicle_id,
            metadata={"changes": update_doc},
        )
    except Exception:
        logger.exception("Failed to write audit log for update_vehicle")
    return vehicle


@api_router.delete("/vehicles/{vehicle_id}")
async def delete_vehicle(vehicle_id: str, request: Request):
    await require_auth(request)
    existing = await db.vehicles.find_one({"vehicle_id": vehicle_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    await db.vehicles.delete_one({"vehicle_id": vehicle_id})
    return {"status": "deleted", "vehicle_id": vehicle_id}


@api_router.post("/vehicles")
async def create_vehicle(vehicle_data: VehicleCreate, request: Request):
    await require_auth(request)
    vehicle = Vehicle(**vehicle_data.model_dump())
    doc = vehicle.model_dump()
    # Ensure created_at exists and is ISO string
    created_at_val = doc.get("created_at")
    if not created_at_val:
        created_at_dt = datetime.now(timezone.utc)
    else:
        created_at_dt = created_at_val
        if isinstance(created_at_dt, str):
            try:
                created_at_dt = datetime.fromisoformat(created_at_dt)
            except Exception:
                created_at_dt = datetime.now(timezone.utc)

    doc["created_at"] = created_at_dt.isoformat()

    # Ensure a stable vehicle_id for consumers/tests
    if not doc.get("vehicle_id"):
        doc["vehicle_id"] = f"vehicle_{uuid.uuid4().hex[:12]}"

    await db.vehicles.insert_one(doc)
    # Return authoritative stored document when possible
    stored = await db.vehicles.find_one({"vehicle_id": doc["vehicle_id"]}, {"_id": 0})
    if not stored:
        doc.pop("_id", None)
        logger.info(f"create_vehicle returning (fallback): {list(doc.keys())}")
        return doc

    logger.info(f"create_vehicle returning stored doc for {stored.get('vehicle_id')}")
    return stored


@api_router.get("/vehicles/decode-vin")
async def decode_vehicle_vin(vin: str, request: Request):
    await require_auth(request)

    normalized_vin = (vin or "").strip().upper()
    if not VIN_RE.match(normalized_vin):
        raise HTTPException(
            status_code=400,
            detail="VIN inválido. Debe tener 17 caracteres alfanuméricos (sin I, O, Q).",
        )

    vpic_url = f"https://vpic.nhtsa.dot.gov/api/vehicles/DecodeVinValuesExtended/{normalized_vin}"

    try:
        async with httpx.AsyncClient(timeout=15.0) as client_http:
            response = await client_http.get(vpic_url, params={"format": "json"})
            response.raise_for_status()
            payload = response.json()
    except httpx.HTTPError:
        raise HTTPException(
            status_code=502,
            detail="No se pudo consultar el decodificador VIN (vPIC).",
        )

    results = payload.get("Results") if isinstance(payload, dict) else None
    decoded = results[0] if isinstance(results, list) and results else {}

    brand = (decoded.get("Make") or "").strip()
    model = (decoded.get("Model") or "").strip()
    model_year_raw = (decoded.get("ModelYear") or "").strip()

    model_year: Optional[int] = None
    if model_year_raw and model_year_raw.isdigit():
        model_year = int(model_year_raw)

    if not brand and not model and model_year is None:
        raise HTTPException(
            status_code=404,
            detail="No se encontraron datos para ese VIN.",
        )

    return {
        "vin": normalized_vin,
        "brand": brand.upper() if brand else "",
        "model": model,
        "year": model_year,
        "color": "No especificado",
        "source": "vpic_nhtsa",
    }


@api_router.get("/vehicles/{vehicle_id}")
async def get_vehicle(vehicle_id: str, request: Request):
    await require_auth(request)
    vehicle = await db.vehicles.find_one({"vehicle_id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    return vehicle


# ============ APPROVALS & NOTIFICATIONS ============


@api_router.post("/approvals")
async def create_approval(payload: Dict[str, Any], request: Request):
    """Create an approval request. Payload must include `type` and `payload`.
    Types: 'delete_vehicle', 'edit_vehicle' (payload contains vehicle_id and changes)
    """
    user = await require_auth(request)
    data = payload or {}
    approval_type = data.get("type")
    approval_payload = data.get("payload") or {}
    # motivo/reason is mandatory for supervisor review
    reason = (data.get("reason") or "").strip()
    if not approval_type or not approval_payload:
        raise HTTPException(status_code=400, detail="Invalid approval request")
    if not reason:
        raise HTTPException(status_code=400, detail="El campo 'reason' (motivo) es obligatorio para la solicitud")

    approval_id = f"appr_{uuid.uuid4().hex[:12]}"
    doc = {
        "approval_id": approval_id,
        "type": approval_type,
        "payload": approval_payload,
        "reason": reason,
        "status": "pending",
        "requester_id": user.user_id,
        "requester_name": user.name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "approver_id": None,
        "approver_name": None,
        "approved_at": None,
    }
    await db.approvals.insert_one(doc)
    return {"message": "Approval request created", "approval_id": approval_id}


@api_router.get("/approvals")
async def list_approvals(request: Request):
    # Only supervisors/gerencia may review approvals
    await require_roles(request, ["gerencia", "supervisor"])
    approvals = await db.approvals.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return approvals


@api_router.put("/approvals/{approval_id}/approve")
async def approve_request(approval_id: str, request: Request):
    approver = await require_roles(request, ["gerencia", "supervisor"])
    approval = await db.approvals.find_one({"approval_id": approval_id}, {"_id": 0})
    if not approval:
        raise HTTPException(status_code=404, detail="Approval not found")
    if approval.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Approval already processed")

    # Apply the requested action
    a_type = approval.get("type")
    payload = approval.get("payload") or {}
    try:
        if a_type == "delete_vehicle":
            vehicle_id = payload.get("vehicle_id")
            if not vehicle_id:
                raise HTTPException(status_code=400, detail="Missing vehicle_id")
            await db.vehicles.delete_one({"vehicle_id": vehicle_id})
            note_text = f"Vehículo {vehicle_id} eliminado por aprobación de {approver.name}"
            # audit
            await audit_service.log_audit_event(
                action="delete_vehicle",
                actor_id=approver.user_id,
                actor_name=approver.name,
                actor_role=approver.role,
                entity="vehicle",
                entity_id=vehicle_id,
                metadata={"approval_id": approval_id, "reason": approval.get("reason")},
            )
        elif a_type == "edit_vehicle":
            vehicle_id = payload.get("vehicle_id")
            changes = payload.get("changes") or {}
            if not vehicle_id or not changes:
                raise HTTPException(status_code=400, detail="Invalid edit payload")
            await db.vehicles.update_one({"vehicle_id": vehicle_id}, {"$set": changes})
            note_text = f"Vehículo {vehicle_id} actualizado por aprobación de {approver.name}"
            # audit
            await audit_service.log_audit_event(
                action="edit_vehicle",
                actor_id=approver.user_id,
                actor_name=approver.name,
                actor_role=approver.role,
                entity="vehicle",
                entity_id=vehicle_id,
                metadata={"changes": changes, "approval_id": approval_id, "reason": approval.get("reason")},
            )
        elif a_type == "edit_customer":
            customer_id = payload.get("customer_id")
            changes = payload.get("changes") or {}
            if not customer_id or not changes:
                raise HTTPException(status_code=400, detail="Invalid edit payload")
            await db.customers.update_one({"customer_id": customer_id}, {"$set": changes})
            note_text = f"Cliente {customer_id} actualizado por aprobación de {approver.name}"
            await audit_service.log_audit_event(
                action="edit_customer",
                actor_id=approver.user_id,
                actor_name=approver.name,
                actor_role=approver.role,
                entity="customer",
                entity_id=customer_id,
                metadata={"changes": changes, "approval_id": approval_id, "reason": approval.get("reason")},
            )
        elif a_type == "delete_customer":
            customer_id = payload.get("customer_id")
            if not customer_id:
                raise HTTPException(status_code=400, detail="Missing customer_id")
            await db.customers.delete_one({"customer_id": customer_id})
            note_text = f"Cliente {customer_id} eliminado por aprobación de {approver.name}"
            await audit_service.log_audit_event(
                action="delete_customer",
                actor_id=approver.user_id,
                actor_name=approver.name,
                actor_role=approver.role,
                entity="customer",
                entity_id=customer_id,
                metadata={"approval_id": approval_id, "reason": approval.get("reason")},
            )
        else:
            raise HTTPException(status_code=400, detail="Unknown approval type")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    # mark approval as approved
    await db.approvals.update_one({"approval_id": approval_id}, {"$set": {
        "status": "approved",
        "approver_id": approver.user_id,
        "approver_name": approver.name,
        "approved_at": datetime.now(timezone.utc).isoformat()
    }})

    # create a notification for requester and supervisors
    # include the original reason in the notification to give context
    reason_text = approval.get("reason") or ""
    full_message = f"{note_text}. Motivo: {reason_text}" if reason_text else note_text
    notif = {
        "notification_id": f"notif_{uuid.uuid4().hex[:12]}",
        "message": full_message,
        "approval_id": approval_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "read": False,
        "recipient_id": approval.get("requester_id") or None,
    }
    await db.notifications.insert_one(notif)

    return {"message": "Approved", "notification_id": notif["notification_id"]}


@api_router.get("/notifications")
async def list_notifications(request: Request):
    user = await require_auth(request)
    # return notifications targeted to the user or global ones (recipient_id == None)
    notes = await db.notifications.find({"$or": [{"recipient_id": None}, {"recipient_id": user.user_id}]}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return notes


@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, request: Request):
    user = await require_auth(request)
    await db.notifications.update_one({"notification_id": notification_id, "$or": [{"recipient_id": None}, {"recipient_id": user.user_id}]}, {"$set": {"read": True}})
    return {"message": "Marked read"}


@api_router.get("/notifications/unread-count")
async def unread_notifications_count(request: Request):
    user = await require_auth(request)
    count = await db.notifications.count_documents({"read": False, "$or": [{"recipient_id": None}, {"recipient_id": user.user_id}]})
    return {"unread": int(count)}


@api_router.delete("/notifications/{notification_id}")
async def delete_notification(notification_id: str, request: Request):
    user = await require_auth(request)
    # only allow deleting notifications targeted to the user or global ones
    res = await db.notifications.delete_one({"notification_id": notification_id, "$or": [{"recipient_id": None}, {"recipient_id": user.user_id}]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found or not allowed")
    try:
        window_event = None
        # notify frontend to refresh unread count
        # (can't actually dispatch from server; frontend triggers fetch on modifications)
    except Exception:
        pass
    return {"message": "deleted"}


@api_router.put("/approvals/{approval_id}/reject")
async def reject_request(approval_id: str, request: Request):
    approver = await require_roles(request, ["gerencia", "supervisor"])
    approval = await db.approvals.find_one({"approval_id": approval_id}, {"_id": 0})
    if not approval:
        raise HTTPException(status_code=404, detail="Approval not found")
    if approval.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Approval already processed")

    await db.approvals.update_one({"approval_id": approval_id}, {"$set": {
        "status": "rejected",
        "approver_id": approver.user_id,
        "approver_name": approver.name,
        "approved_at": datetime.now(timezone.utc).isoformat()
    }})

    # notify requester
    reason_text = approval.get("reason") or ""
    note_text = f"Solicitud {approval_id} rechazada por {approver.name}"
    full_message = f"{note_text}. Motivo: {reason_text}" if reason_text else note_text
    notif = {
        "notification_id": f"notif_{uuid.uuid4().hex[:12]}",
        "message": full_message,
        "approval_id": approval_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "read": False,
        "recipient_id": approval.get("requester_id") or None,
    }
    await db.notifications.insert_one(notif)
    return {"message": "Rejected", "notification_id": notif["notification_id"]}


@api_router.delete("/approvals/{approval_id}")
async def delete_approval(approval_id: str, request: Request):
    # Only supervisors/gerencia may remove approvals
    user = await require_roles(request, ["gerencia", "supervisor"])
    approval = await db.approvals.find_one({"approval_id": approval_id}, {"_id": 0})
    if not approval:
        raise HTTPException(status_code=404, detail="Approval not found")
    # Only allow deletion of already processed approvals (approved or rejected)
    status = approval.get("status")
    if status == "pending":
        raise HTTPException(status_code=400, detail="Cannot delete pending approval")
    await db.approvals.delete_one({"approval_id": approval_id})
    return {"message": "deleted"}


# ============ QUOTATIONS ============


async def generate_quotation_id():
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    count = await db.quotations.count_documents(
        {"quotation_id": {"$regex": f"^QUO-{today}"}}
    )
    return f"QUO-{today}-{count + 1:04d}"


@api_router.get("/quotations")
async def get_quotations(request: Request, status: Optional[str] = None):
    user = await require_auth(request)
    effective_role = resolve_effective_role(user.role)
    query: dict[str, Any] = {}
    if status:
        query["status"] = status
    if effective_role in {"ventas", "cajero"}:
        query["salesperson_id"] = user.user_id
    quotations = await db.quotations.find(query).sort("created_at", -1).to_list(500)
    for quotation in quotations:
        quotation_id = quotation.get("quotation_id")
        if not quotation_id:
            quotation_id = await generate_quotation_id()
            await db.quotations.update_one(
                {"_id": quotation.get("_id")},
                {"$set": {"quotation_id": quotation_id}},
            )
            quotation["quotation_id"] = quotation_id
        quotation["id"] = str(quotation.get("_id"))
        quotation.pop("_id", None)
    return quotations


@api_router.post("/quotations")
async def create_quotation(quot_data: QuotationCreate, request: Request):
    user = await require_auth(request)

    customer = await db.customers.find_one(
        {"customer_id": quot_data.customer_id}, {"_id": 0}
    )
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    # Build items with product info
    items = []
    subtotal = 0
    currency = quot_data.currency or "USD"
    exchange_rate = quot_data.exchange_rate or (36.5 if currency == "NIO" else 1)
    normalized_payment_method = _normalize_method_name(
        quot_data.payment_method or quot_data.payment_type
    )
    discounts_allowed_by_method = _is_discount_allowed(normalized_payment_method)
    for item in quot_data.items:
        product = await db.products.find_one(
            {"product_id": item["product_id"]}, {"_id": 0}
        )
        if not product:
            raise HTTPException(
                status_code=404, detail=f"Product {item['product_id']} not found"
            )

        qty = item["quantity"]
        price = product["price"]
        if currency == "NIO":
            price = price * exchange_rate

        install_type = product.get("installation_type", "optional")
        wants_installation = bool(item.get("with_installation", False))
        if install_type == "required":
            wants_installation = True
        if install_type == "not_available":
            wants_installation = False

        install_price = product.get("installation_price", 0)
        if currency == "NIO":
            install_price = install_price * exchange_rate

        if wants_installation and install_type != "not_available":
            price = price + install_price

        discount = item.get("discount", 0) if discounts_allowed_by_method else 0
        item_subtotal = (price * qty) * (1 - discount / 100)

        items.append(
            QuotationItem(
                product_id=product["product_id"],
                product_name=product["name"],
                quantity=qty,
                unit_price=price,
                with_installation=wants_installation,
                discount=discount,
                subtotal=item_subtotal,
            )
        )
        subtotal += item_subtotal

    apply_iva = quot_data.apply_iva if quot_data.apply_iva is not None else True
    iva_rate = float(quot_data.iva_rate) if quot_data.iva_rate is not None else await _get_billing_iva_rate()
    tax = subtotal * (iva_rate / 100) if apply_iva else 0
    effective_discount_percent = quot_data.discount if discounts_allowed_by_method else 0.0
    await _enforce_seller_global_discount_limits(
        actor=user,
        subtotal=subtotal,
        discount_percent=effective_discount_percent,
        currency=currency,
        exchange_rate=exchange_rate,
    )
    total_discount = subtotal * (effective_discount_percent / 100)
    total = subtotal + tax - total_discount

    quotation = Quotation(
        quotation_id=await generate_quotation_id(),
        customer_id=customer["customer_id"],
        customer_name=customer["name"],
        branch_id=user.branch_id or "default",
        salesperson_id=user.user_id,
        salesperson_name=user.name,
        vehicle_id=quot_data.vehicle_id,
        warehouse_id=quot_data.warehouse_id,
        items=[i.model_dump() for i in items],
        subtotal=round(subtotal, 2),
        tax=round(tax, 2),
        discount=round(total_discount, 2),
        discount_percent=effective_discount_percent,
        apply_iva=apply_iva,
        iva_rate=iva_rate,
        currency=currency,
        exchange_rate=exchange_rate if currency == "NIO" else None,
        discount_codes=(quot_data.discount_codes or []) if discounts_allowed_by_method else [],
        payment_type=normalized_payment_method,
        payment_method=normalized_payment_method,
        credit_days=(quot_data.credit_days or 30) if normalized_payment_method == "credit" else None,
        discounts_blocked_by_method=not discounts_allowed_by_method,
        total=round(total, 2),
        valid_until=datetime.now(timezone.utc) + timedelta(days=quot_data.valid_days),
        notes=quot_data.notes,
    )

    doc = quotation.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    doc["valid_until"] = doc["valid_until"].isoformat()
    await db.quotations.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/quotations/{quotation_id}/status")
async def update_quotation_status(quotation_id: str, status: str, request: Request):
    await require_roles(request, ["gerencia", "supervisor", "ventas", "cajero"])
    if status not in ["pending", "approved", "rejected", "converted"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    result = await db.quotations.update_one(
        {"quotation_id": quotation_id}, {"$set": {"status": status}}
    )
    if result.modified_count == 0:
        try:
            result = await db.quotations.update_one(
                {"_id": ObjectId(quotation_id)}, {"$set": {"status": status}}
            )
        except Exception:
            result = None
    if not result or result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Quotation not found")
    return {"message": "Status updated"}


# ============ SALES ============


async def generate_invoice_number():
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    count = await db.sales.count_documents(
        {"invoice_number": {"$regex": f"^INV-{today}"}}
    )
    return f"INV-{today}-{count + 1:04d}"


def build_sales_visibility_query(user: User) -> Dict[str, Any]:
    """Return Mongo query scope for sales visibility by role.

    - ventas/cajero: only their own sales
    - supervisor: only sales from their branch
    - gerencia: all branches
    """
    effective_role = resolve_effective_role(user.role)

    if effective_role == "gerencia":
        return {}

    if effective_role == "supervisor":
        if user.branch_id:
            return {"branch_id": user.branch_id}
        return {"branch_id": "__no_branch__"}

    if effective_role in {"ventas", "cajero"}:
        return {
            "$or": [
                {"salesperson_id": user.user_id},
                {"seller_id": user.user_id},
                {"created_by": user.user_id},
            ]
        }

    if user.branch_id:
        return {"branch_id": user.branch_id}
    return {"branch_id": "__no_branch__"}


def merge_queries(base: Dict[str, Any], extra: Dict[str, Any]) -> Dict[str, Any]:
    if not base:
        return extra or {}
    if not extra:
        return base or {}
    return {"$and": [base, extra]}


def can_access_sale_for_user(user: User, sale: Dict[str, Any]) -> bool:
    effective_role = resolve_effective_role(user.role)

    if effective_role == "gerencia":
        return True

    if effective_role == "supervisor":
        if not user.branch_id:
            return False
        return sale.get("branch_id") == user.branch_id

    if effective_role in {"ventas", "cajero"}:
        owner_ids = {
            sale.get("salesperson_id"),
            sale.get("seller_id"),
            sale.get("created_by"),
        }
        return user.user_id in owner_ids

    if user.branch_id:
        return sale.get("branch_id") == user.branch_id
    return False


async def get_visible_sale_ids_for_user(user: User, limit: int = 5000) -> List[str]:
    sales_query = build_sales_visibility_query(user)
    sales = await db.sales.find(sales_query, {"_id": 0, "sale_id": 1}).to_list(limit)
    return [s.get("sale_id") for s in sales if s.get("sale_id")]


async def build_work_order_visibility_query(user: User) -> Dict[str, Any]:
    effective_role = resolve_effective_role(user.role)

    if effective_role == "gerencia":
        return {}

    if effective_role == "supervisor":
        if user.branch_id:
            return {"branch_id": user.branch_id}
        return {"branch_id": "__no_branch__"}

    if effective_role in {"ventas", "cajero"}:
        visible_sale_ids = await get_visible_sale_ids_for_user(user)
        seller_filters: List[Dict[str, Any]] = [{"created_by": user.user_id}]
        if visible_sale_ids:
            seller_filters.append({"sale_id": {"$in": visible_sale_ids}})
        if len(seller_filters) == 1:
            return seller_filters[0]
        return {"$or": seller_filters}

    if user.branch_id:
        return {"branch_id": user.branch_id}
    return {"branch_id": "__no_branch__"}


async def create_notification_entry(
    message: str,
    recipient_id: Optional[str] = None,
    metadata: Optional[Dict[str, Any]] = None,
    dedupe_key: Optional[str] = None,
) -> str:
    now_iso = datetime.now(timezone.utc).isoformat()
    notification_id = f"notif_{uuid.uuid4().hex[:12]}"
    doc: Dict[str, Any] = {
        "notification_id": notification_id,
        "message": message,
        "created_at": now_iso,
        "read": False,
        "recipient_id": recipient_id,
    }
    if metadata:
        doc["metadata"] = metadata

    # Idempotent notification creation for retried operations.
    if dedupe_key:
        doc["dedupe_key"] = dedupe_key
        upsert_result = await db.notifications.update_one(
            {"dedupe_key": dedupe_key},
            {"$setOnInsert": doc},
            upsert=True,
        )
        if not upsert_result.upserted_id:
            existing = await db.notifications.find_one(
                {"dedupe_key": dedupe_key}, {"_id": 0, "notification_id": 1}
            )
            if existing and existing.get("notification_id"):
                return str(existing.get("notification_id"))
        return notification_id

    await db.notifications.insert_one(doc)
    return notification_id


async def _notify_branch_reviewers(
    branch_id: str,
    message: str,
    metadata: Optional[Dict[str, Any]] = None,
    dedupe_seed: Optional[str] = None,
) -> List[str]:
    recipients = await db.users.find(
        {
            "role": {"$in": ["gerencia", "recursos_humanos"]},
            "is_active": {"$ne": False},
            "$or": [
                {"branch_id": branch_id},
                {"role": "gerencia"},
            ],
        },
        {"_id": 0, "user_id": 1},
    ).to_list(200)
    created_ids: List[str] = []
    for row in recipients:
        recipient_id = str(row.get("user_id") or "").strip()
        if not recipient_id:
            continue
        dedupe = None
        if dedupe_seed:
            dedupe = f"{dedupe_seed}:{recipient_id}"
        notification_id = await create_notification_entry(
            message=message,
            recipient_id=recipient_id,
            metadata=metadata,
            dedupe_key=dedupe,
        )
        created_ids.append(notification_id)
    return created_ids


async def _close_sale_request_notifications(request_id: str, request_type: str) -> None:
    if not request_id:
        return
    notif_type = "sale_edit_request" if request_type == "edit" else "sale_cancel_request"
    now_iso = _utc_now().isoformat()
    await db.notifications.update_many(
        {
            "metadata.request_id": request_id,
            "metadata.type": notif_type,
            "read": False,
        },
        {
            "$set": {
                "read": True,
                "read_at": now_iso,
                "metadata.request_status": "approved",
            }
        },
    )


def _normalize_dispatch_status(status: Optional[str]) -> str:
    return str(status or "pending").strip().lower()


def _can_transition_dispatch_status(current: str, next_status: str) -> bool:
    transitions = {
        "pending": {"in_progress"},
        "in_progress": {"completed"},
        "completed": set(),
    }
    if current == next_status:
        return True
    return next_status in transitions.get(current, set())


async def pick_available_technician(branch_id: Optional[str]) -> Optional[Dict[str, Any]]:
    tech_query: Dict[str, Any] = {
        "is_active": {"$ne": False},
        "$or": [
            {"is_technician": True},
            {"role": {"$in": ["instalaciones", "electrico", "polarizador", "instalador"]}},
        ],
    }
    technicians = await db.users.find(
        tech_query,
        {"_id": 0, "user_id": 1, "name": 1, "role": 1, "branch_id": 1},
    ).to_list(500)
    if not technicians:
        return None

    # Prefer technicians from the same branch; then branch-agnostic; otherwise any.
    same_branch = [t for t in technicians if str(t.get("branch_id") or "") == str(branch_id or "")]
    no_branch = [t for t in technicians if not t.get("branch_id")]
    candidates = same_branch or no_branch or technicians
    candidate_ids = [c.get("user_id") for c in candidates if c.get("user_id")]
    if not candidate_ids:
        return None

    active_statuses = ["pending", "in_progress", "quality_check"]
    pipeline = [
        {
            "$match": {
                "technician_id": {"$in": candidate_ids},
                "status": {"$in": active_statuses},
            }
        },
        {"$group": {"_id": "$technician_id", "count": {"$sum": 1}}},
    ]
    load_rows = await db.work_orders.aggregate(pipeline).to_list(500)
    load_map = {row.get("_id"): int(row.get("count", 0)) for row in load_rows if row.get("_id")}

    def _sort_key(tech: Dict[str, Any]) -> Any:
        return (load_map.get(tech.get("user_id"), 0), (tech.get("name") or "").lower())

    return sorted(candidates, key=_sort_key)[0]


async def create_dispatch_order_from_sale(
    sale_doc: Dict[str, Any],
    customer: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    sale_id = sale_doc.get("sale_id")
    if not sale_id:
        return None

    existing = await db.dispatch_orders.find_one({"sale_id": sale_id}, {"_id": 0})
    if existing:
        return existing

    sale_items = cast(List[Dict[str, Any]], sale_doc.get("items") or [])
    dispatch_items: List[Dict[str, Any]] = []
    primary_warehouse_id: Optional[str] = None

    for item in sale_items:
        product_id = item.get("product_id")
        if not product_id:
            continue
        product = await db.products.find_one({"product_id": product_id}, {"_id": 0})
        if not product:
            continue
        if product.get("product_type") == "service":
            continue

        warehouse_id = str(item.get("warehouse_id") or "")
        if warehouse_id and not primary_warehouse_id:
            primary_warehouse_id = warehouse_id

        dispatch_items.append(
            {
                "product_id": product_id,
                "product_name": product.get("name") or item.get("product_name") or "Producto",
                "product_sku": product.get("sku"),
                "product_image": (product.get("images") or [None])[0],
                "quantity": item.get("quantity", 1),
                "delivered": False,
                "delivered_by": None,
                "delivered_at": None,
            }
        )

    if not dispatch_items:
        return None

    if not primary_warehouse_id:
        primary_warehouse_id = "wh_main"

    warehouse = await db.warehouses.find_one({"warehouse_id": primary_warehouse_id}, {"_id": 0})
    dispatch_doc: Dict[str, Any] = {
        "dispatch_id": f"DSP-{uuid.uuid4().hex[:8].upper()}",
        "sale_id": sale_id,
        "invoice_number": sale_doc.get("invoice_number", "N/A"),
        "reference_number": sale_doc.get("invoice_number", "N/A"),
        "customer_name": customer.get("name") or sale_doc.get("customer_name", "N/A"),
        "customer_phone": customer.get("phone"),
        "requested_by": sale_doc.get("salesperson_id"),
        "requested_by_name": sale_doc.get("salesperson_name"),
        "warehouse_id": primary_warehouse_id,
        "branch_id": warehouse.get("branch_id") if warehouse else sale_doc.get("branch_id"),
        "warehouse_name": warehouse.get("name") if warehouse else primary_warehouse_id,
        "items": dispatch_items,
        "status": "pending",
        "priority": "high" if sale_doc.get("delivery_required") else "normal",
        "delivery_required": bool(sale_doc.get("delivery_required", False)),
        "notes": f"Venta: {sale_doc.get('invoice_number', 'N/A')}",
        "dispatch_type": "sale",
        "sample_id": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "started_at": None,
        "completed_at": None,
        "total_time_minutes": None,
        "dispatchers": [],
    }
    await db.dispatch_orders.insert_one(dispatch_doc)
    return dispatch_doc


@api_router.get("/sales")
async def get_sales(
    request: Request,
    payment_type: Optional[str] = None,
    payment_status: Optional[str] = None,
):
    user = await require_auth(request)
    query: dict[str, Any] = build_sales_visibility_query(user)
    if payment_type:
        query["payment_type"] = payment_type
    if payment_status:
        query["payment_status"] = payment_status
    # Retrieve full documents (including _id) so we can backfill missing sale_id values
    sales = await db.sales.find(query).sort("created_at", -1).to_list(500)
    updated = False
    for s in sales:
        if not s.get("sale_id"):
            new_id = f"sale_{uuid.uuid4().hex[:12]}"
            # Prefer updating by _id when available
            try:
                if s.get("_id"):
                    await db.sales.update_one({"_id": s.get("_id")}, {"$set": {"sale_id": new_id}})
                elif s.get("invoice_number"):
                    await db.sales.update_one({"invoice_number": s.get("invoice_number")}, {"$set": {"sale_id": new_id}})
                updated = True
            except Exception:
                # ignore update failures
                updated = updated

            # Ensure returned document always contains a non-null sale_id
            try:
                s["sale_id"] = new_id
            except Exception:
                try:
                    s.update({"sale_id": new_id})
                except Exception:
                    # ignore update failures
                    pass

    if updated:
        logger.info("Patched missing sale_id values for returned sales")

    # Remove internal _id before returning to clients
    result_list: List[Dict[str, Any]] = []
    for s in sales:
        s.pop("_id", None)
        result_list.append(s)
    logger.info(f"get_sales returning sale_ids: {[s.get('sale_id') for s in result_list[:5]]}")
    return result_list


@api_router.post("/sales")
async def create_sale(sale_data: SaleCreate, request: Request):
    user = await require_auth(request)
    user_branch_id = str(user.branch_id or "branch_main")

    selected_cash_session_id: Optional[str] = None
    requested_session_id = str(getattr(sale_data, "cash_session_id", "") or "").strip()
    if requested_session_id:
        requested_session = await db.caja_sesiones.find_one(
            {"session_id": requested_session_id},
            {"_id": 0, "session_id": 1, "estado": 1, "branch_id": 1},
        )
        if not requested_session:
            raise HTTPException(status_code=404, detail="Sesión de caja no encontrada")
        if str(requested_session.get("branch_id") or "") != user_branch_id:
            raise HTTPException(status_code=400, detail="La sesión de caja pertenece a otra sucursal")
        if requested_session.get("estado") != "abierta":
            raise HTTPException(status_code=400, detail="No se pueden crear facturas en una sesión de caja cerrada")
        selected_cash_session_id = requested_session_id
    else:
        active_session = await db.caja_sesiones.find_one(
            {
                "branch_id": user_branch_id,
                "estado": "abierta",
            },
            {"_id": 0, "session_id": 1},
            sort=[("opened_at", -1)],
        )
        if active_session and active_session.get("session_id"):
            selected_cash_session_id = str(active_session.get("session_id"))

    idempotency_key = str(sale_data.idempotency_key or "").strip()
    if idempotency_key:
        existing_sale = await db.sales.find_one(
            {
                "salesperson_id": user.user_id,
                "idempotency_key": idempotency_key,
            },
            {"_id": 0},
        )
        if existing_sale:
            return {
                **existing_sale,
                "work_order_created": bool(existing_sale.get("work_order_id")),
                "work_order_id": existing_sale.get("work_order_id"),
                "dispatch_created": bool(existing_sale.get("dispatch_id")),
                "dispatch_id": existing_sale.get("dispatch_id"),
            }

    branch_default_warehouse = await get_default_warehouse_for_branch(user_branch_id)
    branch_default_warehouse_id = (
        str(branch_default_warehouse.get("warehouse_id")) if branch_default_warehouse and branch_default_warehouse.get("warehouse_id") else None
    )

    customer = await db.customers.find_one(
        {"customer_id": sale_data.customer_id}, {"_id": 0}
    )
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    sale_channel = str(
        getattr(sale_data, "sale_channel", "") or customer.get("customer_type") or "minorista"
    ).strip().lower()
    if sale_channel not in {"minorista", "mayorista"}:
        sale_channel = "minorista"

    if sale_data.quotation_id:
        quotation = await db.quotations.find_one(
            {"quotation_id": sale_data.quotation_id}, {"_id": 0}
        )
        if not quotation:
            raise HTTPException(status_code=404, detail="Quotation not found")
        if quotation.get("status") != "approved":
            raise HTTPException(
                status_code=400,
                detail="Quotation is not approved for conversion",
            )
        if quotation.get("customer_id") != sale_data.customer_id:
            raise HTTPException(
                status_code=400,
                detail="Quotation customer does not match sale customer",
            )
        valid_until = quotation.get("valid_until")
        if isinstance(valid_until, str):
            valid_until = datetime.fromisoformat(valid_until.replace("Z", "+00:00"))
        if isinstance(valid_until, datetime) and valid_until < datetime.now(timezone.utc):
            raise HTTPException(status_code=400, detail="Quotation has expired")

    # Build items and check inventory
    items = []
    subtotal = 0
    items_requiring_installation = []
    items_requiring_manager_auth = []

    sample_docs = await db.sample_requests.find(
        {
            "customer_id": sale_data.customer_id,
            "status": {"$in": ["delivered", "return_requested"]},
        },
        {"_id": 0},
    ).sort("created_at", 1).to_list(200)
    sample_by_product: Dict[str, Dict[str, Any]] = {}
    for s in sample_docs:
        if s.get("product_id") not in sample_by_product:
            sample_by_product[s.get("product_id")] = s

    inventory_updates: List[Dict[str, Any]] = []
    sample_usage: List[Dict[str, Any]] = []

    for item in sale_data.items:
        product = await db.products.find_one(
            {"product_id": item["product_id"]}, {"_id": 0}
        )
        if not product:
            raise HTTPException(
                status_code=404, detail=f"Product {item['product_id']} not found"
            )

        requested_warehouse_id = item.get("warehouse_id")
        warehouse_id = str(requested_warehouse_id or branch_default_warehouse_id or "")
        if not warehouse_id:
            raise HTTPException(
                status_code=400,
                detail="No se encontró una bodega activa para la sucursal. Configure una bodega por sucursal antes de vender.",
            )
        await ensure_warehouse_belongs_to_branch(warehouse_id, user_branch_id)

        # Check inventory for physical products
        if product.get("product_type") != "service":
            inv = await db.inventory.find_one(
                {"product_id": item["product_id"], "warehouse_id": warehouse_id}
            )
            sample_doc = sample_by_product.get(item["product_id"])
            sample_qty_available = sample_doc.get("quantity", 1) if sample_doc else 0
            sample_qty_used = min(item["quantity"], sample_qty_available) if sample_doc else 0
            required_qty = max(item["quantity"] - sample_qty_used, 0)

            if required_qty > 0 and (not inv or inv["quantity"] < required_qty):
                raise HTTPException(
                    status_code=400,
                    detail=f"Insufficient inventory for {product['name']}",
                )

            inventory_updates.append(
                {
                    "product_id": item["product_id"],
                    "warehouse_id": warehouse_id,
                    "quantity": required_qty,
                }
            )

            if sample_doc and sample_qty_used > 0:
                sample_usage.append(
                    {
                        "sample_id": sample_doc.get("sample_id"),
                        "quantity": sample_qty_used,
                    }
                )

        qty = item["quantity"]
        price = product["price"]
        discount = item.get("discount", 0)
        item_subtotal = (price * qty) * (1 - discount / 100)

        # Check installation requirements
        install_type = product.get("installation_type", "optional")
        wants_installation = item.get("with_installation", False)

        # Determine display note
        display_note = ""
        if install_type == "not_available":
            display_note = "Solo para llevar"
            if wants_installation:
                # Requires manager authorization
                items_requiring_manager_auth.append(
                    {
                        "product_name": product["name"],
                        "product_id": product["product_id"],
                    }
                )
        elif install_type == "required" or (
            install_type == "optional" and wants_installation
        ):
            items_requiring_installation.append(
                {
                    "product_id": product["product_id"],
                    "product_name": product["name"],
                    "quantity": qty,
                    "installation_price": product.get("installation_price", 0),
                    "installation_time_minutes": product.get(
                        "installation_time_minutes", 60
                    ),
                }
            )

        items.append(
            SaleItem(
                product_id=product["product_id"],
                product_name=product["name"],
                quantity=qty,
                unit_price=price,
                discount=discount,
                subtotal=item_subtotal,
                warehouse_id=warehouse_id,
                installation_type=install_type,
                with_installation=wants_installation or install_type == "required",
                display_note=display_note,
            )
        )
        subtotal += item_subtotal

        # Add installation price if applicable
        if install_type == "required" or (
            install_type == "optional" and wants_installation
        ):
            install_price = product.get("installation_price", 0) * qty
            subtotal += install_price

    iva_rate_percent = await _get_billing_iva_rate()
    iva_rate_decimal = iva_rate_percent / 100.0
    tax = subtotal * iva_rate_decimal
    raw_sale_currency = getattr(sale_data, "currency", "USD")
    sale_currency = _currency_code(raw_sale_currency)
    raw_sale_exchange_rate = getattr(sale_data, "exchange_rate", None)
    try:
        sale_exchange_rate = float(raw_sale_exchange_rate) if raw_sale_exchange_rate is not None else None
    except (TypeError, ValueError):
        sale_exchange_rate = None
    await _enforce_seller_global_discount_limits(
        actor=user,
        subtotal=subtotal,
        discount_percent=sale_data.discount,
        currency=sale_currency,
        exchange_rate=sale_exchange_rate,
    )
    total_discount = subtotal * (sale_data.discount / 100)
    total = subtotal + tax - total_discount

    # Check credit limit for credit sales
    if sale_data.payment_type == "credit":
        available_credit = customer.get("credit_limit", 0) - customer.get(
            "credit_balance", 0
        )
        if total > available_credit:
            raise HTTPException(status_code=400, detail="Exceeds credit limit")

    # Check manager authorization for "solo para llevar" products that want installation
    if items_requiring_manager_auth and not sale_data.manager_authorization_code:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "REQUIRES_MANAGER_AUTH",
                "message": (
                    "Los siguientes productos son 'Solo para llevar' y requieren "
                    "autorización del gerente para instalación"
                ),
                "products": items_requiring_manager_auth,
            },
        )

    # Verify manager authorization if provided
    if items_requiring_manager_auth and sale_data.manager_authorization_code:
        auth_valid = await db.manager_authorizations.find_one(
            {
                "code": sale_data.manager_authorization_code,
                "used": False,
                "expires_at": {"$gt": datetime.now(timezone.utc).isoformat()},
            }
        )
        if not auth_valid:
            raise HTTPException(
                status_code=400, detail="Código de autorización inválido o expirado"
            )
        # Mark authorization as used
        await db.manager_authorizations.update_one(
            {"code": sale_data.manager_authorization_code},
            {"$set": {"used": True, "used_at": datetime.now(timezone.utc).isoformat()}},
        )

    invoice_number = await generate_invoice_number()
    credit_due_date = None
    if sale_data.payment_type == "credit" and sale_data.credit_days:
        credit_due_date = datetime.now(timezone.utc) + timedelta(
            days=sale_data.credit_days
        )

    # Determine if sale needs work order
    needs_work_order = len(items_requiring_installation) > 0
    if needs_work_order:
        await ensure_branch_service_enabled(
            user.branch_id,
            "installations_enabled",
            "Esta sucursal no ofrece servicios de instalaciones. Solo la sucursal central puede procesar instalaciones/polarizados.",
        )

    normalized_payment_method = _normalize_method_name(
        sale_data.payment_method or sale_data.payment_type
    )

    sale = Sale(
        invoice_number=invoice_number,
        quotation_id=sale_data.quotation_id,
        customer_id=customer["customer_id"],
        customer_name=customer["name"],
        branch_id=user_branch_id,
        salesperson_id=user.user_id,
        salesperson_name=user.name,
        items=[i.model_dump() for i in items],
        subtotal=round(subtotal, 2),
        tax=round(tax, 2),
        discount=round(total_discount, 2),
        total=round(total, 2),
        payment_type=sale_data.payment_type,
        payment_status="pending" if sale_data.payment_type != "cash" else "paid",
        payment_method=normalized_payment_method,
        sale_channel=sale_channel,
        credit_due_date=credit_due_date,
        delivery_required=sale_data.delivery_required,
        delivery_address=sale_data.delivery_address,
        delivery_status="pending" if sale_data.delivery_required else None,
        notes=sale_data.notes,
        has_installation=needs_work_order,
        iva_rate=iva_rate_decimal,
        iva_amount=round(tax, 2),
        total_legal=round(total, 2),
        discounts_applied_amount=round(total_discount, 2),
        discounts_blocked_by_method=False,
        retention_rate=0.0,
        retention_amount=0.0,
        net_to_collect=round(total, 2),
        print_format="thermal80",
        retention_receipt_required=False,
        pos_bank_withholding_expected=0.0,
        commercial_terms_locked=False,
        settlement_warnings=[],
    )

    doc = sale.model_dump()
    # Ensure a stable sale_id is present
    if not doc.get("sale_id"):
        doc["sale_id"] = f"sale_{uuid.uuid4().hex[:12]}"
    doc["created_at"] = doc["created_at"].isoformat()
    if doc["credit_due_date"]:
        doc["credit_due_date"] = doc["credit_due_date"].isoformat()
    doc["idempotency_key"] = idempotency_key or None
    doc["cash_session_id"] = selected_cash_session_id
    doc["warehouse_dispatch_status"] = (
        "pending" if sale_data.delivery_required else "not_required"
    )
    doc["workflow_state"] = (
        "dispatch_pending" if sale_data.delivery_required else "created"
    )
    await db.sales.insert_one(doc)

    branch_info = await get_branch_with_policy(user_branch_id)
    branch_name = str(branch_info.get("name") or user_branch_id)

    customer_history_update = {
        "$set": {
            "last_sale_at": doc.get("created_at"),
            "last_sale_branch_id": user_branch_id,
            "last_sale_branch_name": branch_name,
        },
        "$addToSet": {
            "customer_segments": sale_channel,
        },
        "$inc": {
            "total_sales_count": 1,
        },
    }
    await db.customers.update_one({"customer_id": customer["customer_id"]}, customer_history_update)

    history_key = {
        "customer_id": customer["customer_id"],
        "salesperson_history.user_id": user.user_id,
        "salesperson_history.branch_id": user_branch_id,
    }
    history_inc = await db.customers.update_one(
        history_key,
        {
            "$inc": {
                "salesperson_history.$.sales_count": 1,
            },
            "$set": {
                "salesperson_history.$.last_sale_at": doc.get("created_at"),
                "salesperson_history.$.salesperson_name": user.name,
                "salesperson_history.$.branch_name": branch_name,
                "salesperson_history.$.last_sale_id": doc.get("sale_id"),
            },
        },
    )
    if history_inc.modified_count == 0:
        await db.customers.update_one(
            {"customer_id": customer["customer_id"]},
            {
                "$push": {
                    "salesperson_history": {
                        "user_id": user.user_id,
                        "salesperson_name": user.name,
                        "role": user.role,
                        "branch_id": user_branch_id,
                        "branch_name": branch_name,
                        "first_sale_at": doc.get("created_at"),
                        "last_sale_at": doc.get("created_at"),
                        "last_sale_id": doc.get("sale_id"),
                        "sales_count": 1,
                    }
                }
            },
        )

    branch_visit_inc = await db.customers.update_one(
        {
            "customer_id": customer["customer_id"],
            "branch_visit_history.branch_id": user_branch_id,
        },
        {
            "$inc": {
                "branch_visit_history.$.visit_count": 1,
            },
            "$set": {
                "branch_visit_history.$.branch_name": branch_name,
                "branch_visit_history.$.last_visit_at": doc.get("created_at"),
            },
        },
    )
    if branch_visit_inc.modified_count == 0:
        await db.customers.update_one(
            {"customer_id": customer["customer_id"]},
            {
                "$push": {
                    "branch_visit_history": {
                        "branch_id": user_branch_id,
                        "branch_name": branch_name,
                        "first_visit_at": doc.get("created_at"),
                        "last_visit_at": doc.get("created_at"),
                        "visit_count": 1,
                    }
                }
            },
        )

    await audit_service.log_audit_event(
        action="sale_create",
        actor_id=user.user_id,
        actor_name=user.name,
        actor_role=user.role,
        entity="sale",
        entity_id=doc.get("sale_id"),
        branch_id=user_branch_id,
        metadata={
            "customer_id": sale_data.customer_id,
            "total": doc.get("total"),
            "payment_type": doc.get("payment_type"),
            "items_count": len(doc.get("items") or []),
        },
    )

    # Exclude _id for response
    doc.pop("_id", None)

    # Update inventory (exclude items already covered by samples)
    for inv_item in inventory_updates:
        if inv_item.get("warehouse_id") and inv_item.get("quantity", 0) > 0:
            product_id = inv_item.get("product_id")
            warehouse_id = inv_item.get("warehouse_id")
            if not product_id or not warehouse_id:
                continue
            await db.inventory.update_one(
                {
                    "product_id": product_id,
                    "warehouse_id": warehouse_id,
                },
                {
                    "$inc": {"quantity": -inv_item.get("quantity", 0)},
                    "$set": {"last_updated": datetime.now(timezone.utc).isoformat()},
                },
            )
            await audit_service.log_inventory_movement(
                product_id=product_id,
                warehouse_id=warehouse_id,
                quantity_change=-int(inv_item.get("quantity", 0)),
                reason="sale",
                actor=user,
                branch_id=user_branch_id,
                reference_id=doc.get("sale_id"),
                metadata={"customer_id": sale_data.customer_id},
            )

    # Mark samples as consumed when used in sale
    for usage in sample_usage:
        if usage.get("sample_id"):
            await db.sample_requests.update_one(
                {"sample_id": usage.get("sample_id")},
                {
                    "$set": {
                        "status": "consumed",
                        "sale_id": doc.get("sale_id"),
                    }
                },
            )

    # Update credit balance
    if sale_data.payment_type == "credit":
        await db.customers.update_one(
            {"customer_id": customer["customer_id"]},
            {"$inc": {"credit_balance": total}},
        )

    # Update quotation if from quotation
    if sale_data.quotation_id:
        await db.quotations.update_one(
            {"quotation_id": sale_data.quotation_id}, {"$set": {"status": "converted"}}
        )

    # AUTO-GENERATE WORK ORDER if products require installation
    work_order_id = None
    dispatch_id = None
    if needs_work_order:
        # Get vehicle if provided
        vehicle_info = ""
        if sale_data.vehicle_id:
            vehicle = await db.vehicles.find_one(
                {"vehicle_id": sale_data.vehicle_id}, {"_id": 0}
            )
            if vehicle:
                vehicle_info = f"{vehicle['brand']} {vehicle['model']} {vehicle['year']} - {vehicle['plate']}"

        total_install_time = sum(
            i["installation_time_minutes"] for i in items_requiring_installation
        )

        assigned_technician = await pick_available_technician(user_branch_id)

        work_order_doc = {
            "work_order_id": f"wo_{uuid.uuid4().hex[:8]}",
            "sale_id": doc["sale_id"],
            "invoice_number": invoice_number,
            "customer_id": customer["customer_id"],
            "customer_name": customer["name"],
            "vehicle_id": sale_data.vehicle_id,
            "vehicle_info": vehicle_info,
            "branch_id": user_branch_id,
            "items": items_requiring_installation,
            "status": "pending",
            "priority": "normal",
            "estimated_time": total_install_time,
            "actual_time": None,
            "technician_id": assigned_technician.get("user_id") if assigned_technician else None,
            "technician_name": assigned_technician.get("name") if assigned_technician else None,
            "start_time": None,
            "end_time": None,
            "quality_score": None,
            "notes": f"Orden generada automáticamente desde venta {invoice_number}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "auto_generated": True,
        }

        await db.work_orders.insert_one(work_order_doc)
        work_order_id = work_order_doc["work_order_id"]

        if assigned_technician and assigned_technician.get("user_id"):
            await create_notification_entry(
                message=(
                    f"Nueva orden de instalación {work_order_id} asignada: "
                    f"venta {invoice_number} para {customer.get('name', 'cliente')}"
                ),
                recipient_id=assigned_technician.get("user_id"),
                metadata={
                    "type": "work_order_assigned",
                    "work_order_id": work_order_id,
                    "sale_id": doc.get("sale_id"),
                    "invoice_number": invoice_number,
                },
                dedupe_key=f"work_order_assigned:{work_order_id}",
            )
        elif doc.get("salesperson_id"):
            await create_notification_entry(
                message=(
                    f"Orden {work_order_id} creada sin técnico disponible para "
                    f"la venta {invoice_number}."
                ),
                recipient_id=doc.get("salesperson_id"),
                metadata={
                    "type": "work_order_unassigned",
                    "work_order_id": work_order_id,
                    "sale_id": doc.get("sale_id"),
                    "invoice_number": invoice_number,
                },
                dedupe_key=f"work_order_unassigned:{work_order_id}",
            )

        # Update sale with work order reference
        await db.sales.update_one(
            {"sale_id": doc["sale_id"]}, {"$set": {"work_order_id": work_order_id}}
        )
        doc["work_order_id"] = work_order_id

    dispatch_doc = await create_dispatch_order_from_sale(doc, customer)
    if dispatch_doc:
        dispatch_id = dispatch_doc.get("dispatch_id")
        await db.sales.update_one(
            {"sale_id": doc["sale_id"]},
            {"$set": {"dispatch_id": dispatch_id}},
        )
        doc["dispatch_id"] = dispatch_id

    return {
        **doc,
        "work_order_created": work_order_id is not None,
        "work_order_id": work_order_id,
        "dispatch_created": dispatch_id is not None,
        "dispatch_id": dispatch_id,
    }


@api_router.get("/sales/{sale_id}")
async def get_sale(sale_id: str, request: Request):
    await require_auth(request)
    sale = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    sale = cast(Dict[str, Any], sale)
    return sale


@api_router.post("/sales/preview-settlement")
async def preview_sale_settlement(payload: SaleSettlementPreviewRequest, request: Request):
    await require_roles(request, ["gerencia", "supervisor", "ventas", "cajero"])

    customer: Optional[Dict[str, Any]] = None
    if payload.customer_id:
        customer = await db.customers.find_one({"customer_id": payload.customer_id}, {"_id": 0})

    subtotal_base = float(payload.subtotal) if payload.subtotal is not None else _compute_items_subtotal(payload.items)
    retention_profile = payload.retention_profile or _extract_retention_profile_from_customer(customer)

    iva_rate_percent = await _get_billing_iva_rate()

    settlement = _build_sale_settlement(
        subtotal_base=subtotal_base,
        discount_percent=payload.discount_percent,
        discounts_amount=payload.discounts_amount,
        promotions_amount=payload.promotions_amount,
        payment_method=payload.payment_method,
        print_format=payload.print_format,
        apply_iva=payload.apply_iva,
        iva_rate_percent=iva_rate_percent,
        retention_profile=retention_profile,
        retention_rate_hint=payload.retention_rate_hint,
    )

    return {
        "customer_id": payload.customer_id,
        **settlement,
    }


@api_router.patch("/sales/{sale_id}/commercial-terms")
async def update_sale_commercial_terms(
    sale_id: str,
    payload: SaleCommercialTermsUpdate,
    request: Request,
):
    user = await require_roles(request, ["gerencia", "supervisor", "ventas"])

    sale = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    sale = cast(Dict[str, Any], sale)

    if not can_access_sale_for_user(user, sale):
        raise HTTPException(status_code=403, detail="No autorizado para actualizar términos comerciales de esta venta")

    if bool(sale.get("commercial_terms_locked")) and user.role not in {"gerencia", "supervisor"}:
        raise HTTPException(status_code=400, detail="Los términos comerciales ya están bloqueados")

    customer = await db.customers.find_one({"customer_id": sale.get("customer_id")}, {"_id": 0})
    retention_profile = payload.retention_profile or _extract_retention_profile_from_customer(customer)
    iva_rate_percent = await _get_billing_iva_rate()

    subtotal_base = float(sale.get("subtotal") or 0.0)
    discounts_amount = float(sale.get("discount") or 0.0)

    settlement = _build_sale_settlement(
        subtotal_base=subtotal_base,
        discount_percent=0.0,
        discounts_amount=discounts_amount,
        promotions_amount=0.0,
        payment_method=payload.payment_method,
        print_format=payload.print_format,
        apply_iva=payload.apply_iva,
        iva_rate_percent=iva_rate_percent,
        retention_profile=retention_profile,
        retention_rate_hint=payload.retention_rate_hint,
    )

    set_values = {
        "payment_type": settlement["payment_method"],
        "payment_method": settlement["payment_method"],
        "tax": settlement["iva_amount"],
        "discount": settlement["discounts_applied_amount"],
        "total": settlement["total_legal"],
        "iva_rate": settlement["iva_rate"],
        "iva_amount": settlement["iva_amount"],
        "total_legal": settlement["total_legal"],
        "discounts_applied_amount": settlement["discounts_applied_amount"],
        "discounts_blocked_by_method": settlement["discounts_blocked_by_method"],
        "retention_rate": settlement["retention_rate"],
        "retention_amount": settlement["retention_amount"],
        "net_to_collect": settlement["net_to_collect"],
        "print_format": settlement["print_format"],
        "retention_receipt_required": settlement["retention_receipt_required"],
        "pos_bank_withholding_expected": settlement["pos_bank_withholding_expected"],
        "settlement_warnings": settlement["warnings"],
        "commercial_terms_locked": bool(payload.lock_terms),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": user.user_id,
    }

    await db.sales.update_one({"sale_id": sale_id}, {"$set": set_values})

    updated = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0})
    return cast(Dict[str, Any], updated or {**sale, **set_values})


async def _normalize_mixed_payments(
    payments: List[MixedPaymentItem],
    usd_to_nio_rate: float,
) -> Dict[str, Any]:
    normalized_rows: List[Dict[str, Any]] = []
    total_nio = 0.0
    has_card = False
    methods_summary: Dict[str, Dict[str, float]] = {}

    for idx, row in enumerate(payments):
        method = _normalize_method_name(row.metodo)
        currency = _currency_code(row.moneda)
        source_amount = _round4(row.monto_origen)
        if source_amount <= 0:
            raise HTTPException(status_code=400, detail=f"monto_origen inválido en pago #{idx + 1}")

        row_rate = row.tasa_cambio if row.tasa_cambio and row.tasa_cambio > 0 else (
            usd_to_nio_rate if currency == "USD" else 1.0
        )
        row_rate = _round4(row_rate)

        if row.monto_cordobas is not None and float(row.monto_cordobas) > 0:
            amount_nio = _round4(row.monto_cordobas)
        else:
            amount_nio = _round4(source_amount if currency == "NIO" else (source_amount * row_rate))

        total_nio += amount_nio
        has_card = has_card or _is_card_method(method)

        if method not in methods_summary:
            methods_summary[method] = {"NIO": 0.0, "USD": 0.0}
        methods_summary[method][currency] = _round2(methods_summary[method].get(currency, 0.0) + source_amount)

        normalized_rows.append(
            {
                "line_no": idx + 1,
                "payment_method": method,
                "currency": currency,
                "amount_origin": source_amount,
                "exchange_rate": row_rate,
                "amount_nio": _round2(amount_nio),
                "reference": row.referencia_bancaria,
                "audit_note": row.notas_auditoria,
            }
        )

    return {
        "rows": normalized_rows,
        "total_nio": _round2(total_nio),
        "has_card": has_card,
        "methods_summary": methods_summary,
    }


async def _enforce_pos_discount_policy(
    sale: Dict[str, Any],
    *,
    has_card: bool,
    authorization: Optional[PosDiscountAuthorization],
    force_remove_discount: bool,
    actor: User,
) -> Dict[str, Any]:
    current_discount = _round2(sale.get("discounts_applied_amount") or sale.get("discount") or 0.0)
    if not has_card or current_discount <= 0:
        return {
            "sale": sale,
            "warning": None,
            "audit_note": None,
        }

    if authorization and authorization.autorizado_por and authorization.justificacion_interna:
        audit_note = (
            f"POS_DISCOUNT_AUTH|autorizado_por={authorization.autorizado_por}|"
            f"justificacion={authorization.justificacion_interna}|mostrar_al_cliente={authorization.mostrar_al_cliente}"
        )
        await db.sales.update_one(
            {"sale_id": sale.get("sale_id")},
            {
                "$set": {
                    "pos_discount_authorized": True,
                    "pos_discount_authorized_by": authorization.autorizado_por,
                    "pos_discount_justification_internal": authorization.justificacion_interna,
                    "pos_discount_show_to_customer": bool(authorization.mostrar_al_cliente),
                    "pos_discount_authorized_at": datetime.now(timezone.utc).isoformat(),
                    "pos_discount_authorized_actor": actor.user_id,
                }
            },
        )
        updated = await db.sales.find_one({"sale_id": sale.get("sale_id")}, {"_id": 0})
        return {
            "sale": cast(Dict[str, Any], updated or sale),
            "warning": None,
            "audit_note": audit_note,
        }

    if not force_remove_discount:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "POS_DISCOUNT_CONFLICT",
                "message": "El pago con tarjeta anula el descuento. ¿Desea proceder o requiere autorización de gerencia?",
                "requires": ["force_remove_discount=true", "autorizacion_descuento_pos"],
            },
        )

    subtotal_base = _round2(sale.get("subtotal") or 0.0)
    retention_profile = _normalize_retention_profile(sale.get("retention_profile") or "general")
    iva_rate_percent = await _get_billing_iva_rate()
    settlement = _build_sale_settlement(
        subtotal_base=subtotal_base,
        discount_percent=0.0,
        discounts_amount=current_discount,
        promotions_amount=0.0,
        payment_method="card",
        print_format=sale.get("print_format") or "thermal80",
        apply_iva=bool((sale.get("iva_rate") or 0) > 0),
        iva_rate_percent=iva_rate_percent,
        retention_profile=retention_profile,
        retention_rate_hint=sale.get("retention_rate") if sale.get("retention_rate") in {0.01, 0.02} else None,
    )

    updates = {
        "payment_method": settlement["payment_method"],
        "payment_type": settlement["payment_method"],
        "tax": settlement["iva_amount"],
        "discount": settlement["discounts_applied_amount"],
        "total": settlement["total_legal"],
        "iva_rate": settlement["iva_rate"],
        "iva_amount": settlement["iva_amount"],
        "total_legal": settlement["total_legal"],
        "discounts_applied_amount": settlement["discounts_applied_amount"],
        "discounts_blocked_by_method": settlement["discounts_blocked_by_method"],
        "retention_rate": settlement["retention_rate"],
        "retention_amount": settlement["retention_amount"],
        "net_to_collect": settlement["net_to_collect"],
        "print_format": settlement["print_format"],
        "retention_receipt_required": settlement["retention_receipt_required"],
        "pos_bank_withholding_expected": settlement["pos_bank_withholding_expected"],
        "settlement_warnings": settlement["warnings"],
        "pos_discount_removed_by_card": True,
        "pos_discount_removed_at": datetime.now(timezone.utc).isoformat(),
        "pos_discount_removed_by": actor.user_id,
    }
    await db.sales.update_one({"sale_id": sale.get("sale_id")}, {"$set": updates})
    updated = await db.sales.find_one({"sale_id": sale.get("sale_id")}, {"_id": 0})
    return {
        "sale": cast(Dict[str, Any], updated or {**sale, **updates}),
        "warning": "Con tarjeta no aplican descuentos ni promociones; se recalculó el total.",
        "audit_note": "POS_DISCOUNT_AUTO_REMOVED",
    }


def _can_collect_sale_for_user(user: User, sale: Dict[str, Any]) -> bool:
    if user.role == "gerencia":
        return True
    if user.role in {"supervisor", "cajero"}:
        return bool(user.branch_id) and str(user.branch_id) == str(sale.get("branch_id") or "")
    return False


def _sale_effective_state(sale: Dict[str, Any]) -> str:
    explicit_state = str(sale.get("invoice_state") or "").strip().lower()
    if explicit_state in {"cancelled", "anulada", "anulado"}:
        return "cancelled"
    payment_status = str(sale.get("payment_status") or "").strip().lower()
    if payment_status == "paid":
        return "closed"
    return "open"


def _cashier_invoice_summary(sale: Dict[str, Any]) -> Dict[str, Any]:
    due = _round2(float(sale.get("net_to_collect") if sale.get("net_to_collect") is not None else sale.get("total") or 0.0))
    amount_paid = _round2(float(sale.get("amount_paid") or 0.0))
    amount_pending = _round2(float(sale.get("amount_pending") if sale.get("amount_pending") is not None else max(due - amount_paid, 0.0)))
    state = _sale_effective_state(sale)

    return {
        "sale_id": sale.get("sale_id"),
        "invoice_number": sale.get("invoice_number"),
        "customer_id": sale.get("customer_id"),
        "customer_name": sale.get("customer_name"),
        "branch_id": sale.get("branch_id"),
        "created_at": sale.get("created_at"),
        "payment_status": sale.get("payment_status"),
        "invoice_state": state,
        "payment_method": sale.get("payment_method") or sale.get("payment_type"),
        "subtotal": _round2(float(sale.get("subtotal") or 0.0)),
        "discount": _round2(float(sale.get("discount") or 0.0)),
        "discounts_applied_amount": _round2(float(sale.get("discounts_applied_amount") or sale.get("discount") or 0.0)),
        "iva_amount": _round2(float(sale.get("iva_amount") or sale.get("tax") or 0.0)),
        "total_legal": _round2(float(sale.get("total_legal") or sale.get("total") or 0.0)),
        "net_to_collect": due,
        "amount_paid": amount_paid,
        "amount_pending": amount_pending,
        "retention_amount": _round2(float(sale.get("retention_amount") or 0.0)),
        "retention_rate": float(sale.get("retention_rate") or 0.0),
        "cancelled_at": sale.get("cancelled_at"),
        "cancelled_by_name": sale.get("cancelled_by_name"),
        "cancel_reason": sale.get("cancel_reason"),
    }


@api_router.get("/caja/facturas")
async def list_cashier_invoices(
    request: Request,
    tab: str = "abiertas",
    search: str = "",
    branch_id: Optional[str] = None,
    limit: int = 200,
):
    user = await require_roles(request, ["gerencia", "supervisor", "cajero", "ventas", "jefe_vendedores", "jefe_tienda"])
    effective_role = resolve_effective_role(user.role)

    tab_value = str(tab or "abiertas").strip().lower()
    tab_aliases = {
        "abiertas": "open",
        "open": "open",
        "cerradas": "closed",
        "closed": "closed",
        "anuladas": "cancelled",
        "cancelled": "cancelled",
    }
    target_state = tab_aliases.get(tab_value)
    if target_state is None:
        raise HTTPException(status_code=400, detail="tab inválido. Usa abiertas, cerradas o anuladas")

    safe_limit = max(1, min(int(limit or 200), 500))

    query: Dict[str, Any] = {}
    if effective_role == "supervisor":
        query["branch_id"] = str(user.branch_id or "")
    elif effective_role in {"ventas", "cajero"}:
        query["$or"] = [
            {"salesperson_id": user.user_id},
            {"seller_id": user.user_id},
            {"created_by": user.user_id},
        ]
    elif branch_id:
        query["branch_id"] = str(branch_id)
    else:
        query["branch_id"] = str(user.branch_id or "")

    if target_state == "cancelled":
        query["invoice_state"] = "cancelled"
    elif target_state == "closed":
        query["invoice_state"] = {"$ne": "cancelled"}
        query["payment_status"] = "paid"
    else:
        query["invoice_state"] = {"$ne": "cancelled"}
        query["payment_status"] = {"$in": ["pending", "partial"]}

    search_value = str(search or "").strip()
    if search_value:
        search_filter = {
            "$or": [
                {"invoice_number": {"$regex": search_value, "$options": "i"}},
                {"customer_name": {"$regex": search_value, "$options": "i"}},
                {"sale_id": {"$regex": search_value, "$options": "i"}},
            ]
        }
        if "$or" in query:
            query = {"$and": [query, search_filter]}
        else:
            query["$or"] = search_filter["$or"]

    projection = {
        "_id": 0,
        "sale_id": 1,
        "invoice_number": 1,
        "customer_id": 1,
        "customer_name": 1,
        "branch_id": 1,
        "created_at": 1,
        "payment_status": 1,
        "payment_method": 1,
        "payment_type": 1,
        "subtotal": 1,
        "discount": 1,
        "discounts_applied_amount": 1,
        "iva_amount": 1,
        "tax": 1,
        "total_legal": 1,
        "total": 1,
        "net_to_collect": 1,
        "amount_paid": 1,
        "amount_pending": 1,
        "retention_amount": 1,
        "retention_rate": 1,
        "invoice_state": 1,
        "cancelled_at": 1,
        "cancelled_by_name": 1,
        "cancel_reason": 1,
    }

    docs = await db.sales.find(query, projection).sort("created_at", -1).to_list(safe_limit)
    rows = [_cashier_invoice_summary(cast(Dict[str, Any], d)) for d in docs]

    return {
        "tab": target_state,
        "count": len(rows),
        "rows": rows,
    }


@api_router.post("/caja/facturas/{sale_id}/cobrar")
async def collect_cashier_invoice(sale_id: str, payload: CashierInvoiceCollectRequest, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor", "cajero"])

    session = await db.caja_sesiones.find_one({"session_id": payload.sesion_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión de caja no encontrada")
    _validate_session_access(user, cast(Dict[str, Any], session))
    if session.get("estado") != "abierta":
        raise HTTPException(status_code=400, detail="La sesión de caja no está abierta")

    sale = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    if str(sale.get("branch_id") or "") != str(session.get("branch_id") or ""):
        raise HTTPException(status_code=400, detail="La factura no pertenece a la misma sucursal de la sesión")
    if str(sale.get("invoice_state") or "").lower() == "cancelled":
        raise HTTPException(status_code=400, detail="La factura está anulada")

    has_card = any(_is_card_method(p.metodo) for p in payload.pagos)
    discount_amount = _round2(float(sale.get("discounts_applied_amount") or sale.get("discount") or 0.0))
    auth = payload.autorizacion_descuento_pos
    if has_card and discount_amount > 0 and not payload.force_remove_discount:
        if not auth:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "POS_DISCOUNT_CONFLICT",
                    "message": "Con tarjeta y descuento activo requiere autorización de supervisor/gerencia o remover descuento.",
                },
            )

        authorized_user_id = str(auth.autorizado_por or "").strip()
        if not authorized_user_id:
            raise HTTPException(status_code=400, detail="Debe indicar quién autoriza el descuento POS")
        if len(str(auth.justificacion_interna or "").strip()) < 20:
            raise HTTPException(status_code=400, detail="La justificación interna debe tener al menos 20 caracteres")

        authorized_user = await db.users.find_one({"user_id": authorized_user_id}, {"_id": 0})
        if not authorized_user:
            raise HTTPException(status_code=404, detail="Usuario autorizador no encontrado")
        if str(authorized_user.get("role") or "") not in {"supervisor", "gerencia"}:
            raise HTTPException(status_code=403, detail="Solo supervisor o gerencia pueden autorizar descuentos con tarjeta")

        if str(authorized_user.get("role") or "") == "supervisor":
            if str(authorized_user.get("branch_id") or "") != str(sale.get("branch_id") or ""):
                raise HTTPException(status_code=403, detail="El supervisor autorizador debe pertenecer a la misma sucursal")

    await db.sales.update_one(
        {"sale_id": sale_id},
        {"$set": {"cash_session_id": payload.sesion_id, "updated_at": datetime.now(timezone.utc).isoformat()}},
    )

    collect_payload = CashierCollectRequest(
        amount=payload.amount,
        payment_method=payload.payment_method,
        reference=payload.reference,
        notes=payload.notes,
        idempotency_key=payload.idempotency_key,
        received_amount=payload.received_amount,
        force_remove_discount=payload.force_remove_discount,
        pagos=payload.pagos,
        autorizacion_descuento_pos=payload.autorizacion_descuento_pos,
    )
    result = await collect_sale_invoice(sale_id, collect_payload, request)
    return result


@api_router.post("/caja/facturas/{sale_id}/anular")
async def cancel_cashier_invoice(sale_id: str, payload: CashierInvoiceCancelRequest, request: Request):
    user = await require_roles(request, ["gerencia", "recursos_humanos"])

    sale = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    sale = cast(Dict[str, Any], sale)

    if not _can_collect_sale_for_user(user, sale):
        raise HTTPException(status_code=403, detail="No autorizado para anular esta factura")
    if str(sale.get("invoice_state") or "").lower() == "cancelled":
        raise HTTPException(status_code=400, detail="La factura ya está anulada")
    if str(sale.get("payment_status") or "").lower() == "paid":
        raise HTTPException(status_code=400, detail="No se puede anular una factura ya pagada")

    reason = str(payload.motivo or "").strip()
    justification = str(payload.justificacion_interna or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="Debe indicar el motivo de anulación")
    if len(justification) < 20:
        raise HTTPException(status_code=400, detail="La justificación interna debe tener al menos 20 caracteres")

    authorized_by = str(payload.autorizado_por or "").strip()
    if not authorized_by:
        authorized_by = user.user_id

    now_iso = datetime.now(timezone.utc).isoformat()
    await db.sales.update_one(
        {"sale_id": sale_id},
        {
            "$set": {
                "invoice_state": "cancelled",
                "cancel_reason": reason,
                "cancel_justification_internal": justification,
                "cancel_authorized_by": authorized_by or user.user_id,
                "cancelled_by": user.user_id,
                "cancelled_by_name": user.name,
                "cancelled_at": now_iso,
                "updated_at": now_iso,
            }
        },
    )

    return {
        "message": "Factura anulada",
        "sale_id": sale_id,
        "invoice_number": sale.get("invoice_number"),
        "invoice_state": "cancelled",
        "cancelled_at": now_iso,
    }


@api_router.post("/sales/{sale_id}/requests/edit")
async def request_sale_edit(sale_id: str, payload: SaleRequestPayload, request: Request):
    user = await require_roles(request, ["ventas", "supervisor", "cajero", "gerencia", "recursos_humanos"])
    sale = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    sale = cast(Dict[str, Any], sale)

    if not can_access_sale_for_user(user, sale):
        raise HTTPException(status_code=403, detail="No autorizado para solicitar edición de esta factura")

    reason = str(payload.reason or "").strip()
    if len(reason) < 10:
        raise HTTPException(status_code=400, detail="La razón de solicitud debe tener al menos 10 caracteres")

    request_id = f"sreq_{uuid.uuid4().hex[:12]}"
    now_iso = _utc_now().isoformat()
    request_doc = {
        "request_id": request_id,
        "sale_id": sale_id,
        "invoice_number": sale.get("invoice_number"),
        "request_type": "edit",
        "reason": reason,
        "status": "pending",
        "branch_id": sale.get("branch_id"),
        "requested_by": user.user_id,
        "requested_by_name": user.name,
        "created_at": now_iso,
        "resolved_at": None,
        "resolved_by": None,
    }
    await db.sale_requests.insert_one(request_doc)

    await _notify_branch_reviewers(
        branch_id=str(sale.get("branch_id") or ""),
        message=f"Solicitud de edición de factura {sale.get('invoice_number') or sale_id} ({user.name})",
        metadata={
            "type": "sale_edit_request",
            "request_id": request_id,
            "sale_id": sale_id,
            "invoice_number": sale.get("invoice_number"),
            "branch_id": sale.get("branch_id"),
        },
        dedupe_seed=f"sale_edit_request:{request_id}",
    )

    return {"message": "Solicitud de edición enviada", "request_id": request_id}


@api_router.post("/sales/{sale_id}/requests/cancel")
async def request_sale_cancel(sale_id: str, payload: SaleRequestPayload, request: Request):
    user = await require_roles(request, ["ventas", "supervisor", "cajero", "gerencia", "recursos_humanos"])
    sale = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    sale = cast(Dict[str, Any], sale)

    if not can_access_sale_for_user(user, sale):
        raise HTTPException(status_code=403, detail="No autorizado para solicitar anulación de esta factura")

    reason = str(payload.reason or "").strip()
    if len(reason) < 10:
        raise HTTPException(status_code=400, detail="La razón de solicitud debe tener al menos 10 caracteres")

    request_id = f"sreq_{uuid.uuid4().hex[:12]}"
    now_iso = _utc_now().isoformat()
    request_doc = {
        "request_id": request_id,
        "sale_id": sale_id,
        "invoice_number": sale.get("invoice_number"),
        "request_type": "cancel",
        "reason": reason,
        "status": "pending",
        "branch_id": sale.get("branch_id"),
        "requested_by": user.user_id,
        "requested_by_name": user.name,
        "created_at": now_iso,
        "resolved_at": None,
        "resolved_by": None,
    }
    await db.sale_requests.insert_one(request_doc)

    await _notify_branch_reviewers(
        branch_id=str(sale.get("branch_id") or ""),
        message=f"Solicitud de anulación de factura {sale.get('invoice_number') or sale_id} ({user.name})",
        metadata={
            "type": "sale_cancel_request",
            "request_id": request_id,
            "sale_id": sale_id,
            "invoice_number": sale.get("invoice_number"),
            "branch_id": sale.get("branch_id"),
        },
        dedupe_seed=f"sale_cancel_request:{request_id}",
    )

    return {"message": "Solicitud de anulación enviada", "request_id": request_id}


@api_router.post("/sales/requests/{request_id}/approve-edit")
async def approve_sale_edit_request(request_id: str, request: Request):
    approver = await require_roles(request, ["gerencia", "recursos_humanos"])
    req = await db.sale_requests.find_one({"request_id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    if str(req.get("request_type") or "") != "edit":
        raise HTTPException(status_code=400, detail="Solicitud no corresponde a edición")
    req_status = str(req.get("status") or "")
    if req_status != "pending":
        if req_status == "approved":
            await _close_sale_request_notifications(request_id, "edit")
            return {"message": "Solicitud ya estaba aprobada", "request_id": request_id}
        raise HTTPException(status_code=400, detail="Solicitud ya procesada")

    now_iso = _utc_now().isoformat()
    await db.sale_requests.update_one(
        {"request_id": request_id},
        {
            "$set": {
                "status": "approved",
                "resolved_at": now_iso,
                "resolved_by": approver.user_id,
                "resolved_by_name": approver.name,
            }
        },
    )
    await _close_sale_request_notifications(request_id, "edit")

    sale_id = str(req.get("sale_id") or "")
    await db.sales.update_one(
        {"sale_id": sale_id},
        {
            "$set": {
                "edit_request_status": "approved",
                "edit_request_id": request_id,
                "edit_approved_at": now_iso,
                "updated_at": now_iso,
            }
        },
    )

    requester_id = str(req.get("requested_by") or "")
    if requester_id:
        await create_notification_entry(
            message=f"Tu solicitud de edición de factura {req.get('invoice_number') or sale_id} fue aprobada",
            recipient_id=requester_id,
            metadata={
                "type": "sale_edit_request_approved",
                "request_id": request_id,
                "sale_id": sale_id,
            },
            dedupe_key=f"sale_edit_request_approved:{request_id}:{requester_id}",
        )

    return {"message": "Solicitud de edición aprobada", "request_id": request_id}


@api_router.post("/sales/requests/{request_id}/approve-cancel")
async def approve_sale_cancel_request(request_id: str, request: Request):
    approver = await require_roles(request, ["gerencia", "recursos_humanos"])
    req = await db.sale_requests.find_one({"request_id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    if str(req.get("request_type") or "") != "cancel":
        raise HTTPException(status_code=400, detail="Solicitud no corresponde a anulación")
    req_status = str(req.get("status") or "")
    if req_status != "pending":
        if req_status == "approved":
            await _close_sale_request_notifications(request_id, "cancel")
            return {"message": "Solicitud ya estaba aprobada", "request_id": request_id}
        raise HTTPException(status_code=400, detail="Solicitud ya procesada")

    sale_id = str(req.get("sale_id") or "")
    sale = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    sale = cast(Dict[str, Any], sale)
    if str(sale.get("invoice_state") or "").lower() == "cancelled":
        raise HTTPException(status_code=400, detail="La factura ya está anulada")
    if str(sale.get("payment_status") or "").lower() == "paid":
        raise HTTPException(status_code=400, detail="No se puede anular una factura ya pagada")

    now_iso = _utc_now().isoformat()
    reason = str(req.get("reason") or "Solicitud de anulación")
    await db.sales.update_one(
        {"sale_id": sale_id},
        {
            "$set": {
                "invoice_state": "cancelled",
                "cancel_reason": "Solicitud aprobada",
                "cancel_justification_internal": reason,
                "cancel_authorized_by": approver.user_id,
                "cancelled_by": approver.user_id,
                "cancelled_by_name": approver.name,
                "cancelled_at": now_iso,
                "updated_at": now_iso,
            }
        },
    )

    await db.sale_requests.update_one(
        {"request_id": request_id},
        {
            "$set": {
                "status": "approved",
                "resolved_at": now_iso,
                "resolved_by": approver.user_id,
                "resolved_by_name": approver.name,
            }
        },
    )
    await _close_sale_request_notifications(request_id, "cancel")

    requester_id = str(req.get("requested_by") or "")
    if requester_id:
        await create_notification_entry(
            message=f"Tu solicitud de anulación de factura {req.get('invoice_number') or sale_id} fue aprobada",
            recipient_id=requester_id,
            metadata={
                "type": "sale_cancel_request_approved",
                "request_id": request_id,
                "sale_id": sale_id,
            },
            dedupe_key=f"sale_cancel_request_approved:{request_id}:{requester_id}",
        )

    return {"message": "Solicitud de anulación aprobada", "request_id": request_id}


@api_router.post("/cashier/invoices/{sale_id}/collect")
async def collect_sale_invoice(sale_id: str, payload: CashierCollectRequest, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor", "cajero"])

    sale = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    sale = cast(Dict[str, Any], sale)

    if not _can_collect_sale_for_user(user, sale):
        raise HTTPException(status_code=403, detail="No autorizado para cobrar esta factura")

    if sale.get("payment_status") == "paid":
        raise HTTPException(status_code=400, detail="La factura ya está pagada")

    method_to_use = _normalize_method_name(payload.payment_method or sale.get("payment_method") or sale.get("payment_type"))
    if sale.get("payment_method") and method_to_use != _normalize_method_name(sale.get("payment_method")) and not payload.force_remove_discount:
        raise HTTPException(status_code=400, detail="El cajero no puede cambiar el método de pago de la factura")

    idempotency_key = str(payload.idempotency_key or "").strip()
    if idempotency_key:
        existing = await db.invoice_payments.find_one(
            {"sale_id": sale_id, "idempotency_key": idempotency_key},
            {"_id": 0},
        )
        if existing:
            return existing

    due = _round2(float(sale.get("net_to_collect") if sale.get("net_to_collect") is not None else sale.get("total") or 0.0))
    paid_so_far_docs = await db.invoice_payments.find(
        {"sale_id": sale_id, "status": "paid"},
        {"_id": 0, "amount": 1},
    ).to_list(2000)
    paid_so_far = _round2(sum(float(p.get("amount") or 0.0) for p in paid_so_far_docs))
    pending = _round2(due - paid_so_far)

    usd_to_nio_rate = await _get_usd_to_nio_rate()

    mixed_rows: List[Dict[str, Any]] = []
    mixed_warning: Optional[str] = None
    mixed_audit_note: Optional[str] = None

    if payload.pagos:
        normalized_mixed = await _normalize_mixed_payments(payload.pagos, usd_to_nio_rate)
        pos_policy = await _enforce_pos_discount_policy(
            sale,
            has_card=bool(normalized_mixed.get("has_card")),
            authorization=payload.autorizacion_descuento_pos,
            force_remove_discount=bool(payload.force_remove_discount),
            actor=user,
        )
        sale = cast(Dict[str, Any], pos_policy.get("sale") or sale)
        mixed_warning = cast(Optional[str], pos_policy.get("warning"))
        mixed_audit_note = cast(Optional[str], pos_policy.get("audit_note"))

        due = _round2(float(sale.get("net_to_collect") if sale.get("net_to_collect") is not None else sale.get("total") or 0.0))
        pending = _round2(due - paid_so_far)

        amount = _round2(normalized_mixed.get("total_nio") or 0.0)
        if amount <= 0:
            raise HTTPException(status_code=400, detail="El total de pagos mixtos debe ser mayor a 0")
        mixed_rows = cast(List[Dict[str, Any]], normalized_mixed.get("rows") or [])
    else:
        amount = _round2(float(payload.amount or 0.0))
        if amount <= 0:
            raise HTTPException(status_code=400, detail="amount debe ser mayor a 0")

    if amount > pending:
        raise HTTPException(status_code=400, detail=f"El cobro excede el pendiente. Pendiente actual: C${pending:.2f}")

    received = _round2(float(payload.received_amount or 0.0)) if payload.received_amount is not None else None
    change_amount = 0.0
    if not payload.pagos and method_to_use == "cash":
        if received is not None and received < amount:
            raise HTTPException(status_code=400, detail="received_amount no puede ser menor al monto cobrado")
        if received is not None:
            change_amount = _round2(received - amount)

    now_iso = datetime.now(timezone.utc).isoformat()
    payment_doc = {
        "payment_id": f"ipay_{uuid.uuid4().hex[:10]}",
        "sale_id": sale_id,
        "invoice_number": sale.get("invoice_number"),
        "branch_id": sale.get("branch_id"),
        "amount": amount,
        "payment_method": method_to_use,
        "reference": payload.reference,
        "notes": payload.notes,
        "received_amount": received,
        "change_amount": change_amount,
        "cashier_id": user.user_id,
        "cashier_name": user.name,
        "idempotency_key": idempotency_key or None,
        "status": "paid",
        "created_at": now_iso,
        "session_id": sale.get("cash_session_id"),
        "cash_session_id": sale.get("cash_session_id"),
        "currency": "NIO",
        "amount_origin": amount,
        "exchange_rate": 1.0,
        "notes_auditoria": mixed_audit_note,
    }

    docs_to_insert: List[Dict[str, Any]] = []
    if mixed_rows:
        for row in mixed_rows:
            row_doc = {
                "payment_id": f"ipay_{uuid.uuid4().hex[:10]}",
                "sale_id": sale_id,
                "invoice_number": sale.get("invoice_number"),
                "branch_id": sale.get("branch_id"),
                "amount": _round2(row.get("amount_nio") or 0.0),
                "payment_method": row.get("payment_method"),
                "reference": row.get("reference"),
                "notes": payload.notes,
                "received_amount": None,
                "change_amount": 0.0,
                "cashier_id": user.user_id,
                "cashier_name": user.name,
                "idempotency_key": idempotency_key or None,
                "status": "paid",
                "created_at": now_iso,
                "session_id": sale.get("cash_session_id"),
                "cash_session_id": sale.get("cash_session_id"),
                "currency": row.get("currency"),
                "amount_origin": _round2(row.get("amount_origin") or 0.0),
                "exchange_rate": _round4(row.get("exchange_rate") or 1.0),
                "amount_nio_internal": _round4(row.get("amount_nio") or 0.0),
                "notes_auditoria": row.get("audit_note") or mixed_audit_note,
            }
            docs_to_insert.append(row_doc)
    else:
        docs_to_insert.append(payment_doc)

    if docs_to_insert:
        await db.invoice_payments.insert_many(docs_to_insert)

    new_paid = _round2(paid_so_far + amount)
    new_pending = _round2(due - new_paid)
    new_status = "paid" if new_pending <= 0.0001 else "partial"

    update_sale_set = {
        "payment_status": new_status,
        "payment_method": sale.get("payment_method") or method_to_use,
        "payment_type": sale.get("payment_type") or method_to_use,
        "amount_paid": new_paid,
        "amount_pending": max(new_pending, 0.0),
        "collected_at": now_iso,
        "collected_by": user.user_id,
    }
    if method_to_use == "card":
        total_legal = float(sale.get("total_legal") or sale.get("total") or 0.0)
        update_sale_set["pos_bank_withholding_expected"] = _round2(total_legal * 0.015)

    await db.sales.update_one({"sale_id": sale_id}, {"$set": update_sale_set})

    pending_usd_suggested = _round2((new_pending / usd_to_nio_rate) if usd_to_nio_rate > 0 else 0.0)

    safe_payment_doc = dict(payment_doc)
    safe_payment_doc.pop("_id", None)
    safe_payment_doc.pop("notes_auditoria", None)

    return {
        **safe_payment_doc,
        "sale_payment_status": new_status,
        "amount_paid_total": new_paid,
        "amount_pending": max(new_pending, 0.0),
        "amount_pending_usd_suggested": pending_usd_suggested,
        "usd_to_nio_rate": usd_to_nio_rate,
        "mixed_payments": [
            {
                k: v
                for k, v in d.items()
                if k not in {"_id", "notes_auditoria"}
            }
            for d in docs_to_insert
        ],
        "warnings": [w for w in [mixed_warning] if w],
    }


@api_router.post("/facturacion/pagar")
async def pay_invoice_facturacion(payload: FacturacionPagarRequest, request: Request):
    sale = await db.sales.find_one(
        {
            "$or": [
                {"sale_id": payload.factura_id},
                {"invoice_number": payload.factura_id},
            ]
        },
        {"_id": 0, "sale_id": 1},
    )
    if not sale or not sale.get("sale_id"):
        raise HTTPException(status_code=404, detail="Factura no encontrada")

    auth = payload.autorizacion_descuento_pos
    has_card = any(_is_card_method(p.metodo) for p in payload.pagos)
    if has_card and payload.descuento_aplicado and payload.descuento_aplicado > 0 and auth is None and not payload.force_remove_discount:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "POS_DISCOUNT_CONFLICT",
                "message": "El pago con tarjeta anula el descuento. ¿Desea proceder o requiere autorización de gerencia?",
            },
        )

    collect_payload = CashierCollectRequest(
        amount=0.0,
        payment_method=payload.pagos[0].metodo if payload.pagos else None,
        notes=payload.notas,
        idempotency_key=payload.idempotency_key,
        force_remove_discount=payload.force_remove_discount,
        pagos=payload.pagos,
        autorizacion_descuento_pos=payload.autorizacion_descuento_pos,
    )
    return await collect_sale_invoice(str(sale.get("sale_id")), collect_payload, request)


def _build_retention_receipt_pdf_bytes(receipt: Dict[str, Any], sale: Dict[str, Any]) -> bytes:
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    c.setFont("Helvetica-Bold", 13)
    c.drawString(50, 760, "Comprobante de Retención IR")
    c.setFont("Helvetica", 10)
    lines = [
        f"Correlativo: {receipt.get('receipt_number', 'N/A')}",
        f"Fecha: {receipt.get('created_at', '')}",
        f"Factura: {sale.get('invoice_number', 'N/A')}",
        f"Venta ID: {sale.get('sale_id', 'N/A')}",
        f"Cliente: {sale.get('customer_name', 'N/A')}",
        f"Sucursal: {sale.get('branch_id', 'N/A')}",
        f"Subtotal base: C${float(receipt.get('subtotal_base') or 0.0):,.2f}",
        f"Tasa retención: {float(receipt.get('retention_rate') or 0.0) * 100:.2f}%",
        f"Monto retenido: C${float(receipt.get('retention_amount') or 0.0):,.2f}",
        f"Total legal factura: C${float(sale.get('total_legal') or sale.get('total') or 0.0):,.2f}",
        f"Neto a cobrar: C${float(sale.get('net_to_collect') or sale.get('total') or 0.0):,.2f}",
        "",
        "Documento generado automáticamente por ERP.",
    ]
    y = 730
    for line in lines:
        c.drawString(50, y, line)
        y -= 18
    c.showPage()
    c.save()
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


@api_router.post("/invoices/{sale_id}/retention-receipt")
async def create_retention_receipt(sale_id: str, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor", "cajero"])
    sale = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    sale = cast(Dict[str, Any], sale)

    if user.role in {"supervisor", "cajero"}:
        if not user.branch_id or str(user.branch_id) != str(sale.get("branch_id") or ""):
            raise HTTPException(status_code=403, detail="No autorizado para generar comprobante de esta sucursal")

    retention_amount = float(sale.get("retention_amount") or 0.0)
    if retention_amount <= 0:
        raise HTTPException(status_code=400, detail="La factura no tiene retención aplicable")

    existing = await db.retention_receipts.find_one({"sale_id": sale_id}, {"_id": 0, "pdf_base64": 0})
    if existing:
        return existing

    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    count = await db.retention_receipts.count_documents({"receipt_number": {"$regex": f"^RET-{today}-"}})
    receipt_number = f"RET-{today}-{count + 1:04d}"

    now_iso = datetime.now(timezone.utc).isoformat()
    receipt_doc = {
        "receipt_id": f"rret_{uuid.uuid4().hex[:10]}",
        "receipt_number": receipt_number,
        "sale_id": sale_id,
        "invoice_number": sale.get("invoice_number"),
        "branch_id": sale.get("branch_id"),
        "customer_id": sale.get("customer_id"),
        "customer_name": sale.get("customer_name"),
        "subtotal_base": float(sale.get("subtotal") or 0.0) - float(sale.get("discounts_applied_amount") or sale.get("discount") or 0.0),
        "retention_rate": float(sale.get("retention_rate") or 0.0),
        "retention_amount": retention_amount,
        "created_by": user.user_id,
        "created_by_name": user.name,
        "created_at": now_iso,
    }

    pdf_bytes = _build_retention_receipt_pdf_bytes(receipt_doc, sale)
    receipt_doc["pdf_base64"] = base64.b64encode(pdf_bytes).decode("utf-8")
    await db.retention_receipts.insert_one(receipt_doc)

    await db.sales.update_one(
        {"sale_id": sale_id},
        {
            "$set": {
                "retention_receipt_id": receipt_doc["receipt_id"],
                "retention_receipt_number": receipt_doc["receipt_number"],
                "retention_receipt_required": True,
            }
        },
    )

    response_doc = dict(receipt_doc)
    response_doc.pop("pdf_base64", None)
    return response_doc


# ============ STRIPE PAYMENTS ============


@api_router.post("/payments/checkout")
async def create_checkout(checkout_req: CheckoutRequest, request: Request):
    await require_auth(request)
    try:
        from emergentintegrations.payments.stripe.checkout import (  # type: ignore[import]
            CheckoutSessionRequest,
            StripeCheckout,
        )
    except Exception:
        raise HTTPException(status_code=501, detail="Stripe integration not available")

    sale = await db.sales.find_one({"sale_id": checkout_req.sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    host_url = checkout_req.origin_url
    webhook_url = (
        f"{os.environ.get('REACT_APP_BACKEND_URL', host_url)}/api/webhook/stripe"
    )

    stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url=webhook_url)

    success_url = (
        f"{host_url}/sales/{sale['sale_id']}?session_id={{CHECKOUT_SESSION_ID}}"
    )
    cancel_url = f"{host_url}/sales/{sale['sale_id']}"

    checkout_request = CheckoutSessionRequest(
        amount=float(sale["total"]),
        currency="usd",
        success_url=success_url,
        cancel_url=cancel_url,
        metadata={"sale_id": sale["sale_id"], "invoice": sale["invoice_number"]},
    )

    session = await stripe_checkout.create_checkout_session(checkout_request)

    # Create payment transaction
    txn_doc = {
        "transaction_id": f"txn_{uuid.uuid4().hex[:8]}",
        "sale_id": sale["sale_id"],
        "amount": float(sale["total"]),
        "currency": "usd",
        "payment_method": "stripe",
        "stripe_session_id": session.session_id,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.payment_transactions.insert_one(txn_doc)

    # Update sale with session id
    await db.sales.update_one(
        {"sale_id": sale["sale_id"]},
        {"$set": {"stripe_session_id": session.session_id}},
    )

    return {"url": session.url, "session_id": session.session_id}


@api_router.get("/payments/status/{session_id}")
async def get_payment_status(session_id: str, request: Request):
    await require_auth(request)
    try:
        from emergentintegrations.payments.stripe.checkout import StripeCheckout  # type: ignore[import]
    except Exception:
        raise HTTPException(status_code=501, detail="Stripe integration not available")

    stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url="")
    status = await stripe_checkout.get_checkout_status(session_id)

    # Update transaction and sale if paid
    if status.payment_status == "paid":
        txn = await db.payment_transactions.find_one({"stripe_session_id": session_id})
        if txn and txn["status"] != "paid":
            await db.payment_transactions.update_one(
                {"stripe_session_id": session_id}, {"$set": {"status": "paid"}}
            )
            await db.sales.update_one(
                {"stripe_session_id": session_id}, {"$set": {"payment_status": "paid"}}
            )

    return {
        "status": status.status,
        "payment_status": status.payment_status,
        "amount_total": status.amount_total,
        "currency": status.currency,
    }


@api_router.post("/webhook/stripe")
async def stripe_webhook(request: Request):
    try:
        from emergentintegrations.payments.stripe.checkout import StripeCheckout  # type: ignore[import]
    except Exception:
        raise HTTPException(status_code=501, detail="Stripe integration not available")

    body = await request.body()
    signature = request.headers.get("Stripe-Signature")

    stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url="")
    webhook_response = await stripe_checkout.handle_webhook(body, signature)

    if webhook_response.payment_status == "paid":
        await db.payment_transactions.update_one(
            {"stripe_session_id": webhook_response.session_id},
            {"$set": {"status": "paid"}},
        )
        await db.sales.update_one(
            {"stripe_session_id": webhook_response.session_id},
            {"$set": {"payment_status": "paid"}},
        )

    return {"received": True}


# ============ WORK ORDERS ============


@api_router.get("/work-orders")
async def get_work_orders(
    request: Request, status: Optional[str] = None, branch_id: Optional[str] = None
):
    user = await require_auth(request)
    query: dict[str, Any] = {}
    if status:
        query["status"] = status
    if branch_id:
        query["branch_id"] = branch_id

    visibility_query = await build_work_order_visibility_query(user)
    query = merge_queries(query, visibility_query)

    if user.role == "instalaciones":
        query = merge_queries(query, {"technician_id": user.user_id})

    work_orders = (
        await db.work_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    )
    return work_orders


@api_router.post("/work-orders")
async def create_work_order(wo_data: WorkOrderCreate, request: Request):
    user = await require_auth(request)

    await ensure_branch_service_enabled(
        user.branch_id,
        "installations_enabled",
        "Esta sucursal no ofrece servicios de instalaciones. Solo la sucursal central puede crear órdenes de trabajo.",
    )

    customer = await db.customers.find_one(
        {"customer_id": wo_data.customer_id}, {"_id": 0}
    )
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    vehicle = await db.vehicles.find_one({"vehicle_id": wo_data.vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")

    vehicle_info = (
        f"{vehicle['brand']} {vehicle['model']} {vehicle['year']} - {vehicle['plate']}"
    )

    work_order = WorkOrder(
        sale_id=wo_data.sale_id,
        customer_id=customer["customer_id"],
        customer_name=customer["name"],
        vehicle_id=vehicle["vehicle_id"],
        vehicle_info=vehicle_info,
        branch_id=user.branch_id or "default",
        items=wo_data.items,
        priority=wo_data.priority,
        estimated_time=wo_data.estimated_time,
        notes=wo_data.notes,
    )

    doc = work_order.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.work_orders.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/work-orders/{work_order_id}")
async def update_work_order(
    work_order_id: str, update: WorkOrderStatusUpdate, request: Request
):
    await require_auth(request)

    wo = await db.work_orders.find_one({"work_order_id": work_order_id})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")

    updates = {}

    if update.status:
        updates["status"] = update.status
        if update.status == "in_progress" and not wo.get("start_time"):
            updates["start_time"] = datetime.now(timezone.utc).isoformat()
        if update.status in ["completed", "delivered"] and not wo.get("end_time"):
            updates["end_time"] = datetime.now(timezone.utc).isoformat()
            if wo.get("start_time"):
                start = datetime.fromisoformat(wo["start_time"])
                end = datetime.now(timezone.utc)
                updates["actual_time"] = int((end - start).total_seconds() / 60)

    if update.technician_id:
        tech = await db.users.find_one({"user_id": update.technician_id}, {"_id": 0})
        if tech:
            updates["technician_id"] = tech["user_id"]
            updates["technician_name"] = tech["name"]

    if update.quality_score is not None:
        updates["quality_score"] = update.quality_score
    if update.quality_notes:
        updates["quality_notes"] = update.quality_notes
    if update.notes:
        updates["notes"] = update.notes

    await db.work_orders.update_one({"work_order_id": work_order_id}, {"$set": updates})
    return {"message": "Work order updated"}


@api_router.get("/work-orders/{work_order_id}")
async def get_work_order(work_order_id: str, request: Request):
    await require_auth(request)
    wo = await db.work_orders.find_one({"work_order_id": work_order_id}, {"_id": 0})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")
    wo = cast(Dict[str, Any], wo)
    return wo


# ============ QUALITY CONTROL ============


@api_router.get("/quality-control")
async def get_quality_controls(
    request: Request,
    technician_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """Get quality control records with optional filters"""
    await require_roles(request, ["gerencia", "supervisor"])

    query: dict[str, Any] = {}
    if technician_id:
        query["technician_id"] = technician_id
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        if "created_at" in query:
            query["created_at"]["$lte"] = date_to
        else:
            query["created_at"] = {"$lte": date_to}

    qcs = (
        await db.quality_controls.find(query, {"_id": 0})
        .sort("created_at", -1)
        .to_list(500)
    )
    return qcs


@api_router.get("/quality-control/pending")
async def get_pending_quality_checks(request: Request):
    """Get work orders pending quality check"""
    user = await require_roles(request, ["gerencia", "supervisor"])

    # Get work orders in "quality_check" status that don't have a QC record yet
    query: Dict[str, Any] = {"status": "quality_check"}
    if user.role == "supervisor":
        if user.branch_id:
            query["branch_id"] = user.branch_id
        else:
            query["branch_id"] = "__no_branch__"

    work_orders = (
        await db.work_orders.find(query, {"_id": 0})
        .sort("end_time", 1)
        .to_list(100)
    )

    # Filter out ones that already have QC
    pending = []
    for wo in work_orders:
        existing_qc = await db.quality_controls.find_one(
            {"work_order_id": wo["work_order_id"]}
        )
        if not existing_qc:
            pending.append(wo)

    return pending


@api_router.post("/quality-control")
async def create_quality_control(qc_data: QualityControlCreate, request: Request):
    """Create a quality control record for a work order"""
    inspector = await require_roles(request, ["gerencia", "supervisor"])

    # Get work order
    wo = await db.work_orders.find_one(
        {"work_order_id": qc_data.work_order_id}, {"_id": 0}
    )
    if not wo:
        raise HTTPException(status_code=404, detail="Orden de trabajo no encontrada")

    # Check if QC already exists
    existing = await db.quality_controls.find_one(
        {"work_order_id": qc_data.work_order_id}
    )
    if existing:
        raise HTTPException(
            status_code=400, detail="Ya existe un control de calidad para esta orden"
        )

    # Calculate average rating
    ratings = [
        qc_data.overall_rating,
        qc_data.cleanliness_rating,
        qc_data.functionality_rating,
        qc_data.finish_rating,
        qc_data.safety_rating,
    ]
    rating_values = [r for r in ratings if r is not None]
    avg_rating = sum(rating_values) / len(rating_values) if rating_values else 0

    qc_doc = {
        "qc_id": f"qc_{uuid.uuid4().hex[:8]}",
        "work_order_id": qc_data.work_order_id,
        "technician_id": wo.get("technician_id", ""),
        "technician_name": wo.get("technician_name", "Sin asignar"),
        "inspector_id": inspector.user_id,
        "inspector_name": inspector.name,
        "overall_rating": qc_data.overall_rating,
        "cleanliness_rating": qc_data.cleanliness_rating,
        "functionality_rating": qc_data.functionality_rating,
        "finish_rating": qc_data.finish_rating,
        "safety_rating": qc_data.safety_rating,
        "average_rating": round(avg_rating, 2),
        "checklist": qc_data.checklist,
        "comments": qc_data.comments,
        "photos": qc_data.photos,
        "approved": qc_data.approved,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.quality_controls.insert_one(qc_doc)

    # Update work order with QC info
    new_status = "completed" if qc_data.approved else "quality_check"
    await db.work_orders.update_one(
        {"work_order_id": qc_data.work_order_id},
        {
            "$set": {
                "quality_score": qc_data.overall_rating,
                "quality_notes": qc_data.comments,
                "status": new_status,
            }
        },
    )

    qc_doc.pop("_id", None)
    return qc_doc


@api_router.get("/quality-control/checklist-template")
async def get_qc_checklist_template(request: Request):
    """Get standard checklist items for quality control"""
    await require_auth(request)

    return {
        "categories": [
            {
                "name": "Limpieza",
                "items": [
                    "Área de trabajo limpia",
                    "Herramientas guardadas",
                    "Vehículo limpio interior",
                    "Vehículo limpio exterior",
                    "Sin residuos de instalación",
                ],
            },
            {
                "name": "Funcionamiento",
                "items": [
                    "Producto funciona correctamente",
                    "Todas las funciones probadas",
                    "Sin ruidos anormales",
                    "Conexiones eléctricas seguras",
                    "Configuración del cliente aplicada",
                ],
            },
            {
                "name": "Acabados",
                "items": [
                    "Instalación estéticamente correcta",
                    "Sin cables visibles",
                    "Paneles bien colocados",
                    "Sin rayones o daños",
                    "Molduras en su lugar",
                ],
            },
            {
                "name": "Seguridad",
                "items": [
                    "No interfiere con airbags",
                    "Fusibles correctos instalados",
                    "Cableado protegido",
                    "Sin obstrucción de visibilidad",
                    "Sistema de alarma probado",
                ],
            },
        ]
    }


@api_router.get("/quality-control/{qc_id}")
async def get_quality_control(qc_id: str, request: Request):
    """Get a specific quality control record"""
    await require_roles(request, ["gerencia", "supervisor"])

    qc = await db.quality_controls.find_one({"qc_id": qc_id}, {"_id": 0})
    if not qc:
        raise HTTPException(status_code=404, detail="Control de calidad no encontrado")
    return qc


@api_router.get("/quality-control/work-order/{work_order_id}")
async def get_quality_control_by_work_order(work_order_id: str, request: Request):
    """Get quality control for a specific work order"""
    await require_auth(request)

    qc = await db.quality_controls.find_one(
        {"work_order_id": work_order_id}, {"_id": 0}
    )
    if not qc:
        raise HTTPException(
            status_code=404, detail="No hay control de calidad para esta orden"
        )
    return qc


@api_router.get("/quality-control/stats/technicians")
async def get_technician_quality_stats(request: Request):
    """Get quality statistics by technician"""
    await require_roles(request, ["gerencia", "supervisor"])

    # Get all QC records
    qcs = await db.quality_controls.find({}, {"_id": 0}).to_list(10000)

    # Group by technician
    stats: dict[str, Any] = {}
    for qc in qcs:
        tech_id = qc.get("technician_id", "unknown")
        if tech_id not in stats:
            stats[tech_id] = {
                "technician_id": tech_id,
                "technician_name": qc.get("technician_name", "Sin asignar"),
                "total_inspections": 0,
                "approved_count": 0,
                "rejected_count": 0,
                "total_rating": 0,
                "ratings": {
                    "overall": [],
                    "cleanliness": [],
                    "functionality": [],
                    "finish": [],
                    "safety": [],
                },
            }

        stats[tech_id]["total_inspections"] += 1
        if qc.get("approved"):
            stats[tech_id]["approved_count"] += 1
        else:
            stats[tech_id]["rejected_count"] += 1

        stats[tech_id]["total_rating"] += qc.get("overall_rating", 0)
        stats[tech_id]["ratings"]["overall"].append(qc.get("overall_rating", 0))
        stats[tech_id]["ratings"]["cleanliness"].append(qc.get("cleanliness_rating", 0))
        stats[tech_id]["ratings"]["functionality"].append(
            qc.get("functionality_rating", 0)
        )
        stats[tech_id]["ratings"]["finish"].append(qc.get("finish_rating", 0))
        stats[tech_id]["ratings"]["safety"].append(qc.get("safety_rating", 0))

    # Calculate averages
    result = []
    for tech_id, data in stats.items():
        if data["total_inspections"] > 0:
            data["average_rating"] = round(
                data["total_rating"] / data["total_inspections"], 2
            )
            data["approval_rate"] = round(
                (data["approved_count"] / data["total_inspections"]) * 100, 1
            )
            for key in data["ratings"]:
                if data["ratings"][key]:
                    data["ratings"][key] = round(
                        sum(data["ratings"][key]) / len(data["ratings"][key]), 2
                    )
                else:
                    data["ratings"][key] = 0
        result.append(data)

    # Sort by average rating
    result.sort(key=lambda x: x.get("average_rating", 0), reverse=True)
    return result


# ============ KDS (Kitchen Display System) ============


@api_router.get("/kds/orders")
async def get_kds_orders(request: Request, branch_id: Optional[str] = None):
    user = await require_auth(request)
    query: dict[str, Any] = {"status": {"$in": ["pending", "in_progress", "quality_check"]}}
    if branch_id:
        query["branch_id"] = branch_id

    visibility_query = await build_work_order_visibility_query(user)
    query = merge_queries(query, visibility_query)

    orders = (
        await db.work_orders.find(query, {"_id": 0})
        .sort([("priority", -1), ("created_at", 1)])
        .to_list(100)
    )

    return orders


@api_router.get("/kds/warehouse")
async def get_kds_warehouse(request: Request, warehouse_id: Optional[str] = None):
    await require_auth(request)

    # Get pending dispatch orders for warehouse
    query: dict[str, Any] = {"status": {"$in": ["pending", "in_progress"]}}
    if warehouse_id:
        query["warehouse_id"] = warehouse_id

    dispatches = (
        await db.dispatch_orders.find(query, {"_id": 0})
        .sort([("priority", -1), ("created_at", 1)])
        .to_list(200)
    )
    return dispatches


# ============ PROMOTIONS ============


@api_router.get("/promotions")
async def get_promotions(request: Request, active_only: bool = True):
    await require_auth(request)
    query: dict[str, Any] = {}
    if active_only:
        now = datetime.now(timezone.utc).isoformat()
        query = {
            "is_active": True,
            "start_date": {"$lte": now},
            "end_date": {"$gte": now},
        }
    promotions = await db.promotions.find(query, {"_id": 0}).to_list(100)
    return promotions


@api_router.post("/promotions")
async def create_promotion(promo: Promotion, request: Request):
    await require_roles(request, ["gerencia", "supervisor"])
    doc = promo.model_dump()
    # Ensure promotion_id exists for later references
    if not doc.get("promotion_id"):
        doc["promotion_id"] = f"promo_{uuid.uuid4().hex[:10]}"

    # Normalize dates: accept datetime or ISO strings
    try:
        if "start_date" in doc and isinstance(doc["start_date"], datetime):
            doc["start_date"] = doc["start_date"].isoformat()
    except Exception:
        pass
    try:
        if "end_date" in doc and isinstance(doc["end_date"], datetime):
            doc["end_date"] = doc["end_date"].isoformat()
    except Exception:
        pass

    await db.promotions.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/promotions/{promotion_id}")
async def update_promotion(
    promotion_id: str, updates: Dict[str, Any], request: Request
):
    await require_roles(request, ["gerencia", "supervisor"])
    if "start_date" in updates and isinstance(updates["start_date"], datetime):
        updates["start_date"] = updates["start_date"].isoformat()
    if "end_date" in updates and isinstance(updates["end_date"], datetime):
        updates["end_date"] = updates["end_date"].isoformat()
    result = await db.promotions.update_one(
        {"promotion_id": promotion_id}, {"$set": updates}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Promotion not found")
    return {"message": "Promotion updated"}


@api_router.delete("/promotions/{promotion_id}")
async def delete_promotion(promotion_id: str, request: Request):
    await require_roles(request, ["gerencia"])
    result = await db.promotions.delete_one({"promotion_id": promotion_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Promotion not found")
    return {"message": "Promotion deleted"}


# ============ DELIVERY ============


class DeliveryStatusUpdate(BaseModel):
    status: str  # pending, assigned, in_transit, delivered, failed
    driver_id: Optional[str] = None
    notes: Optional[str] = None
    delivery_proof: Optional[str] = None  # URL of delivery proof image


@api_router.get("/deliveries")
async def get_deliveries(
    request: Request, status: Optional[str] = None, driver_id: Optional[str] = None
):
    user = await require_auth(request)
    query: dict[str, Any] = {"delivery_required": True}
    if status:
        query["delivery_status"] = status
    if driver_id:
        query["delivery_driver_id"] = driver_id
    elif user.role == "transporte":
        # Drivers see their assigned deliveries or unassigned ones
        query["$or"] = [
            {"delivery_driver_id": user.user_id},
            {"delivery_driver_id": None, "delivery_status": "pending"},
        ]

    deliveries = (
        await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    )

    # Enrich with customer data
    for d in deliveries:
        customer = await db.customers.find_one(
            {"customer_id": d["customer_id"]}, {"_id": 0}
        )
        if customer:
            d["customer"] = customer

    return deliveries


@api_router.put("/deliveries/{sale_id}")
async def update_delivery(sale_id: str, update: DeliveryStatusUpdate, request: Request):
    await require_auth(request)

    sale = await db.sales.find_one({"sale_id": sale_id})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    sale = cast(Dict[str, Any], sale)

    updates = {"delivery_status": update.status}

    if update.driver_id:
        driver = await db.users.find_one({"user_id": update.driver_id}, {"_id": 0})
        if driver:
            updates["delivery_driver_id"] = driver["user_id"]
            updates["delivery_driver_name"] = driver["name"]

    if update.status == "in_transit":
        updates["delivery_started_at"] = datetime.now(timezone.utc).isoformat()

    if update.status == "delivered":
        updates["delivery_completed_at"] = datetime.now(timezone.utc).isoformat()
        if update.delivery_proof:
            updates["delivery_proof"] = update.delivery_proof

    if update.notes:
        updates["delivery_notes"] = update.notes

    await db.sales.update_one({"sale_id": sale_id}, {"$set": updates})
    return {"message": "Delivery updated"}


@api_router.post("/deliveries/{sale_id}/assign")
async def assign_delivery(sale_id: str, driver_id: str, request: Request):
    await require_roles(request, ["gerencia", "supervisor", "transporte"])

    driver = await db.users.find_one({"user_id": driver_id}, {"_id": 0})
    if not driver:
        raise HTTPException(status_code=404, detail="Driver not found")

    result = await db.sales.update_one(
        {"sale_id": sale_id, "delivery_required": True},
        {
            "$set": {
                "delivery_driver_id": driver["user_id"],
                "delivery_driver_name": driver["name"],
                "delivery_status": "assigned",
            }
        },
    )
    if result.modified_count == 0:
        raise HTTPException(
            status_code=404, detail="Sale not found or delivery not required"
        )
    return {"message": f"Delivery assigned to {driver['name']}"}


@api_router.get("/deliveries/drivers")
async def get_available_drivers(request: Request):
    await require_auth(request)
    drivers = await db.users.find(
        {"role": "transporte", "is_active": True}, {"_id": 0}
    ).to_list(100)
    return drivers


# ============ CREDIT PAYMENTS ============


# ============ CASH REGISTER / ARQUEO ============


class CashDenominationItem(FlexibleModel):
    valor_nominal: Optional[float] = None
    denominacion: Optional[float] = None
    cantidad: int = 0
    tipo: Optional[str] = "billete"
    moneda: str = "NIO"
    subtotal: Optional[float] = None


class CashOpenRequest(FlexibleModel):
    usuario_id: Optional[str] = None
    caja_id: str
    denominaciones: List[CashDenominationItem] = Field(default_factory=list)
    tipo_cambio_usd_nio: Optional[float] = None
    observaciones: Optional[str] = None


class CashMovementCreate(FlexibleModel):
    sesion_id: str
    tipo: str  # entrada | salida
    monto: float
    moneda: str = "NIO"
    referencia: Optional[str] = None
    observaciones: Optional[str] = None


class CashCloseRequest(FlexibleModel):
    sesion_id: str
    conteo_fisico: List[CashDenominationItem] = Field(default_factory=list)
    observaciones: Optional[str] = None
    otros_pagos: Optional[Dict[str, float]] = None
    tipo_cambio_usd_nio: Optional[float] = None


class CashExpenseCreate(FlexibleModel):
    sesion_id: str
    monto: float
    moneda: str = "NIO"
    concepto: str
    beneficiario: Optional[str] = None
    numero_vale: Optional[str] = None
    referencia: Optional[str] = None
    observaciones: Optional[str] = None


class CashBlindArqueoRequest(FlexibleModel):
    sesion_id: str
    conteo_fisico: List[CashDenominationItem] = Field(default_factory=list)
    tipo_cambio_usd_nio: Optional[float] = None


def _parse_dt_utc(value: Any) -> Optional[datetime]:
    if not value:
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
    except Exception:
        return None


def _to_iso_utc(value: datetime) -> str:
    dt = value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc).isoformat()


def _normalize_payment_type(value: Any) -> str:
    return str(value or "").strip().lower()


def _is_cash_payment_type(value: Any) -> bool:
    normalized = _normalize_payment_type(value)
    return normalized in {"cash", "efectivo"}


def _validate_session_access(user: User, session_doc: Dict[str, Any]) -> None:
    session_branch_id = str(session_doc.get("branch_id") or "")
    if user.role == "gerencia":
        return
    if user.role in {"supervisor", "cajero"}:
        if not user.branch_id or str(user.branch_id) != session_branch_id:
            raise HTTPException(status_code=403, detail="No autorizado para esta sesión de caja")
        return
    raise HTTPException(status_code=403, detail="Forbidden")


async def _get_usd_to_nio_rate_with_source(rate_override: Optional[float] = None) -> Tuple[float, str]:
    if rate_override and rate_override > 0:
        return float(rate_override), "override"

    try:
        billing = await _get_billing_settings_doc()
        exchange_doc = billing.get("exchange") or {}
        selected_rate, selected_source = _select_effective_billing_rate(exchange_doc, _utc_now())
        if selected_rate and selected_rate > 0:
            return float(selected_rate), selected_source
    except Exception:
        pass

    rate_doc = await db.exchange_rates.find_one(
        {"from_currency": "USD", "to_currency": "NIO"}, {"_id": 0}
    )
    if rate_doc and rate_doc.get("rate"):
        try:
            return float(rate_doc.get("rate")), "database"
        except Exception:
            pass

    inverse_doc = await db.exchange_rates.find_one(
        {"from_currency": "NIO", "to_currency": "USD"}, {"_id": 0}
    )
    if inverse_doc and inverse_doc.get("rate"):
        try:
            inverse = float(inverse_doc.get("rate"))
            if inverse > 0:
                return 1.0 / inverse, "inverse"
        except Exception:
            pass

    return 36.5, "default"


async def _get_usd_to_nio_rate(rate_override: Optional[float] = None) -> float:
    rate, _ = await _get_usd_to_nio_rate_with_source(rate_override)
    return rate


async def _normalize_denominations(
    rows: List[CashDenominationItem],
    usd_to_nio_rate: float,
) -> Dict[str, Any]:
    normalized_rows: List[Dict[str, Any]] = []
    totals_by_currency: Dict[str, float] = {"NIO": 0.0, "USD": 0.0}

    for idx, raw in enumerate(rows):
        nominal = raw.valor_nominal if raw.valor_nominal is not None else raw.denominacion
        if nominal is None:
            raise HTTPException(status_code=400, detail=f"Denominación inválida en fila {idx + 1}: falta valor_nominal")
        try:
            nominal_value = float(nominal)
        except Exception:
            raise HTTPException(status_code=400, detail=f"Denominación inválida en fila {idx + 1}")

        if nominal_value <= 0:
            raise HTTPException(status_code=400, detail=f"valor_nominal debe ser mayor a 0 en fila {idx + 1}")

        qty = int(raw.cantidad or 0)
        if qty < 0:
            raise HTTPException(status_code=400, detail=f"cantidad no puede ser negativa en fila {idx + 1}")

        currency = str(raw.moneda or "NIO").strip().upper()
        if currency not in {"NIO", "USD"}:
            raise HTTPException(status_code=400, detail=f"Moneda no soportada en fila {idx + 1}: {currency}")

        subtotal = round(nominal_value * qty, 2)
        totals_by_currency[currency] = round(totals_by_currency.get(currency, 0.0) + subtotal, 2)

        normalized_rows.append(
            {
                "line_no": idx + 1,
                "valor_nominal": nominal_value,
                "cantidad": qty,
                "tipo": str(raw.tipo or "billete").strip().lower() or "billete",
                "moneda": currency,
                "subtotal": subtotal,
            }
        )

    total_nio_equiv = round(totals_by_currency.get("NIO", 0.0) + (totals_by_currency.get("USD", 0.0) * usd_to_nio_rate), 2)

    return {
        "rows": normalized_rows,
        "totals_by_currency": totals_by_currency,
        "total_nio_equiv": total_nio_equiv,
    }


async def _compute_cash_theoretical(session_doc: Dict[str, Any], as_of: Optional[datetime] = None) -> Dict[str, Any]:
    session_id = str(session_doc.get("session_id") or "")
    branch_id = str(session_doc.get("branch_id") or "")
    opened_at = _parse_dt_utc(session_doc.get("opened_at"))
    if not opened_at:
        opened_at = datetime.now(timezone.utc)

    as_of_dt = as_of or datetime.now(timezone.utc)
    usd_to_nio_rate = await _get_usd_to_nio_rate(session_doc.get("usd_to_nio_rate"))

    opening_nio_equiv = float(session_doc.get("opening_total_nio_equiv") or 0.0)
    opening_totals = cast(Dict[str, Any], session_doc.get("opening_totals") or {})
    opening_nio = _round2(float(opening_totals.get("NIO") or 0.0))
    opening_usd = _round2(float(opening_totals.get("USD") or 0.0))

    payment_docs = await db.invoice_payments.find(
        {
            "$or": [
                {"session_id": session_id},
                {"cash_session_id": session_id},
            ],
            "status": "paid",
            "created_at": {
                "$gte": _to_iso_utc(opened_at),
                "$lte": _to_iso_utc(as_of_dt),
            },
        },
        {
            "_id": 0,
            "payment_method": 1,
            "currency": 1,
            "amount": 1,
            "amount_origin": 1,
            "notes_auditoria": 1,
        },
    ).to_list(10000)

    method_currency_summary: Dict[str, Dict[str, float]] = {
        "cash": {"NIO": 0.0, "USD": 0.0},
        "card": {"NIO": 0.0, "USD": 0.0},
        "transfer": {"NIO": 0.0, "USD": 0.0},
    }

    pos_discount_notes: List[str] = []
    cash_sales_nio_currency = 0.0
    cash_sales_usd_currency = 0.0

    if payment_docs:
        for p in payment_docs:
            method = _normalize_method_name(p.get("payment_method"))
            if method not in method_currency_summary:
                method_currency_summary[method] = {"NIO": 0.0, "USD": 0.0}
            currency = _currency_code(p.get("currency"))

            origin_amount = p.get("amount_origin")
            if origin_amount is None:
                if currency == "NIO":
                    origin_amount = p.get("amount") or 0.0
                else:
                    amount_nio = float(p.get("amount") or 0.0)
                    origin_amount = amount_nio / usd_to_nio_rate if usd_to_nio_rate > 0 else 0.0

            amount_origin = _round2(float(origin_amount or 0.0))
            method_currency_summary[method][currency] = _round2(method_currency_summary[method].get(currency, 0.0) + amount_origin)

            if method == "cash":
                if currency == "NIO":
                    cash_sales_nio_currency = _round2(cash_sales_nio_currency + amount_origin)
                else:
                    cash_sales_usd_currency = _round2(cash_sales_usd_currency + amount_origin)

            note = str(p.get("notes_auditoria") or "").strip()
            if note and "POS_DISCOUNT" in note and note not in pos_discount_notes:
                pos_discount_notes.append(note)

    sales_docs = await db.sales.find(
        {
            "branch_id": branch_id,
            "payment_status": {"$in": ["paid", "partial"]},
            "$or": [
                {"cash_session_id": session_id},
                {
                    "cash_session_id": {"$exists": False},
                    "created_at": {
                        "$gte": _to_iso_utc(opened_at),
                        "$lte": _to_iso_utc(as_of_dt),
                    },
                },
            ],
        },
        {"_id": 0, "sale_id": 1, "total": 1, "payment_type": 1},
    ).to_list(5000)

    cash_sales_total = round(
        sum(float(s.get("total") or 0.0) for s in sales_docs if _is_cash_payment_type(s.get("payment_type"))),
        2,
    )

    if not payment_docs:
        cash_sales_nio_currency = cash_sales_total
        cash_sales_usd_currency = 0.0

    returns_docs = await db.returns.find(
        {
            "branch_id": branch_id,
            "status": "completed",
            "created_at": {
                "$gte": _to_iso_utc(opened_at),
                "$lte": _to_iso_utc(as_of_dt),
            },
        },
        {"_id": 0, "sale_id": 1, "total_refund": 1},
    ).to_list(5000)

    sale_ids = [str(r.get("sale_id")) for r in returns_docs if r.get("sale_id")]
    sale_map: Dict[str, Dict[str, Any]] = {}
    if sale_ids:
        linked_sales = await db.sales.find(
            {"sale_id": {"$in": sale_ids}},
            {"_id": 0, "sale_id": 1, "payment_type": 1},
        ).to_list(len(sale_ids) + 5)
        sale_map = {str(s.get("sale_id")): s for s in linked_sales if s.get("sale_id")}

    cash_returns_total = 0.0
    cash_returns_nio_currency = 0.0
    cash_returns_usd_currency = 0.0
    for ret in returns_docs:
        sale = sale_map.get(str(ret.get("sale_id") or ""))
        if sale and _is_cash_payment_type(sale.get("payment_type")):
            refund = float(ret.get("total_refund") or 0.0)
            cash_returns_total += refund
            currency = _currency_code(sale.get("currency") or "NIO")
            if currency == "USD":
                cash_returns_usd_currency = _round2(cash_returns_usd_currency + (refund / usd_to_nio_rate if usd_to_nio_rate > 0 else 0.0))
            else:
                cash_returns_nio_currency = _round2(cash_returns_nio_currency + refund)
    cash_returns_total = round(cash_returns_total, 2)

    movement_docs = await db.caja_movimientos.find(
        {
            "session_id": session_id,
            "status": "active",
            "created_at": {
                "$gte": _to_iso_utc(opened_at),
                "$lte": _to_iso_utc(as_of_dt),
            },
        },
        {"_id": 0, "tipo": 1, "monto": 1, "moneda": 1},
    ).to_list(5000)

    entradas_nio = 0.0
    salidas_nio = 0.0
    entradas_nio_currency = 0.0
    entradas_usd_currency = 0.0
    salidas_nio_currency = 0.0
    salidas_usd_currency = 0.0
    for mov in movement_docs:
        currency = str(mov.get("moneda") or "NIO").upper()
        amount = float(mov.get("monto") or 0.0)
        amount_nio = amount if currency == "NIO" else round(amount * usd_to_nio_rate, 2)
        if str(mov.get("tipo") or "").lower() == "entrada":
            entradas_nio += amount_nio
            if currency == "USD":
                entradas_usd_currency = _round2(entradas_usd_currency + amount)
            else:
                entradas_nio_currency = _round2(entradas_nio_currency + amount)
        elif str(mov.get("tipo") or "").lower() == "salida":
            salidas_nio += amount_nio
            if currency == "USD":
                salidas_usd_currency = _round2(salidas_usd_currency + amount)
            else:
                salidas_nio_currency = _round2(salidas_nio_currency + amount)

    entradas_nio = round(entradas_nio, 2)
    salidas_nio = round(salidas_nio, 2)
    saldo_teorico = round(opening_nio_equiv + cash_sales_total - cash_returns_total + entradas_nio - salidas_nio, 2)

    expected_nio_currency = _round2(opening_nio + cash_sales_nio_currency - cash_returns_nio_currency + entradas_nio_currency - salidas_nio_currency)
    expected_usd_currency = _round2(opening_usd + cash_sales_usd_currency - cash_returns_usd_currency + entradas_usd_currency - salidas_usd_currency)

    physical_totals = cast(Dict[str, Any], session_doc.get("closing_totals") or {"NIO": 0.0, "USD": 0.0})
    physical_nio_currency = _round2(float(physical_totals.get("NIO") or 0.0))
    physical_usd_currency = _round2(float(physical_totals.get("USD") or 0.0))

    diff_nio_currency = _round2(physical_nio_currency - expected_nio_currency)
    diff_usd_currency = _round2(physical_usd_currency - expected_usd_currency)

    def _diff_label(v: float, symbol: str) -> str:
        if v > 0:
            return f"+ {symbol}{abs(v):,.2f} (Sobrante)"
        if v < 0:
            return f"- {symbol}{abs(v):,.2f} (Faltante)"
        return f"{symbol}0.00 (Cuadrado)"

    resumen_entradas = {
        "efectivo": {
            "nio": _round2(method_currency_summary.get("cash", {}).get("NIO", 0.0)),
            "usd": _round2(method_currency_summary.get("cash", {}).get("USD", 0.0)),
        },
        "tarjeta": {
            "nio": _round2(method_currency_summary.get("card", {}).get("NIO", 0.0)),
            "usd": _round2(method_currency_summary.get("card", {}).get("USD", 0.0)),
        },
        "transferencia": {
            "nio": _round2(method_currency_summary.get("transfer", {}).get("NIO", 0.0)),
            "usd": _round2(method_currency_summary.get("transfer", {}).get("USD", 0.0)),
        },
    }
    resumen_entradas["totales"] = {
        "nio": _round2(sum(v.get("nio", 0.0) for v in resumen_entradas.values() if isinstance(v, dict))),
        "usd": _round2(sum(v.get("usd", 0.0) for v in resumen_entradas.values() if isinstance(v, dict))),
    }

    comparativo = {
        "nio": {
            "fondo_inicial": opening_nio,
            "ventas_efectivo": cash_sales_nio_currency,
            "salidas_caja": salidas_nio_currency,
            "saldo_esperado": expected_nio_currency,
            "conteo_fisico": physical_nio_currency,
            "diferencia": diff_nio_currency,
            "diferencia_label": _diff_label(diff_nio_currency, "C$"),
        },
        "usd": {
            "fondo_inicial": opening_usd,
            "ventas_efectivo": cash_sales_usd_currency,
            "salidas_caja": salidas_usd_currency,
            "saldo_esperado": expected_usd_currency,
            "conteo_fisico": physical_usd_currency,
            "diferencia": diff_usd_currency,
            "diferencia_label": _diff_label(diff_usd_currency, "$"),
        },
    }

    return {
        "opening_total_nio_equiv": round(opening_nio_equiv, 2),
        "cash_sales_total": cash_sales_total,
        "cash_returns_total": cash_returns_total,
        "manual_entries_total_nio": entradas_nio,
        "manual_exits_total_nio": salidas_nio,
        "saldo_teorico_nio": saldo_teorico,
        "usd_to_nio_rate": usd_to_nio_rate,
        "expected_by_currency": {
            "NIO": expected_nio_currency,
            "USD": expected_usd_currency,
        },
        "resumen_entradas": resumen_entradas,
        "comparativo": comparativo,
        "pos_discount_audit_notes": pos_discount_notes,
    }


async def _get_cash_close_report_data(session_doc: Dict[str, Any]) -> Dict[str, Any]:
    session_id = str(session_doc.get("session_id") or "")
    branch_id = str(session_doc.get("branch_id") or "")
    opened_at = _parse_dt_utc(session_doc.get("opened_at")) or datetime.now(timezone.utc)
    closed_at = _parse_dt_utc(session_doc.get("closed_at")) or datetime.now(timezone.utc)

    theoretical = await _compute_cash_theoretical(session_doc, as_of=closed_at)

    payments = await db.invoice_payments.find(
        {
            "$or": [
                {"session_id": session_id},
                {"cash_session_id": session_id},
            ],
            "status": "paid",
            "created_at": {
                "$gte": _to_iso_utc(opened_at),
                "$lte": _to_iso_utc(closed_at),
            },
        },
        {"_id": 0},
    ).to_list(10000)

    expenses = await db.caja_egresos.find(
        {"session_id": session_id, "status": "active"},
        {"_id": 0},
    ).sort("created_at", 1).to_list(5000)

    pos_alerts: List[Dict[str, Any]] = []
    for p in payments:
        note = str(p.get("notes_auditoria") or "")
        if "POS_DISCOUNT" in note:
            pos_alerts.append(
                {
                    "invoice_number": p.get("invoice_number"),
                    "sale_id": p.get("sale_id"),
                    "justificacion": note,
                    "payment_method": p.get("payment_method"),
                    "amount": p.get("amount"),
                    "currency": p.get("currency"),
                }
            )

    sales_docs = await db.sales.find(
        {
            "branch_id": branch_id,
            "$or": [
                {"cash_session_id": session_id},
                {
                    "cash_session_id": {"$exists": False},
                    "created_at": {
                        "$gte": _to_iso_utc(opened_at),
                        "$lte": _to_iso_utc(closed_at),
                    },
                },
            ],
        },
        {
            "_id": 0,
            "sale_id": 1,
            "invoice_number": 1,
            "customer_name": 1,
            "total_legal": 1,
            "iva_amount": 1,
            "retention_amount": 1,
            "retention_rate": 1,
            "payment_method": 1,
            "payment_type": 1,
        },
    ).to_list(10000)

    iva_total = _round2(sum(float(s.get("iva_amount") or 0.0) for s in sales_docs))
    retention_1_total = _round2(
        sum(
            float(s.get("retention_amount") or 0.0)
            for s in sales_docs
            if abs(float(s.get("retention_rate") or 0.0) - 0.01) < 0.0001
        )
    )
    retention_2_total = _round2(
        sum(
            float(s.get("retention_amount") or 0.0)
            for s in sales_docs
            if abs(float(s.get("retention_rate") or 0.0) - 0.02) < 0.0001
        )
    )

    closed_details = await db.caja_conteo_detalle.find(
        {"session_id": session_id, "conteo_tipo": "cierre"},
        {"_id": 0},
    ).sort("line_no", 1).to_list(2000)

    branch = await db.branches.find_one({"branch_id": branch_id}, {"_id": 0, "name": 1})
    branch_name = str((branch or {}).get("name") or branch_id)

    return {
        "session": {
            "session_id": session_id,
            "branch_id": branch_id,
            "branch_name": branch_name,
            "caja_id": session_doc.get("caja_id"),
            "opened_at": session_doc.get("opened_at"),
            "closed_at": session_doc.get("closed_at"),
            "opened_by_name": session_doc.get("opened_by_name"),
            "closed_by_name": session_doc.get("closed_by_name"),
            "estado": session_doc.get("estado"),
            "diferencia_tipo": session_doc.get("diferencia_tipo"),
            "diferencia_nio": session_doc.get("diferencia_nio"),
        },
        "theoretical": theoretical,
        "payments": payments,
        "expenses": expenses,
        "pos_alerts": pos_alerts,
        "iva_retention": {
            "iva_total": iva_total,
            "retention_1_total": retention_1_total,
            "retention_2_total": retention_2_total,
        },
        "sales": sales_docs,
        "closing_denominations": closed_details,
    }


async def _resolve_cash_close_report_recipients(branch_id: str) -> List[str]:
    env_recipients = [
        r.strip()
        for r in str(os.environ.get("CASH_CLOSE_REPORT_EMAILS", "")).split(",")
        if r.strip()
    ]
    recipients: List[str] = []
    for r in env_recipients:
        if EMAIL_RE.match(r) and r not in recipients:
            recipients.append(r)

    if recipients:
        return recipients

    managers = await db.users.find(
        {
            "role": {"$in": ["gerencia", "supervisor"]},
            "$or": [
                {"role": "gerencia"},
                {"branch_id": branch_id},
            ],
            "email": {"$exists": True, "$ne": ""},
        },
        {"_id": 0, "email": 1},
    ).to_list(100)

    for m in managers:
        email = str(m.get("email") or "").strip().lower()
        if email and EMAIL_RE.match(email) and email not in recipients:
            recipients.append(email)

    return recipients


def _build_cash_close_email_html(report: Dict[str, Any]) -> str:
    session = cast(Dict[str, Any], report.get("session") or {})
    theo = cast(Dict[str, Any], report.get("theoretical") or {})
    cmp_data = cast(Dict[str, Any], theo.get("comparativo") or {})
    resumen_entradas = cast(Dict[str, Any], theo.get("resumen_entradas") or {})
    iva_data = cast(Dict[str, Any], report.get("iva_retention") or {})

    nio_cmp = cast(Dict[str, Any], cmp_data.get("nio") or {})
    usd_cmp = cast(Dict[str, Any], cmp_data.get("usd") or {})
    card = cast(Dict[str, Any], resumen_entradas.get("tarjeta") or {})
    transfer = cast(Dict[str, Any], resumen_entradas.get("transferencia") or {})

    pos_alert_rows = ""
    for row in cast(List[Dict[str, Any]], report.get("pos_alerts") or [])[:50]:
        pos_alert_rows += (
            f"<li>{row.get('invoice_number','N/A')} - {row.get('justificacion','')}</li>"
        )
    if not pos_alert_rows:
        pos_alert_rows = "<li>Sin alertas POS con descuento.</li>"

    total_gastos_nio = _round2(
        sum(
            float(e.get("monto") or 0.0)
            if _currency_code(e.get("moneda")) == "NIO"
            else float(e.get("monto") or 0.0) * float(theo.get("usd_to_nio_rate") or 35.0)
            for e in cast(List[Dict[str, Any]], report.get("expenses") or [])
        )
    )

    return f"""
    <html><body>
      <h2>Cierre de Caja - {session.get('branch_name','Sucursal')}</h2>
      <p>Sesion: {session.get('session_id','N/A')}<br/>Estado: {session.get('diferencia_tipo','N/A')}</p>
      <h3>Resumen Efectivo (Esperado vs Real)</h3>
      <ul>
        <li>NIO Esperado: C${float(nio_cmp.get('saldo_esperado') or 0):,.2f} | Real: C${float(nio_cmp.get('conteo_fisico') or 0):,.2f} | {nio_cmp.get('diferencia_label','')}</li>
        <li>USD Esperado: ${float(usd_cmp.get('saldo_esperado') or 0):,.2f} | Real: ${float(usd_cmp.get('conteo_fisico') or 0):,.2f} | {usd_cmp.get('diferencia_label','')}</li>
      </ul>
      <h3>Ventas No Efectivas</h3>
      <ul>
        <li>Tarjetas: C${float(card.get('nio') or 0):,.2f} | USD ${float(card.get('usd') or 0):,.2f}</li>
        <li>Transferencias: C${float(transfer.get('nio') or 0):,.2f} | USD ${float(transfer.get('usd') or 0):,.2f}</li>
      </ul>
      <h3>Alertas de Auditoria (Descuento + POS)</h3>
      <ul>{pos_alert_rows}</ul>
      <h3>Gastos del Turno</h3>
      <p>Total Egresos (equiv NIO): C${total_gastos_nio:,.2f}</p>
      <h3>Resumen IVA y Retenciones</h3>
      <ul>
        <li>IVA total: C${float(iva_data.get('iva_total') or 0):,.2f}</li>
        <li>Retencion IR 1%: C${float(iva_data.get('retention_1_total') or 0):,.2f}</li>
        <li>Retencion IR 2%: C${float(iva_data.get('retention_2_total') or 0):,.2f}</li>
      </ul>
    </body></html>
    """


@api_router.post("/caja/apertura")
async def open_cash_session(payload: CashOpenRequest, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor", "cajero"])
    branch_id = str(user.branch_id or "")
    if not branch_id:
        raise HTTPException(status_code=400, detail="El usuario no tiene sucursal asignada")

    requested_user_id = str(payload.usuario_id or user.user_id)
    if user.role == "cajero" and requested_user_id != user.user_id:
        raise HTTPException(status_code=403, detail="El cajero solo puede abrir su propia sesión")

    already_open = await db.caja_sesiones.find_one(
        {
            "branch_id": branch_id,
            "caja_id": payload.caja_id,
            "estado": "abierta",
        },
        {"_id": 0, "session_id": 1},
    )
    if already_open:
        raise HTTPException(status_code=400, detail="Ya existe una sesión de caja abierta para esta caja")

    usd_to_nio_rate = await _get_usd_to_nio_rate(payload.tipo_cambio_usd_nio)
    normalized = await _normalize_denominations(payload.denominaciones, usd_to_nio_rate)
    now = datetime.now(timezone.utc)
    session_id = f"caja_{uuid.uuid4().hex[:10]}"

    session_doc = {
        "session_id": session_id,
        "branch_id": branch_id,
        "caja_id": str(payload.caja_id).strip(),
        "usuario_id": requested_user_id,
        "opened_by": user.user_id,
        "opened_by_name": user.name,
        "estado": "abierta",
        "opened_at": _to_iso_utc(now),
        "closed_at": None,
        "usd_to_nio_rate": usd_to_nio_rate,
        "opening_totals": normalized["totals_by_currency"],
        "opening_total_nio_equiv": normalized["total_nio_equiv"],
        "lock_invoicing": False,
        "observaciones_apertura": payload.observaciones,
    }
    await db.caja_sesiones.insert_one(session_doc)

    detail_docs = []
    for row in normalized["rows"]:
        detail_docs.append(
            {
                "session_id": session_id,
                "conteo_tipo": "apertura",
                "created_at": _to_iso_utc(now),
                **row,
            }
        )
    if detail_docs:
        await db.caja_conteo_detalle.insert_many(detail_docs)

    return {
        "message": "Caja abierta",
        "session_id": session_id,
        "estado": "abierta",
        "opening_totals": normalized["totals_by_currency"],
        "opening_total_nio_equiv": normalized["total_nio_equiv"],
        "usd_to_nio_rate": usd_to_nio_rate,
    }


@api_router.post("/caja/movimiento")
async def create_cash_movement(payload: CashMovementCreate, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor", "cajero"])
    session = await db.caja_sesiones.find_one({"session_id": payload.sesion_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión de caja no encontrada")
    _validate_session_access(user, session)
    if session.get("estado") != "abierta":
        raise HTTPException(status_code=400, detail="La sesión de caja no está abierta")

    movement_type = str(payload.tipo or "").strip().lower()
    if movement_type not in {"entrada", "salida"}:
        raise HTTPException(status_code=400, detail="tipo debe ser 'entrada' o 'salida'")

    if payload.monto <= 0:
        raise HTTPException(status_code=400, detail="monto debe ser mayor a 0")

    currency = str(payload.moneda or "NIO").strip().upper()
    if currency not in {"NIO", "USD"}:
        raise HTTPException(status_code=400, detail="moneda debe ser NIO o USD")

    now = datetime.now(timezone.utc)
    movement_doc = {
        "movement_id": f"cmov_{uuid.uuid4().hex[:10]}",
        "session_id": payload.sesion_id,
        "branch_id": session.get("branch_id"),
        "caja_id": session.get("caja_id"),
        "tipo": movement_type,
        "monto": round(float(payload.monto), 2),
        "moneda": currency,
        "referencia": payload.referencia,
        "observaciones": payload.observaciones,
        "created_by": user.user_id,
        "created_by_name": user.name,
        "created_at": _to_iso_utc(now),
        "status": "active",
    }
    await db.caja_movimientos.insert_one(movement_doc)
    movement_doc.pop("_id", None)
    return movement_doc


@api_router.get("/caja/arqueo/{sesion_id}")
async def get_cash_arqueo(sesion_id: str, request: Request, include_theoretical: bool = False):
    user = await require_roles(request, ["gerencia", "supervisor", "cajero"])
    session = await db.caja_sesiones.find_one({"session_id": sesion_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión de caja no encontrada")
    _validate_session_access(user, session)

    as_of = _parse_dt_utc(session.get("closed_at")) if session.get("estado") == "cerrada" else datetime.now(timezone.utc)
    theoretical = await _compute_cash_theoretical(session, as_of=as_of)

    blind_mode = user.role == "cajero" and session.get("estado") == "abierta"
    blind_submitted = bool(session.get("blind_count_submitted_at"))
    can_show_theoretical = (not blind_mode) or include_theoretical or blind_submitted

    if not can_show_theoretical:
        theoretical = {
            "usd_to_nio_rate": theoretical.get("usd_to_nio_rate"),
            "blind_mode": True,
            "message": "Cierre ciego activo. Ingrese conteo físico para ver comparación.",
        }

    return {
        "session_id": session.get("session_id"),
        "estado": session.get("estado"),
        "branch_id": session.get("branch_id"),
        "caja_id": session.get("caja_id"),
        "opened_at": session.get("opened_at"),
        "closed_at": session.get("closed_at"),
        "opening_totals": session.get("opening_totals", {}),
        "opening_total_nio_equiv": session.get("opening_total_nio_equiv", 0),
        "blind_mode": blind_mode,
        "blind_count_submitted": blind_submitted,
        **theoretical,
    }


@api_router.post("/caja/arqueo/preview-fisico")
async def preview_blind_arqueo(payload: CashBlindArqueoRequest, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor", "cajero"])
    session = await db.caja_sesiones.find_one({"session_id": payload.sesion_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión de caja no encontrada")
    _validate_session_access(user, session)
    if session.get("estado") != "abierta":
        raise HTTPException(status_code=400, detail="La sesión de caja no está abierta")

    usd_to_nio_rate = await _get_usd_to_nio_rate(payload.tipo_cambio_usd_nio or session.get("usd_to_nio_rate"))
    normalized = await _normalize_denominations(payload.conteo_fisico, usd_to_nio_rate)
    theoretical = await _compute_cash_theoretical(session, as_of=datetime.now(timezone.utc))

    expected_by_currency = cast(Dict[str, Any], theoretical.get("expected_by_currency") or {})
    expected_nio = _round2(float(expected_by_currency.get("NIO") or 0.0))
    expected_usd = _round2(float(expected_by_currency.get("USD") or 0.0))
    physical_nio = _round2(float(normalized["totals_by_currency"].get("NIO") or 0.0))
    physical_usd = _round2(float(normalized["totals_by_currency"].get("USD") or 0.0))

    await db.caja_sesiones.update_one(
        {"session_id": payload.sesion_id},
        {
            "$set": {
                "blind_count_submitted_at": datetime.now(timezone.utc).isoformat(),
                "blind_preview_totals": normalized["totals_by_currency"],
            }
        },
    )

    return {
        "session_id": payload.sesion_id,
        "expected_by_currency": {"NIO": expected_nio, "USD": expected_usd},
        "physical_by_currency": {"NIO": physical_nio, "USD": physical_usd},
        "difference_by_currency": {
            "NIO": _round2(physical_nio - expected_nio),
            "USD": _round2(physical_usd - expected_usd),
        },
        "saldo_teorico_nio": theoretical.get("saldo_teorico_nio"),
        "total_fisico_nio": normalized["total_nio_equiv"],
    }


@api_router.post("/caja/egresos")
async def create_cash_expense(payload: CashExpenseCreate, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor", "cajero"])
    session = await db.caja_sesiones.find_one({"session_id": payload.sesion_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión de caja no encontrada")
    _validate_session_access(user, session)
    if session.get("estado") != "abierta":
        raise HTTPException(status_code=400, detail="La sesión de caja no está abierta")

    currency = _currency_code(payload.moneda)
    amount = _round2(payload.monto)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="monto debe ser mayor a 0")

    theoretical = await _compute_cash_theoretical(session, as_of=datetime.now(timezone.utc))
    expected = cast(Dict[str, Any], theoretical.get("expected_by_currency") or {})
    available_in_currency = _round2(float(expected.get(currency) or 0.0))

    # Regla de seguridad: no hacer conversiones implícitas entre monedas para egresos.
    if amount > available_in_currency:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Saldo insuficiente en {currency}. Disponible: {available_in_currency:.2f}. "
                "No se permite conversión automática entre monedas para egresos. "
                "Realice primero un movimiento interno de cambio de moneda."
            ),
        )

    now_iso = datetime.now(timezone.utc).isoformat()
    expense_doc = {
        "expense_id": f"cexp_{uuid.uuid4().hex[:10]}",
        "session_id": payload.sesion_id,
        "branch_id": session.get("branch_id"),
        "caja_id": session.get("caja_id"),
        "monto": amount,
        "moneda": currency,
        "concepto": payload.concepto,
        "beneficiario": payload.beneficiario,
        "numero_vale": payload.numero_vale,
        "referencia": payload.referencia,
        "observaciones": payload.observaciones,
        "created_by": user.user_id,
        "created_by_name": user.name,
        "created_at": now_iso,
        "status": "active",
    }
    await db.caja_egresos.insert_one(expense_doc)

    movement_doc = {
        "movement_id": f"cmov_{uuid.uuid4().hex[:10]}",
        "session_id": payload.sesion_id,
        "branch_id": session.get("branch_id"),
        "caja_id": session.get("caja_id"),
        "tipo": "salida",
        "monto": amount,
        "moneda": currency,
        "referencia": payload.referencia,
        "observaciones": payload.observaciones,
        "created_by": user.user_id,
        "created_by_name": user.name,
        "created_at": now_iso,
        "status": "active",
        "movement_category": "petty_expense",
        "linked_expense_id": expense_doc["expense_id"],
        "numero_vale": payload.numero_vale,
    }
    await db.caja_movimientos.insert_one(movement_doc)

    expense_doc.pop("_id", None)
    return expense_doc


@api_router.post("/caja/cierre")
async def close_cash_session(payload: CashCloseRequest, request: Request, background_tasks: BackgroundTasks):
    user = await require_roles(request, ["gerencia", "supervisor", "cajero"])
    session = await db.caja_sesiones.find_one({"session_id": payload.sesion_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión de caja no encontrada")
    _validate_session_access(user, session)
    if session.get("estado") != "abierta":
        raise HTTPException(status_code=400, detail="La sesión de caja ya está cerrada")

    now = datetime.now(timezone.utc)
    usd_to_nio_rate = await _get_usd_to_nio_rate(payload.tipo_cambio_usd_nio or session.get("usd_to_nio_rate"))
    normalized = await _normalize_denominations(payload.conteo_fisico, usd_to_nio_rate)
    theoretical = await _compute_cash_theoretical(session, as_of=now)

    total_fisico_nio = normalized["total_nio_equiv"]
    saldo_teorico = float(theoretical.get("saldo_teorico_nio") or 0.0)
    diferencia = round(total_fisico_nio - saldo_teorico, 2)

    expected_by_currency = cast(Dict[str, Any], theoretical.get("expected_by_currency") or {})
    expected_nio_currency = _round2(float(expected_by_currency.get("NIO") or 0.0))
    expected_usd_currency = _round2(float(expected_by_currency.get("USD") or 0.0))
    physical_nio_currency = _round2(float(normalized["totals_by_currency"].get("NIO") or 0.0))
    physical_usd_currency = _round2(float(normalized["totals_by_currency"].get("USD") or 0.0))
    diff_nio_currency = _round2(physical_nio_currency - expected_nio_currency)
    diff_usd_currency = _round2(physical_usd_currency - expected_usd_currency)

    if diferencia > 0:
        diferencia_tipo = "sobrante"
    elif diferencia < 0:
        diferencia_tipo = "faltante"
    else:
        diferencia_tipo = "cuadrado"

    await db.caja_sesiones.update_one(
        {"session_id": payload.sesion_id},
        {
            "$set": {
                "estado": "cerrada",
                "closed_at": _to_iso_utc(now),
                "closed_by": user.user_id,
                "closed_by_name": user.name,
                "usd_to_nio_rate": usd_to_nio_rate,
                "closing_totals": normalized["totals_by_currency"],
                "closing_total_nio_equiv": total_fisico_nio,
                "saldo_teorico_nio": saldo_teorico,
                "diferencia_nio": diferencia,
                "diferencia_tipo": diferencia_tipo,
                "expected_by_currency": {
                    "NIO": expected_nio_currency,
                    "USD": expected_usd_currency,
                },
                "physical_by_currency": {
                    "NIO": physical_nio_currency,
                    "USD": physical_usd_currency,
                },
                "difference_by_currency": {
                    "NIO": diff_nio_currency,
                    "USD": diff_usd_currency,
                },
                "otros_pagos": payload.otros_pagos or {},
                "observaciones_cierre": payload.observaciones,
                "lock_invoicing": True,
            }
        },
    )

    detail_docs = []
    for row in normalized["rows"]:
        detail_docs.append(
            {
                "session_id": payload.sesion_id,
                "conteo_tipo": "cierre",
                "created_at": _to_iso_utc(now),
                **row,
            }
        )
    if detail_docs:
        await db.caja_conteo_detalle.insert_many(detail_docs)

    closed_session = await db.caja_sesiones.find_one({"session_id": payload.sesion_id}, {"_id": 0})
    if closed_session:
        report_data = await _get_cash_close_report_data(cast(Dict[str, Any], closed_session))
        recipients = await _resolve_cash_close_report_recipients(str(closed_session.get("branch_id") or ""))
        status_label = str((closed_session or {}).get("diferencia_tipo") or "cuadrado").capitalize()
        subject = (
            f"⚠️ Cierre de Caja - Sucursal {report_data['session'].get('branch_name','N/A')} - "
            f"{datetime.now(timezone.utc).strftime('%Y-%m-%d')} - Estado: {status_label}"
        )
        html = _build_cash_close_email_html(report_data)
        for recipient in recipients:
            background_tasks.add_task(send_email_notification, recipient, subject, html)
        await db.caja_sesiones.update_one(
            {"session_id": payload.sesion_id},
            {
                "$set": {
                    "after_close_cashier_notified_at": datetime.now(timezone.utc).isoformat(),
                    "after_close_cashier_recipients": recipients,
                }
            },
        )

    return {
        "message": "Caja cerrada",
        "session_id": payload.sesion_id,
        "estado": "cerrada",
        "saldo_teorico_nio": saldo_teorico,
        "total_fisico_nio": total_fisico_nio,
        "diferencia_nio": diferencia,
        "diferencia_tipo": diferencia_tipo,
        "closing_totals": normalized["totals_by_currency"],
        "expected_by_currency": {
            "NIO": expected_nio_currency,
            "USD": expected_usd_currency,
        },
        "difference_by_currency": {
            "NIO": diff_nio_currency,
            "USD": diff_usd_currency,
        },
        "resumen_entradas": theoretical.get("resumen_entradas", {}),
        "comparativo": theoretical.get("comparativo", {}),
        "pos_discount_audit_notes": theoretical.get("pos_discount_audit_notes", []),
        "after_close_cashier_notified": True,
    }


@api_router.get("/caja/cierre/{session_id}/reporte-gerencia")
async def get_cash_close_management_report(session_id: str, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor"])
    session = await db.caja_sesiones.find_one({"session_id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión de caja no encontrada")
    if user.role == "supervisor" and str(user.branch_id or "") != str(session.get("branch_id") or ""):
        raise HTTPException(status_code=403, detail="No autorizado para esta sucursal")
    if session.get("estado") != "cerrada":
        raise HTTPException(status_code=400, detail="La sesión debe estar cerrada para generar reporte gerencial")

    report_data = await _get_cash_close_report_data(cast(Dict[str, Any], session))
    return report_data


@api_router.get("/caja/cierre/{session_id}/excel")
async def export_cash_close_excel(session_id: str, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor"])
    session = await db.caja_sesiones.find_one({"session_id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión de caja no encontrada")
    if user.role == "supervisor" and str(user.branch_id or "") != str(session.get("branch_id") or ""):
        raise HTTPException(status_code=403, detail="No autorizado para esta sucursal")
    if session.get("estado") != "cerrada":
        raise HTTPException(status_code=400, detail="La sesión debe estar cerrada para exportar")

    report_data = await _get_cash_close_report_data(cast(Dict[str, Any], session))
    wb = Workbook()

    # Hoja 1: Resumen General
    ws1 = wb.active
    ws1.title = "Resumen General"
    session_info = cast(Dict[str, Any], report_data.get("session") or {})
    theo = cast(Dict[str, Any], report_data.get("theoretical") or {})
    iva_ret = cast(Dict[str, Any], report_data.get("iva_retention") or {})
    ws1.append(["Campo", "Valor"])
    ws1.append(["Sesion", session_info.get("session_id")])
    ws1.append(["Sucursal", session_info.get("branch_name")])
    ws1.append(["Caja", session_info.get("caja_id")])
    ws1.append(["Apertura", session_info.get("opened_at")])
    ws1.append(["Cierre", session_info.get("closed_at")])
    ws1.append(["Estado", session_info.get("diferencia_tipo")])
    ws1.append(["Saldo teorico NIO", theo.get("saldo_teorico_nio")])
    ws1.append(["IVA cobrado", iva_ret.get("iva_total")])
    ws1.append(["Retencion IR 1%", iva_ret.get("retention_1_total")])
    ws1.append(["Retencion IR 2%", iva_ret.get("retention_2_total")])

    # Hoja 2: Detalle de Pagos
    ws2 = wb.create_sheet("Detalle de Pagos")
    ws2.append([
        "Factura",
        "Sale ID",
        "Metodo",
        "Moneda",
        "Monto Origen",
        "Tipo Cambio",
        "Monto NIO",
        "Referencia",
        "Cajero",
        "Fecha",
        "Nota Auditoria",
    ])
    for p in cast(List[Dict[str, Any]], report_data.get("payments") or []):
        ws2.append([
            p.get("invoice_number"),
            p.get("sale_id"),
            p.get("payment_method"),
            p.get("currency", "NIO"),
            p.get("amount_origin", p.get("amount")),
            p.get("exchange_rate", 1),
            p.get("amount"),
            p.get("reference"),
            p.get("cashier_name"),
            p.get("created_at"),
            p.get("notes_auditoria"),
        ])

    # Hoja 3: Gastos y Vales
    ws3 = wb.create_sheet("Gastos y Vales")
    ws3.append([
        "Vale",
        "Concepto",
        "Beneficiario",
        "Moneda",
        "Monto",
        "Referencia",
        "Observaciones",
        "Creado por",
        "Fecha",
    ])
    for e in cast(List[Dict[str, Any]], report_data.get("expenses") or []):
        ws3.append([
            e.get("numero_vale"),
            e.get("concepto"),
            e.get("beneficiario"),
            e.get("moneda"),
            e.get("monto"),
            e.get("referencia"),
            e.get("observaciones"),
            e.get("created_by_name"),
            e.get("created_at"),
        ])

    # Hoja 4: Arqueo Fisico
    ws4 = wb.create_sheet("Arqueo Fisico")
    ws4.append(["Moneda", "Tipo", "Denominacion", "Cantidad", "Subtotal"])
    for r in cast(List[Dict[str, Any]], report_data.get("closing_denominations") or []):
        ws4.append([
            r.get("moneda"),
            r.get("tipo"),
            r.get("valor_nominal"),
            r.get("cantidad"),
            r.get("subtotal"),
        ])

    # widths
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 48)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"cierre_caja_{session_id}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


class CreditPaymentCreate(BaseModel):
    sale_id: str
    amount: float
    payment_method: str = "cash"  # cash, transfer, card
    reference: Optional[str] = None
    notes: Optional[str] = None


class CreditPayment(BaseModel):
    payment_id: str = Field(default_factory=lambda: f"cpay_{uuid.uuid4().hex[:8]}")
    sale_id: str
    amount: float
    payment_method: str
    reference: Optional[str] = None
    notes: Optional[str] = None
    received_by: str
    received_by_name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


@api_router.get("/credit/pending")
async def get_pending_credits(request: Request):
    await require_roles(request, ["gerencia", "supervisor", "ventas", "cajero"])

    sales = (
        await db.sales.find(
            {"payment_type": "credit", "payment_status": {"$ne": "paid"}}, {"_id": 0}
        )
        .sort("credit_due_date", 1)
        .to_list(500)
    )

    # Enrich with customer and payments data
    for sale in sales:
        customer = await db.customers.find_one(
            {"customer_id": sale["customer_id"]}, {"_id": 0}
        )
        if customer:
            sale["customer"] = customer

        payments = await db.credit_payments.find(
            {"sale_id": sale["sale_id"]}, {"_id": 0}
        ).to_list(100)
        sale["payments"] = payments
        sale["amount_paid"] = sum(p["amount"] for p in payments)
        sale["amount_pending"] = sale["total"] - sale["amount_paid"]

    return sales


@api_router.post("/credit/payment")
async def register_credit_payment(payment_data: CreditPaymentCreate, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor", "ventas", "cajero"])

    sale = await db.sales.find_one({"sale_id": payment_data.sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    if sale["payment_type"] != "credit":
        raise HTTPException(status_code=400, detail="Sale is not a credit sale")

    # Get existing payments
    existing_payments = await db.credit_payments.find(
        {"sale_id": payment_data.sale_id}, {"_id": 0}
    ).to_list(100)
    total_paid = sum(p["amount"] for p in existing_payments)
    pending = sale["total"] - total_paid

    if payment_data.amount > pending:
        raise HTTPException(
            status_code=400, detail=f"Payment exceeds pending amount: ${pending:.2f}"
        )

    payment = CreditPayment(
        sale_id=payment_data.sale_id,
        amount=payment_data.amount,
        payment_method=payment_data.payment_method,
        reference=payment_data.reference,
        notes=payment_data.notes,
        received_by=user.user_id,
        received_by_name=user.name,
    )

    doc = payment.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.credit_payments.insert_one(doc)
    doc.pop("_id", None)

    # Update customer credit balance
    await db.customers.update_one(
        {"customer_id": sale["customer_id"]},
        {"$inc": {"credit_balance": -payment_data.amount}},
    )

    # Check if fully paid
    new_total_paid = total_paid + payment_data.amount
    if new_total_paid >= sale["total"]:
        await db.sales.update_one(
            {"sale_id": payment_data.sale_id}, {"$set": {"payment_status": "paid"}}
        )
    else:
        await db.sales.update_one(
            {"sale_id": payment_data.sale_id}, {"$set": {"payment_status": "partial"}}
        )

    return doc


@api_router.get("/credit/payments/{sale_id}")
async def get_sale_payments(sale_id: str, request: Request):
    await require_auth(request)
    payments = (
        await db.credit_payments.find({"sale_id": sale_id}, {"_id": 0})
        .sort("created_at", -1)
        .to_list(100)
    )
    return payments


@api_router.get("/credit/customer/{customer_id}")
async def get_customer_credit_summary(customer_id: str, request: Request):
    await require_auth(request)

    customer = await db.customers.find_one({"customer_id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    # Get all credit sales
    credit_sales = await db.sales.find(
        {"customer_id": customer_id, "payment_type": "credit"}, {"_id": 0}
    ).to_list(500)

    # Calculate totals
    total_credit = sum(s["total"] for s in credit_sales)

    # Get all payments
    all_payments = []
    for sale in credit_sales:
        payments = await db.credit_payments.find(
            {"sale_id": sale["sale_id"]}, {"_id": 0}
        ).to_list(100)
        sale["payments"] = payments
        sale["amount_paid"] = sum(p["amount"] for p in payments)
        all_payments.extend(payments)

    total_paid = sum(p["amount"] for p in all_payments)

    return {
        "customer": customer,
        "credit_limit": customer.get("credit_limit", 0),
        "credit_balance": customer.get("credit_balance", 0),
        "total_credit_purchases": total_credit,
        "total_paid": total_paid,
        "pending_amount": total_credit - total_paid,
        "credit_sales": credit_sales,
    }


# ============ REPORTS ============


@api_router.get("/reports/sales")
async def get_sales_report(
    request: Request, start_date: str, end_date: str, branch_id: Optional[str] = None
):
    await require_roles(request, ["gerencia", "supervisor"])

    query: dict[str, Any] = {"created_at": {"$gte": start_date, "$lte": end_date}}
    if branch_id:
        query["branch_id"] = branch_id

    sales = await db.sales.find(query, {"_id": 0}).to_list(10000)

    total_sales = sum(s["total"] for s in sales)
    total_count = len(sales)
    by_payment_type = {}
    by_salesperson = {}

    for s in sales:
        pt = s["payment_type"]
        by_payment_type[pt] = by_payment_type.get(pt, 0) + s["total"]

        sp = s["salesperson_name"]
        if sp not in by_salesperson:
            by_salesperson[sp] = {"total": 0, "count": 0}
        by_salesperson[sp]["total"] += s["total"]
        by_salesperson[sp]["count"] += 1

    return {
        "total_sales": round(total_sales, 2),
        "total_count": total_count,
        "by_payment_type": by_payment_type,
        "by_salesperson": by_salesperson,
    }


@api_router.get("/reports/installations")
async def get_installations_report(request: Request, start_date: str, end_date: str):
    await require_roles(request, ["gerencia", "supervisor"])

    query: dict[str, Any] = {
        "created_at": {"$gte": start_date, "$lte": end_date},
        "status": "completed",
    }

    work_orders = await db.work_orders.find(query, {"_id": 0}).to_list(10000)

    by_technician = {}
    total_time = 0
    quality_scores = []

    for wo in work_orders:
        tech = wo.get("technician_name", "Sin asignar")
        if tech not in by_technician:
            by_technician[tech] = {
                "count": 0,
                "total_time": 0,
                "avg_quality": 0,
                "quality_scores": [],
            }
        by_technician[tech]["count"] += 1
        if wo.get("actual_time"):
            by_technician[tech]["total_time"] += wo["actual_time"]
            total_time += wo["actual_time"]
        if wo.get("quality_score"):
            by_technician[tech]["quality_scores"].append(wo["quality_score"])
            quality_scores.append(wo["quality_score"])

    for tech in by_technician:
        scores = by_technician[tech]["quality_scores"]
        by_technician[tech]["avg_quality"] = (
            round(sum(scores) / len(scores), 1) if scores else 0
        )
        del by_technician[tech]["quality_scores"]

    return {
        "total_installations": len(work_orders),
        "total_time_minutes": total_time,
        "avg_quality": (
            round(sum(quality_scores) / len(quality_scores), 1) if quality_scores else 0
        ),
        "by_technician": by_technician,
    }


# ============ PRINTING ============

def _currency_symbol(code: str) -> str:
    try:
        return CURRENCIES.get(code, {}).get("symbol", "$")
    except Exception:
        return "$"


def _format_money(amount: float, currency: str) -> str:
    symbol = _currency_symbol(currency)
    try:
        return f"{symbol}{amount:,.2f}"
    except Exception:
        return f"{symbol}{amount}"


def _safe_date(value: Optional[str]) -> str:
    if not value:
        return ""
    return str(value).replace("T", " ").split("+")[0]


def _load_logo_image(logo_url: Optional[str]) -> Optional[ImageReader]:
    if not logo_url:
        return None
    try:
        if logo_url.startswith("data:image"):
            header, data = logo_url.split(",", 1)
            return ImageReader(BytesIO(base64.b64decode(data)))
        if logo_url.startswith("http://") or logo_url.startswith("https://"):
            resp = httpx.get(logo_url, timeout=5.0)
            resp.raise_for_status()
            return ImageReader(BytesIO(resp.content))
        # local file path
        path = Path(logo_url)
        if path.exists():
            return ImageReader(str(path))
    except Exception:
        logger.exception("Failed to load logo image")
    return None


def _draw_document_pdf(
    p: canvas.Canvas,
    doc_title: str,
    doc_number: str,
    doc_date: str,
    company: dict,
    customer: dict,
    vehicle: Optional[dict],
    items: list,
    currency: str,
    iva_rate: float,
    apply_iva: bool,
    totals: dict,
    notes: Optional[str] = None,
):
    width, height = letter
    safe_items = items if isinstance(items, list) else list(items or [])

    # Header bar
    p.setFillColor(colors.HexColor("#EAF1F9"))
    p.rect(0, height - 120, width, 90, stroke=0, fill=1)

    # Brand / Logo
    logo = _load_logo_image(company.get("logo_url"))
    if logo:
        p.drawImage(logo, 50, height - 90, width=90, height=40, preserveAspectRatio=True, mask='auto')
        brand_x = 150
    else:
        brand_x = 50
    p.setFillColor(colors.HexColor("#1F2937"))
    p.setFont("Helvetica-Bold", 18)
    p.drawString(brand_x, height - 70, company.get("name", "MUNDO DE ACCESORIOS"))
    p.setFont("Helvetica", 9)
    p.setFillColor(colors.HexColor("#6B7280"))
    p.drawString(50, height - 85, company.get("tagline", "Sistema ERP"))

    # Document info
    p.setFillColor(colors.HexColor("#1F2937"))
    p.setFont("Helvetica-Bold", 10)
    p.drawRightString(width - 50, height - 60, "Date")
    p.drawRightString(width - 50, height - 80, f"{doc_title} #")
    p.setFont("Helvetica", 10)
    p.drawRightString(width - 110, height - 60, _safe_date(doc_date))
    p.drawRightString(width - 110, height - 80, doc_number)

    # Parties block (two columns)
    left_x = 50
    right_x = width / 2 + 20
    left_y = height - 135
    right_y = height - 135

    p.setFont("Helvetica-Bold", 10)
    p.drawString(left_x, left_y, company.get("legal_name", "Supplier Company"))
    left_y -= 14
    p.setFont("Helvetica", 9)
    if company.get("tax_id"):
        p.drawString(left_x, left_y, f"RUC: {company.get('tax_id')}")
        left_y -= 12
    if company.get("vat"):
        p.drawString(left_x, left_y, f"VAT: {company.get('vat')}")
        left_y -= 12
    if company.get("address"):
        p.drawString(left_x, left_y, f"Dirección: {company.get('address')}")
        left_y -= 12
    if company.get("city"):
        p.drawString(left_x, left_y, f"Ciudad: {company.get('city')}")
        left_y -= 12
    if company.get("country"):
        p.drawString(left_x, left_y, f"País: {company.get('country')}")
        left_y -= 12
    if company.get("phone"):
        p.drawString(left_x, left_y, f"Tel: {company.get('phone')}")
        left_y -= 12
    if company.get("email"):
        p.drawString(left_x, left_y, f"Email: {company.get('email')}")
        left_y -= 12

    p.setFont("Helvetica-Bold", 10)
    p.drawString(right_x, right_y, "Cliente")
    right_y -= 14
    p.setFont("Helvetica", 9)
    if customer.get("name"):
        p.drawString(right_x, right_y, f"Nombre: {customer.get('name')}")
        right_y -= 12
    if customer.get("tax_id"):
        p.drawString(right_x, right_y, f"RUC: {customer.get('tax_id')}")
        right_y -= 12
    if customer.get("phone"):
        p.drawString(right_x, right_y, f"Tel: {customer.get('phone')}")
        right_y -= 12
    if customer.get("email"):
        p.drawString(right_x, right_y, f"Email: {customer.get('email')}")
        right_y -= 12
    if customer.get("address"):
        p.drawString(right_x, right_y, f"Dirección: {customer.get('address')}")
        right_y -= 12

    if vehicle:
        left_y -= 6
        p.setFont("Helvetica-Bold", 10)
        p.drawString(left_x, left_y, "Vehículo")
        left_y -= 14
        p.setFont("Helvetica", 9)
        vehicle_line = f"{vehicle.get('brand','')} {vehicle.get('model','')} {vehicle.get('year','')}"
        if vehicle_line.strip():
            p.drawString(left_x, left_y, vehicle_line.strip())
            left_y -= 12
        if vehicle.get("plate"):
            p.drawString(left_x, left_y, f"Placa: {vehicle.get('plate')}")
            left_y -= 12
        vin = vehicle.get("vin") or vehicle.get("chasis")
        if vin:
            p.drawString(left_x, left_y, f"Chasis: {vin}")
            left_y -= 12
        if vehicle.get("color"):
            p.drawString(left_x, left_y, f"Color: {vehicle.get('color')}")
            left_y -= 12

    # Table header
    y = min(left_y, right_y) - 10
    p.setFillColor(colors.HexColor("#E6EAF2"))
    p.rect(50, y, width - 100, 18, stroke=0, fill=1)
    p.setFillColor(colors.HexColor("#374151"))
    p.setFont("Helvetica-Bold", 9)
    p.drawString(55, y + 5, "#")
    p.drawString(70, y + 5, "Product details")
    p.drawRightString(360, y + 5, "Price")
    p.drawRightString(430, y + 5, "Qty")
    p.drawRightString(width - 60, y + 5, "Subtotal")
    y -= 12

    p.setFont("Helvetica", 9)
    index = 1
    for item in safe_items:
        if not item:
            continue
        if y < 110:
            p.showPage()
            y = height - 80
        unit_price = float(item.get("unit_price", 0) or 0)
        qty = float(item.get("quantity", 0) or 0)
        discount = float(item.get("discount", 0) or 0)
        line_subtotal = unit_price * qty * (1 - discount / 100)
        p.setFillColor(colors.HexColor("#111827"))
        p.drawString(55, y, f"{index}.")
        p.drawString(70, y, str(item.get("product_name", ""))[:45])
        p.drawRightString(360, y, _format_money(unit_price, currency))
        p.drawRightString(430, y, f"{int(qty)}")
        p.drawRightString(width - 60, y, _format_money(line_subtotal, currency))
        y -= 14
        index += 1

    # Totals block
    y -= 10
    p.setFont("Helvetica-Bold", 10)
    p.setFillColor(colors.HexColor("#374151"))
    p.drawRightString(width - 120, y, "Subtotal:")
    p.drawRightString(width - 60, y, _format_money(totals.get("subtotal", 0), currency))
    y -= 14
    p.drawRightString(width - 120, y, "IVA total:")
    p.drawRightString(width - 60, y, _format_money(totals.get("tax", 0), currency))
    y -= 20
    p.setFillColor(colors.HexColor("#5B6BD6"))
    p.rect(width - 170, y - 4, 110, 20, stroke=0, fill=1)
    p.setFillColor(colors.white)
    p.drawRightString(width - 60, y + 2, _format_money(totals.get("total", 0), currency))
    p.setFillColor(colors.white)
    p.drawString(width - 160, y + 2, "Total:")

    # Notes
    if notes:
        y -= 40
        p.setFillColor(colors.HexColor("#374151"))
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, y, "Notes")
        p.setFont("Helvetica", 9)
        p.drawString(50, y - 14, str(notes)[:120])



@api_router.get("/print/thermal/test")
async def test_thermal_printer(request: Request):
    """Generate test print commands for thermal printer"""
    await require_auth(request)

    ESC = b"\x1b"
    GS = b"\x1d"

    commands = bytearray()
    commands.extend(ESC + b"@")  # Init
    commands.extend(ESC + b"a\x01")  # Center
    commands.extend(ESC + b"!\x10")  # Double height
    commands.extend("MUNDO DE ACCESORIOS\n".encode("cp437"))
    commands.extend(ESC + b"!")  # Normal
    commands.extend("Prueba de Impresion\n".encode("cp437"))
    commands.extend(b"-" * 32 + b"\n")
    commands.extend(ESC + b"a")  # Left
    commands.extend(
        f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n".encode("cp437")
    )
    commands.extend("Impresora configurada correctamente\n".encode("cp437"))
    commands.extend(ESC + b"d\x03")  # Feed
    commands.extend(GS + b"V")  # Cut

    import base64

    return {
        "commands_base64": base64.b64encode(bytes(commands)).decode(),
        "message": "Test print commands generated",
    }


@api_router.get("/print/thermal/{sale_id}")
async def get_thermal_receipt(sale_id: str, request: Request):
    await require_auth(request)

    logger.info(f"get_thermal_receipt: lookup sale_id={sale_id}")
    # Accept common falsy placeholders and fallback to latest sale
    if not sale_id or str(sale_id).lower() in ("none", "null", ""):
        logger.info("get_thermal_receipt: received empty sale_id, falling back to latest sale")
        latest = await db.sales.find({}).sort("created_at", -1).limit(1).to_list(length=1)
        if latest and latest[0].get("sale_id"):
            sale_id = latest[0].get("sale_id")
            logger.info(f"get_thermal_receipt: using fallback sale_id={sale_id}")
    sale = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0})
    if not sale:
        # Fallback searches to be resilient to differing stored formats
        try:
            # try invoice_number match
            alt = await db.sales.find_one({"invoice_number": sale_id}, {"_id": 0})
            if alt:
                sale = alt
        except Exception:
            pass

    if not sale:
        # try loose substring match on sale_id (if stored differently)
        try:
            cursor = db.sales.find({}, {"_id": 0, "sale_id": 1}).limit(20)
            candidates = await cursor.to_list(length=20)
            for c in candidates:
                sid = c.get("sale_id")
                if sid and str(sale_id) in str(sid):
                    sale = await db.sales.find_one({"sale_id": sid}, {"_id": 0})
                    if sale:
                        break
        except Exception:
            sale = None

    if not sale:
        logger.info(f"get_thermal_receipt: sale {sale_id} not found after fallbacks")
        raise HTTPException(status_code=404, detail="Sale not found")

    sale = cast(Dict[str, Any], sale)

    # Generate 80mm thermal receipt format (text-based for thermal printers)
    width = 48  # characters for 80mm
    lines = []

    lines.append("=" * width)
    lines.append("MUNDO DE ACCESORIOS".center(width))
    lines.append("=" * width)
    lines.append(f"Factura: {sale['invoice_number']}")
    lines.append(f"Fecha: {sale['created_at'][:10]}")
    lines.append(f"Cliente: {sale['customer_name']}")
    lines.append("-" * width)

    items = cast(List[Dict[str, Any]], sale.get("items", []))
    for item in items:
        name = item["product_name"][:20]
        qty = item["quantity"]
        price = item["subtotal"]
        lines.append(f"{name:<20} {qty:>3} x ${price:>8.2f}")

    lines.append("-" * width)
    lines.append(f"{'Subtotal:':<30} ${sale['subtotal']:>8.2f}")
    lines.append(f"{'IVA 12%:':<30} ${sale['tax']:>8.2f}")
    if sale["discount"] > 0:
        lines.append(f"{'Descuento:':<30} -${sale['discount']:>7.2f}")
    lines.append(f"{'TOTAL:':<30} ${sale['total']:>8.2f}")
    lines.append("=" * width)
    lines.append("¡Gracias por su compra!".center(width))
    lines.append("")

    content = "\n".join(lines)
    return Response(content=content, media_type="text/plain")


@api_router.get("/print/work-order/{work_order_id}")
async def get_work_order_thermal(work_order_id: str, request: Request):
    await require_auth(request)

    wo = await db.work_orders.find_one({"work_order_id": work_order_id}, {"_id": 0})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")

    width = 48
    lines = []

    lines.append("=" * width)
    lines.append("ORDEN DE TRABAJO".center(width))
    lines.append("=" * width)
    lines.append(f"OT: {wo['work_order_id']}")
    lines.append(f"Fecha: {wo['created_at'][:10]}")
    lines.append(f"Cliente: {wo['customer_name']}")
    lines.append(f"Vehículo: {wo['vehicle_info']}")
    lines.append(f"Prioridad: {wo['priority'].upper()}")
    lines.append("-" * width)
    lines.append("TRABAJOS A REALIZAR:")

    items = cast(List[Dict[str, Any]], wo.get("items", []))
    for item in items:
        lines.append(f"  - {item.get('description', item.get('product_name', 'N/A'))}")

    lines.append("-" * width)
    lines.append(f"Tiempo estimado: {wo['estimated_time']} min")
    if wo.get("technician_name"):
        lines.append(f"Técnico: {wo['technician_name']}")
    lines.append("=" * width)

    content = "\n".join(lines)
    return Response(content=content, media_type="text/plain")


@api_router.get("/print/invoice-pdf/{sale_id}")
async def get_invoice_pdf(sale_id: str, request: Request):
    await require_auth(request)

    sale = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    customer = await db.customers.find_one(
        {"customer_id": sale["customer_id"]}, {"_id": 0}
    )
    vehicle = None
    if sale.get("vehicle_id"):
        vehicle = await db.vehicles.find_one({"vehicle_id": sale["vehicle_id"]}, {"_id": 0})
    branch = None
    if sale.get("branch_id"):
        branch = await db.branches.find_one({"branch_id": sale["branch_id"]}, {"_id": 0})

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)

    currency = sale.get("currency", "USD")
    iva_rate = sale.get("iva_rate", 12)
    apply_iva = sale.get("apply_iva", True)

    company = {
        "name": (branch or {}).get("company_name") or os.environ.get("COMPANY_NAME", "MUNDO DE ACCESORIOS"),
        "tagline": os.environ.get("COMPANY_TAGLINE", "Sistema ERP"),
        "legal_name": (branch or {}).get("company_legal_name") or (branch or {}).get("name") or os.environ.get("COMPANY_LEGAL_NAME", "MUNDO DE ACCESORIOS"),
        "tax_id": (branch or {}).get("company_tax_id") or "",
        "vat": (branch or {}).get("company_vat") or "",
        "address": (branch or {}).get("company_address") or (branch or {}).get("address") or "",
        "city": (branch or {}).get("company_city") or "",
        "country": (branch or {}).get("company_country") or "",
        "phone": (branch or {}).get("company_phone") or (branch or {}).get("phone") or "",
        "email": (branch or {}).get("company_email") or "",
        "logo_url": (branch or {}).get("logo_url") or "",
    }
    customer_info = {
        "name": sale.get("customer_name", ""),
        "tax_id": customer.get("tax_id") if customer else "",
        "address": customer.get("address") if customer else "",
        "phone": customer.get("phone") if customer else "",
        "email": customer.get("email") if customer else "",
    }

    totals = {
        "subtotal": sale.get("subtotal", 0),
        "tax": sale.get("tax", 0),
        "total": sale.get("total", 0),
    }

    _draw_document_pdf(
        p,
        "Invoice",
        sale.get("invoice_number", ""),
        sale.get("created_at", ""),
        company,
        customer_info,
        vehicle,
        sale.get("items") or [],
        currency,
        iva_rate,
        apply_iva,
        totals,
        notes=sale.get("notes"),
    )

    p.save()
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=factura_{sale['invoice_number']}.pdf"
        },
    )


@api_router.get("/print/quotation-pdf/{quotation_id}")
async def get_quotation_pdf(quotation_id: str, request: Request):
    """Generate PDF for a quotation with company header."""
    await require_auth(request)

    quotation = await db.quotations.find_one({"quotation_id": quotation_id}, {"_id": 0})
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")
    quotation = cast(Dict[str, Any], quotation)

    customer = await db.customers.find_one(
        {"customer_id": quotation.get("customer_id")}, {"_id": 0}
    )
    vehicle = None
    if quotation.get("vehicle_id"):
        vehicle = await db.vehicles.find_one({"vehicle_id": quotation["vehicle_id"]}, {"_id": 0})
    branch = None
    if quotation.get("branch_id"):
        branch = await db.branches.find_one({"branch_id": quotation["branch_id"]}, {"_id": 0})

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)

    currency = quotation.get("currency", "USD")
    iva_rate = quotation.get("iva_rate", 12)
    apply_iva = quotation.get("apply_iva", True)

    company = {
        "name": (branch or {}).get("company_name") or os.environ.get("COMPANY_NAME", "MUNDO DE ACCESORIOS"),
        "tagline": os.environ.get("COMPANY_TAGLINE", "Sistema ERP"),
        "legal_name": (branch or {}).get("company_legal_name") or (branch or {}).get("name") or os.environ.get("COMPANY_LEGAL_NAME", "MUNDO DE ACCESORIOS"),
        "tax_id": (branch or {}).get("company_tax_id") or "",
        "vat": (branch or {}).get("company_vat") or "",
        "address": (branch or {}).get("company_address") or (branch or {}).get("address") or "",
        "city": (branch or {}).get("company_city") or "",
        "country": (branch or {}).get("company_country") or "",
        "phone": (branch or {}).get("company_phone") or (branch or {}).get("phone") or "",
        "email": (branch or {}).get("company_email") or "",
        "logo_url": (branch or {}).get("logo_url") or "",
    }
    customer_info = {
        "name": quotation.get("customer_name", ""),
        "tax_id": customer.get("tax_id") if customer else "",
        "address": customer.get("address") if customer else "",
        "phone": customer.get("phone") if customer else "",
        "email": customer.get("email") if customer else "",
    }

    totals = {
        "subtotal": quotation.get("subtotal", 0),
        "tax": quotation.get("tax", 0),
        "total": quotation.get("total", 0),
    }

    _draw_document_pdf(
        p,
        "Quotation",
        quotation_id,
        quotation.get("created_at", ""),
        company,
        customer_info,
        vehicle,
        quotation.get("items") or [],
        currency,
        iva_rate,
        apply_iva,
        totals,
        notes=quotation.get("notes"),
    )

    p.save()
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=cotizacion_{quotation_id}.pdf"
        },
    )


# ============ DASHBOARD STATS ============


@api_router.get("/dashboard/stats")
async def get_dashboard_stats(request: Request):
    user = await require_auth(request)

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    month_prefix = datetime.now(timezone.utc).strftime("%Y-%m")
    sales_scope_query = build_sales_visibility_query(user)

    # Sales today
    sales_today_query = merge_queries(
        sales_scope_query, {"created_at": {"$regex": f"^{today}"}}
    )
    sales_today = await db.sales.find(sales_today_query, {"_id": 0}).to_list(1000)

    total_sales_today = sum(s["total"] for s in sales_today)

    # Pending work orders
    pending_wo = await db.work_orders.count_documents(
        {"status": {"$in": ["pending", "in_progress"]}}
    )

    # Low stock items
    low_stock = await db.inventory.count_documents(
        {"$expr": {"$lte": ["$quantity", "$min_stock"]}}
    )

    # Pending deliveries
    pending_deliveries = await db.sales.count_documents(
        merge_queries(sales_scope_query, {"delivery_status": "pending"})
    )

    # Deliveries completed today
    delivered_today_query = merge_queries(
        sales_scope_query,
        {
            "delivery_status": "delivered",
            "delivery_completed_at": {"$regex": f"^{today}"},
        },
    )
    delivered_today_sales = await db.sales.find(
        delivered_today_query,
        {"_id": 0, "created_at": 1, "delivery_completed_at": 1, "sale_id": 1},
    ).to_list(1000)
    deliveries_completed_today = len(delivered_today_sales)

    deliveries_completed_month = await db.sales.count_documents(
        merge_queries(
            sales_scope_query,
            {
                "delivery_status": "delivered",
                "delivery_completed_at": {"$regex": f"^{month_prefix}"},
            },
        )
    )

    delivered_month_sales = await db.sales.find(
        merge_queries(
            sales_scope_query,
            {
                "delivery_status": "delivered",
                "delivery_completed_at": {"$regex": f"^{month_prefix}"},
            },
        ),
        {"_id": 0, "created_at": 1, "delivery_completed_at": 1, "sale_id": 1},
    ).to_list(3000)

    # Average delivery time in minutes (today)
    def _avg_delivery_minutes(sales: List[Dict[str, Any]]) -> float:
        delivery_minutes: list[float] = []
        for sale in sales:
            created_at_raw = sale.get("created_at")
            completed_at_raw = sale.get("delivery_completed_at")
            if not created_at_raw or not completed_at_raw:
                continue
            try:
                created_dt = datetime.fromisoformat(str(created_at_raw).replace("Z", "+00:00"))
                completed_dt = datetime.fromisoformat(str(completed_at_raw).replace("Z", "+00:00"))
            except Exception:
                continue
            delta_minutes = (completed_dt - created_dt).total_seconds() / 60
            if delta_minutes >= 0:
                delivery_minutes.append(delta_minutes)

        return (
            round(sum(delivery_minutes) / len(delivery_minutes), 1)
            if delivery_minutes
            else 0
        )

    avg_delivery_minutes = _avg_delivery_minutes(delivered_today_sales)
    avg_delivery_minutes_month = _avg_delivery_minutes(delivered_month_sales)

    # Returns today (created)
    returns_today = await db.returns.count_documents({"created_at": {"$regex": f"^{today}"}})

    # Returns completed today
    returns_completed_today = await db.returns.count_documents(
        {
            "status": "completed",
            "completed_at": {"$regex": f"^{today}"},
        }
    )

    returns_completed_month = await db.returns.count_documents(
        {
            "status": "completed",
            "completed_at": {"$regex": f"^{month_prefix}"},
        }
    )

    # Credit pending
    credit_pending = await db.sales.find(
        merge_queries(
            sales_scope_query,
            {"payment_type": "credit", "payment_status": {"$ne": "paid"}},
        ),
        {"_id": 0},
    ).to_list(1000)
    total_credit_pending = sum(s["total"] for s in credit_pending)

    return {
        "sales_today": {
            "count": len(sales_today),
            "total": round(total_sales_today, 2),
        },
        "pending_work_orders": pending_wo,
        "low_stock_items": low_stock,
        "pending_deliveries": pending_deliveries,
        "deliveries_completed_today": deliveries_completed_today,
        "deliveries_completed_month": deliveries_completed_month,
        "returns_today": returns_today,
        "returns_completed_today": returns_completed_today,
        "returns_completed_month": returns_completed_month,
        "avg_delivery_minutes": avg_delivery_minutes,
        "avg_delivery_minutes_month": avg_delivery_minutes_month,
        "credit_pending": round(total_credit_pending, 2),
    }


# ============ SEED DATA ============


@api_router.post("/seed")
async def seed_data(request: Request):
    """Seed initial data with automotive accessories"""

    core_seed_report: Dict[str, Any]
    try:
        # This powers the Dashboard "Cargar datos de pruebas" button as a recovery action
        # to bring back missing core data after full rebuilds.
        core_seed_report = await apply_core_seed_data(overwrite=False, only_when_empty=False)
    except Exception:
        logger.exception("Failed applying core seed data during /seed")
        core_seed_report = {
            "applied": False,
            "reason": "error",
            "collections": {},
        }

    # Branches
    branches = [
        {
            "branch_id": "branch_main",
            "name": "Mundo de Accesorios",
            "address": "Av. Principal 123",
            "phone": "555-0001",
            "is_active": True,
            "company_name": "Mundo de Accesorios",
            "company_legal_name": "Mundo de Accesorios",
            "branch_kind": "central",
            "sales_channels": ["minorista", "mayorista"],
            "service_policy": {
                "installations_enabled": True,
                "tint_enabled": True,
                "local_delivery_enabled": True,
                "departmental_delivery_enabled": True,
            },
            "company_tax_id": "",
            "company_vat": "",
            "company_address": "Av. Principal 123",
            "company_city": "",
            "company_country": "",
            "company_phone": "555-0001",
            "company_email": "",
            "company_website": "",
        },
        {
            "branch_id": "branch_north",
            "name": "TopCar El Calvario",
            "address": "Calle Norte 456",
            "phone": "555-0002",
            "is_active": True,
            "company_name": "TopCar",
            "company_legal_name": "TopCar",
            "branch_kind": "wholesale",
            "sales_channels": ["minorista", "mayorista"],
            "service_policy": {
                "installations_enabled": False,
                "tint_enabled": False,
                "local_delivery_enabled": True,
                "departmental_delivery_enabled": True,
            },
            "company_tax_id": "",
            "company_vat": "",
            "company_address": "Calle Norte 456",
            "company_city": "",
            "company_country": "",
            "company_phone": "555-0002",
            "company_email": "",
            "company_website": "",
        },
        {
            "branch_id": "branch_south",
            "name": "TopCar La Tigre",
            "address": "Av. Sur 789",
            "phone": "555-0003",
            "is_active": True,
            "company_name": "TopCar",
            "company_legal_name": "TopCar",
            "branch_kind": "wholesale",
            "sales_channels": ["minorista", "mayorista"],
            "service_policy": {
                "installations_enabled": False,
                "tint_enabled": False,
                "local_delivery_enabled": True,
                "departmental_delivery_enabled": True,
            },
            "company_tax_id": "",
            "company_vat": "",
            "company_address": "Av. Sur 789",
            "company_city": "",
            "company_country": "",
            "company_phone": "555-0003",
            "company_email": "",
            "company_website": "",
        },
    ]
    for b in branches:
        await db.branches.update_one(
            {"branch_id": b["branch_id"]}, {"$set": b}, upsert=True
        )

    # Warehouses
    warehouses = [
        {
            "warehouse_id": "wh_main",
            "name": "Bodega Central",
            "address": "Zona Industrial 1",
            "branch_id": "branch_main",
            "is_active": True,
        },
        {
            "warehouse_id": "wh_topcar_calvario",
            "name": "Bodega TopCar El Calvario",
            "address": "Calle Norte 456",
            "branch_id": "branch_north",
            "is_active": True,
        },
        {
            "warehouse_id": "wh_topcar_tigre",
            "name": "Bodega TopCar La Tigre",
            "address": "Av. Sur 789",
            "branch_id": "branch_south",
            "is_active": True,
        },
    ]
    for w in warehouses:
        await db.warehouses.update_one(
            {"warehouse_id": w["warehouse_id"]}, {"$set": w}, upsert=True
        )

    await db.warehouses.update_many(
        {
            "warehouse_id": {
                "$in": [
                    "wh_north1",
                    "wh_north2",
                    "wh_south1",
                    "wh_south2",
                    "wh_east",
                    "wh_west",
                    "wh_express",
                ]
            }
        },
        {"$set": {"is_active": False}},
    )

    # Products - Accesorios No Eléctricos
    products = [
        # Defensas y Protección
        {
            "product_id": "prod_def_001",
            "sku": "DEF-TOY-001",
            "name": "Defensa Delantera Toyota Hilux 2016-2023",
            "description": "Defensa tubular de acero inoxidable con protección de cárter",
            "category": "accesorios_no_electricos",
            "subcategory": "Defensas",
            "brand": "ARB",
            "price": 450.00,
            "cost": 280.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1558618666-fcd25c85cd64?w=400"
            ],
            "compatibility": {
                "brands": ["Toyota"],
                "models": ["Hilux"],
                "year_from": 2016,
                "year_to": 2023,
                "vehicle_types": ["Camioneta Doble Cabina", "Camioneta 1 Cabina"],
            },
            "installation_required": True,
            "installation_price": 50.00,
            "installation_time_minutes": 90,
            "warranty_months": 24,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_bar_001",
            "sku": "BAR-UNI-001",
            "name": "Barra Antivuelco Universal Pickup",
            "description": "Barra antivuelco cromada con luz LED integrada",
            "category": "accesorios_no_electricos",
            "subcategory": "Barras Antivuelcos",
            "brand": "Westin",
            "price": 320.00,
            "cost": 180.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1609521263047-f8f205293f24?w=400"
            ],
            "compatibility": {
                "vehicle_types": [
                    "Camioneta Doble Cabina",
                    "Camioneta 1 Cabina",
                    "Pickup",
                ]
            },
            "installation_required": True,
            "installation_price": 40.00,
            "installation_time_minutes": 60,
            "warranty_months": 12,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_can_001",
            "sku": "CAN-001",
            "name": "Canastero de Techo Universal",
            "description": "Canastero de aluminio negro mate con capacidad 100kg",
            "category": "accesorios_no_electricos",
            "subcategory": "Canasteros",
            "brand": "Thule",
            "price": 280.00,
            "cost": 160.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1533473359331-0135ef1b58bf?w=400"
            ],
            "compatibility": {},
            "installation_required": True,
            "installation_price": 35.00,
            "installation_time_minutes": 45,
            "warranty_months": 24,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_lon_001",
            "sku": "LON-HIL-001",
            "name": "Lona Enrollable Toyota Hilux",
            "description": "Lona de lona gruesa impermeable con sistema de enrollado manual",
            "category": "accesorios_no_electricos",
            "subcategory": "Lonas para Tinas",
            "brand": "Extang",
            "price": 180.00,
            "cost": 95.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1544636331-e26879cd4d9b?w=400"
            ],
            "compatibility": {
                "brands": ["Toyota"],
                "models": ["Hilux"],
                "year_from": 2016,
                "vehicle_types": ["Camioneta Doble Cabina"],
            },
            "installation_required": True,
            "installation_price": 25.00,
            "installation_time_minutes": 30,
            "warranty_months": 12,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_snk_001",
            "sku": "SNK-UNI-001",
            "name": "Snorkel Universal 4x4",
            "description": "Snorkel de polietileno con filtro de aire para vadeo",
            "category": "accesorios_no_electricos",
            "subcategory": "Snorkels",
            "brand": "Safari",
            "price": 350.00,
            "cost": 200.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1519641471654-76ce0107ad1b?w=400"
            ],
            "compatibility": {
                "vehicle_types": ["SUV", "Camioneta Doble Cabina", "Camioneta 1 Cabina"]
            },
            "installation_required": True,
            "installation_price": 80.00,
            "installation_time_minutes": 120,
            "warranty_months": 24,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_alf_001",
            "sku": "ALF-UNI-001",
            "name": "Alfombras de Goma Universal 4pcs",
            "description": "Set de 4 alfombras de goma de alto tráfico antideslizantes",
            "category": "accesorios_no_electricos",
            "subcategory": "Alfombras",
            "brand": "WeatherTech",
            "price": 45.00,
            "cost": 22.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1503376780353-7e6692767b70?w=400"
            ],
            "compatibility": {},
            "installation_required": False,
            "installation_price": 0,
            "installation_time_minutes": 0,
            "warranty_months": 12,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_for_001",
            "sku": "FOR-TIM-001",
            "name": "Forro de Timón Cuero Negro",
            "description": "Forro de timón de cuero genuino con costuras rojas deportivas",
            "category": "accesorios_no_electricos",
            "subcategory": "Forros de Timón",
            "brand": "Sparco",
            "price": 35.00,
            "cost": 15.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1489824904134-891ab64532f1?w=400"
            ],
            "compatibility": {},
            "installation_required": False,
            "installation_price": 0,
            "installation_time_minutes": 0,
            "warranty_months": 6,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        # Accesorios Electrónicos
        {
            "product_id": "prod_rad_001",
            "sku": "RAD-AND-001",
            "name": 'Radio Android 10" Toyota Universal',
            "description": "Pantalla Android 13 con GPS, WiFi, Bluetooth, CarPlay y Android Auto",
            "category": "accesorios_electronicos",
            "subcategory": "Pantallas Android",
            "brand": "Teyes",
            "price": 280.00,
            "cost": 150.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1558618666-fcd25c85cd64?w=400"
            ],
            "compatibility": {"brands": ["Toyota"], "year_from": 2012},
            "installation_required": True,
            "installation_price": 45.00,
            "installation_time_minutes": 90,
            "warranty_months": 12,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_rad_002",
            "sku": "RAD-PIO-001",
            "name": 'Pioneer AVH-Z9250BT 7"',
            "description": 'Pantalla táctil Pioneer 7" con DVD, Bluetooth, CarPlay',
            "category": "accesorios_electronicos",
            "subcategory": "Pantallas Pioneer",
            "brand": "Pioneer",
            "price": 420.00,
            "cost": 280.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1558618666-fcd25c85cd64?w=400"
            ],
            "compatibility": {},
            "installation_required": True,
            "installation_price": 50.00,
            "installation_time_minutes": 120,
            "warranty_months": 12,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_par_001",
            "sku": "PAR-JBL-001",
            "name": 'Parlantes JBL 6.5" 300W Par',
            "description": "Par de parlantes coaxiales JBL Club 6520 con tweeter incorporado",
            "category": "accesorios_electronicos",
            "subcategory": "Parlantes",
            "brand": "JBL",
            "price": 85.00,
            "cost": 50.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1545454675-3531b543be5d?w=400"
            ],
            "compatibility": {},
            "installation_required": True,
            "installation_price": 25.00,
            "installation_time_minutes": 45,
            "warranty_months": 12,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_amp_001",
            "sku": "AMP-TAR-001",
            "name": "Amplificador Taramps MD3000 3000W",
            "description": "Amplificador monoblock clase D para subwoofer",
            "category": "accesorios_electronicos",
            "subcategory": "Amplificadores",
            "brand": "Taramps",
            "price": 320.00,
            "cost": 190.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1545454675-3531b543be5d?w=400"
            ],
            "compatibility": {},
            "installation_required": True,
            "installation_price": 60.00,
            "installation_time_minutes": 90,
            "warranty_months": 12,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_led_001",
            "sku": "LED-BAR-001",
            "name": 'Barra LED 32" 180W Combo',
            "description": "Barra de luces LED con combo spot/flood para off-road",
            "category": "accesorios_electronicos",
            "subcategory": "Luces LED",
            "brand": "Rigid",
            "price": 150.00,
            "cost": 80.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1558618666-fcd25c85cd64?w=400"
            ],
            "compatibility": {},
            "installation_required": True,
            "installation_price": 45.00,
            "installation_time_minutes": 60,
            "warranty_months": 24,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_cam_001",
            "sku": "CAM-REV-001",
            "name": "Cámara de Reversa HD con Líneas Guía",
            "description": "Cámara trasera HD 170° con visión nocturna y líneas de parqueo dinámicas",
            "category": "accesorios_electronicos",
            "subcategory": "Cámaras de Reversa",
            "brand": "Kenwood",
            "price": 45.00,
            "cost": 22.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1558618666-fcd25c85cd64?w=400"
            ],
            "compatibility": {},
            "installation_required": True,
            "installation_price": 35.00,
            "installation_time_minutes": 60,
            "warranty_months": 12,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_sen_001",
            "sku": "SEN-PAR-001",
            "name": "Sensores de Parqueo 8pcs con Display",
            "description": "Kit 8 sensores (4 delanteros + 4 traseros) con pantalla LED y buzzer",
            "category": "accesorios_electronicos",
            "subcategory": "Sensores de Aparcamiento",
            "brand": "Steelmate",
            "price": 65.00,
            "cost": 32.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1558618666-fcd25c85cd64?w=400"
            ],
            "compatibility": {},
            "installation_required": True,
            "installation_price": 50.00,
            "installation_time_minutes": 90,
            "warranty_months": 12,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_alr_001",
            "sku": "ALR-VIP-001",
            "name": "Alarma Viper 5906V con GPS",
            "description": "Sistema de seguridad con arranque remoto, GPS y app móvil",
            "category": "accesorios_electronicos",
            "subcategory": "Alarmas",
            "brand": "Viper",
            "price": 280.00,
            "cost": 160.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1558618666-fcd25c85cd64?w=400"
            ],
            "compatibility": {},
            "installation_required": True,
            "installation_price": 80.00,
            "installation_time_minutes": 180,
            "warranty_months": 12,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_dvr_001",
            "sku": "DVR-DUA-001",
            "name": "DVR Dual Cámara Dashcam 4K",
            "description": "Grabador DVR con cámara frontal 4K y trasera 1080p, GPS y WiFi",
            "category": "accesorios_electronicos",
            "subcategory": "DVR",
            "brand": "Viofo",
            "price": 180.00,
            "cost": 95.00,
            "product_type": "product",
            "images": [
                "https://images.unsplash.com/photo-1558618666-fcd25c85cd64?w=400"
            ],
            "compatibility": {},
            "installation_required": True,
            "installation_price": 40.00,
            "installation_time_minutes": 60,
            "warranty_months": 12,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        # Polarizados
        {
            "product_id": "prod_pol_001",
            "sku": "POL-SED-COM",
            "name": "Polarizado Completo Sedán",
            "description": "Polarizado profesional película 20% todos los vidrios",
            "category": "polarizados",
            "subcategory": "Polarizado Completo",
            "brand": "3M",
            "price": 120.00,
            "cost": 40.00,
            "product_type": "service",
            "images": [
                "https://images.unsplash.com/photo-1494976388531-d1058494cdd8?w=400"
            ],
            "compatibility": {"vehicle_types": ["Sedán", "Hatchback"]},
            "polarizado_type": "3M FX Premium 20%",
            "window_options": ["completo"],
            "installation_required": True,
            "installation_price": 0,
            "installation_time_minutes": 120,
            "warranty_months": 60,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_pol_002",
            "sku": "POL-SUV-COM",
            "name": "Polarizado Completo SUV/Camioneta",
            "description": "Polarizado profesional película 20% todos los vidrios SUV grande",
            "category": "polarizados",
            "subcategory": "Polarizado Completo",
            "brand": "3M",
            "price": 180.00,
            "cost": 55.00,
            "product_type": "service",
            "images": [
                "https://images.unsplash.com/photo-1519641471654-76ce0107ad1b?w=400"
            ],
            "compatibility": {
                "vehicle_types": ["SUV", "Camioneta Doble Cabina", "Camioneta 1 Cabina"]
            },
            "polarizado_type": "3M FX Premium 20%",
            "window_options": ["completo"],
            "installation_required": True,
            "installation_price": 0,
            "installation_time_minutes": 180,
            "warranty_months": 60,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_pol_003",
            "sku": "POL-CAM-COM",
            "name": "Polarizado Completo Camión/Microbús",
            "description": "Polarizado profesional para camiones y microbuses",
            "category": "polarizados",
            "subcategory": "Polarizado Completo",
            "brand": "3M",
            "price": 250.00,
            "cost": 80.00,
            "product_type": "service",
            "images": [
                "https://images.unsplash.com/photo-1544636331-e26879cd4d9b?w=400"
            ],
            "compatibility": {
                "vehicle_types": [
                    "Camión Pequeño",
                    "Camión Mediano",
                    "Camión Grande",
                    "Microbús Pasajeros",
                    "Microbús Carga",
                ]
            },
            "polarizado_type": "3M FX Premium 20%",
            "window_options": ["completo"],
            "installation_required": True,
            "installation_price": 0,
            "installation_time_minutes": 240,
            "warranty_months": 60,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_pol_004",
            "sku": "POL-DEL-001",
            "name": "Polarizado Solo Vidrios Delanteros",
            "description": "Polarizado de vidrios laterales delanteros únicamente",
            "category": "polarizados",
            "subcategory": "Vidrios Delanteros",
            "brand": "3M",
            "price": 35.00,
            "cost": 12.00,
            "product_type": "service",
            "images": [
                "https://images.unsplash.com/photo-1494976388531-d1058494cdd8?w=400"
            ],
            "compatibility": {},
            "polarizado_type": "3M FX Premium 35%",
            "window_options": ["delanteros"],
            "installation_required": True,
            "installation_price": 0,
            "installation_time_minutes": 30,
            "warranty_months": 60,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_pol_005",
            "sku": "POL-FRA-SUP",
            "name": "Franja Superior Parabrisas",
            "description": "Franja de polarizado en parte superior del parabrisas",
            "category": "polarizados",
            "subcategory": "Franja Superior",
            "brand": "3M",
            "price": 25.00,
            "cost": 8.00,
            "product_type": "service",
            "images": [
                "https://images.unsplash.com/photo-1494976388531-d1058494cdd8?w=400"
            ],
            "compatibility": {},
            "polarizado_type": "3M FX Premium Gradiente",
            "window_options": ["franja_superior"],
            "installation_required": True,
            "installation_price": 0,
            "installation_time_minutes": 20,
            "warranty_months": 60,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_pol_006",
            "sku": "POL-LIM-001",
            "name": "Limpieza/Remoción de Polarizado",
            "description": "Servicio de remoción de polarizado viejo y limpieza de vidrios",
            "category": "polarizados",
            "subcategory": "Limpieza de Polarizado",
            "brand": "Servicio",
            "price": 45.00,
            "cost": 10.00,
            "product_type": "service",
            "images": [],
            "compatibility": {},
            "window_options": ["completo"],
            "installation_required": True,
            "installation_price": 0,
            "installation_time_minutes": 90,
            "warranty_months": 0,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        # Servicios
        {
            "product_id": "prod_srv_001",
            "sku": "SRV-INS-BAS",
            "name": "Servicio de Instalación Básica",
            "description": "Instalación de accesorios simples (alfombras, forros, etc.)",
            "category": "servicios",
            "subcategory": "Instalación Básica",
            "brand": "Servicio",
            "price": 15.00,
            "cost": 5.00,
            "product_type": "service",
            "images": [],
            "compatibility": {},
            "installation_required": False,
            "installation_price": 0,
            "installation_time_minutes": 30,
            "warranty_months": 1,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_srv_002",
            "sku": "SRV-INS-COM",
            "name": "Servicio de Instalación Compleja",
            "description": "Instalación de accesorios que requieren modificación (snorkels, defensas)",
            "category": "servicios",
            "subcategory": "Instalación Compleja",
            "brand": "Servicio",
            "price": 50.00,
            "cost": 15.00,
            "product_type": "service",
            "images": [],
            "compatibility": {},
            "installation_required": False,
            "installation_price": 0,
            "installation_time_minutes": 120,
            "warranty_months": 3,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_srv_003",
            "sku": "SRV-REV-HOR",
            "name": "Servicio de Revisión por Hora",
            "description": "Diagnóstico y revisión general de accesorios instalados",
            "category": "servicios",
            "subcategory": "Revisión por Hora",
            "brand": "Servicio",
            "price": 25.00,
            "cost": 8.00,
            "product_type": "service_hourly",
            "images": [],
            "compatibility": {},
            "hourly_rate": 25.00,
            "installation_required": False,
            "installation_price": 0,
            "installation_time_minutes": 60,
            "warranty_months": 0,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        {
            "product_id": "prod_srv_004",
            "sku": "SRV-DIA-ELE",
            "name": "Diagnóstico Electrónico",
            "description": "Diagnóstico de sistemas electrónicos con escáner OBD2",
            "category": "servicios",
            "subcategory": "Diagnóstico Electrónico",
            "brand": "Servicio",
            "price": 35.00,
            "cost": 10.00,
            "product_type": "service",
            "images": [],
            "compatibility": {},
            "installation_required": False,
            "installation_price": 0,
            "installation_time_minutes": 45,
            "warranty_months": 0,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
    ]

    for p in products:
        await db.products.update_one(
            {"product_id": p["product_id"]}, {"$set": p}, upsert=True
        )

    # Inventory
    for p in products:
        if p.get("product_type") == "product":  # Only physical products have inventory
            for w in warehouses[:4]:
                inv = {
                    "inventory_id": f"inv_{p['product_id']}_{w['warehouse_id']}",
                    "product_id": p["product_id"],
                    "warehouse_id": w["warehouse_id"],
                    "quantity": 20,
                    "min_stock": 5,
                    "last_updated": datetime.now(timezone.utc).isoformat(),
                }
                await db.inventory.update_one(
                    {"product_id": p["product_id"], "warehouse_id": w["warehouse_id"]},
                    {"$set": inv},
                    upsert=True,
                )

    # Customers & Vehicles test dataset for Dashboard "Cargar Datos Prueba"
    # Requested behavior: clean previous customers/vehicles and rebuild complete sample set.
    deleted_vehicles = await db.vehicles.delete_many({})
    deleted_customers = await db.customers.delete_many({})

    def _phone(prefix: str, idx: int) -> str:
        return f"{prefix}{(idx % 90) + 10:02d}-{((idx * 137) % 10000):04d}"

    def _plate(seed: int) -> str:
        left = 100 + (seed % 900)
        right = 100 + ((seed * 7) % 900)
        return f"M {left:03d} {right:03d}"

    def _chasis(seed: str) -> str:
        # 17-character VIN-safe string (avoids I/O/Q by using UUID hex)
        return uuid.uuid5(uuid.NAMESPACE_DNS, seed).hex.upper()[:17]

    natural_rows = [
        ("Juan Carlos", "Perez", "juan.perez88@example.com", "Barrio Altagracia, Managua"),
        ("Maria Fernanda", "Lopez", "maria.lopez88@example.com", "Colonia Centroamerica, Managua"),
        ("Luis Alberto", "Gonzalez", "luis.gonzalez88@example.com", "Villa Fontana, Managua"),
        ("Ana Lucia", "Martinez", "ana.martinez88@example.com", "Bello Horizonte, Managua"),
        ("Carlos Enrique", "Ruiz", "carlos.ruiz88@example.com", "Carretera a Masaya, km 8, Managua"),
        ("Sofia Elena", "Castillo", "sofia.castillo88@example.com", "Reparto San Juan, Managua"),
        ("Pedro Jose", "Torres", "pedro.torres88@example.com", "Altamira, Managua"),
        ("Adriana Maria", "Mora", "adriana.mora88@example.com", "Las Colinas, Managua"),
        ("Miguel Angel", "Rios", "miguel.rios88@example.com", "Ciudad Jardin, Managua"),
        ("Rosa Maria", "Zelaya", "rosa.zelaya88@example.com", "Bolonia, Managua"),
        ("Jorge Luis", "Ortega", "jorge.ortega88@example.com", "Reparto Las Palmas, Managua"),
        ("Teresa Isabel", "Blandon", "teresa.blandon88@example.com", "Linda Vista, Managua"),
        ("Roberto Antonio", "Navarro", "roberto.navarro88@example.com", "Monsenor Lezcano, Managua"),
        ("Carmen Julia", "Luna", "carmen.luna88@example.com", "Batahola Norte, Managua"),
        ("Francisco Javier", "Hernandez", "francisco.hernandez88@example.com", "Colonia Miguel Bonilla, Managua"),
        ("Daniela Sofia", "Vargas", "daniela.vargas88@example.com", "Reparto Schick, Managua"),
        ("Ricardo Andres", "Mejia", "ricardo.mejia88@example.com", "Bello Horizonte Oeste, Managua"),
        ("Lucia Alejandra", "Gaitan", "lucia.gaitan88@example.com", "Las Brisas, Managua"),
        ("Mario Augusto", "Flores", "mario.flores88@example.com", "Larreynaga, Managua"),
        ("Patricia Elena", "Silva", "patricia.silva88@example.com", "Mercado Oriental sector sur, Managua"),
    ]

    company_rows = [
        ("Claro Nicaragua", "contacto@claro.com.ni", "Pista Juan Pablo II, Managua"),
        ("Tigo Nicaragua", "contacto@tigo.com.ni", "Carretera a Masaya, Managua"),
        ("Banco Lafise", "contacto@lafise.com", "Edificio Corporativo Lafise, Managua"),
        ("BAC Credomatic Nicaragua", "contacto@baccredomatic.com", "Galerias Santo Domingo, Managua"),
        ("Banpro Grupo Promerica", "contacto@banpro.com.ni", "Edificio Banpro, Managua"),
        ("Casa Pellas", "contacto@casapellas.com", "Carretera Norte, Managua"),
        ("Grupo Q Nicaragua", "contacto@grupoq.com", "Carretera a Masaya, Managua"),
        ("DHL Nicaragua", "contacto@dhl.com", "Carretera Norte, Managua"),
        ("PriceSmart Nicaragua", "contacto@pricesmart.com", "Carretera a Masaya, Managua"),
        ("Walmart Nicaragua", "contacto@walmart.com", "Metrocentro, Managua"),
        ("Supermercados La Colonia", "contacto@lacolonia.com.ni", "Oficinas Centrales, Managua"),
        ("Disnorte Dissur", "contacto@disnorte-dissur.com", "Carretera Norte, Managua"),
        ("ENACAL", "contacto@enacal.com.ni", "Bolonia, Managua"),
        ("EPN Nicaragua", "contacto@epn.gob.ni", "Avenida Bolivar, Managua"),
        ("Cargill Nicaragua", "contacto@cargill.com", "Las Mercedes, Managua"),
        ("NESTLE Nicaragua", "contacto@ni.nestle.com", "Carretera a Masaya, Managua"),
        ("Grupo Lala Nicaragua", "contacto@lala.com", "Carretera Norte, Managua"),
        ("Comtech Nicaragua", "contacto@comtech.com.ni", "Villa Fontana, Managua"),
        ("SINSA", "contacto@sinsa.com.ni", "Pista Suburbana, Managua"),
        ("Farmacias Kielsa Nicaragua", "contacto@kielsa.com", "Carretera a Masaya, Managua"),
    ]

    colors = ["BLANCO", "NEGRO", "GRIS", "AZUL", "ROJO", "PLATA"]
    natural_brands = ["NISSAN", "HONDA", "SUZUKI"]
    company_brands = ["KIA", "HYUNDAI", "TOYOTA"]
    models = {
        "NISSAN": ["SENTRA", "VERSA", "X-TRAIL"],
        "HONDA": ["CIVIC", "CR-V", "FIT"],
        "SUZUKI": ["SWIFT", "VITARA", "JIMNY"],
        "KIA": ["RIO", "SPORTAGE", "SORENTO"],
        "HYUNDAI": ["ACCENT", "TUCSON", "SANTA FE"],
        "TOYOTA": ["COROLLA", "YARIS", "HILUX"],
    }

    customers: list[dict[str, Any]] = []
    vehicles: list[dict[str, Any]] = []

    vehicle_seed = 1
    seed_tag = "dashboard_seed_managua_v2"

    for idx, (first_name, last_name, email, address) in enumerate(natural_rows, start=1):
        customer_id = f"seed_nat_{idx:03d}"
        customers.append(
            {
                "customer_id": customer_id,
                "name": f"{first_name} {last_name}",
                "first_name": first_name,
                "last_name": last_name,
                "customer_type": "natural",
                "tax_id": "",
                "phone_prefix": "+505",
                "phone": _phone("88", idx),
                "email": email,
                "address": address,
                "credit_limit": 5000,
                "credit_balance": 0,
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "source_seed": seed_tag,
            }
        )

        for brand_idx, brand in enumerate(natural_brands):
            model = models[brand][idx % len(models[brand])]
            vehicles.append(
                {
                    "vehicle_id": f"seed_vehicle_{vehicle_seed:04d}",
                    "customer_id": customer_id,
                    "plate": _plate(vehicle_seed),
                    "chasis": _chasis(f"nat-{idx}-{brand}-{brand_idx}"),
                    "vin": _chasis(f"nat-{idx}-{brand}-{brand_idx}"),
                    "brand": brand,
                    "model": model,
                    "year": 2015 + ((idx + brand_idx) % 10),
                    "vehicle_type": "sedan" if brand_idx == 0 else ("suv" if brand_idx == 1 else "pickup"),
                    "color": colors[(idx + brand_idx) % len(colors)],
                    "doors": 4,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "source_seed": seed_tag,
                }
            )
            vehicle_seed += 1

    for idx, (company_name, email, address) in enumerate(company_rows, start=1):
        customer_id = f"seed_emp_{idx:03d}"
        customers.append(
            {
                "customer_id": customer_id,
                "name": company_name,
                "first_name": company_name,
                "last_name": "",
                "customer_type": "empresa",
                "tax_id": f"J{5000 + idx:013d}",
                "phone_prefix": "+505",
                "phone": _phone("78", idx),
                "email": email,
                "address": address,
                "credit_limit": 20000,
                "credit_balance": 0,
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "source_seed": seed_tag,
            }
        )

        for brand_idx, brand in enumerate(company_brands):
            model = models[brand][idx % len(models[brand])]
            vehicles.append(
                {
                    "vehicle_id": f"seed_vehicle_{vehicle_seed:04d}",
                    "customer_id": customer_id,
                    "plate": _plate(vehicle_seed),
                    "chasis": _chasis(f"emp-{idx}-{brand}-{brand_idx}"),
                    "vin": _chasis(f"emp-{idx}-{brand}-{brand_idx}"),
                    "brand": brand,
                    "model": model,
                    "year": 2018 + ((idx + brand_idx) % 7),
                    "vehicle_type": "sedan" if brand != "HYUNDAI" else "suv",
                    "color": colors[(idx + brand_idx + 2) % len(colors)],
                    "doors": 4,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "source_seed": seed_tag,
                }
            )
            vehicle_seed += 1

    if customers:
        await db.customers.insert_many(customers)
    if vehicles:
        await db.vehicles.insert_many(vehicles)

    return {
        "message": "Data seeded successfully",
        "products": len(products),
        "categories": list(PRODUCT_CATEGORIES.keys()),
        "deleted": {
            "customers": deleted_customers.deleted_count,
            "vehicles": deleted_vehicles.deleted_count,
        },
        "seeded": {
            "customers": len(customers),
            "vehicles": len(vehicles),
            "natural_customers": len(natural_rows),
            "company_customers": len(company_rows),
        },
        "seed_preview": {
            "customer_natural": {k: v for k, v in customers[0].items() if k != "_id"}
            if customers
            else None,
            "customer_company": {
                k: v
                for k, v in customers[len(natural_rows)].items()
                if k != "_id"
            }
            if len(customers) > len(natural_rows)
            else None,
            "vehicle": {k: v for k, v in vehicles[0].items() if k != "_id"}
            if vehicles
            else None,
        },
        "core_seed": core_seed_report,
    }


# ============ NOTIFICATIONS & EMAIL (SendGrid) ============


class EmailNotification(BaseModel):
    notification_id: str = Field(
        default_factory=lambda: f"notif_{uuid.uuid4().hex[:8]}"
    )
    recipient_id: str
    recipient_email: str
    notification_type: str  # invoice, low_stock, work_order_update, delivery_update
    subject: str
    content: str
    related_id: Optional[str] = None  # sale_id, work_order_id, etc.
    status: str = "pending"  # pending, sent, failed
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    sent_at: Optional[datetime] = None


# Email sending function (will use SendGrid when API key is provided)
async def send_email_notification(
    to_email: str,
    subject: str,
    html_content: str,
    attachment_data: Optional[bytes] = None,
    attachment_name: Optional[str] = None,
):
    """Send email via SendGrid if configured, otherwise log"""
    sendgrid_key = os.environ.get("SENDGRID_API_KEY")
    sender_email = os.environ.get("SENDER_EMAIL", "noreply@mundodeaccesorios.com")

    if sendgrid_key and sendgrid_key != "your_sendgrid_api_key":
        try:
            import base64

            import sendgrid  # type: ignore[import]
            from sendgrid.helpers.mail import (  # type: ignore[import]
                Attachment,
                Disposition,
                FileContent,
                FileName,
                FileType,
                Mail,
            )

            message = Mail(
                from_email=sender_email,
                to_emails=to_email,
                subject=subject,
                html_content=html_content,
            )

            if attachment_data and attachment_name:
                encoded = base64.b64encode(attachment_data).decode()
                attachment = Attachment()
                attachment.file_content = FileContent(encoded)
                attachment.file_name = FileName(attachment_name)
                attachment.file_type = FileType("application/pdf")
                attachment.disposition = Disposition("attachment")
                message.attachment = attachment

            sg = sendgrid.SendGridAPIClient(api_key=sendgrid_key)
            response = sg.send(message)
            return response.status_code == 202
        except Exception as e:
            logger.error(f"SendGrid error: {e}")
            return False
    else:
        logger.info(f"Email would be sent to {to_email}: {subject}")
        return True  # Simulated success when no SendGrid key


@api_router.post("/notifications/send-invoice/{sale_id}")
async def send_invoice_email(
    sale_id: str, request: Request, background_tasks: BackgroundTasks
):
    """Send invoice PDF via email to customer"""
    await require_roles(request, ["gerencia", "supervisor", "ventas", "cajero"])

    sale = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    customer = await db.customers.find_one(
        {"customer_id": sale["customer_id"]}, {"_id": 0}
    )
    if not customer or not customer.get("email"):
        raise HTTPException(status_code=400, detail="Customer has no email")

    # Generate PDF
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)

    vehicle = None
    if sale.get("vehicle_id"):
        vehicle = await db.vehicles.find_one({"vehicle_id": sale["vehicle_id"]}, {"_id": 0})
    branch = None
    if sale.get("branch_id"):
        branch = await db.branches.find_one({"branch_id": sale["branch_id"]}, {"_id": 0})

    currency = sale.get("currency", "USD")
    iva_rate = sale.get("iva_rate", 12)
    apply_iva = sale.get("apply_iva", True)

    company = {
        "name": (branch or {}).get("company_name") or os.environ.get("COMPANY_NAME", "MUNDO DE ACCESORIOS"),
        "tagline": os.environ.get("COMPANY_TAGLINE", "Sistema ERP"),
        "legal_name": (branch or {}).get("company_legal_name") or (branch or {}).get("name") or os.environ.get("COMPANY_LEGAL_NAME", "MUNDO DE ACCESORIOS"),
        "tax_id": (branch or {}).get("company_tax_id") or "",
        "vat": (branch or {}).get("company_vat") or "",
        "address": (branch or {}).get("company_address") or (branch or {}).get("address") or "",
        "city": (branch or {}).get("company_city") or "",
        "country": (branch or {}).get("company_country") or "",
        "phone": (branch or {}).get("company_phone") or (branch or {}).get("phone") or "",
        "email": (branch or {}).get("company_email") or "",
        "logo_url": (branch or {}).get("logo_url") or "",
    }
    customer_info = {
        "name": sale.get("customer_name", ""),
        "tax_id": customer.get("tax_id") if customer else "",
        "address": customer.get("address") if customer else "",
        "phone": customer.get("phone") if customer else "",
        "email": customer.get("email") if customer else "",
    }

    totals = {
        "subtotal": sale.get("subtotal", 0),
        "tax": sale.get("tax", 0),
        "total": sale.get("total", 0),
    }

    _draw_document_pdf(
        p,
        "Invoice",
        sale.get("invoice_number", ""),
        sale.get("created_at", ""),
        company,
        customer_info,
        vehicle,
        sale.get("items", []),
        currency,
        iva_rate,
        apply_iva,
        totals,
        notes=sale.get("notes"),
    )

    p.save()
    pdf_data = buffer.getvalue()

    # Load invoice HTML template from file to avoid very long inline strings
    tpl_path = ROOT_DIR / "templates" / "invoice.html"
    try:
        tpl = tpl_path.read_text(encoding="utf-8")
    except Exception:
        tpl = (
            "<html><body><h2>Factura {invoice_number}</h2>"
            "<p>Estimado/a {customer_name},</p>"
            "<p>Adjunto encontrará su factura por la compra realizada.</p>"
            "<p>Total: ${total}</p>"
            "<p><em>MUNDO DE ACCESORIOS</em></p></body></html>"
        )

    html_content = tpl.format(
        invoice_number=sale.get("invoice_number", "N/A"),
        customer_name=customer.get("name", ""),
        total=f"{sale.get('total', 0):.2f}",
        created_at=sale.get("created_at", "")[:10],
        payment_status=sale.get("payment_status", "").upper(),
    )
    background_tasks.add_task(
        send_email_notification,
        customer["email"],
        f"Factura {sale['invoice_number']} - MUNDO DE ACCESORIOS",
        html_content,
        pdf_data,
        f"factura_{sale['invoice_number']}.pdf",
    )

    # Log notification
    notif_doc = {
        "notification_id": f"notif_{uuid.uuid4().hex[:8]}",
        "recipient_id": customer["customer_id"],
        "recipient_email": customer["email"],
        "notification_type": "invoice",
        "subject": f"Factura {sale['invoice_number']}",
        "content": "Invoice email sent",
        "related_id": sale_id,
        "status": "sent",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "sent_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.notifications.insert_one(notif_doc)

    return {"message": f"Invoice sent to {customer['email']}"}


# ============ LOW STOCK ALERTS ============


@api_router.get("/alerts/low-stock")
async def get_low_stock_alerts(request: Request):
    """Get all products with low stock"""
    await require_roles(request, ["gerencia", "supervisor", "bodegas"])

    # Find inventory items where quantity <= min_stock
    pipeline = [
        {"$match": {"$expr": {"$lte": ["$quantity", "$min_stock"]}}},
        {
            "$lookup": {
                "from": "products",
                "localField": "product_id",
                "foreignField": "product_id",
                "as": "product",
            }
        },
        {"$unwind": {"path": "$product", "preserveNullAndEmptyArrays": True}},
        {
            "$lookup": {
                "from": "warehouses",
                "localField": "warehouse_id",
                "foreignField": "warehouse_id",
                "as": "warehouse",
            }
        },
        {"$unwind": {"path": "$warehouse", "preserveNullAndEmptyArrays": True}},
        {"$project": {"_id": 0, "product._id": 0, "warehouse._id": 0}},
    ]

    alerts = await db.inventory.aggregate(pipeline).to_list(500)
    return alerts


@api_router.post("/alerts/send-low-stock")
async def send_low_stock_alerts(request: Request, background_tasks: BackgroundTasks):
    """Send email alerts for all low stock items to supervisors"""
    await require_roles(request, ["gerencia"])

    # Get low stock items
    low_stock = await db.inventory.find(
        {"$expr": {"$lte": ["$quantity", "$min_stock"]}}, {"_id": 0}
    ).to_list(500)

    if not low_stock:
        return {"message": "No low stock items"}

    # Get product and warehouse details
    alerts = []
    for item in low_stock:
        product = await db.products.find_one(
            {"product_id": item["product_id"]}, {"_id": 0}
        )
        warehouse = await db.warehouses.find_one(
            {"warehouse_id": item["warehouse_id"]}, {"_id": 0}
        )
        if product:
            alerts.append(
                {
                    "product_name": product["name"],
                    "sku": product["sku"],
                    "warehouse": (
                        warehouse["name"] if warehouse else item["warehouse_id"]
                    ),
                    "current_stock": item["quantity"],
                    "min_stock": item["min_stock"],
                }
            )

    # Generate HTML email
    # Build rows in a loop to avoid very long single-line comprehensions
    rows = []
    for a in alerts:
        rows.append(
            (
                "<tr>"
                "<td>{sku}</td>"
                "<td>{name}</td>"
                "<td>{warehouse}</td>"
                "<td style='color:red;font-weight:bold;'>{current}</td>"
                "<td>{min_stock}</td>"
                "</tr>"
            ).format(
                sku=a.get("sku"),
                name=a.get("product_name"),
                warehouse=a.get("warehouse"),
                current=a.get("current_stock"),
                min_stock=a.get("min_stock"),
            )
        )

    items_html = "".join(rows)

    html_content = f"""
    <html>
    <body style="font-family: Arial, sans-serif;">
        <h2>⚠️ Alerta de Stock Bajo</h2>
        <p>Los siguientes productos tienen stock por debajo del mínimo:</p>
        <table style="border-collapse: collapse; width: 100%;">
            <tr style="background-color: #f2f2f2;">
                <th style="padding: 8px; border: 1px solid #ddd;">SKU</th>
                <th style="padding: 8px; border: 1px solid #ddd;">Producto</th>
                <th style="padding: 8px; border: 1px solid #ddd;">Bodega</th>
                <th style="padding: 8px; border: 1px solid #ddd;">Stock Actual</th>
                <th style="padding: 8px; border: 1px solid #ddd;">Stock Mínimo</th>
            </tr>
            {items_html}
        </table>
        <p><em>Este es un mensaje automático del sistema MUNDO DE ACCESORIOS</em></p>
    </body>
    </html>
    """

    # Get admin/supervisor emails
    admins = await db.users.find(
        {"role": {"$in": ["gerencia", "supervisor"]}, "is_active": True}, {"_id": 0}
    ).to_list(100)

    for admin in admins:
        if admin.get("email"):
            background_tasks.add_task(
                send_email_notification,
                admin["email"],
                f"⚠️ Alerta: {len(alerts)} productos con stock bajo - MUNDO DE ACCESORIOS",
                html_content,
            )

    return {
        "message": f"Alerts sent to {len(admins)} supervisors for {len(alerts)} products"
    }


# ============ RETURNS/DEVOLUCIONES ============


class ReturnCreate(BaseModel):
    sale_id: str
    items: List[Dict[str, Any]]  # [{product_id, quantity, reason}]
    return_type: str = "refund"  # refund, exchange, store_credit
    notes: Optional[str] = None


class Return(BaseModel):
    return_id: str = Field(default_factory=lambda: f"ret_{uuid.uuid4().hex[:8]}")
    sale_id: str
    invoice_number: str
    customer_id: str
    customer_name: str
    items: List[Dict[str, Any]]
    return_type: str
    total_refund: float
    status: str = "pending"  # pending, approved, completed, rejected
    processed_by: Optional[str] = None
    processed_by_name: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    completed_at: Optional[datetime] = None


@api_router.get("/returns")
async def get_returns(request: Request, status: Optional[str] = None):
    user = await require_roles(request, ["gerencia", "supervisor", "ventas", "cajero"])
    query: dict[str, Any] = {}
    if status:
        query["status"] = status
    returns = (
        await db.returns.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    )

    if user.role == "gerencia":
        return returns

    sale_ids = list({r.get("sale_id") for r in returns if r.get("sale_id")})
    sale_map: Dict[str, Dict[str, Any]] = {}
    if sale_ids:
        sales = await db.sales.find(
            {"sale_id": {"$in": sale_ids}},
            {"_id": 0, "sale_id": 1, "branch_id": 1, "salesperson_id": 1, "seller_id": 1, "created_by": 1},
        ).to_list(len(sale_ids) + 5)
        sale_map = {s.get("sale_id"): s for s in sales if s.get("sale_id")}

    filtered_returns: List[Dict[str, Any]] = []
    for ret in returns:
        sale = sale_map.get(ret.get("sale_id"))
        if not sale:
            continue
        if can_access_sale_for_user(user, sale):
            filtered_returns.append(ret)

    return filtered_returns


@api_router.post("/returns")
async def create_return(return_data: ReturnCreate, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor", "ventas", "cajero"])

    sale = await db.sales.find_one({"sale_id": return_data.sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    sale = cast(Dict[str, Any], sale)

    if not can_access_sale_for_user(user, sale):
        raise HTTPException(status_code=403, detail="No autorizado para crear devolución de esta venta")

    # Validate items and calculate refund
    total_refund = 0
    validated_items = []

    sale_items = cast(List[Dict[str, Any]], sale.get("items", []))
    for item in return_data.items:
        sale_item = next(
            (i for i in sale_items if i["product_id"] == item["product_id"]), None
        )
        if not sale_item:
            raise HTTPException(
                status_code=400, detail=f"Product {item['product_id']} not in this sale"
            )

        if item["quantity"] > sale_item["quantity"]:
            raise HTTPException(
                status_code=400, detail="Return quantity exceeds purchased quantity"
            )

        item_refund = (sale_item["unit_price"] * item["quantity"]) * (
            1 - sale_item.get("discount", 0) / 100
        )
        total_refund += item_refund

        validated_items.append(
            {
                "product_id": item["product_id"],
                "product_name": sale_item["product_name"],
                "quantity": item["quantity"],
                "unit_price": sale_item["unit_price"],
                "refund_amount": round(item_refund, 2),
                "reason": item.get("reason", ""),
                "warehouse_id": sale_item.get("warehouse_id"),
            }
        )

    # Add tax to refund
    total_refund += total_refund * 0.12

    return_doc = Return(
        sale_id=sale["sale_id"],
        invoice_number=sale["invoice_number"],
        customer_id=sale["customer_id"],
        customer_name=sale["customer_name"],
        items=validated_items,
        return_type=return_data.return_type,
        total_refund=round(total_refund, 2),
        notes=return_data.notes,
    )

    doc = return_doc.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    doc["branch_id"] = sale.get("branch_id")
    doc["created_by"] = user.user_id
    doc["created_by_name"] = user.name
    await db.returns.insert_one(doc)
    doc.pop("_id", None)

    return doc


@api_router.put("/returns/{return_id}/approve")
async def approve_return(return_id: str, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor"])

    ret = await db.returns.find_one({"return_id": return_id})
    if not ret:
        raise HTTPException(status_code=404, detail="Return not found")
    ret = cast(Dict[str, Any], ret)

    sale = await db.sales.find_one({"sale_id": ret.get("sale_id")}, {"_id": 0})
    if sale and not can_access_sale_for_user(user, cast(Dict[str, Any], sale)):
        raise HTTPException(status_code=403, detail="No autorizado para procesar esta devolución")

    if ret["status"] != "pending":
        raise HTTPException(status_code=400, detail="Return already processed")

    # Return items to inventory
    items = cast(List[Dict[str, Any]], ret.get("items", []))
    for item in items:
        warehouse_id = item.get("warehouse_id", "wh_main")
        await db.inventory.update_one(
            {"product_id": item["product_id"], "warehouse_id": warehouse_id},
            {
                "$inc": {"quantity": item["quantity"]},
                "$set": {"last_updated": datetime.now(timezone.utc).isoformat()},
            },
            upsert=True,
        )
        await audit_service.log_inventory_movement(
            product_id=item["product_id"],
            warehouse_id=warehouse_id,
            quantity_change=int(item.get("quantity", 0)),
            reason="return_approved_in",
            actor=user,
            branch_id=user.branch_id,
            reference_id=return_id,
            metadata={"sale_id": ret.get("sale_id")},
        )

    # Update customer credit if it was a credit sale
    sale = await db.sales.find_one({"sale_id": ret["sale_id"]}, {"_id": 0})
    if sale and sale.get("payment_type") == "credit":
        await db.customers.update_one(
            {"customer_id": ret["customer_id"]},
            {"$inc": {"credit_balance": -ret["total_refund"]}},
        )

    await db.returns.update_one(
        {"return_id": return_id},
        {
            "$set": {
                "status": "completed",
                "processed_by": user.user_id,
                "processed_by_name": user.name,
                "completed_at": datetime.now(timezone.utc).isoformat(),
            }
        },
    )

    return {"message": "Return approved and processed"}


@api_router.put("/returns/{return_id}/reject")
async def reject_return(return_id: str, reason: str, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor"])

    ret = await db.returns.find_one({"return_id": return_id}, {"_id": 0})
    if not ret:
        raise HTTPException(status_code=404, detail="Return not found")

    sale = await db.sales.find_one({"sale_id": ret.get("sale_id")}, {"_id": 0})
    if sale and not can_access_sale_for_user(user, cast(Dict[str, Any], sale)):
        raise HTTPException(status_code=403, detail="No autorizado para rechazar esta devolución")

    result = await db.returns.update_one(
        {"return_id": return_id, "status": "pending"},
        {
            "$set": {
                "status": "rejected",
                "processed_by": user.user_id,
                "processed_by_name": user.name,
                "rejection_reason": reason,
                "completed_at": datetime.now(timezone.utc).isoformat(),
            }
        },
    )

    if result.modified_count == 0:
        raise HTTPException(
            status_code=404, detail="Return not found or already processed"
        )

    return {"message": "Return rejected"}


# ============ INSTALLATION CALENDAR ============


class CalendarEvent(BaseModel):
    event_id: str = Field(default_factory=lambda: f"evt_{uuid.uuid4().hex[:8]}")
    title: str
    event_type: str  # work_order, delivery, appointment
    related_id: Optional[str] = None
    start_time: datetime
    end_time: datetime
    branch_id: Optional[str] = None
    technician_id: Optional[str] = None
    technician_name: Optional[str] = None
    customer_name: Optional[str] = None
    vehicle_info: Optional[str] = None
    status: str = "scheduled"  # scheduled, in_progress, completed, cancelled
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


@api_router.get("/calendar")
async def get_calendar_events(
    request: Request,
    start_date: str,
    end_date: str,
    branch_id: Optional[str] = None,
    technician_id: Optional[str] = None,
):
    user = await require_auth(request)

    query: dict[str, Any] = {"start_time": {"$gte": start_date, "$lte": end_date + "T23:59:59"}}
    if branch_id:
        query["branch_id"] = branch_id
    elif user.branch_id:
        query["branch_id"] = user.branch_id
    if technician_id:
        query["technician_id"] = technician_id
    elif user.role == "instalaciones":
        query["technician_id"] = user.user_id

    events = (
        await db.calendar_events.find(query, {"_id": 0})
        .sort("start_time", 1)
        .to_list(500)
    )
    return events


@api_router.post("/calendar")
async def create_calendar_event(event: CalendarEvent, request: Request):
    await require_roles(request, ["gerencia", "supervisor", "instalaciones"])

    doc = event.model_dump()
    doc["start_time"] = doc["start_time"].isoformat()
    doc["end_time"] = doc["end_time"].isoformat()
    doc["created_at"] = doc["created_at"].isoformat()

    await db.calendar_events.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/calendar/{event_id}")
async def update_calendar_event(
    event_id: str, updates: Dict[str, Any], request: Request
):
    await require_roles(request, ["gerencia", "supervisor", "instalaciones"])

    if "start_time" in updates and isinstance(updates["start_time"], datetime):
        updates["start_time"] = updates["start_time"].isoformat()
    if "end_time" in updates and isinstance(updates["end_time"], datetime):
        updates["end_time"] = updates["end_time"].isoformat()

    result = await db.calendar_events.update_one(
        {"event_id": event_id}, {"$set": updates}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    return {"message": "Event updated"}


@api_router.delete("/calendar/{event_id}")
async def delete_calendar_event(event_id: str, request: Request):
    await require_roles(request, ["gerencia", "supervisor"])
    result = await db.calendar_events.delete_one({"event_id": event_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    return {"message": "Event deleted"}


# Auto-create calendar events from work orders
@api_router.post("/calendar/sync-work-orders")
async def sync_work_orders_to_calendar(request: Request):
    await require_roles(request, ["gerencia", "supervisor"])

    # Get work orders without calendar events
    work_orders = await db.work_orders.find(
        {"status": {"$in": ["pending", "in_progress"]}}, {"_id": 0}
    ).to_list(500)

    synced = 0
    for wo in work_orders:
        # Check if event already exists
        existing = await db.calendar_events.find_one(
            {"related_id": wo["work_order_id"]}
        )
        if existing:
            continue

        # Create calendar event
        start_time = (
            datetime.fromisoformat(wo["created_at"])
            if isinstance(wo["created_at"], str)
            else wo["created_at"]
        )
        end_time = start_time + timedelta(minutes=wo.get("estimated_time", 60))

        event_doc = {
            "event_id": f"evt_{uuid.uuid4().hex[:8]}",
            "title": f"OT: {wo['work_order_id'][:12]}",
            "event_type": "work_order",
            "related_id": wo["work_order_id"],
            "start_time": start_time.isoformat(),
            "end_time": end_time.isoformat(),
            "branch_id": wo.get("branch_id"),
            "technician_id": wo.get("technician_id"),
            "technician_name": wo.get("technician_name"),
            "customer_name": wo.get("customer_name"),
            "vehicle_info": wo.get("vehicle_info"),
            "status": "scheduled" if wo["status"] == "pending" else "in_progress",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.calendar_events.insert_one(event_doc)
        synced += 1

    return {"message": f"Synced {synced} work orders to calendar"}


# ============ WARRANTY TRACKING ============


class WarrantyClaim(BaseModel):
    claim_id: str = Field(default_factory=lambda: f"wcl_{uuid.uuid4().hex[:8]}")
    vehicle_id: str
    customer_id: str
    customer_name: str
    vehicle_info: str
    sale_id: str
    invoice_number: str
    product_id: str
    product_name: str
    purchase_date: str
    warranty_end_date: str
    claim_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    issue_description: str
    status: str = "pending"  # pending, approved, in_repair, completed, denied
    resolution: Optional[str] = None
    work_order_id: Optional[str] = None
    notes: Optional[str] = None


@api_router.get("/warranties/vehicle/{vehicle_id}")
async def get_vehicle_warranty_history(vehicle_id: str, request: Request):
    """Get all warranty-covered purchases and claims for a vehicle"""
    await require_auth(request)

    vehicle = await db.vehicles.find_one({"vehicle_id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")

    customer = await db.customers.find_one(
        {"customer_id": vehicle["customer_id"]}, {"_id": 0}
    )

    # Get all sales for this customer
    sales = await db.sales.find(
        {"customer_id": vehicle["customer_id"]}, {"_id": 0}
    ).to_list(500)

    warranty_items = []
    now = datetime.now(timezone.utc)

    for sale in sales:
        purchase_date = (
            datetime.fromisoformat(sale["created_at"].replace("Z", "+00:00"))
            if isinstance(sale["created_at"], str)
            else sale["created_at"]
        )

        for item in sale["items"]:
            product = await db.products.find_one(
                {"product_id": item["product_id"]}, {"_id": 0}
            )
            if not product:
                continue

            warranty_months = product.get("warranty_months", 12)
            warranty_end = purchase_date + timedelta(days=warranty_months * 30)
            is_active = warranty_end > now

            warranty_items.append(
                {
                    "sale_id": sale["sale_id"],
                    "invoice_number": sale["invoice_number"],
                    "product_id": item["product_id"],
                    "product_name": item["product_name"],
                    "purchase_date": purchase_date.isoformat(),
                    "warranty_months": warranty_months,
                    "warranty_end_date": warranty_end.isoformat(),
                    "is_warranty_active": is_active,
                    "days_remaining": (
                        max(0, (warranty_end - now).days) if is_active else 0
                    ),
                }
            )

    # Get warranty claims
    claims = (
        await db.warranty_claims.find({"vehicle_id": vehicle_id}, {"_id": 0})
        .sort("claim_date", -1)
        .to_list(100)
    )

    return {
        "vehicle": vehicle,
        "customer": customer,
        "warranty_items": warranty_items,
        "claims": claims,
    }


@api_router.post("/warranties/claim")
async def create_warranty_claim(
    request: Request,
    vehicle_id: str,
    sale_id: str,
    product_id: str,
    issue_description: str,
):
    await require_auth(request)

    vehicle = await db.vehicles.find_one({"vehicle_id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")

    customer = await db.customers.find_one(
        {"customer_id": vehicle["customer_id"]}, {"_id": 0}
    )
    sale = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    sale = cast(Dict[str, Any], sale)

    sale_items = cast(List[Dict[str, Any]], sale.get("items", []))
    sale_item = next((i for i in sale_items if i["product_id"] == product_id), None)
    if not sale_item:
        raise HTTPException(status_code=400, detail="Product not in this sale")

    product = await db.products.find_one({"product_id": product_id}, {"_id": 0})
    warranty_months = product.get("warranty_months", 12) if product else 12

    purchase_date = (
        datetime.fromisoformat(sale["created_at"].replace("Z", "+00:00"))
        if isinstance(sale["created_at"], str)
        else sale["created_at"]
    )
    warranty_end = purchase_date + timedelta(days=warranty_months * 30)

    if warranty_end < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Warranty has expired")

    vehicle_info = (
        f"{vehicle['brand']} {vehicle['model']} {vehicle['year']} - {vehicle['plate']}"
    )

    claim = WarrantyClaim(
        vehicle_id=vehicle_id,
        customer_id=vehicle["customer_id"],
        customer_name=customer["name"] if customer else "Unknown",
        vehicle_info=vehicle_info,
        sale_id=sale_id,
        invoice_number=sale["invoice_number"],
        product_id=product_id,
        product_name=sale_item["product_name"],
        purchase_date=purchase_date.isoformat(),
        warranty_end_date=warranty_end.isoformat(),
        issue_description=issue_description,
    )

    doc = claim.model_dump()
    doc["claim_date"] = doc["claim_date"].isoformat()
    await db.warranty_claims.insert_one(doc)
    doc.pop("_id", None)

    return doc


@api_router.get("/warranties/claims")
async def get_warranty_claims(request: Request, status: Optional[str] = None):
    await require_roles(request, ["gerencia", "supervisor", "instalaciones"])
    query: dict[str, Any] = {}
    if status:
        query["status"] = status
    claims = (
        await db.warranty_claims.find(query, {"_id": 0})
        .sort("claim_date", -1)
        .to_list(500)
    )
    return claims


@api_router.put("/warranties/claims/{claim_id}")
async def update_warranty_claim(
    claim_id: str, updates: Dict[str, Any], request: Request
):
    await require_roles(request, ["gerencia", "supervisor"])

    result = await db.warranty_claims.update_one(
        {"claim_id": claim_id}, {"$set": updates}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Claim not found")
    return {"message": "Claim updated"}


# ============ PRODUCTIVITY REPORTS ============


@api_router.get("/reports/productivity")
async def get_productivity_report(request: Request, start_date: str, end_date: str):
    """Get detailed productivity report by salesperson and technician"""
    await require_roles(request, ["gerencia", "supervisor"])

    # Sales by salesperson
    sales = await db.sales.find(
        {"created_at": {"$gte": start_date, "$lte": end_date + "T23:59:59"}}, {"_id": 0}
    ).to_list(10000)

    salesperson_stats: dict[str, Any] = {}
    for sale in sales:
        sp = sale.get("salesperson_name", "Unknown")
        if sp not in salesperson_stats:
            salesperson_stats[sp] = {
                "name": sp,
                "user_id": sale.get("salesperson_id"),
                "total_sales": 0,
                "sales_count": 0,
                "cash_sales": 0,
                "credit_sales": 0,
                "avg_sale_value": 0,
            }
        salesperson_stats[sp]["total_sales"] += sale["total"]
        salesperson_stats[sp]["sales_count"] += 1
        if sale["payment_type"] == "cash":
            salesperson_stats[sp]["cash_sales"] += sale["total"]
        else:
            salesperson_stats[sp]["credit_sales"] += sale["total"]

    for sp in salesperson_stats.values():
        if sp["sales_count"] > 0:
            sp["avg_sale_value"] = round(sp["total_sales"] / sp["sales_count"], 2)
        sp["total_sales"] = round(sp["total_sales"], 2)
        sp["cash_sales"] = round(sp["cash_sales"], 2)
        sp["credit_sales"] = round(sp["credit_sales"], 2)

    # Work orders by technician
    work_orders = await db.work_orders.find(
        {
            "created_at": {"$gte": start_date, "$lte": end_date + "T23:59:59"},
            "status": "completed",
        },
        {"_id": 0},
    ).to_list(10000)

    technician_stats: dict[str, Any] = {}
    for wo in work_orders:
        tech = wo.get("technician_name", "Sin asignar")
        if tech not in technician_stats:
            technician_stats[tech] = {
                "name": tech,
                "user_id": wo.get("technician_id"),
                "completed_orders": 0,
                "total_time_minutes": 0,
                "avg_time_minutes": 0,
                "quality_scores": [],
                "avg_quality": 0,
            }
        technician_stats[tech]["completed_orders"] += 1
        if wo.get("actual_time"):
            technician_stats[tech]["total_time_minutes"] += wo["actual_time"]
        if wo.get("quality_score"):
            technician_stats[tech]["quality_scores"].append(wo["quality_score"])

    for tech in technician_stats.values():
        if tech["completed_orders"] > 0:
            tech["avg_time_minutes"] = round(
                tech["total_time_minutes"] / tech["completed_orders"], 1
            )
        if tech["quality_scores"]:
            tech["avg_quality"] = round(
                sum(tech["quality_scores"]) / len(tech["quality_scores"]), 1
            )
        del tech["quality_scores"]

    return {
        "period": {"start": start_date, "end": end_date},
        "salespeople": list(salesperson_stats.values()),
        "technicians": list(technician_stats.values()),
        "summary": {
            "total_sales": round(
                sum(s["total_sales"] for s in salesperson_stats.values()), 2
            ),
            "total_orders": len(sales),
            "completed_installations": len(work_orders),
            "top_salesperson": (
                max(salesperson_stats.values(), key=lambda x: x["total_sales"])["name"]
                if salesperson_stats
                else None
            ),
            "top_technician": (
                max(technician_stats.values(), key=lambda x: x["completed_orders"])[
                    "name"
                ]
                if technician_stats
                else None
            ),
        },
    }


# ============ EXPORT REPORTS ============


@api_router.get("/reports/export/sales")
async def export_sales_report(
    request: Request, start_date: str, end_date: str, format: str = "csv"
):
    """Export sales report as CSV or Excel"""
    await require_roles(request, ["gerencia", "supervisor"])

    sales = (
        await db.sales.find(
            {"created_at": {"$gte": start_date, "$lte": end_date + "T23:59:59"}},
            {"_id": 0},
        )
        .sort("created_at", -1)
        .to_list(10000)
    )

    if format == "csv":
        import csv
        from io import StringIO

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(
            [
                "Factura",
                "Fecha",
                "Cliente",
                "Vendedor",
                "Tipo Pago",
                "Subtotal",
                "IVA",
                "Descuento",
                "Total",
                "Estado",
            ]
        )

        for sale in sales:
            writer.writerow(
                [
                    sale["invoice_number"],
                    sale["created_at"][:10],
                    sale["customer_name"],
                    sale["salesperson_name"],
                    sale["payment_type"],
                    sale["subtotal"],
                    sale["tax"],
                    sale["discount"],
                    sale["total"],
                    sale["payment_status"],
                ]
            )

        content = output.getvalue()
        return Response(
            content=content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=ventas_{start_date}_{end_date}.csv"
            },
        )

    elif format == "pdf":
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter

        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, "Reporte de Ventas")
        p.setFont("Helvetica", 10)
        p.drawString(50, height - 70, f"Período: {start_date} al {end_date}")

        total_sales = sum(s["total"] for s in sales)
        p.drawString(50, height - 90, f"Total de Ventas: ${total_sales:,.2f}")
        p.drawString(250, height - 90, f"Cantidad: {len(sales)} facturas")

        y = height - 130
        p.setFont("Helvetica-Bold", 8)
        p.drawString(50, y, "Factura")
        p.drawString(120, y, "Fecha")
        p.drawString(180, y, "Cliente")
        p.drawString(320, y, "Vendedor")
        p.drawString(420, y, "Total")
        p.drawString(480, y, "Estado")
        p.line(50, y - 5, 550, y - 5)

        y -= 15
        p.setFont("Helvetica", 7)
        for sale in sales[:50]:  # Limit to 50 for PDF
            p.drawString(50, y, sale["invoice_number"])
            p.drawString(120, y, sale["created_at"][:10])
            p.drawString(180, y, sale["customer_name"][:20])
            p.drawString(320, y, sale["salesperson_name"][:15])
            p.drawString(420, y, f"${sale['total']:,.2f}")
            p.drawString(480, y, sale["payment_status"])
            y -= 12
            if y < 50:
                p.showPage()
                y = height - 50

        p.save()
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=ventas_{start_date}_{end_date}.pdf"
            },
        )

    else:
        raise HTTPException(status_code=400, detail="Format must be 'csv' or 'pdf'")


# ============ REAL-TIME NOTIFICATIONS (WebSocket-like polling) ============


@api_router.get("/notifications/recent")
async def get_recent_notifications(request: Request, since: Optional[str] = None):
    """Get recent notifications for real-time updates"""
    user = await require_auth(request)

    query = {}
    if since:
        query["created_at"] = {"$gt": since}

    # Get notifications relevant to user's role
    if user.role == "instalaciones":
        query["$or"] = [
            {"notification_type": "work_order_update", "recipient_id": user.user_id},
            {"notification_type": {"$in": ["work_order_assigned"]}},
        ]
    elif user.role == "transporte":
        query["notification_type"] = {"$in": ["delivery_assigned", "delivery_update"]}
    elif user.role in {"ventas", "cajero"}:
        query["notification_type"] = {"$in": ["sale_completed", "quotation_approved"]}

    notifications = (
        await db.notifications.find(query, {"_id": 0})
        .sort("created_at", -1)
        .limit(20)
        .to_list(20)
    )
    return notifications


@api_router.post("/notifications/mark-read")
async def mark_notifications_read(request: Request, notification_ids: List[str]):
    await require_auth(request)

    await db.notifications.update_many(
        {"notification_id": {"$in": notification_ids}},
        {"$set": {"read": True, "read_at": datetime.now(timezone.utc).isoformat()}},
    )
    return {"message": f"Marked {len(notification_ids)} notifications as read"}


# Create notification helper
async def create_notification(
    recipient_id: str,
    recipient_email: str,
    notification_type: str,
    subject: str,
    content: str,
    related_id: Optional[str] = None,
):
    doc = {
        "notification_id": f"notif_{uuid.uuid4().hex[:8]}",
        "recipient_id": recipient_id,
        "recipient_email": recipient_email,
        "notification_type": notification_type,
        "subject": subject,
        "content": content,
        "related_id": related_id,
        "status": "sent",
        "read": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.notifications.insert_one(doc)
    return doc


# ============ DASHBOARD BY ROLE ============


@api_router.get("/dashboard/role-stats")
async def get_role_dashboard(request: Request):
    """Get dashboard stats specific to user's role"""
    user = await require_auth(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    stats: dict[str, Any] = {"role": user.role}

    if user.role in ["gerencia", "supervisor"]:
        # Gerencia: all branches | Supervisor: own branch only
        sales_scope_query = build_sales_visibility_query(user)
        sales_today = await db.sales.find(
            merge_queries(sales_scope_query, {"created_at": {"$regex": f"^{today}"}}),
            {"_id": 0},
        ).to_list(1000)
        stats["sales_today"] = {
            "count": len(sales_today),
            "total": round(sum(s["total"] for s in sales_today), 2),
        }
        stats["pending_work_orders"] = await db.work_orders.count_documents(
            {"status": {"$in": ["pending", "in_progress"]}}
        )
        stats["low_stock_items"] = await db.inventory.count_documents(
            {"$expr": {"$lte": ["$quantity", "$min_stock"]}}
        )
        stats["pending_deliveries"] = await db.sales.count_documents(
            merge_queries(sales_scope_query, {"delivery_status": "pending"})
        )
        stats["pending_credits"] = await db.sales.count_documents(
            merge_queries(
                sales_scope_query,
                {"payment_type": "credit", "payment_status": {"$ne": "paid"}},
            )
        )
        stats["pending_returns"] = await db.returns.count_documents(
            {"status": "pending"}
        )
        stats["warranty_claims"] = await db.warranty_claims.count_documents(
            {"status": "pending"}
        )

    elif user.role in {"ventas", "cajero"}:
        # Salesperson stats
        my_sales = await db.sales.find(
            {"salesperson_id": user.user_id, "created_at": {"$regex": f"^{today}"}},
            {"_id": 0},
        ).to_list(100)
        stats["my_sales_today"] = {
            "count": len(my_sales),
            "total": round(sum(s["total"] for s in my_sales), 2),
        }
        stats["my_pending_quotations"] = await db.quotations.count_documents(
            {"salesperson_id": user.user_id, "status": "pending"}
        )

    elif user.role == "instalaciones":
        # Technician stats
        stats["my_assigned_orders"] = await db.work_orders.count_documents(
            {
                "technician_id": user.user_id,
                "status": {"$in": ["pending", "in_progress"]},
            }
        )
        stats["available_orders"] = await db.work_orders.count_documents(
            {"technician_id": None, "status": "pending"}
        )
        completed_today = await db.work_orders.find(
            {
                "technician_id": user.user_id,
                "status": "completed",
                "end_time": {"$regex": f"^{today}"},
            },
            {"_id": 0},
        ).to_list(100)
        stats["completed_today"] = len(completed_today)

    elif user.role == "transporte":
        # Driver stats
        stats["my_pending_deliveries"] = await db.sales.count_documents(
            {
                "delivery_driver_id": user.user_id,
                "delivery_status": {"$in": ["pending", "assigned", "in_transit"]},
            }
        )
        stats["unassigned_deliveries"] = await db.sales.count_documents(
            {
                "delivery_required": True,
                "delivery_driver_id": None,
                "delivery_status": "pending",
            }
        )
        delivered_today = await db.sales.find(
            {
                "delivery_driver_id": user.user_id,
                "delivery_status": "delivered",
                "delivery_completed_at": {"$regex": f"^{today}"},
            },
            {"_id": 0},
        ).to_list(100)
        stats["delivered_today"] = len(delivered_today)

    elif user.role == "bodegas":
        # Warehouse stats
        query = {}
        if user.warehouse_id:
            query["warehouse_id"] = user.warehouse_id
        stats["total_items"] = await db.inventory.count_documents(query)
        stats["low_stock_items"] = await db.inventory.count_documents(
            {**query, "$expr": {"$lte": ["$quantity", "$min_stock"]}}
        )
        stats["pending_transfers"] = 0  # TODO: implement transfer tracking

    return stats


# ============ PUSH NOTIFICATIONS (PWA) ============


class PushSubscription(BaseModel):
    endpoint: str
    keys: Dict[str, str]  # p256dh, auth
    user_id: Optional[str] = None


@api_router.post("/push/subscribe")
async def subscribe_push_notifications(
    subscription: PushSubscription, request: Request
):
    """Subscribe to push notifications"""
    user = await require_auth(request)

    doc = {
        "subscription_id": f"push_{uuid.uuid4().hex[:8]}",
        "user_id": user.user_id,
        "endpoint": subscription.endpoint,
        "keys": subscription.keys,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "is_active": True,
    }

    # Upsert by endpoint
    await db.push_subscriptions.update_one(
        {"endpoint": subscription.endpoint}, {"$set": doc}, upsert=True
    )

    return {"message": "Subscribed to push notifications"}


@api_router.delete("/push/unsubscribe")
async def unsubscribe_push_notifications(endpoint: str, request: Request):
    """Unsubscribe from push notifications"""
    await require_auth(request)

    await db.push_subscriptions.delete_one({"endpoint": endpoint})
    return {"message": "Unsubscribed from push notifications"}


@api_router.post("/push/send")
async def send_push_notification(
    request: Request,
    user_id: Optional[str] = None,
    role: Optional[str] = None,
    title: str = "MUNDO DE ACCESORIOS",
    body: str = "",
    url: Optional[str] = None,
):
    """Send push notification to users (admin only)"""
    await require_roles(request, ["gerencia", "supervisor"])

    query: dict[str, Any] = {"is_active": True}
    if user_id:
        query["user_id"] = user_id
    elif role:
        # Get users with this role
        users = await db.users.find({"role": role}, {"user_id": 1}).to_list(1000)
        user_ids = [u["user_id"] for u in users]
        query["user_id"] = {"$in": user_ids}

    subscriptions = await db.push_subscriptions.find(query, {"_id": 0}).to_list(1000)

    # Store notification for polling fallback
    notification_doc = {
        "notification_id": f"notif_{uuid.uuid4().hex[:8]}",
        "title": title,
        "body": body,
        "url": url,
        "target_user_id": user_id,
        "target_role": role,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "type": "push",
    }
    await db.notifications.insert_one(notification_doc)

    # In production, you would use pywebpush here
    # For now, we store for polling
    return {
        "message": f"Notification queued for {len(subscriptions)} subscribers",
        "notification_id": notification_doc["notification_id"],
    }


@api_router.get("/push/notifications")
async def get_push_notifications(
    request: Request, since: Optional[str] = None, limit: int = 20
):
    """Get recent notifications for polling fallback"""
    user = await require_auth(request)

    query: dict[str, Any] = {
        "$or": [
            {"target_user_id": user.user_id},
            {"target_user_id": None, "target_role": user.role},
            {"target_user_id": None, "target_role": None},
        ]
    }

    if since:
        query["created_at"] = {"$gt": since}

    notifications = (
        await db.notifications.find(query, {"_id": 0})
        .sort("created_at", -1)
        .limit(limit)
        .to_list(limit)
    )
    return notifications


# ============ THERMAL PRINTER (ESC/POS) ============


class ThermalPrintJob(BaseModel):
    print_type: str  # receipt, ticket, order
    sale_id: Optional[str] = None
    work_order_id: Optional[str] = None
    custom_content: Optional[str] = None


@api_router.post("/print/thermal")
async def generate_thermal_print(job: ThermalPrintJob, request: Request):
    """Generate ESC/POS commands for thermal printer (80mm)"""
    await require_auth(request)

    # ESC/POS Commands
    ESC = b"\x1b"
    GS = b"\x1d"
    INIT = ESC + b"@"  # Initialize printer
    ALIGN_CENTER = ESC + b"a\x01"
    ALIGN_LEFT = ESC + b"a\x00"
    ALIGN_RIGHT = ESC + b"a\x02"
    BOLD_ON = ESC + b"E\x01"
    BOLD_OFF = ESC + b"E\x00"
    DOUBLE_HEIGHT = ESC + b"!\x10"
    NORMAL_SIZE = ESC + b"!\x00"
    CUT = GS + b"V\x00"  # Full cut
    FEED = ESC + b"d\x03"  # Feed 3 lines

    commands = bytearray()
    commands.extend(INIT)

    if job.print_type == "receipt" and job.sale_id:
        sale = await db.sales.find_one({"sale_id": job.sale_id}, {"_id": 0})
        if not sale:
            raise HTTPException(status_code=404, detail="Sale not found")
        sale = cast(Dict[str, Any], sale)

        # Header
        commands.extend(ALIGN_CENTER)
        commands.extend(DOUBLE_HEIGHT)
        commands.extend("MUNDO DE ACCESORIOS\n".encode("cp437", errors="replace"))
        commands.extend(NORMAL_SIZE)
        commands.extend("Sistema de Facturacion\n".encode("cp437", errors="replace"))
        commands.extend(b"-" * 32 + b"\n")

        # Invoice info
        commands.extend(ALIGN_LEFT)
        commands.extend(
            f"Factura: {sale['invoice_number']}\n".encode("cp437", errors="replace")
        )
        commands.extend(
            f"Fecha: {sale['created_at'][:16]}\n".encode("cp437", errors="replace")
        )
        commands.extend(
            f"Cliente: {sale['customer_name'][:20]}\n".encode("cp437", errors="replace")
        )
        commands.extend(
            f"Vendedor: {sale['salesperson_name'][:15]}\n".encode(
                "cp437", errors="replace"
            )
        )
        commands.extend(b"-" * 32 + b"\n")

        # Items
        sale_items = cast(List[Dict[str, Any]], sale.get("items", []))
        for item in sale_items:
            name = item["product_name"][:18]
            qty = item["quantity"]
            price = item["subtotal"]
            commands.extend(f"{name}\n".encode("cp437", errors="replace"))
            commands.extend(
                f"  {qty} x ${item['unit_price']:.2f} = ${price:.2f}\n".encode(
                    "cp437", errors="replace"
                )
            )

        # Totals
        commands.extend(b"-" * 32 + b"\n")
        commands.extend(ALIGN_RIGHT)
        commands.extend(
            f"Subtotal: ${sale['subtotal']:.2f}\n".encode("cp437", errors="replace")
        )
        commands.extend(
            f"IVA 12%: ${sale['tax']:.2f}\n".encode("cp437", errors="replace")
        )
        if sale.get("discount", 0) > 0:
            commands.extend(
                f"Descuento: -${sale['discount']:.2f}\n".encode(
                    "cp437", errors="replace"
                )
            )
        commands.extend(BOLD_ON)
        commands.extend(
            f"TOTAL: ${sale['total']:.2f}\n".encode("cp437", errors="replace")
        )
        commands.extend(BOLD_OFF)

        # Footer
        commands.extend(ALIGN_CENTER)
        commands.extend(b"\n")
        commands.extend(
            f"Pago: {sale['payment_type'].upper()}\n".encode("cp437", errors="replace")
        )
        commands.extend("Gracias por su compra!\n".encode("cp437", errors="replace"))

    elif job.print_type == "order" and job.work_order_id:
        wo = await db.work_orders.find_one(
            {"work_order_id": job.work_order_id}, {"_id": 0}
        )
        if not wo:
            raise HTTPException(status_code=404, detail="Work order not found")
        wo = cast(Dict[str, Any], wo)

        # Header
        commands.extend(ALIGN_CENTER)
        commands.extend(DOUBLE_HEIGHT)
        commands.extend("ORDEN DE TRABAJO\n".encode("cp437", errors="replace"))
        commands.extend(NORMAL_SIZE)
        commands.extend(b"-" * 32 + b"\n")

        # Order info
        commands.extend(ALIGN_LEFT)
        commands.extend(BOLD_ON)
        commands.extend(
            f"OT: {wo['work_order_id'][:15]}\n".encode("cp437", errors="replace")
        )
        commands.extend(BOLD_OFF)
        commands.extend(
            f"Fecha: {wo['created_at'][:16]}\n".encode("cp437", errors="replace")
        )
        commands.extend(
            f"Cliente: {wo.get('customer_name', 'N/A')[:20]}\n".encode(
                "cp437", errors="replace"
            )
        )
        commands.extend(
            f"Vehiculo: {wo.get('vehicle_info', 'N/A')[:25]}\n".encode(
                "cp437", errors="replace"
            )
        )
        commands.extend(
            f"Tecnico: {wo.get('technician_name', 'Sin asignar')[:15]}\n".encode(
                "cp437", errors="replace"
            )
        )
        commands.extend(b"-" * 32 + b"\n")

        # Items
        wo_items = cast(List[Dict[str, Any]], wo.get("items", []))
        for item in wo_items:
            commands.extend(
                f"- {item.get('product_name', item.get('description', 'Item'))[:28]}\n".encode(
                    "cp437", errors="replace"
                )
            )

        if wo.get("notes"):
            commands.extend(b"\n")
            commands.extend(
                f"Notas: {wo['notes'][:50]}\n".encode("cp437", errors="replace")
            )

    elif job.print_type == "ticket" and job.custom_content:
        commands.extend(ALIGN_CENTER)
        commands.extend(job.custom_content.encode("cp437", errors="replace"))

    # Cut paper
    commands.extend(FEED)
    commands.extend(CUT)

    import base64

    return {
        "commands_base64": base64.b64encode(bytes(commands)).decode(),
        "commands_hex": bytes(commands).hex(),
        "print_type": job.print_type,
        "instructions": "Send base64 decoded data directly to thermal printer via USB/Serial/Network",
    }


CURRENCIES = {
    "USD": {"symbol": "$", "name": "Dólar Estadounidense", "decimal_places": 2},
    "NIO": {"symbol": "C$", "name": "Córdoba Nicaragüense", "decimal_places": 2},
    "EUR": {"symbol": "€", "name": "Euro", "decimal_places": 2},
    "MXN": {"symbol": "$", "name": "Peso Mexicano", "decimal_places": 2},
    "COP": {"symbol": "$", "name": "Peso Colombiano", "decimal_places": 0},
    "PEN": {"symbol": "S/", "name": "Sol Peruano", "decimal_places": 2},
    "CLP": {"symbol": "$", "name": "Peso Chileno", "decimal_places": 0},
    "ARS": {"symbol": "$", "name": "Peso Argentino", "decimal_places": 2},
    "BRL": {"symbol": "R$", "name": "Real Brasileño", "decimal_places": 2},
}


class ExchangeRate(BaseModel):
    from_currency: str
    to_currency: str
    rate: float
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


@api_router.get("/currencies")
async def get_currencies():
    """Get available currencies"""
    return {"currencies": CURRENCIES}


@api_router.get("/currencies/rates")
async def get_exchange_rates(base: str = "USD"):
    """Get exchange rates from base currency"""
    if base not in CURRENCIES:
        raise HTTPException(status_code=400, detail="Invalid base currency")

    # Get stored rates or return defaults
    rates = await db.exchange_rates.find({"from_currency": base}, {"_id": 0}).to_list(
        100
    )

    if not rates:
        # Default rates (should be updated via external API or manually)
        default_rates = {
            "USD": {
                "EUR": 0.92,
                "MXN": 17.15,
                "COP": 3950,
                "PEN": 3.72,
                "CLP": 880,
                "ARS": 850,
                "BRL": 4.95,
            },
            "EUR": {
                "USD": 1.09,
                "MXN": 18.65,
                "COP": 4300,
                "PEN": 4.05,
                "CLP": 957,
                "ARS": 925,
                "BRL": 5.38,
            },
        }
        return {"base": base, "rates": default_rates.get(base, {}), "source": "default"}

    return {
        "base": base,
        "rates": {r["to_currency"]: r["rate"] for r in rates},
        "source": "database",
    }


@api_router.get("/currencies/usd-nio-effective")
async def get_effective_usd_nio_rate():
    """Get the effective USD -> NIO exchange rate used system-wide."""
    rate, source = await _get_usd_to_nio_rate_with_source()
    return {
        "pair": "USD_NIO",
        "rate": rate,
        "source": source,
    }


@api_router.put("/currencies/rates")
async def update_exchange_rate(rate: ExchangeRate, request: Request):
    """Update exchange rate (admin only)"""
    await require_roles(request, ["gerencia"])

    if rate.from_currency not in CURRENCIES or rate.to_currency not in CURRENCIES:
        raise HTTPException(status_code=400, detail="Invalid currency")

    doc = rate.model_dump()
    doc["updated_at"] = doc["updated_at"].isoformat()

    await db.exchange_rates.update_one(
        {"from_currency": rate.from_currency, "to_currency": rate.to_currency},
        {"$set": doc},
        upsert=True,
    )

    return {"message": "Exchange rate updated"}


@api_router.get("/currencies/convert")
async def convert_currency(amount: float, from_currency: str, to_currency: str):
    """Convert amount between currencies"""
    if from_currency not in CURRENCIES or to_currency not in CURRENCIES:
        raise HTTPException(status_code=400, detail="Invalid currency")

    if from_currency == to_currency:
        return {
            "amount": amount,
            "converted": amount,
            "from": from_currency,
            "to": to_currency,
            "rate": 1,
        }

    # Centralized effective path for USD<->NIO, shared with cashier/sales/dashboard.
    if from_currency == "USD" and to_currency == "NIO":
        rate = await _get_usd_to_nio_rate()
    elif from_currency == "NIO" and to_currency == "USD":
        usd_to_nio = await _get_usd_to_nio_rate()
        rate = 1 / usd_to_nio if usd_to_nio > 0 else 0
    else:
        # Get rate
        rate_doc = await db.exchange_rates.find_one(
            {"from_currency": from_currency, "to_currency": to_currency}, {"_id": 0}
        )

        if rate_doc:
            rate = rate_doc["rate"]
        else:
            # Try inverse in database
            inverse_doc = await db.exchange_rates.find_one(
                {"from_currency": to_currency, "to_currency": from_currency}, {"_id": 0}
            )
            if inverse_doc:
                rate = 1 / inverse_doc["rate"]
            else:
                # Fall back to default hardcoded rates if present
                default_rates = {
                    "USD": {
                        "EUR": 0.92,
                        "MXN": 17.15,
                        "COP": 3950,
                        "PEN": 3.72,
                        "CLP": 880,
                        "ARS": 850,
                        "BRL": 4.95,
                    },
                    "EUR": {
                        "USD": 1.09,
                        "MXN": 18.65,
                        "COP": 4300,
                        "PEN": 4.05,
                        "CLP": 957,
                        "ARS": 925,
                        "BRL": 5.38,
                        "NIO": 38.0,
                    },
                }
                fr = default_rates.get(from_currency, {})
                if to_currency in fr:
                    rate = fr[to_currency]
                else:
                    raise HTTPException(status_code=404, detail="Exchange rate not found")

    converted = round(amount * rate, CURRENCIES[to_currency]["decimal_places"])

    return {
        "amount": amount,
        "converted": converted,
        "from": from_currency,
        "to": to_currency,
        "rate": rate,
        "symbol": CURRENCIES[to_currency]["symbol"],
    }


@api_router.get("/settings/currency")
async def get_system_currency(request: Request):
    """Get system default currency"""
    settings = await db.settings.find_one({"type": "system"}, {"_id": 0})
    return {
        "currency": settings.get("default_currency", "USD") if settings else "USD",
        "currencies": CURRENCIES,
    }


@api_router.put("/settings/currency")
async def set_system_currency(currency: str, request: Request):
    """Set system default currency (admin only)"""
    await require_roles(request, ["gerencia"])

    if currency not in CURRENCIES:
        raise HTTPException(status_code=400, detail="Invalid currency")

    await db.settings.update_one(
        {"type": "system"},
        {
            "$set": {
                "default_currency": currency,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
        },
        upsert=True,
    )

    return {"message": f"System currency set to {currency}"}


@api_router.get("/settings/discount-policy/seller")
async def get_seller_discount_policy(request: Request):
    await require_auth(request)
    return await _get_seller_global_discount_policy()


@api_router.put("/settings/discount-policy/seller")
async def update_seller_discount_policy(payload: SellerGlobalDiscountPolicyPayload, request: Request):
    await require_roles(request, ["gerencia", "recursos_humanos"])

    max_percent = float(payload.max_percent or 0.0)
    max_amount_nio = float(payload.max_amount_nio or 0.0)

    if max_percent <= 0 or max_percent > 100:
        raise HTTPException(status_code=400, detail="max_percent debe estar entre 0 y 100")
    if max_amount_nio <= 0:
        raise HTTPException(status_code=400, detail="max_amount_nio debe ser mayor a cero")

    doc = {
        "type": "seller_global_discount_policy",
        "enabled": bool(payload.enabled),
        "max_percent": round(max_percent, 4),
        "max_amount_nio": round(max_amount_nio, 2),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.settings.update_one(
        {"type": "seller_global_discount_policy"},
        {"$set": doc},
        upsert=True,
    )

    return doc


ALLOWED_THEME_MODES = {"light", "dark", "system"}
ALLOWED_THEME_SKINS = {
    "atlas",
    "forge",
    "aurora",
    "blossom",
    "prism",
    "wild-paws",
    "spectrum-01",
    "spectrum-02",
    "spectrum-03",
    "spectrum-04",
    "spectrum-05",
    "spectrum-06",
    "spectrum-07",
    "spectrum-08",
    "spectrum-09",
    "spectrum-10",
    "spectrum-11",
    "spectrum-12",
    "spectrum-13",
    "spectrum-14",
    "github",
}
DEFAULT_WATERMARK_OPACITY = 0.11
MIN_WATERMARK_OPACITY = 0.0
MAX_WATERMARK_OPACITY = 0.3


def _normalize_watermark_opacity(value: Any) -> float:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        numeric = DEFAULT_WATERMARK_OPACITY
    return max(MIN_WATERMARK_OPACITY, min(MAX_WATERMARK_OPACITY, numeric))


async def _get_system_settings_doc() -> Dict[str, Any]:
    return await db.settings.find_one({"type": "system"}, {"_id": 0}) or {"type": "system"}


def _serialize_appearance_settings(doc: Optional[Dict[str, Any]]) -> Dict[str, float]:
    data = doc or {}
    return {
        "watermark_opacity": _normalize_watermark_opacity(data.get("watermark_opacity", DEFAULT_WATERMARK_OPACITY)),
    }


@api_router.get("/settings/appearance/public")
async def get_public_appearance_settings():
    doc = await _get_system_settings_doc()
    return _serialize_appearance_settings(doc)


@api_router.get("/settings/appearance")
async def get_appearance_settings(request: Request):
    await require_auth(request)
    doc = await _get_system_settings_doc()
    return _serialize_appearance_settings(doc)


@api_router.put("/settings/appearance")
async def update_appearance_settings(payload: AppearanceSettings, request: Request):
    await require_roles(request, ["gerencia"])
    watermark_opacity = _normalize_watermark_opacity(payload.watermark_opacity)
    await db.settings.update_one(
        {"type": "system"},
        {"$set": {"type": "system", "watermark_opacity": watermark_opacity}},
        upsert=True,
    )
    return {"watermark_opacity": watermark_opacity}


@api_router.get("/settings/theme")
async def get_theme_settings(request: Request):
    user = await require_auth(request)
    mode = getattr(user, "theme_mode", None) or "system"
    skin = getattr(user, "theme_skin", None) or "atlas"
    custom = getattr(user, "theme_custom", None) or {}
    return {"mode": mode, "skin": skin, "custom": custom}


@api_router.put("/settings/theme")
async def update_theme_settings(payload: ThemeSettings, request: Request):
    user = await require_auth(request)
    mode = payload.mode or "system"
    skin = payload.skin or "atlas"
    current_custom = getattr(user, "theme_custom", None) or {}
    custom = payload.custom if payload.custom is not None else current_custom

    if mode not in ALLOWED_THEME_MODES:
        raise HTTPException(status_code=400, detail="Invalid theme mode")
    if skin not in ALLOWED_THEME_SKINS:
        raise HTTPException(status_code=400, detail="Invalid theme skin")

    if not isinstance(custom, dict):
        raise HTTPException(status_code=400, detail="Invalid theme custom")

    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": {"theme_mode": mode, "theme_skin": skin, "theme_custom": custom}},
    )
    return {"mode": mode, "skin": skin, "custom": custom}


@api_router.get("/settings/billing")
async def get_billing_settings(request: Request):
    await require_roles(request, ["gerencia", "recursos_humanos"])
    doc = await _get_billing_settings_doc()
    exchange_doc = doc.get("exchange") or {}
    effective_rate, source = _select_effective_billing_rate(exchange_doc, _utc_now())
    reasons = sorted(
        list(doc.get("cancel_reasons") or []),
        key=lambda r: (int(r.get("sort_order") or 9999), str(r.get("reason") or "")),
    )
    return {
        "exchange": {
            "official_rate": float(exchange_doc.get("official_rate") or 36.5),
            "rules": list(exchange_doc.get("rules") or []),
            "effective_rate": effective_rate,
            "effective_source": source,
        },
        "iva_rate": float(doc.get("iva_rate") or DEFAULT_BILLING_IVA_RATE),
        "cancel_reasons": reasons,
        "updated_at": doc.get("updated_at"),
    }


@api_router.get("/settings/billing/cancel-reasons/public")
async def get_public_cancel_reasons(request: Request):
    await require_auth(request)
    doc = await _get_billing_settings_doc()
    reasons = [r for r in list(doc.get("cancel_reasons") or []) if bool(r.get("active", True))]
    reasons.sort(key=lambda r: (int(r.get("sort_order") or 9999), str(r.get("reason") or "")))
    return {"reasons": reasons}


@api_router.get("/settings/billing/iva/public")
async def get_public_billing_iva(request: Request):
    await require_auth(request)
    iva_rate = await _get_billing_iva_rate()
    return {"iva_rate": iva_rate}


@api_router.put("/settings/billing/exchange")
async def update_billing_exchange(payload: BillingExchangeUpdatePayload, request: Request):
    await require_roles(request, ["gerencia", "recursos_humanos"])
    rate = float(payload.official_rate or 0)
    if rate <= 0:
        raise HTTPException(status_code=400, detail="La tasa oficial debe ser mayor a cero")

    doc = await _get_billing_settings_doc()
    exchange_doc = doc.setdefault("exchange", {})
    exchange_doc["official_rate"] = rate
    await _save_billing_settings_doc(doc)
    return {"message": "Tasa oficial actualizada", "official_rate": rate}


@api_router.put("/settings/billing/iva")
async def update_billing_iva(payload: BillingIvaUpdatePayload, request: Request):
    await require_roles(request, ["gerencia", "recursos_humanos"])
    iva_rate = float(payload.iva_rate or 0)
    if iva_rate <= 0:
        raise HTTPException(status_code=400, detail="El IVA debe ser mayor a cero")
    if iva_rate > 100:
        raise HTTPException(status_code=400, detail="El IVA no puede ser mayor a 100")

    doc = await _get_billing_settings_doc()
    doc["iva_rate"] = iva_rate
    await _save_billing_settings_doc(doc)
    return {"message": "IVA actualizado", "iva_rate": iva_rate}


@api_router.post("/settings/billing/exchange/rules")
async def create_billing_exchange_rule(payload: BillingExchangeRulePayload, request: Request):
    await require_roles(request, ["gerencia", "recursos_humanos"])
    cadence = str(payload.cadence or "").strip().lower()
    if cadence not in ALLOWED_BILLING_CADENCE:
        raise HTTPException(status_code=400, detail="Cadencia inválida")
    if float(payload.rate or 0) <= 0:
        raise HTTPException(status_code=400, detail="La tasa de regla debe ser mayor a cero")

    rule = {
        "id": _new_entity_id("rate_rule"),
        "name": _clean_vehicle_setting_text(payload.name or "Regla"),
        "cadence": cadence,
        "rate": float(payload.rate),
        "start_at": payload.start_at.isoformat() if payload.start_at else None,
        "end_at": payload.end_at.isoformat() if payload.end_at else None,
        "active": bool(payload.active),
        "created_at": _utc_now().isoformat(),
    }

    doc = await _get_billing_settings_doc()
    exchange_doc = doc.setdefault("exchange", {})
    rules = list(exchange_doc.get("rules") or [])
    rules.append(rule)
    exchange_doc["rules"] = rules
    await _save_billing_settings_doc(doc)
    return rule


@api_router.put("/settings/billing/exchange/rules/{rule_id}")
async def update_billing_exchange_rule(rule_id: str, payload: BillingExchangeRulePayload, request: Request):
    await require_roles(request, ["gerencia", "recursos_humanos"])
    cadence = str(payload.cadence or "").strip().lower()
    if cadence not in ALLOWED_BILLING_CADENCE:
        raise HTTPException(status_code=400, detail="Cadencia inválida")
    if float(payload.rate or 0) <= 0:
        raise HTTPException(status_code=400, detail="La tasa de regla debe ser mayor a cero")

    doc = await _get_billing_settings_doc()
    exchange_doc = doc.setdefault("exchange", {})
    rules = list(exchange_doc.get("rules") or [])
    target = next((r for r in rules if str(r.get("id")) == rule_id), None)
    if not target:
        raise HTTPException(status_code=404, detail="Regla no encontrada")

    target.update(
        {
            "name": _clean_vehicle_setting_text(payload.name or target.get("name") or "Regla"),
            "cadence": cadence,
            "rate": float(payload.rate),
            "start_at": payload.start_at.isoformat() if payload.start_at else None,
            "end_at": payload.end_at.isoformat() if payload.end_at else None,
            "active": bool(payload.active),
            "updated_at": _utc_now().isoformat(),
        }
    )
    await _save_billing_settings_doc(doc)
    return target


@api_router.delete("/settings/billing/exchange/rules/{rule_id}")
async def delete_billing_exchange_rule(rule_id: str, request: Request):
    await require_roles(request, ["gerencia", "recursos_humanos"])
    doc = await _get_billing_settings_doc()
    exchange_doc = doc.setdefault("exchange", {})
    rules = list(exchange_doc.get("rules") or [])
    before = len(rules)
    exchange_doc["rules"] = [r for r in rules if str(r.get("id")) != rule_id]
    if len(exchange_doc["rules"]) == before:
        raise HTTPException(status_code=404, detail="Regla no encontrada")
    await _save_billing_settings_doc(doc)
    return {"message": "Regla eliminada"}


@api_router.post("/settings/billing/cancel-reasons")
async def create_billing_cancel_reason(payload: BillingCancelReasonPayload, request: Request):
    await require_roles(request, ["gerencia", "recursos_humanos"])
    reason = _clean_vehicle_setting_text(payload.reason)
    if not reason:
        raise HTTPException(status_code=400, detail="Motivo requerido")

    doc = await _get_billing_settings_doc()
    reasons = list(doc.get("cancel_reasons") or [])
    if any(_canonical_vehicle_setting_text(r.get("reason")) == _canonical_vehicle_setting_text(reason) for r in reasons):
        raise HTTPException(status_code=409, detail="Motivo ya existe")

    reason_doc = {
        "id": _new_entity_id("reason"),
        "reason": reason,
        "active": bool(payload.active),
        "sort_order": int(payload.sort_order) if payload.sort_order is not None else (len(reasons) + 1),
    }
    reasons.append(reason_doc)
    reasons.sort(key=lambda r: (int(r.get("sort_order") or 9999), str(r.get("reason") or "")))
    doc["cancel_reasons"] = reasons
    await _save_billing_settings_doc(doc)
    return reason_doc


@api_router.put("/settings/billing/cancel-reasons/{reason_id}")
async def update_billing_cancel_reason(reason_id: str, payload: BillingCancelReasonPayload, request: Request):
    await require_roles(request, ["gerencia", "recursos_humanos"])
    reason = _clean_vehicle_setting_text(payload.reason)
    if not reason:
        raise HTTPException(status_code=400, detail="Motivo requerido")

    doc = await _get_billing_settings_doc()
    reasons = list(doc.get("cancel_reasons") or [])
    target = next((r for r in reasons if str(r.get("id")) == reason_id), None)
    if not target:
        raise HTTPException(status_code=404, detail="Motivo no encontrado")

    for other in reasons:
        if str(other.get("id")) != reason_id and _canonical_vehicle_setting_text(other.get("reason")) == _canonical_vehicle_setting_text(reason):
            raise HTTPException(status_code=409, detail="Motivo ya existe")

    target["reason"] = reason
    target["active"] = bool(payload.active)
    if payload.sort_order is not None:
        target["sort_order"] = int(payload.sort_order)
    reasons.sort(key=lambda r: (int(r.get("sort_order") or 9999), str(r.get("reason") or "")))
    doc["cancel_reasons"] = reasons
    await _save_billing_settings_doc(doc)
    return target


@api_router.delete("/settings/billing/cancel-reasons/{reason_id}")
async def delete_billing_cancel_reason(reason_id: str, request: Request):
    await require_roles(request, ["gerencia", "recursos_humanos"])
    doc = await _get_billing_settings_doc()
    reasons = list(doc.get("cancel_reasons") or [])
    before = len(reasons)
    reasons = [r for r in reasons if str(r.get("id")) != reason_id]
    if len(reasons) == before:
        raise HTTPException(status_code=404, detail="Motivo no encontrado")
    doc["cancel_reasons"] = reasons
    await _save_billing_settings_doc(doc)
    return {"message": "Motivo eliminado"}


VEHICLE_SETTINGS_DOC_TYPE = "vehicle_catalog_settings"
ALLOWED_FUEL_CODES = {"G", "D", "H", "E", "G/D", "D/G"}
BILLING_SETTINGS_DOC_TYPE = "billing_settings"
DEFAULT_BILLING_IVA_RATE = 15.0
ALLOWED_BILLING_CADENCE = {"daily", "weekly", "monthly", "custom"}


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _billing_default_cancel_reasons() -> List[Dict[str, Any]]:
    return [
        {"id": _new_entity_id("reason"), "reason": "Error de digitación en factura", "active": True, "sort_order": 1},
        {"id": _new_entity_id("reason"), "reason": "Precio o descuento aplicado incorrectamente", "active": True, "sort_order": 2},
        {"id": _new_entity_id("reason"), "reason": "Cliente desistió de la compra", "active": True, "sort_order": 3},
        {"id": _new_entity_id("reason"), "reason": "Pago rechazado o no confirmado", "active": True, "sort_order": 4},
        {"id": _new_entity_id("reason"), "reason": "Producto sin disponibilidad real", "active": True, "sort_order": 5},
        {"id": _new_entity_id("reason"), "reason": "Factura duplicada", "active": True, "sort_order": 6},
        {"id": _new_entity_id("reason"), "reason": "Datos fiscales del cliente incorrectos", "active": True, "sort_order": 7},
        {"id": _new_entity_id("reason"), "reason": "Otro Justifique", "active": True, "sort_order": 999},
    ]


async def _get_billing_settings_doc() -> Dict[str, Any]:
    doc = await db.settings.find_one({"type": BILLING_SETTINGS_DOC_TYPE}, {"_id": 0})
    if not doc:
        doc = {
            "type": BILLING_SETTINGS_DOC_TYPE,
            "exchange": {
                "official_rate": 36.5,
                "rules": [],
            },
            "iva_rate": DEFAULT_BILLING_IVA_RATE,
            "cancel_reasons": _billing_default_cancel_reasons(),
            "updated_at": _utc_now().isoformat(),
        }
    exchange = doc.setdefault("exchange", {})
    exchange.setdefault("official_rate", 36.5)
    exchange.setdefault("rules", [])
    doc.setdefault("iva_rate", DEFAULT_BILLING_IVA_RATE)
    doc.setdefault("cancel_reasons", _billing_default_cancel_reasons())
    return doc


async def _get_billing_iva_rate() -> float:
    try:
        doc = await _get_billing_settings_doc()
        iva_rate = float(doc.get("iva_rate") or DEFAULT_BILLING_IVA_RATE)
        if iva_rate > 0:
            return iva_rate
    except Exception:
        pass
    return DEFAULT_BILLING_IVA_RATE


async def _save_billing_settings_doc(doc: Dict[str, Any]) -> None:
    doc["type"] = BILLING_SETTINGS_DOC_TYPE
    doc["updated_at"] = _utc_now().isoformat()
    await db.settings.update_one(
        {"type": BILLING_SETTINGS_DOC_TYPE},
        {"$set": doc},
        upsert=True,
    )


def _parse_iso_datetime(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
    except Exception:
        return None


def _rule_matches_now(rule: Dict[str, Any], now: datetime) -> bool:
    if not bool(rule.get("active", True)):
        return False
    cadence = str(rule.get("cadence") or "").strip().lower()
    if cadence not in ALLOWED_BILLING_CADENCE:
        return False

    start_at = _parse_iso_datetime(rule.get("start_at"))
    end_at = _parse_iso_datetime(rule.get("end_at"))

    if cadence == "daily":
        return True

    if cadence == "weekly":
        if not start_at:
            return False
        if end_at and now > end_at:
            return False
        return now >= start_at

    if cadence == "monthly":
        if not start_at:
            return False
        if now.year < start_at.year:
            return False
        if now.year == start_at.year and now.month < start_at.month:
            return False
        if end_at:
            if now.year > end_at.year:
                return False
            if now.year == end_at.year and now.month > end_at.month:
                return False
        return True

    # custom
    if start_at and now < start_at:
        return False
    if end_at and now > end_at:
        return False
    return bool(start_at or end_at)


def _select_effective_billing_rate(exchange_doc: Dict[str, Any], now: datetime) -> Tuple[float, str]:
    official_rate = float(exchange_doc.get("official_rate") or 36.5)
    rules = list(exchange_doc.get("rules") or [])
    if not rules:
        return official_rate, "billing_official"

    priority = {"custom": 4, "monthly": 3, "weekly": 2, "daily": 1}
    matched: List[Dict[str, Any]] = []
    for rule in rules:
        try:
            if _rule_matches_now(rule, now):
                matched.append(rule)
        except Exception:
            continue

    if not matched:
        return official_rate, "billing_official"

    matched.sort(
        key=lambda r: (
            priority.get(str(r.get("cadence") or "").lower(), 0),
            _parse_iso_datetime(r.get("start_at")) or datetime.min.replace(tzinfo=timezone.utc),
        ),
        reverse=True,
    )
    top = matched[0]
    rate = float(top.get("rate") or official_rate)
    return rate, f"billing_rule:{str(top.get('id') or 'unknown')}"


def _clean_vehicle_setting_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _canonical_vehicle_setting_text(value: str) -> str:
    clean = _clean_vehicle_setting_text(value)
    normalized = unicodedata.normalize("NFD", clean)
    without_accents = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    return without_accents.casefold()


def _validate_brand_name_format(value: str) -> str:
    brand = _clean_vehicle_setting_text(value).upper()
    if not re.fullmatch(r"[A-Z0-9][A-Z0-9 .&/+-]*", brand):
        raise HTTPException(
            status_code=400,
            detail="Formato de marca inválido. Usa MAYÚSCULAS y caracteres alfanuméricos (ej: TOYOTA, LAND ROVER)",
        )
    return brand


def _validate_year_value_format(value: str) -> str:
    year_value = _clean_vehicle_setting_text(value)
    if not re.fullmatch(r"~?\d{4}(?:-(?:\d{4}|Presente|\d{4}\+))?|\d{4}\+", year_value):
        raise HTTPException(
            status_code=400,
            detail="Formato de año inválido. Usa: 2024, 2004-2015, 2004-Presente, 2000-2010+, ~2000-2004",
        )
    return year_value


def _validate_model_name_format(value: str) -> str:
    model_name = _clean_vehicle_setting_text(value)
    if not re.fullmatch(r".+\([^)]+\)", model_name):
        raise HTTPException(
            status_code=400,
            detail="Formato de modelo inválido. Usa: Modelo (Generación), ej: Hilux (AN10/20)",
        )
    return model_name


def _validate_variation_format(value: str) -> str:
    variation = _clean_vehicle_setting_text(value)
    fuel_match = re.search(r"\[([A-Z/]+)\]$", variation)
    if not fuel_match:
        raise HTTPException(
            status_code=400,
            detail="Formato de variación inválido. Debe terminar con combustible [G], [D], [H], [E], etc.",
        )
    if fuel_match.group(1) not in ALLOWED_FUEL_CODES:
        raise HTTPException(
            status_code=400,
            detail="Código de combustible inválido. Usa [G], [D], [H], [E], [G/D] o [D/G]",
        )
    if not re.search(r"\b\d(?:\.\d)?L\b", variation):
        raise HTTPException(
            status_code=400,
            detail="La variación debe incluir cilindrada (ej: 3.0L 1KD-FTV [D])",
        )
    return variation


def _new_entity_id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:10]}"


async def _get_vehicle_settings_doc() -> Dict[str, Any]:
    doc = await db.settings.find_one({"type": VEHICLE_SETTINGS_DOC_TYPE}, {"_id": 0})
    if not doc:
        doc = {
            "type": VEHICLE_SETTINGS_DOC_TYPE,
            "brands": [],
            "colors": [],
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
    doc.setdefault("brands", [])
    doc.setdefault("colors", [])
    return doc


async def _save_vehicle_settings_doc(doc: Dict[str, Any]) -> None:
    doc["type"] = VEHICLE_SETTINGS_DOC_TYPE
    doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.settings.update_one(
        {"type": VEHICLE_SETTINGS_DOC_TYPE},
        {"$set": doc},
        upsert=True,
    )


def _find_brand(doc: Dict[str, Any], brand_id: str) -> Optional[Dict[str, Any]]:
    return next((b for b in doc.get("brands", []) if b.get("id") == brand_id), None)


def _find_year(brand: Dict[str, Any], year_id: str) -> Optional[Dict[str, Any]]:
    return next((y for y in brand.get("years", []) if y.get("id") == year_id), None)


def _find_model(year_item: Dict[str, Any], model_id: str) -> Optional[Dict[str, Any]]:
    return next((m for m in year_item.get("models", []) if m.get("id") == model_id), None)


@api_router.get("/settings/vehicles")
async def get_vehicle_settings(request: Request):
    await require_auth(request)
    doc = await _get_vehicle_settings_doc()
    return {
        "brands": doc.get("brands", []),
        "colors": doc.get("colors", []),
        "updated_at": doc.get("updated_at"),
    }


@api_router.post("/settings/vehicles/brands")
async def create_vehicle_brand(payload: VehicleSettingsNamePayload, request: Request):
    await require_roles(request, ["gerencia"])
    name = _validate_brand_name_format(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Brand name is required")

    doc = await _get_vehicle_settings_doc()
    exists = any(_canonical_vehicle_setting_text(b.get("name", "")) == _canonical_vehicle_setting_text(name) for b in doc["brands"])
    if exists:
        raise HTTPException(status_code=409, detail="Brand already exists")

    new_brand = {"id": _new_entity_id("brand"), "name": name, "years": []}
    doc["brands"].append(new_brand)
    doc["brands"].sort(key=lambda b: _clean_vehicle_setting_text(b.get("name", "")).lower())
    await _save_vehicle_settings_doc(doc)
    return new_brand


@api_router.put("/settings/vehicles/brands/{brand_id}")
async def update_vehicle_brand(brand_id: str, payload: VehicleSettingsNamePayload, request: Request):
    await require_roles(request, ["gerencia"])
    name = _validate_brand_name_format(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Brand name is required")

    doc = await _get_vehicle_settings_doc()
    brand = _find_brand(doc, brand_id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")

    for other in doc["brands"]:
        if other.get("id") != brand_id and _canonical_vehicle_setting_text(other.get("name", "")) == _canonical_vehicle_setting_text(name):
            raise HTTPException(status_code=409, detail="Brand already exists")

    brand["name"] = name
    doc["brands"].sort(key=lambda b: _clean_vehicle_setting_text(b.get("name", "")).lower())
    await _save_vehicle_settings_doc(doc)
    return brand


@api_router.delete("/settings/vehicles/brands/{brand_id}")
async def delete_vehicle_brand(brand_id: str, request: Request):
    await require_roles(request, ["gerencia"])
    doc = await _get_vehicle_settings_doc()
    before = len(doc["brands"])
    doc["brands"] = [b for b in doc["brands"] if b.get("id") != brand_id]
    if len(doc["brands"]) == before:
        raise HTTPException(status_code=404, detail="Brand not found")
    await _save_vehicle_settings_doc(doc)
    return {"message": "Brand deleted"}


@api_router.post("/settings/vehicles/brands/{brand_id}/years")
async def create_vehicle_year(brand_id: str, payload: VehicleSettingsYearPayload, request: Request):
    await require_roles(request, ["gerencia"])
    year_value = _validate_year_value_format(payload.year)
    if not year_value:
        raise HTTPException(status_code=400, detail="Year is required")

    doc = await _get_vehicle_settings_doc()
    brand = _find_brand(doc, brand_id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    brand.setdefault("years", [])

    exists = any(_canonical_vehicle_setting_text(y.get("value", "")) == _canonical_vehicle_setting_text(year_value) for y in brand["years"])
    if exists:
        raise HTTPException(status_code=409, detail="Year already exists for this brand")

    new_year = {"id": _new_entity_id("year"), "value": year_value, "models": []}
    brand["years"].append(new_year)
    brand["years"].sort(key=lambda y: _clean_vehicle_setting_text(y.get("value", "")).lower(), reverse=True)
    await _save_vehicle_settings_doc(doc)
    return new_year


@api_router.put("/settings/vehicles/brands/{brand_id}/years/{year_id}")
async def update_vehicle_year(brand_id: str, year_id: str, payload: VehicleSettingsYearPayload, request: Request):
    await require_roles(request, ["gerencia"])
    year_value = _validate_year_value_format(payload.year)
    if not year_value:
        raise HTTPException(status_code=400, detail="Year is required")

    doc = await _get_vehicle_settings_doc()
    brand = _find_brand(doc, brand_id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    year_item = _find_year(brand, year_id)
    if not year_item:
        raise HTTPException(status_code=404, detail="Year not found")

    for other in brand.get("years", []):
        if other.get("id") != year_id and _canonical_vehicle_setting_text(other.get("value", "")) == _canonical_vehicle_setting_text(year_value):
            raise HTTPException(status_code=409, detail="Year already exists for this brand")

    year_item["value"] = year_value
    brand["years"].sort(key=lambda y: _clean_vehicle_setting_text(y.get("value", "")).lower(), reverse=True)
    await _save_vehicle_settings_doc(doc)
    return year_item


@api_router.delete("/settings/vehicles/brands/{brand_id}/years/{year_id}")
async def delete_vehicle_year(brand_id: str, year_id: str, request: Request):
    await require_roles(request, ["gerencia"])
    doc = await _get_vehicle_settings_doc()
    brand = _find_brand(doc, brand_id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    years = brand.get("years", [])
    before = len(years)
    brand["years"] = [y for y in years if y.get("id") != year_id]
    if len(brand["years"]) == before:
        raise HTTPException(status_code=404, detail="Year not found")
    await _save_vehicle_settings_doc(doc)
    return {"message": "Year deleted"}


@api_router.post("/settings/vehicles/brands/{brand_id}/years/{year_id}/models")
async def create_vehicle_model(brand_id: str, year_id: str, payload: VehicleSettingsNamePayload, request: Request):
    await require_roles(request, ["gerencia"])
    model_name = _validate_model_name_format(payload.name)
    if not model_name:
        raise HTTPException(status_code=400, detail="Model name is required")

    doc = await _get_vehicle_settings_doc()
    brand = _find_brand(doc, brand_id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    year_item = _find_year(brand, year_id)
    if not year_item:
        raise HTTPException(status_code=404, detail="Year not found")
    year_item.setdefault("models", [])

    exists = any(_canonical_vehicle_setting_text(m.get("name", "")) == _canonical_vehicle_setting_text(model_name) for m in year_item["models"])
    if exists:
        raise HTTPException(status_code=409, detail="Model already exists for this year")

    new_model = {"id": _new_entity_id("model"), "name": model_name, "variations": []}
    year_item["models"].append(new_model)
    year_item["models"].sort(key=lambda m: _clean_vehicle_setting_text(m.get("name", "")).lower())
    await _save_vehicle_settings_doc(doc)
    return new_model


@api_router.put("/settings/vehicles/brands/{brand_id}/years/{year_id}/models/{model_id}")
async def update_vehicle_model(
    brand_id: str,
    year_id: str,
    model_id: str,
    payload: VehicleSettingsNamePayload,
    request: Request,
):
    await require_roles(request, ["gerencia"])
    model_name = _validate_model_name_format(payload.name)
    if not model_name:
        raise HTTPException(status_code=400, detail="Model name is required")

    doc = await _get_vehicle_settings_doc()
    brand = _find_brand(doc, brand_id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    year_item = _find_year(brand, year_id)
    if not year_item:
        raise HTTPException(status_code=404, detail="Year not found")
    model_item = _find_model(year_item, model_id)
    if not model_item:
        raise HTTPException(status_code=404, detail="Model not found")

    for other in year_item.get("models", []):
        if other.get("id") != model_id and _canonical_vehicle_setting_text(other.get("name", "")) == _canonical_vehicle_setting_text(model_name):
            raise HTTPException(status_code=409, detail="Model already exists for this year")

    model_item["name"] = model_name
    year_item["models"].sort(key=lambda m: _clean_vehicle_setting_text(m.get("name", "")).lower())
    await _save_vehicle_settings_doc(doc)
    return model_item


@api_router.delete("/settings/vehicles/brands/{brand_id}/years/{year_id}/models/{model_id}")
async def delete_vehicle_model(brand_id: str, year_id: str, model_id: str, request: Request):
    await require_roles(request, ["gerencia"])
    doc = await _get_vehicle_settings_doc()
    brand = _find_brand(doc, brand_id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    year_item = _find_year(brand, year_id)
    if not year_item:
        raise HTTPException(status_code=404, detail="Year not found")

    models = year_item.get("models", [])
    before = len(models)
    year_item["models"] = [m for m in models if m.get("id") != model_id]
    if len(year_item["models"]) == before:
        raise HTTPException(status_code=404, detail="Model not found")
    await _save_vehicle_settings_doc(doc)
    return {"message": "Model deleted"}


@api_router.post("/settings/vehicles/brands/{brand_id}/years/{year_id}/models/{model_id}/variations")
async def create_vehicle_variation(
    brand_id: str,
    year_id: str,
    model_id: str,
    payload: VehicleSettingsVariationPayload,
    request: Request,
):
    await require_roles(request, ["gerencia"])
    variation_value = _validate_variation_format(payload.value)
    if not variation_value:
        raise HTTPException(status_code=400, detail="Variation is required")

    doc = await _get_vehicle_settings_doc()
    brand = _find_brand(doc, brand_id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    year_item = _find_year(brand, year_id)
    if not year_item:
        raise HTTPException(status_code=404, detail="Year not found")
    model_item = _find_model(year_item, model_id)
    if not model_item:
        raise HTTPException(status_code=404, detail="Model not found")
    model_item.setdefault("variations", [])

    exists = any(_canonical_vehicle_setting_text(v.get("value", "")) == _canonical_vehicle_setting_text(variation_value) for v in model_item["variations"])
    if exists:
        raise HTTPException(status_code=409, detail="Variation already exists")

    new_variation = {"id": _new_entity_id("var"), "value": variation_value}
    model_item["variations"].append(new_variation)
    model_item["variations"].sort(key=lambda v: _clean_vehicle_setting_text(v.get("value", "")).lower())
    await _save_vehicle_settings_doc(doc)
    return new_variation


@api_router.put("/settings/vehicles/brands/{brand_id}/years/{year_id}/models/{model_id}/variations/{variation_id}")
async def update_vehicle_variation(
    brand_id: str,
    year_id: str,
    model_id: str,
    variation_id: str,
    payload: VehicleSettingsVariationPayload,
    request: Request,
):
    await require_roles(request, ["gerencia"])
    variation_value = _validate_variation_format(payload.value)
    if not variation_value:
        raise HTTPException(status_code=400, detail="Variation is required")

    doc = await _get_vehicle_settings_doc()
    brand = _find_brand(doc, brand_id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    year_item = _find_year(brand, year_id)
    if not year_item:
        raise HTTPException(status_code=404, detail="Year not found")
    model_item = _find_model(year_item, model_id)
    if not model_item:
        raise HTTPException(status_code=404, detail="Model not found")
    variations = model_item.get("variations", [])
    variation = next((v for v in variations if v.get("id") == variation_id), None)
    if not variation:
        raise HTTPException(status_code=404, detail="Variation not found")

    for other in variations:
        if other.get("id") != variation_id and _canonical_vehicle_setting_text(other.get("value", "")) == _canonical_vehicle_setting_text(variation_value):
            raise HTTPException(status_code=409, detail="Variation already exists")

    variation["value"] = variation_value
    variations.sort(key=lambda v: _clean_vehicle_setting_text(v.get("value", "")).lower())
    await _save_vehicle_settings_doc(doc)
    return variation


@api_router.delete("/settings/vehicles/brands/{brand_id}/years/{year_id}/models/{model_id}/variations/{variation_id}")
async def delete_vehicle_variation(
    brand_id: str,
    year_id: str,
    model_id: str,
    variation_id: str,
    request: Request,
):
    await require_roles(request, ["gerencia"])
    doc = await _get_vehicle_settings_doc()
    brand = _find_brand(doc, brand_id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    year_item = _find_year(brand, year_id)
    if not year_item:
        raise HTTPException(status_code=404, detail="Year not found")
    model_item = _find_model(year_item, model_id)
    if not model_item:
        raise HTTPException(status_code=404, detail="Model not found")

    variations = model_item.get("variations", [])
    before = len(variations)
    model_item["variations"] = [v for v in variations if v.get("id") != variation_id]
    if len(model_item["variations"]) == before:
        raise HTTPException(status_code=404, detail="Variation not found")
    await _save_vehicle_settings_doc(doc)
    return {"message": "Variation deleted"}


@api_router.post("/settings/vehicles/colors")
async def create_vehicle_color(payload: VehicleSettingsColorPayload, request: Request):
    await require_roles(request, ["gerencia"])
    value = _clean_vehicle_setting_text(payload.value)
    if not value:
        raise HTTPException(status_code=400, detail="Color is required")

    doc = await _get_vehicle_settings_doc()
    exists = any(_canonical_vehicle_setting_text(c.get("value", "")) == _canonical_vehicle_setting_text(value) for c in doc["colors"])
    if exists:
        raise HTTPException(status_code=409, detail="Color already exists")

    color = {"id": _new_entity_id("color"), "value": value}
    doc["colors"].append(color)
    doc["colors"].sort(key=lambda c: _clean_vehicle_setting_text(c.get("value", "")).lower())
    await _save_vehicle_settings_doc(doc)
    return color


@api_router.put("/settings/vehicles/colors/{color_id}")
async def update_vehicle_color(color_id: str, payload: VehicleSettingsColorPayload, request: Request):
    await require_roles(request, ["gerencia"])
    value = _clean_vehicle_setting_text(payload.value)
    if not value:
        raise HTTPException(status_code=400, detail="Color is required")

    doc = await _get_vehicle_settings_doc()
    color = next((c for c in doc["colors"] if c.get("id") == color_id), None)
    if not color:
        raise HTTPException(status_code=404, detail="Color not found")
    for other in doc["colors"]:
        if other.get("id") != color_id and _canonical_vehicle_setting_text(other.get("value", "")) == _canonical_vehicle_setting_text(value):
            raise HTTPException(status_code=409, detail="Color already exists")

    color["value"] = value
    doc["colors"].sort(key=lambda c: _clean_vehicle_setting_text(c.get("value", "")).lower())
    await _save_vehicle_settings_doc(doc)
    return color


@api_router.delete("/settings/vehicles/colors/{color_id}")
async def delete_vehicle_color(color_id: str, request: Request):
    await require_roles(request, ["gerencia"])
    doc = await _get_vehicle_settings_doc()
    before = len(doc["colors"])
    doc["colors"] = [c for c in doc["colors"] if c.get("id") != color_id]
    if len(doc["colors"]) == before:
        raise HTTPException(status_code=404, detail="Color not found")
    await _save_vehicle_settings_doc(doc)
    return {"message": "Color deleted"}


# ============ WORK ORDER NOTIFICATIONS FOR TECHNICIANS ============


@api_router.post("/work-orders/{work_order_id}/notify")
async def notify_work_order_update(
    work_order_id: str, request: Request, background_tasks: BackgroundTasks
):
    """Send notification about work order update to assigned technician"""
    await require_auth(request)

    wo = await db.work_orders.find_one({"work_order_id": work_order_id}, {"_id": 0})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")

    if not wo.get("technician_id"):
        return {"message": "No technician assigned"}

    # Create notification
    notification = {
        "notification_id": f"notif_{uuid.uuid4().hex[:8]}",
        "target_user_id": wo["technician_id"],
        "target_role": None,
        "title": "Orden de Trabajo Actualizada",
        "body": f"OT {work_order_id[:12]} - {wo.get('status', 'actualizada')}",
        "url": f"/work-orders?id={work_order_id}",
        "type": "work_order_update",
        "related_id": work_order_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "read": False,
    }

    await db.notifications.insert_one(notification)

    return {"message": "Notification sent to technician"}


@api_router.get("/work-orders/technician/pending")
async def get_technician_pending_orders(request: Request):
    """Get pending work orders for the authenticated technician"""
    user = await require_auth(request)

    query: dict[str, Any] = {"status": {"$in": ["pending", "in_progress"]}}
    visibility_query = await build_work_order_visibility_query(user)
    query = merge_queries(query, visibility_query)

    if user.role == "instalaciones":
        query = merge_queries(
            query,
            {"$or": [{"technician_id": user.user_id}, {"technician_id": None}]},
        )

    orders = (
        await db.work_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    )

    # Add unread notification count
    unread = await db.notifications.count_documents(
        {"target_user_id": user.user_id, "read": False}
    )

    return {"orders": orders, "unread_notifications": unread}


# ============ MANAGER AUTHORIZATION ============


@api_router.post("/auth/manager/generate-code")
async def generate_manager_auth_code(
    request: Request, reason: str = "Instalación de producto solo para llevar"
):
    """Generate authorization code for special operations (manager only)"""
    user = await require_roles(request, ["gerencia"])

    code = f"AUTH-{uuid.uuid4().hex[:8].upper()}"
    expires_at = datetime.now(timezone.utc) + timedelta(hours=1)

    doc = {
        "code": code,
        "generated_by": user.user_id,
        "generated_by_name": user.name,
        "reason": reason,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "used": False,
        "used_at": None,
    }

    await db.manager_authorizations.insert_one(doc)

    return {
        "code": code,
        "expires_at": expires_at.isoformat(),
        "valid_for_minutes": 60,
        "reason": reason,
    }


@api_router.get("/auth/manager/pending")
async def get_pending_authorizations(request: Request):
    """Get pending authorization requests"""
    await require_roles(request, ["gerencia", "supervisor"])

    auths = (
        await db.manager_authorizations.find(
            {
                "used": False,
                "expires_at": {"$gt": datetime.now(timezone.utc).isoformat()},
            },
            {"_id": 0},
        )
        .sort("created_at", -1)
        .to_list(50)
    )

    return auths


# ============ PRODUCT IMPORT ============


class ProductImportItem(BaseModel):
    sku: str
    name: str
    description: Optional[str] = None
    category: str
    subcategory: Optional[str] = None
    brand: str
    price: float
    cost: float = 0
    installation_type: str = "optional"  # required, optional, not_available
    installation_price: float = 0
    installation_time_minutes: int = 60
    warranty_months: int = 12
    image_url: Optional[str] = None


@api_router.post("/products/import")
async def import_products(
    products: List[ProductImportItem],
    request: Request,
    warehouse_id: str = "wh_main",
    initial_stock: int = 10,
):
    """Import multiple products at once"""
    user = await require_roles(request, ["gerencia", "supervisor"])

    imported = 0
    errors = []

    for p in products:
        try:
            # Check if SKU already exists
            existing = await db.products.find_one({"sku": p.sku})
            if existing:
                errors.append({"sku": p.sku, "error": "SKU already exists"})
                continue

            product_doc = {
                "product_id": f"prod_{uuid.uuid4().hex[:8]}",
                "sku": p.sku,
                "name": p.name,
                "description": p.description or "",
                "category": p.category,
                "subcategory": p.subcategory or "",
                "brand": p.brand,
                "price": p.price,
                "cost": p.cost,
                "product_type": "product",
                "images": [p.image_url] if p.image_url else [],
                "compatibility": None,
                "installation_required": p.installation_type == "required",
                "installation_type": p.installation_type,
                "installation_price": p.installation_price,
                "installation_time_minutes": p.installation_time_minutes,
                "requires_manager_auth_for_install": p.installation_type
                == "not_available",
                "warranty_months": p.warranty_months,
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }

            await db.products.insert_one(product_doc)

            # Create initial inventory
            inventory_doc = {
                "inventory_id": f"inv_{uuid.uuid4().hex[:8]}",
                "product_id": product_doc["product_id"],
                "warehouse_id": warehouse_id,
                "quantity": initial_stock,
                "min_stock": 5,
                "last_updated": datetime.now(timezone.utc).isoformat(),
            }
            await db.inventory.insert_one(inventory_doc)
            await audit_service.log_inventory_movement(
                product_id=product_doc["product_id"],
                warehouse_id=warehouse_id,
                quantity_change=int(initial_stock),
                reason="initial_stock_import",
                actor=user,
                branch_id=user.branch_id,
                reference_id=inventory_doc["inventory_id"],
                metadata={"sku": p.sku},
            )

            imported += 1

        except Exception as e:
            errors.append({"sku": p.sku, "error": str(e)})

    return {"imported": imported, "errors": errors, "total": len(products)}


@api_router.post("/products/import/csv")
async def import_products_csv(
    request: Request,
    file: UploadFile,
    warehouse_id: str = "wh_main",
    initial_stock: int = 10,
):
    """Import products from CSV file"""
    await require_roles(request, ["gerencia", "supervisor"])

    import csv
    from io import StringIO

    content = await file.read()
    decoded = content.decode("utf-8")

    reader = csv.DictReader(StringIO(decoded))

    products = []
    for row in reader:
        products.append(
            ProductImportItem(
                sku=row.get("sku", row.get("SKU", "")),
                name=row.get("name", row.get("nombre", "")),
                description=row.get("description", row.get("descripcion", "")),
                category=row.get(
                    "category", row.get("categoria", "accesorios_electronicos")
                ),
                subcategory=row.get("subcategory", row.get("subcategoria", "")),
                brand=row.get("brand", row.get("marca", "")),
                price=float(row.get("price", row.get("precio", 0))),
                cost=float(row.get("cost", row.get("costo", 0))),
                installation_type=row.get(
                    "installation_type", row.get("tipo_instalacion", "optional")
                ),
                installation_price=float(
                    row.get("installation_price", row.get("precio_instalacion", 0))
                ),
                installation_time_minutes=int(
                    row.get("installation_time", row.get("tiempo_instalacion", 60))
                ),
                warranty_months=int(
                    row.get("warranty_months", row.get("garantia_meses", 12))
                ),
                image_url=row.get("image_url", row.get("imagen", None)),
            )
        )

    # Use the bulk import
    result = await import_products(products, request, warehouse_id, initial_stock)
    return result


@api_router.get("/products/import/template")
async def get_import_template():
    """Get CSV template for product import"""
    template_path = ROOT_DIR / "data" / "product_template.csv"
    try:
        template = template_path.read_text(encoding="utf-8")
    except Exception:
        template = (
            "sku,name,description,category,subcategory,brand,price,cost,installation_type,"
            "installation_price,installation_time,warranty_months,image_url\n"
            "AMP-001,Amplificador 4 Canales,Amplificador de alta potencia,accesorios_electronicos,"
            "Amplificadores,DS18,299.99,150.00,required,50.00,90,24,https://example.com/amp.jpg"
        )

    return Response(
        content=template,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=plantilla_productos.csv"},
    )


# ============ DEMO PRODUCTS (Car Audio) ============


@api_router.post("/products/seed-demo")
async def seed_demo_products(request: Request, warehouse_id: str = "wh_main"):
    """Seed demo products similar to car audio store"""
    user = await require_roles(request, ["gerencia"])

    # Load demo products from external file to avoid extremely long inline data structures
    demo_products = []
    demo_file = ROOT_DIR / "data" / "demo_products.py"
    if demo_file.exists():
        ns: dict = {}
        try:
            exec(demo_file.read_text(encoding="utf-8"), ns)
            demo_products = ns.get("demo_products", [])
        except Exception:
            demo_products = []
    else:
        demo_products = []

    imported = 0
    for p in demo_products:
        # Check if SKU exists
        existing = await db.products.find_one({"sku": p["sku"]})
        if existing:
            continue

        # compute price tiers for seeded demo products
        try:
            base_price_seed = float(p.get("price", 0) or 0)
        except Exception:
            base_price_seed = 0.0
        precio1 = round(base_price_seed, 2)
        precio2 = round(base_price_seed * 1.05, 2)
        precio3 = round(base_price_seed * 1.1, 2)
        precio_vip = round(base_price_seed * 0.9, 2)

        product_doc = {
            "product_id": f"prod_{uuid.uuid4().hex[:8]}",
            "sku": p["sku"],
            "name": p["name"],
            "description": p["description"],
            "category": p["category"],
            "subcategory": p["subcategory"],
            "brand": p["brand"],
            # store precio tiers and keep `price` pointing to precio1 for UI
            "precio1": precio1,
            "precio2": precio2,
            "precio3": precio3,
            "precio_vip": precio_vip,
            "price": precio1,
            "cost": p["cost"],
            "product_type": "service" if p["category"] == "servicios" else "product",
            "images": [p["image_url"]] if p["image_url"] else [],
            "compatibility": None,
            "installation_required": p["installation_type"] == "required",
            "installation_type": p["installation_type"],
            "installation_price": p["installation_price"],
            "installation_time_minutes": p["installation_time_minutes"],
            "requires_manager_auth_for_install": p["installation_type"]
            == "not_available",
            "warranty_months": p["warranty_months"],
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }

        await db.products.insert_one(product_doc)

        # Create inventory for physical products
        if p["category"] != "servicios":
            inventory_doc = {
                "inventory_id": f"inv_{uuid.uuid4().hex[:8]}",
                "product_id": product_doc["product_id"],
                "warehouse_id": warehouse_id,
                "quantity": 20,
                "min_stock": 5,
                "last_updated": datetime.now(timezone.utc).isoformat(),
            }
            await db.inventory.insert_one(inventory_doc)
            await audit_service.log_inventory_movement(
                product_id=product_doc["product_id"],
                warehouse_id=warehouse_id,
                quantity_change=20,
                reason="initial_stock_seed",
                actor=user,
                branch_id=user.branch_id,
                reference_id=inventory_doc["inventory_id"],
                metadata={"sku": p.get("sku")},
            )

        imported += 1

    return {
        "message": "Demo products created",
        "imported": imported,
        "total_available": len(demo_products),
        "categories": ["electronicos", "no_electricos", "polarizados", "servicios"],
    }


# ============ DISPATCH ORDERS (DESPACHOS) ============


class DispatchItem(BaseModel):
    product_id: str
    product_name: str
    product_sku: Optional[str] = None
    product_image: Optional[str] = None
    quantity: int
    delivered: bool = False
    delivered_by: Optional[str] = None
    delivered_at: Optional[str] = None


class DispatchOrder(BaseModel):
    dispatch_id: str
    sale_id: str
    invoice_number: str
    customer_name: str
    customer_phone: Optional[str] = None
    requested_by: Optional[str] = None
    requested_by_name: Optional[str] = None
    warehouse_id: str
    warehouse_name: Optional[str] = None
    items: List[DispatchItem]
    status: str = "pending"  # pending, in_progress, completed
    priority: str = "normal"  # low, normal, high, urgent
    notes: Optional[str] = None
    dispatch_type: Optional[str] = None  # sale, sample_out, sample_return
    reference_number: Optional[str] = None
    sample_id: Optional[str] = None
    created_at: str
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    total_time_minutes: Optional[int] = None
    dispatchers: List[str] = []


class DispatchItemDelivery(BaseModel):
    product_id: str
    dispatcher_id: str
    notes: Optional[str] = None


class SampleRequestCreate(BaseModel):
    customer_id: str
    product_id: str
    warehouse_id: str
    quantity: int = 1
    notes: Optional[str] = None


class SampleRequest(BaseModel):
    sample_id: str = Field(default_factory=lambda: f"smp_{uuid.uuid4().hex[:8]}")
    customer_id: str
    customer_name: str
    customer_phone: Optional[str] = None
    product_id: str
    product_name: str
    product_sku: Optional[str] = None
    product_image: Optional[str] = None
    warehouse_id: str
    warehouse_name: Optional[str] = None
    quantity: int = 1
    status: str = "requested"  # requested, delivered, return_requested, returned, consumed, cancelled
    requested_by: str
    requested_by_name: str
    dispatch_out_id: Optional[str] = None
    dispatch_return_id: Optional[str] = None
    sale_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    delivered_at: Optional[str] = None
    returned_at: Optional[str] = None
    notes: Optional[str] = None


@api_router.post("/samples/request")
async def request_sample(sample: SampleRequestCreate, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor", "ventas"])

    customer = await db.customers.find_one({"customer_id": sample.customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    product = await db.products.find_one({"product_id": sample.product_id}, {"_id": 0})
    warehouse = await db.warehouses.find_one({"warehouse_id": sample.warehouse_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    if product.get("product_type") == "service":
        raise HTTPException(status_code=400, detail="Samples are only available for physical products")

    quantity = max(1, int(sample.quantity or 1))

    inv = await db.inventory.find_one(
        {"product_id": sample.product_id, "warehouse_id": sample.warehouse_id}
    )
    if not inv or inv.get("quantity", 0) < quantity:
        raise HTTPException(status_code=400, detail="No hay muestras disponibles en bodega")

    existing = await db.sample_requests.find_one(
        {
            "customer_id": sample.customer_id,
            "product_id": sample.product_id,
            "status": {"$in": ["requested", "delivered", "return_requested"]},
        },
        {"_id": 0},
    )
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe una muestra activa para este cliente y producto")

    sample_doc = SampleRequest(
        customer_id=sample.customer_id,
        customer_name=customer.get("name", "N/A"),
        customer_phone=customer.get("phone"),
        product_id=sample.product_id,
        product_name=product.get("name", "Producto"),
        product_sku=product.get("sku"),
        product_image=(product.get("images") or [None])[0],
        warehouse_id=sample.warehouse_id,
        warehouse_name=warehouse.get("name") if warehouse else sample.warehouse_id,
        quantity=quantity,
        requested_by=user.user_id,
        requested_by_name=user.name,
        notes=sample.notes,
    ).model_dump()

    sample_doc["created_at"] = sample_doc["created_at"].isoformat()

    dispatch_id = f"DSP-{uuid.uuid4().hex[:8].upper()}"
    reference_number = f"MUESTRA-{sample_doc['sample_id'][-6:].upper()}"
    dispatch_doc = {
        "dispatch_id": dispatch_id,
        "sale_id": None,
        "invoice_number": reference_number,
        "reference_number": reference_number,
        "customer_name": customer.get("name", "N/A"),
        "customer_phone": customer.get("phone"),
        "requested_by": user.user_id,
        "requested_by_name": user.name,
        "warehouse_id": sample.warehouse_id,
        "branch_id": warehouse.get("branch_id") if warehouse else None,
        "warehouse_name": warehouse.get("name") if warehouse else sample.warehouse_id,
        "items": [
            {
                "product_id": sample.product_id,
                "product_name": product.get("name", "Producto"),
                "product_sku": product.get("sku"),
                "product_image": (product.get("images") or [None])[0],
                "quantity": quantity,
                "delivered": False,
                "delivered_by": None,
                "delivered_at": None,
            }
        ],
        "status": "pending",
        "priority": "normal",
        "notes": f"Muestra solicitada por {user.name}",
        "dispatch_type": "sample_out",
        "sample_id": sample_doc["sample_id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "started_at": None,
        "completed_at": None,
        "total_time_minutes": None,
        "dispatchers": [],
    }

    await db.sample_requests.insert_one(sample_doc)
    await db.dispatch_orders.insert_one(dispatch_doc)

    await db.sample_requests.update_one(
        {"sample_id": sample_doc["sample_id"]},
        {"$set": {"dispatch_out_id": dispatch_id}},
    )

    sample_doc["dispatch_out_id"] = dispatch_id
    sample_response = {**sample_doc}
    sample_response.pop("_id", None)
    dispatch_response = {**dispatch_doc}
    dispatch_response.pop("_id", None)

    return {**sample_response, "dispatch": dispatch_response}


@api_router.get("/samples")
async def get_samples(
    request: Request,
    status: Optional[str] = None,
    customer_id: Optional[str] = None,
    product_id: Optional[str] = None,
):
    user = await require_auth(request)
    query: dict[str, Any] = {}
    if status:
        query["status"] = status
    if customer_id:
        query["customer_id"] = customer_id
    if product_id:
        query["product_id"] = product_id
    if user.role == "ventas":
        query["requested_by"] = user.user_id

    samples = await db.sample_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return samples


@api_router.post("/samples/{sample_id}/return")
async def return_sample(sample_id: str, request: Request):
    user = await require_roles(request, ["gerencia", "supervisor", "ventas"])

    sample = await db.sample_requests.find_one({"sample_id": sample_id})
    if not sample:
        raise HTTPException(status_code=404, detail="Sample not found")

    if sample.get("status") != "delivered":
        raise HTTPException(status_code=400, detail="Solo se pueden devolver muestras entregadas")

    dispatch_id = f"DSP-{uuid.uuid4().hex[:8].upper()}"
    reference_number = f"DEV-{sample_id[-6:].upper()}"
    dispatch_doc = {
        "dispatch_id": dispatch_id,
        "sale_id": None,
        "invoice_number": reference_number,
        "reference_number": reference_number,
        "customer_name": sample.get("customer_name", "N/A"),
        "customer_phone": sample.get("customer_phone"),
        "requested_by": user.user_id,
        "requested_by_name": user.name,
        "warehouse_id": sample.get("warehouse_id"),
        "branch_id": sample.get("branch_id"),
        "warehouse_name": sample.get("warehouse_name") or sample.get("warehouse_id"),
        "items": [
            {
                "product_id": sample.get("product_id"),
                "product_name": sample.get("product_name", "Producto"),
                "product_sku": sample.get("product_sku"),
                "product_image": sample.get("product_image"),
                "quantity": sample.get("quantity", 1),
                "delivered": False,
                "delivered_by": None,
                "delivered_at": None,
            }
        ],
        "status": "pending",
        "priority": "normal",
        "notes": f"Devolución de muestra solicitada por {user.name}",
        "dispatch_type": "sample_return",
        "sample_id": sample_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "started_at": None,
        "completed_at": None,
        "total_time_minutes": None,
        "dispatchers": [],
    }

    await db.dispatch_orders.insert_one(dispatch_doc)
    await db.sample_requests.update_one(
        {"sample_id": sample_id},
        {
            "$set": {
                "status": "return_requested",
                "dispatch_return_id": dispatch_id,
            }
        },
    )

    sample["status"] = "return_requested"
    sample["dispatch_return_id"] = dispatch_id
    sample.pop("_id", None)
    return {**sample, "dispatch": {**dispatch_doc, "_id": None}}


@api_router.get("/dispatch")
async def get_dispatch_orders(
    request: Request,
    warehouse_id: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """Get dispatch orders for warehouse"""
    user = await require_roles(request, ["gerencia", "supervisor", "bodegas"])

    allowed_warehouses: Optional[List[str]] = None
    if user.role != "gerencia":
        if user.warehouse_id:
            allowed_warehouses = [str(user.warehouse_id)]
        elif user.branch_id:
            branch_warehouses = await db.warehouses.find(
                {"branch_id": user.branch_id},
                {"_id": 0, "warehouse_id": 1},
            ).to_list(200)
            allowed_warehouses = [str(item.get("warehouse_id")) for item in branch_warehouses if item.get("warehouse_id")]
        else:
            allowed_warehouses = []

    query = {}
    if warehouse_id:
        if allowed_warehouses is not None and warehouse_id not in allowed_warehouses:
            return []
        query["warehouse_id"] = warehouse_id
    elif allowed_warehouses is not None:
        if not allowed_warehouses:
            return []
        query["warehouse_id"] = {"$in": allowed_warehouses}
    if status:
        query["status"] = status
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        if "created_at" in query:
            query["created_at"]["$lte"] = date_to
        else:
            query["created_at"] = {"$lte": date_to}

    dispatches: List[Dict[str, Any]] = (
        await db.dispatch_orders.find(query).sort("created_at", -1).to_list(500)
    )

    async def _enrich_dispatch(dispatch: Dict[str, Any]) -> Dict[str, Any]:
        items = dispatch.get("items", []) or []
        for item in items:
            if not item.get("product_sku") or not item.get("product_image"):
                product = await db.products.find_one(
                    {"product_id": item.get("product_id")}, {"_id": 0}
                )
                if product:
                    item.setdefault("product_sku", product.get("sku"))
                    item.setdefault("product_image", (product.get("images") or [None])[0])

        if not dispatch.get("requested_by_name"):
            if dispatch.get("dispatch_type") in ["sample_out", "sample_return"] and dispatch.get("sample_id"):
                sample = await db.sample_requests.find_one(
                    {"sample_id": dispatch.get("sample_id")}, {"_id": 0}
                )
                if sample:
                    dispatch["requested_by"] = sample.get("requested_by")
                    dispatch["requested_by_name"] = sample.get("requested_by_name")
            elif dispatch.get("sale_id"):
                sale = await db.sales.find_one(
                    {"sale_id": dispatch.get("sale_id")}, {"_id": 0}
                )
                if sale:
                    dispatch["requested_by"] = sale.get("salesperson_id")
                    dispatch["requested_by_name"] = sale.get("salesperson_name")
        if not dispatch.get("requested_by_name") and dispatch.get("requested_by"):
            requester = await db.users.find_one(
                {"user_id": dispatch.get("requested_by")}, {"_id": 0, "name": 1}
            )
            if requester:
                dispatch["requested_by_name"] = requester.get("name")

        dispatch["items"] = items
        return dispatch

    enriched = [await _enrich_dispatch(d) for d in dispatches]
    return [{**d, "_id": None} for d in enriched]


@api_router.get("/dispatch/{dispatch_id}")
async def get_dispatch_order(request: Request, dispatch_id: str):
    """Get single dispatch order"""
    user = await require_roles(request, ["gerencia", "supervisor", "bodegas"])

    dispatch = await db.dispatch_orders.find_one({"dispatch_id": dispatch_id})
    if not dispatch:
        raise HTTPException(status_code=404, detail="Dispatch order not found")

    dispatch_warehouse = str(dispatch.get("warehouse_id") or "")
    if user.role != "gerencia":
        if user.warehouse_id and dispatch_warehouse != str(user.warehouse_id):
            raise HTTPException(status_code=403, detail="No autorizado para ver despachos de otra bodega")
        if not user.warehouse_id and user.branch_id:
            warehouse_doc = await db.warehouses.find_one({"warehouse_id": dispatch_warehouse}, {"_id": 0, "branch_id": 1})
            warehouse_branch = str((warehouse_doc or {}).get("branch_id") or "")
            if warehouse_branch and warehouse_branch != str(user.branch_id):
                raise HTTPException(status_code=403, detail="No autorizado para ver despachos de otra sucursal")

    dispatch = cast(Dict[str, Any], dispatch)

    items: List[Dict[str, Any]] = dispatch.get("items", []) or []
    for item in items:
        if not item.get("product_sku") or not item.get("product_image"):
            product = await db.products.find_one(
                {"product_id": item.get("product_id")}, {"_id": 0}
            )
            if product:
                item.setdefault("product_sku", product.get("sku"))
                item.setdefault("product_image", (product.get("images") or [None])[0])

    if not dispatch.get("requested_by_name"):
        if dispatch.get("dispatch_type") in ["sample_out", "sample_return"] and dispatch.get("sample_id"):
            sample = await db.sample_requests.find_one(
                {"sample_id": dispatch.get("sample_id")}, {"_id": 0}
            )
            if sample:
                dispatch["requested_by"] = sample.get("requested_by")
                dispatch["requested_by_name"] = sample.get("requested_by_name")
        elif dispatch.get("sale_id"):
            sale = await db.sales.find_one({"sale_id": dispatch.get("sale_id")}, {"_id": 0})
            if sale:
                dispatch["requested_by"] = sale.get("salesperson_id")
                dispatch["requested_by_name"] = sale.get("salesperson_name")
    if not dispatch.get("requested_by_name") and dispatch.get("requested_by"):
        requester = await db.users.find_one(
            {"user_id": dispatch.get("requested_by")}, {"_id": 0, "name": 1}
        )
        if requester:
            dispatch["requested_by_name"] = requester.get("name")

    dispatch["items"] = items
    return {**dispatch, "_id": None}


@api_router.post("/dispatch/from-sale/{sale_id}")
async def create_dispatch_from_sale(request: Request, sale_id: str):
    """Create dispatch order from a sale"""
    await require_roles(request, ["gerencia", "supervisor", "ventas", "bodegas"])

    sale = await db.sales.find_one({"sale_id": sale_id})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    # Check if dispatch already exists
    existing = await db.dispatch_orders.find_one({"sale_id": sale_id})
    if existing:
        raise HTTPException(
            status_code=400, detail="Dispatch order already exists for this sale"
        )

    # Get customer info
    customer = await db.customers.find_one({"customer_id": sale.get("customer_id")})

    # Get warehouse info
    warehouse_id = sale.get("items", [{}])[0].get("warehouse_id", "wh_main")
    warehouse = await db.warehouses.find_one({"warehouse_id": warehouse_id})

    # Build dispatch items from sale items (only products, not services)
    dispatch_items = []
    sale_items = cast(List[Dict[str, Any]], sale.get("items", []))
    for item in sale_items:
        product = await db.products.find_one({"product_id": item.get("product_id")})
        if product and product.get("product_type") != "service":
            dispatch_items.append(
                {
                    "product_id": item.get("product_id"),
                    "product_name": product.get(
                        "name", item.get("product_name", "Producto")
                    ),
                    "product_sku": product.get("sku"),
                    "product_image": (product.get("images") or [None])[0],
                    "quantity": item.get("quantity", 1),
                    "delivered": False,
                    "delivered_by": None,
                    "delivered_at": None,
                }
            )

    if not dispatch_items:
        raise HTTPException(
            status_code=400, detail="No dispatchable items in this sale"
        )

    dispatch_doc = {
        "dispatch_id": f"DSP-{uuid.uuid4().hex[:8].upper()}",
        "sale_id": sale_id,
        "invoice_number": sale.get("invoice_number", "N/A"),
        "reference_number": sale.get("invoice_number", "N/A"),
        "customer_name": (
            customer.get("name") if customer else sale.get("customer_name", "N/A")
        ),
        "customer_phone": customer.get("phone") if customer else None,
        "requested_by": sale.get("salesperson_id"),
        "requested_by_name": sale.get("salesperson_name"),
        "warehouse_id": warehouse_id,
        "branch_id": warehouse.get("branch_id") if warehouse else sale.get("branch_id"),
        "warehouse_name": warehouse.get("name") if warehouse else warehouse_id,
        "items": dispatch_items,
        "status": "pending",
        "priority": "high" if sale.get("delivery_required") else "normal",
        "delivery_required": bool(sale.get("delivery_required", False)),
        "notes": f"Venta: {sale.get('invoice_number')}",
        "dispatch_type": "sale",
        "sample_id": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "started_at": None,
        "completed_at": None,
        "total_time_minutes": None,
        "dispatchers": [],
    }

    await db.dispatch_orders.insert_one(dispatch_doc)

    return {**dispatch_doc, "_id": None}


@api_router.put("/dispatch/{dispatch_id}/start")
async def start_dispatch(request: Request, dispatch_id: str):
    """Start working on a dispatch order"""
    user = await require_roles(request, ["gerencia", "supervisor", "bodegas"])

    dispatch = await db.dispatch_orders.find_one({"dispatch_id": dispatch_id})
    if not dispatch:
        raise HTTPException(status_code=404, detail="Dispatch order not found")

    dispatch_warehouse = str(dispatch.get("warehouse_id") or "")
    if user.role == "bodegas" and user.warehouse_id and dispatch_warehouse != str(user.warehouse_id):
        raise HTTPException(status_code=403, detail="No autorizado para iniciar despachos de otra bodega")

    current_status = _normalize_dispatch_status(dispatch.get("status"))
    if current_status == "completed":
        raise HTTPException(status_code=400, detail="Dispatch already completed")
    if not _can_transition_dispatch_status(current_status, "in_progress"):
        raise HTTPException(
            status_code=400, detail="Dispatch already started or completed"
        )

    await db.dispatch_orders.update_one(
        {"dispatch_id": dispatch_id},
        {
            "$set": {
                "status": "in_progress",
                "started_at": datetime.now(timezone.utc).isoformat(),
            }
        },
    )

    if dispatch.get("dispatch_type") == "sale" and dispatch.get("sale_id"):
        await db.sales.update_one(
            {"sale_id": dispatch.get("sale_id")},
            {
                "$set": {
                    "warehouse_dispatch_status": "in_progress",
                    "workflow_state": "dispatch_in_progress",
                }
            },
        )

    return {"message": "Dispatch started", "dispatch_id": dispatch_id}


@api_router.put("/dispatch/{dispatch_id}/deliver-item")
async def deliver_dispatch_item(
    request: Request, dispatch_id: str, delivery: DispatchItemDelivery
):
    """Mark an item as delivered"""
    actor = await require_roles(request, ["gerencia", "supervisor", "bodegas"])

    dispatch = await db.dispatch_orders.find_one({"dispatch_id": dispatch_id})
    if not dispatch:
        raise HTTPException(status_code=404, detail="Dispatch order not found")

    dispatch_status = _normalize_dispatch_status(dispatch.get("status"))
    if dispatch_status == "completed":
        return {
            "message": "Dispatch already completed",
            "all_completed": True,
            "dispatch_id": dispatch_id,
        }
    if dispatch_status != "in_progress":
        raise HTTPException(
            status_code=400,
            detail="Dispatch must be started before delivering items",
        )

    dispatch_warehouse = str(dispatch.get("warehouse_id") or "")
    if actor.role == "bodegas" and actor.warehouse_id and dispatch_warehouse != str(actor.warehouse_id):
        raise HTTPException(status_code=403, detail="No autorizado para despachar productos de otra bodega")

    # Get dispatcher name
    dispatcher = await db.users.find_one({"user_id": delivery.dispatcher_id})
    dispatcher_name = dispatcher.get("name") if dispatcher else delivery.dispatcher_id

    # Update the specific item
    items = cast(List[Dict[str, Any]], dispatch.get("items", []))
    item_found = False
    already_delivered = False
    for item in items:
        if item["product_id"] == delivery.product_id:
            if item.get("delivered"):
                already_delivered = True
                item_found = True
                break
            item["delivered"] = True
            item["delivered_by"] = dispatcher_name
            item["delivered_at"] = datetime.now(timezone.utc).isoformat()
            item_found = True
            break

    if not item_found:
        raise HTTPException(status_code=404, detail="Item not found in dispatch")

    if already_delivered:
        all_delivered = all(item.get("delivered", False) for item in items)
        return {
            "message": "Item already delivered",
            "all_completed": all_delivered,
            "dispatch_id": dispatch_id,
        }

    # Add dispatcher to list if not already there
    dispatchers = cast(List[str], dispatch.get("dispatchers", []))
    if dispatcher_name not in dispatchers:
        dispatchers.append(dispatcher_name)

    # Check if all items are delivered
    all_delivered = all(item.get("delivered", False) for item in items)

    update_data: Dict[str, Any] = {"items": items, "dispatchers": dispatchers}

    if all_delivered:
        started_at = dispatch.get("started_at")
        completed_at = datetime.now(timezone.utc)
        total_minutes = 0
        if started_at:
            start_time = datetime.fromisoformat(started_at.replace("Z", "+00:00"))
            total_minutes = int((completed_at - start_time).total_seconds() / 60)

        update_data["status"] = "completed"
        update_data["completed_at"] = completed_at.isoformat()
        update_data["total_time_minutes"] = total_minutes

        dispatch_type = dispatch.get("dispatch_type")
        if dispatch_type in ["sample_out", "sample_return"]:
            sample_id = dispatch.get("sample_id")
            if sample_id:
                if dispatch_type == "sample_out":
                    warehouse_id = str(dispatch.get("warehouse_id") or "wh_main")
                    for dispatch_item in items:
                        product_id = dispatch_item.get("product_id")
                        if not product_id:
                            continue
                        await db.inventory.update_one(
                            {
                                "product_id": product_id,
                                "warehouse_id": warehouse_id,
                            },
                            {
                                "$inc": {"quantity": -dispatch_item.get("quantity", 1)},
                                "$set": {
                                    "last_updated": datetime.now(timezone.utc).isoformat()
                                },
                            },
                        )
                        await audit_service.log_inventory_movement(
                            product_id=str(product_id),
                            warehouse_id=warehouse_id,
                            quantity_change=-int(dispatch_item.get("quantity", 1)),
                            reason="sample_dispatch_out",
                            actor=actor,
                            branch_id=None,
                            reference_id=dispatch_id,
                            metadata={"sample_id": sample_id},
                        )
                    await db.sample_requests.update_one(
                        {"sample_id": sample_id},
                        {
                            "$set": {
                                "status": "delivered",
                                "delivered_at": completed_at.isoformat(),
                            }
                        },
                    )

                if dispatch_type == "sample_return":
                    warehouse_id = str(dispatch.get("warehouse_id") or "wh_main")
                    for dispatch_item in items:
                        product_id = dispatch_item.get("product_id")
                        if not product_id:
                            continue
                        await db.inventory.update_one(
                            {
                                "product_id": product_id,
                                "warehouse_id": warehouse_id,
                            },
                            {
                                "$inc": {"quantity": dispatch_item.get("quantity", 1)},
                                "$set": {
                                    "last_updated": datetime.now(timezone.utc).isoformat()
                                },
                            },
                        )
                        await audit_service.log_inventory_movement(
                            product_id=str(product_id),
                            warehouse_id=warehouse_id,
                            quantity_change=int(dispatch_item.get("quantity", 1)),
                            reason="sample_dispatch_return",
                            actor=actor,
                            branch_id=None,
                            reference_id=dispatch_id,
                            metadata={"sample_id": sample_id},
                        )
                    await db.sample_requests.update_one(
                        {"sample_id": sample_id},
                        {
                            "$set": {
                                "status": "returned",
                                "returned_at": completed_at.isoformat(),
                            }
                        },
                    )

        if dispatch_type == "sale":
            sale_id = dispatch.get("sale_id")
            sale_invoice = dispatch.get("invoice_number")
            salesperson_id = dispatch.get("requested_by")
            completed_by_name = actor.name or dispatcher_name or "Bodega"
            sale_record = None
            if sale_id:
                sale_record = await db.sales.find_one({"sale_id": sale_id}, {"_id": 0, "delivery_required": 1})
            requires_delivery = bool(
                dispatch.get("delivery_required")
                or (sale_record or {}).get("delivery_required")
            )

            if sale_id:
                sale_update: Dict[str, Any] = {
                    "warehouse_dispatch_status": "completed",
                    "warehouse_dispatch_completed_at": completed_at.isoformat(),
                }
                if requires_delivery:
                    sale_update["delivery_status"] = "ready_for_delivery"
                    sale_update["workflow_state"] = "ready_for_delivery"
                else:
                    sale_update["workflow_state"] = "fulfilled"
                await db.sales.update_one({"sale_id": sale_id}, {"$set": sale_update})

            if salesperson_id:
                await create_notification_entry(
                    message=(
                        f"Despacho completado para factura {sale_invoice}. "
                        f"Producto listo para entrega al cliente."
                    ),
                    recipient_id=salesperson_id,
                    metadata={
                        "type": "dispatch_completed",
                        "sale_id": sale_id,
                        "invoice_number": sale_invoice,
                        "dispatch_id": dispatch_id,
                        "completed_by": completed_by_name,
                    },
                    dedupe_key=f"dispatch_completed:{dispatch_id}",
                )

    await db.dispatch_orders.update_one(
        {"dispatch_id": dispatch_id}, {"$set": update_data}
    )

    return {
        "message": "Item delivered",
        "all_completed": all_delivered,
        "dispatch_id": dispatch_id,
    }


@api_router.get("/dispatch/stats/performance")
async def get_dispatch_performance(request: Request, days: int = 30):
    """Get dispatch performance statistics"""
    await require_roles(request, ["gerencia", "supervisor"])

    date_from = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()

    completed = await db.dispatch_orders.find(
        {"status": "completed", "completed_at": {"$gte": date_from}}
    ).to_list(1000)

    total_completed = len(completed)
    total_time = sum(d.get("total_time_minutes", 0) for d in completed)
    avg_time = total_time / total_completed if total_completed > 0 else 0

    # Group by dispatcher
    dispatcher_stats = {}
    for d in completed:
        for dispatcher in d.get("dispatchers", []):
            if dispatcher not in dispatcher_stats:
                dispatcher_stats[dispatcher] = {"count": 0, "total_time": 0}
            dispatcher_stats[dispatcher]["count"] += 1
            dispatcher_stats[dispatcher]["total_time"] += d.get(
                "total_time_minutes", 0
            ) / len(d.get("dispatchers", [1]))

    return {
        "period_days": days,
        "total_completed": total_completed,
        "average_time_minutes": round(avg_time, 1),
        "dispatchers": [
            {
                "name": name,
                "completed": stats["count"],
                "avg_time": (
                    round(stats["total_time"] / stats["count"], 1)
                    if stats["count"] > 0
                    else 0
                ),
            }
            for name, stats in dispatcher_stats.items()
        ],
    }


@api_router.get("/dispatch/dispatchers/list")
async def get_dispatchers(request: Request):
    """Get list of available dispatchers"""
    await require_auth(request)

    async def _ensure_default_dispatchers() -> None:
        dispatcher_pin_base = os.environ.get("DEFAULT_DISPATCHER_PIN_BASE", "7701")
        base_pin = int(dispatcher_pin_base) if dispatcher_pin_base.isdigit() else 7701
        dispatcher_names = [
            "Despachador 1",
            "Despachador 2",
            "Despachador 3",
            "Despachador 4",
            "Despachador 5",
        ]
        for idx, name in enumerate(dispatcher_names, start=0):
            email = f"dispatcher_{idx + 1}@pin.local"
            pin_value = str(base_pin + idx).zfill(PIN_LENGTH)[-PIN_LENGTH:]
            existing_user = await db.users.find_one(
                {"$or": [{"email": email}, {"name": name}]}, {"_id": 0}
            )
            if existing_user:
                updates: Dict[str, Any] = {
                    "role": "bodegas",
                    "is_active": True,
                    "is_pin_user": True,
                }
                if not existing_user.get("pin_hash"):
                    updates["pin_hash"] = hash_pin(pin_value)
                    updates["pin_index"] = compute_pin_index(pin_value)
                    updates["pin_last_set_at"] = datetime.now(timezone.utc).isoformat()
                else:
                    try:
                        if not verify_pin_hash(pin_value, existing_user.get("pin_hash")):
                            updates["pin_hash"] = hash_pin(pin_value)
                            updates["pin_index"] = compute_pin_index(pin_value)
                            updates["pin_last_set_at"] = datetime.now(timezone.utc).isoformat()
                    except Exception:
                        updates["pin_hash"] = hash_pin(pin_value)
                        updates["pin_index"] = compute_pin_index(pin_value)
                        updates["pin_last_set_at"] = datetime.now(timezone.utc).isoformat()

                await db.users.update_one(
                    {"user_id": existing_user.get("user_id")}, {"$set": updates}
                )
                continue

            dispatcher_id = f"user_{uuid.uuid4().hex[:12]}"
            await db.users.insert_one(
                {
                    "user_id": dispatcher_id,
                    "email": email,
                    "name": name,
                    "role": "bodegas",
                    "branch_id": None,
                    "warehouse_id": None,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "is_active": True,
                    "is_pin_user": True,
                    "pin_hash": hash_pin(pin_value),
                    "pin_index": compute_pin_index(pin_value),
                    "pin_last_set_at": datetime.now(timezone.utc).isoformat(),
                    "failed_pin_attempts": 0,
                }
            )

    await _ensure_default_dispatchers()

    # Get users with bodegas role
    dispatchers = await db.users.find(
        {
            "$and": [
                {
                    "$or": [
                        {"role": "bodegas"},
                        {"role": "bodega"},
                        {"role": {"$regex": "^bodegas?$", "$options": "i"}},
                    ]
                },
                {"$or": [{"is_active": True}, {"is_active": {"$exists": False}}]},
            ]
        },
        {"_id": 0, "user_id": 1, "name": 1, "role": 1},
    ).to_list(200)

    return [
        {"id": d.get("user_id"), "name": d.get("name"), "role": d.get("role")}
        for d in dispatchers
    ]


# ============ TINT ORDERS (POLARIZADOS) ============


class TintWindowSpec(BaseModel):
    window_type: str  # frontal, trasero, lateral_izq, lateral_der, franja_superior, franja_inferior
    material: Optional[str] = None
    shade_percentage: Optional[int] = None  # 5%, 15%, 20%, 35%, 50%, 70%
    width_cm: Optional[float] = None
    height_cm: Optional[float] = None
    status: str = "pending"  # pending, in_progress, completed
    technician_id: Optional[str] = None
    completed_at: Optional[str] = None


class TintOrder(BaseModel):
    tint_order_id: str
    sale_id: Optional[str] = None
    work_order_id: Optional[str] = None
    customer_id: str
    customer_name: str
    vehicle_id: str
    vehicle_info: dict  # plate, brand, model, year, color
    windows: List[TintWindowSpec]
    total_material_used: Optional[float] = None
    status: str = "pending"  # pending, in_progress, quality_check, completed
    priority: str = "normal"
    notes: Optional[str] = None
    created_at: str
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    quality_rating: Optional[int] = None


class TintOrderCreate(BaseModel):
    sale_id: Optional[str] = None
    customer_id: str
    vehicle_id: str
    windows: List[dict]
    notes: Optional[str] = None
    priority: str = "normal"


class TintWindowUpdate(BaseModel):
    window_type: str
    material: Optional[str] = None
    shade_percentage: Optional[int] = None
    width_cm: Optional[float] = None
    height_cm: Optional[float] = None
    status: Optional[str] = None
    technician_id: Optional[str] = None


@api_router.get("/tint-orders")
async def get_tint_orders(
    request: Request,
    status: Optional[str] = None,
    technician_id: Optional[str] = None,
    date_from: Optional[str] = None,
):
    """Get tint orders"""
    await require_auth(request)

    query = {}
    if status:
        query["status"] = status
    if technician_id:
        query["windows.technician_id"] = technician_id
    if date_from:
        query["created_at"] = {"$gte": date_from}

    orders = await db.tint_orders.find(query).sort("created_at", -1).to_list(500)
    return [{**o, "_id": None} for o in orders]


@api_router.get("/tint-orders/{tint_order_id}")
async def get_tint_order(request: Request, tint_order_id: str):
    """Get single tint order"""
    await require_auth(request)

    order = await db.tint_orders.find_one({"tint_order_id": tint_order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Tint order not found")

    return {**order, "_id": None}


@api_router.post("/tint-orders")
async def create_tint_order(request: Request, order: TintOrderCreate):
    """Create a new tint order"""
    user = await require_roles(request, ["gerencia", "supervisor", "ventas"])

    await ensure_branch_service_enabled(
        user.branch_id,
        "tint_enabled",
        "Esta sucursal no ofrece servicio de polarizados. Solo la sucursal central puede crear órdenes de polarizado.",
    )

    # Get customer info
    customer = await db.customers.find_one({"customer_id": order.customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    # Get vehicle info
    vehicle = await db.vehicles.find_one({"vehicle_id": order.vehicle_id})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")

    # Build windows list
    windows = []
    for w in order.windows:
        windows.append(
            {
                "window_type": w.get("window_type"),
                "material": w.get("material"),
                "shade_percentage": w.get("shade_percentage"),
                "width_cm": w.get("width_cm"),
                "height_cm": w.get("height_cm"),
                "status": "pending",
                "technician_id": None,
                "completed_at": None,
            }
        )

    tint_doc = {
        "tint_order_id": f"TINT-{uuid.uuid4().hex[:8].upper()}",
        "sale_id": order.sale_id,
        "work_order_id": None,
        "customer_id": order.customer_id,
        "customer_name": customer.get("name"),
        "vehicle_id": order.vehicle_id,
        "vehicle_info": {
            "plate": vehicle.get("plate"),
            "brand": vehicle.get("brand"),
            "model": vehicle.get("model"),
            "year": vehicle.get("year"),
            "color": vehicle.get("color"),
        },
        "windows": windows,
        "total_material_used": None,
        "status": "pending",
        "priority": order.priority,
        "notes": order.notes,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "started_at": None,
        "completed_at": None,
        "quality_rating": None,
    }

    await db.tint_orders.insert_one(tint_doc)

    return {**tint_doc, "_id": None}


@api_router.put("/tint-orders/{tint_order_id}/start")
async def start_tint_order(request: Request, tint_order_id: str):
    """Start working on a tint order"""
    await require_roles(request, ["gerencia", "supervisor", "instalaciones"])

    order = await db.tint_orders.find_one({"tint_order_id": tint_order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Tint order not found")

    await db.tint_orders.update_one(
        {"tint_order_id": tint_order_id},
        {
            "$set": {
                "status": "in_progress",
                "started_at": datetime.now(timezone.utc).isoformat(),
            }
        },
    )

    return {"message": "Tint order started", "tint_order_id": tint_order_id}


@api_router.put("/tint-orders/{tint_order_id}/window")
async def update_tint_window(
    request: Request, tint_order_id: str, update: TintWindowUpdate
):
    """Update a specific window in tint order"""
    await require_roles(request, ["gerencia", "supervisor", "instalaciones"])

    order = await db.tint_orders.find_one({"tint_order_id": tint_order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Tint order not found")

    windows = cast(List[Dict[str, Any]], order.get("windows", []))
    window_found = False

    for window in windows:
        if window["window_type"] == update.window_type:
            if update.material is not None:
                window["material"] = update.material
            if update.shade_percentage is not None:
                window["shade_percentage"] = update.shade_percentage
            if update.width_cm is not None:
                window["width_cm"] = update.width_cm
            if update.height_cm is not None:
                window["height_cm"] = update.height_cm
            if update.status is not None:
                window["status"] = update.status
                if update.status == "completed":
                    window["completed_at"] = datetime.now(timezone.utc).isoformat()
            if update.technician_id is not None:
                window["technician_id"] = update.technician_id
            window_found = True
            break

    if not window_found:
        raise HTTPException(status_code=404, detail="Window type not found")

    # Check if all windows are completed
    all_completed = all(w.get("status") == "completed" for w in windows)

    update_data: Dict[str, Any] = {"windows": windows}
    if all_completed:
        update_data["status"] = "quality_check"

    await db.tint_orders.update_one(
        {"tint_order_id": tint_order_id}, {"$set": update_data}
    )

    return {"message": "Window updated", "all_completed": all_completed}


@api_router.put("/tint-orders/{tint_order_id}/complete")
async def complete_tint_order(
    request: Request,
    tint_order_id: str,
    quality_rating: int = 5,
    total_material: float = 0,
):
    """Complete a tint order after quality check"""
    await require_roles(request, ["gerencia", "supervisor"])

    order = await db.tint_orders.find_one({"tint_order_id": tint_order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Tint order not found")

    await db.tint_orders.update_one(
        {"tint_order_id": tint_order_id},
        {
            "$set": {
                "status": "completed",
                "completed_at": datetime.now(timezone.utc).isoformat(),
                "quality_rating": quality_rating,
                "total_material_used": total_material,
            }
        },
    )

    return {"message": "Tint order completed", "tint_order_id": tint_order_id}


@api_router.get("/tint-orders/materials/list")
async def get_tint_materials(request: Request):
    """Get available tint materials"""
    await require_auth(request)

    return {
        "materials": [
            {
                "id": "ceramic",
                "name": "Polarizado Cerámico",
                "description": "Alta calidad, rechazo de calor superior",
            },
            {
                "id": "carbon",
                "name": "Carbón Premium",
                "description": "Durabilidad y estética",
            },
            {
                "id": "metallic",
                "name": "Polarizado Metálico",
                "description": "Buena protección UV",
            },
            {
                "id": "dyed",
                "name": "Polarizado Teñido",
                "description": "Económico, buena apariencia",
            },
            {
                "id": "hybrid",
                "name": "Polarizado Híbrido",
                "description": "Combinación de tecnologías",
            },
        ],
        "shade_options": [5, 15, 20, 35, 50, 70],
        "window_types": [
            {"id": "frontal", "name": "Parabrisas Frontal"},
            {"id": "trasero", "name": "Vidrio Trasero"},
            {"id": "lateral_conductor", "name": "Lateral Conductor"},
            {"id": "lateral_copiloto", "name": "Lateral Copiloto"},
            {"id": "lateral_trasero_izq", "name": "Lateral Trasero Izq"},
            {"id": "lateral_trasero_der", "name": "Lateral Trasero Der"},
            {"id": "franja_superior", "name": "Franja Superior"},
            {"id": "franja_inferior", "name": "Franja Inferior"},
            {"id": "quemacocos", "name": "Quemacocos/Sunroof"},
        ],
    }


# ============ INVENTORY ============
from backend.routes.inventory import get_inventory_router
from backend.routes.human_resources import get_human_resources_router

inventory_router = get_inventory_router(
    db,
    audit_service,
    require_auth,
    require_roles,
    InventoryUpdate,
)
api_router.include_router(inventory_router)

human_resources_router = get_human_resources_router(
    db,
    audit_service,
    require_auth,
    require_roles,
    verify_pin_hash,
)
api_router.include_router(human_resources_router)

app.include_router(api_router)

cors_origins_env = os.environ.get("CORS_ORIGINS")
if cors_origins_env:
    cors_origins = [o.strip() for o in cors_origins_env.split(",") if o.strip()]
else:
    # ADVERTENCIA: modificar estos orígenes afecta el login (PIN) desde el frontend.
    # Si cambias puertos o dominios, actualiza el frontend y el docker-compose.
    cors_origins = [
        "http://127.0.0.1:3000",
        "http://localhost:3000",
        "http://127.0.0.1:3001",
        "http://localhost:3001",
        "http://localhost:8001",
        "http://127.0.0.1:8001"
    ]

cors_origins = list(dict.fromkeys(cors_origins))

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=cors_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)


frontend_build_dir = ROOT_DIR.parent / "frontend" / "build"
if frontend_build_dir.exists():
    static_dir = frontend_build_dir / "static"
    if static_dir.exists():
        app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        if full_path.startswith("api/"):
            raise HTTPException(status_code=404, detail="Not Found")

        if not full_path:
            return FileResponse(frontend_build_dir / "index.html")

        candidate = frontend_build_dir / full_path
        if candidate.exists() and candidate.is_file():
            return FileResponse(candidate)

        return FileResponse(frontend_build_dir / "index.html")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
